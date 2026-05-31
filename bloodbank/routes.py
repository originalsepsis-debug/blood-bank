from flask import Blueprint, render_template, request, redirect, session, jsonify, current_app, send_from_directory, Response, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from .db import db, one, all_rows, exec_sql, ph, is_pg, scalar, insert_and_get_id, reset_postgres_sequences
from .auth import login_user, current_user, require_login, require_role, require_csrf
import json
import time
import os
import urllib.parse
import urllib.request
import html
import hashlib
import sqlite3
import tempfile
from io import BytesIO
from datetime import datetime, date
import re

bp = Blueprint("main", __name__)

def login_client_key():
    return (request.headers.get("X-Forwarded-For") or request.remote_addr or "unknown").split(",")[0].strip()

def login_blocked(key):
    """DB-backed login throttling, shared across gunicorn workers."""
    now = time.time()
    try:
        with db() as conn:
            p = ph()
            rec = one(conn, f"SELECT fail_count, blocked_until FROM login_attempts WHERE client_key={p}", (key,))
            if rec and float(rec.get("blocked_until") or 0) > now:
                return int(float(rec.get("blocked_until") or 0) - now)
            if rec and float(rec.get("blocked_until") or 0) <= now and int(rec.get("fail_count") or 0) == 0:
                exec_sql(conn, f"DELETE FROM login_attempts WHERE client_key={p}", (key,))
    except Exception:
        return 0
    return 0

def login_failure(key):
    now = time.time()
    try:
        with db() as conn:
            p = ph()
            rec = one(conn, f"SELECT fail_count, blocked_until FROM login_attempts WHERE client_key={p}", (key,))
            count = 1
            blocked_until = 0
            if rec and float(rec.get("blocked_until") or 0) <= now:
                count = int(rec.get("fail_count") or 0) + 1
            if count >= 5:
                blocked_until = now + 15 * 60
                count = 0
            if rec:
                exec_sql(conn, f"UPDATE login_attempts SET fail_count={p}, blocked_until={p}, updated_at=CURRENT_TIMESTAMP WHERE client_key={p}", (count, blocked_until, key))
            else:
                exec_sql(conn, f"INSERT INTO login_attempts(client_key, fail_count, blocked_until) VALUES ({p},{p},{p})", (key, count, blocked_until))
    except Exception:
        pass

def login_success(key):
    try:
        with db() as conn:
            exec_sql(conn, f"DELETE FROM login_attempts WHERE client_key={ph()}", (key,))
    except Exception:
        pass

def record_login_event(username, success, reason="", user=None):
    """Append-only login history for admins/transfusiologists."""
    try:
        ip = login_client_key()
        ua = (request.headers.get("User-Agent") or "")[:500]
        p = ph()
        with db() as conn:
            exec_sql(conn, f"""INSERT INTO login_events(username,user_id,role,full_name,ip,user_agent,success,reason)
                VALUES ({p},{p},{p},{p},{p},{p},{p},{p})""",
                ((username or "").strip()[:120],
                 user.get("id") if user else None,
                 user.get("role","") if user else "",
                 user.get("full_name","") if user else "",
                 ip, ua, 1 if success else 0, str(reason or "")[:300]))
    except Exception:
        pass


def ok(**kw):
    return jsonify(ok=True, **kw)


def err(msg, code=400):
    return jsonify(ok=False, error=msg), code


def getj():
    return request.get_json(silent=True) or {}


def today_iso():
    return date.today().isoformat()



def clean_text(v, max_len=500):
    return str(v or "").strip()[:max_len]

ABO_ALIASES = {
    "0": "O(I)", "O": "O(I)", "О": "O(I)", "O(I)": "O(I)", "0(I)": "O(I)", "I": "O(I)",
    "A": "A(II)", "А": "A(II)", "A(II)": "A(II)", "II": "A(II)",
    "B": "B(III)", "В": "B(III)", "B(III)": "B(III)", "III": "B(III)",
    "AB": "AB(IV)", "АВ": "AB(IV)", "AB(IV)": "AB(IV)", "IV": "AB(IV)",
}
RH_ALIASES = {"+": "Rh+", "RH+": "Rh+", "Rh+": "Rh+", "POS": "Rh+", "позитивний": "Rh+", "-": "Rh-", "RH-": "Rh-", "Rh-": "Rh-", "NEG": "Rh-", "негативний": "Rh-"}

DEFAULT_COMPONENTS = [
    {"name":"Цільна кров", "category":"rbc"},
    {"name":"Еритроцити", "category":"rbc"},
    {"name":"Еритроцити в додатковому розчині", "category":"rbc"},
    {"name":"Еритроцити з видаленим лейкоцитарно-тромбоцитарним шаром", "category":"rbc"},
    {"name":"Еритроцити з видаленим лейкоцитарно-тромбоцитарним шаром у додатковому розчині", "category":"rbc"},
    {"name":"Еритроцити збіднені на лейкоцити", "category":"rbc"},
    {"name":"Еритроцити збіднені на лейкоцити у додатковому розчині", "category":"rbc"},
    {"name":"Еритроцити відмиті", "category":"rbc"},
    {"name":"Плазма свіжозаморожена", "category":"plasma"},
    {"name":"Плазма свіжозаморожена збіднена на кріопреципітат", "category":"plasma"},
    {"name":"Плазма свіжозаморожена оброблена методом патогенредукції", "category":"plasma"},
    {"name":"Тромбоцити аферез", "category":"platelets"},
    {"name":"Тромбоцити аферез оброблені методом патогенредукції", "category":"platelets"},
    {"name":"Тромбоцити відновлені", "category":"platelets"},
    {"name":"Кріопреципітат", "category":"cryo"},
]

REACTION_TYPES = {
    "none": "Реакції немає",
    "fever": "Підвищення температури",
    "chills": "Озноб",
    "urticaria": "Кропив’янка / висип",
    "hypotension": "Гіпотензія",
    "dyspnea": "Задишка",
    "back_pain": "Біль у попереку",
    "hemolysis_suspected": "Підозра на гемоліз",
    "other": "Інше",
}
REACTION_SEVERITY = {"none": "Немає", "mild": "Легка", "moderate": "Середня", "severe": "Тяжка"}

def normalize_reaction_payload(j):
    rt = (j.get("reaction_type") or "none").strip()
    sev = (j.get("reaction_severity") or ("none" if rt == "none" else "mild")).strip()
    note = clean_text(j.get("reaction_note") or j.get("reaction") or "", 1000)
    if rt not in REACTION_TYPES:
        rt = "other"
    if sev not in REACTION_SEVERITY:
        sev = "mild" if rt != "none" else "none"
    if rt == "none":
        sev = "none"
        if not note:
            note = "Реакції немає"
    return rt, sev, note

def reaction_label(rt, sev="", note=""):
    base = REACTION_TYPES.get(rt or "none", rt or "Реакція")
    s = REACTION_SEVERITY.get(sev or "", sev or "")
    parts = [base]
    if s and sev != "none":
        parts.append(f"тяжкість: {s}")
    if note:
        parts.append(str(note))
    return " | ".join(parts)


def normalize_abo(v):
    key = str(v or "").strip().upper().replace(" ", "")
    return ABO_ALIASES.get(key, str(v or "").strip() or "O(I)")

def normalize_rh(v):
    raw = str(v or "").strip()
    key = raw.upper().replace(" ", "")
    return RH_ALIASES.get(key, raw or "Rh+")

def component_category(name):
    n = str(name or "").lower()
    if "плаз" in n:
        return "plasma"
    if "тромбо" in n:
        return "platelets"
    if "кріо" in n or "крио" in n:
        return "cryo"
    if "ерит" in n or "цільна" in n or "кров" in n:
        return "rbc"
    return "other"

def normalize_component(conn, value):
    name = clean_text(value, 180)
    if not name:
        return ""
    p = ph()
    exact = one(conn, f"SELECT name FROM component_catalog WHERE active=1 AND LOWER(name)=LOWER({p}) LIMIT 1", (name,))
    if exact:
        return exact["name"]
    # V7.1.4: strict catalog enforcement. Unknown free-text components would corrupt reports.
    return ""


def require_catalog_component(conn, value):
    normalized = normalize_component(conn, value)
    if not normalized:
        raise ValueError("Компонент має бути вибраний з довідника")
    return normalized

def unit_matches_request(req, unit):
    """Conservative compatibility. Exact component name; RBC/plasma ABO rules; exact Rh for safety."""
    if (unit.get("component_type") or "") != (req.get("component_type") or ""):
        return False, "інший компонент"
    if normalize_rh(unit.get("rh")) != normalize_rh(req.get("rh")):
        return False, "інший Rh"
    donor = normalize_abo(unit.get("abo"))
    recipient = normalize_abo(req.get("abo"))
    cat = component_category(req.get("component_type"))
    if cat == "rbc":
        allowed = {"O(I)": {"O(I)", "A(II)", "B(III)", "AB(IV)"}, "A(II)": {"A(II)", "AB(IV)"}, "B(III)": {"B(III)", "AB(IV)"}, "AB(IV)": {"AB(IV)"}}
        return recipient in allowed.get(donor, {donor}), "ABO несумісність для еритроцитарного компонента"
    if cat == "plasma":
        allowed = {"AB(IV)": {"O(I)", "A(II)", "B(III)", "AB(IV)"}, "A(II)": {"A(II)", "O(I)"}, "B(III)": {"B(III)", "O(I)"}, "O(I)": {"O(I)"}}
        return recipient in allowed.get(donor, {donor}), "ABO несумісність для плазми"
    # Platelets/cryo/unknown stay exact by design until validated locally.
    return donor == recipient, "ABO має бути точним для цього компонента"

def audit_diff(old, new, fields):
    changes = {}
    old = old or {}
    new = new or {}
    for f in fields:
        ov = old.get(f, "")
        nv = new.get(f, "")
        if str(ov) != str(nv):
            changes[f] = {"old": ov, "new": nv}
    return json.dumps(changes, ensure_ascii=False)[:1800]


def count_rows(conn, table, where="1=1", params=()):
    return int(scalar(conn, f"SELECT COUNT(*) FROM {table} WHERE {where}", params) or 0)


def delivered_count_sql():
    return "SELECT COUNT(*) FROM stock_units su WHERE su.request_id=requests.id AND su.status IN ('issued','used')"


def used_count_sql():
    return "SELECT COUNT(*) FROM stock_units su WHERE su.request_id=requests.id AND su.status='used'"



def pdf_font_status():
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
    ]
    found = [{"path": fp, "exists": os.path.exists(fp)} for fp in candidates]
    chosen = next((x["path"] for x in found if x["exists"]), "")
    return {"chosen": chosen, "font_name": "AppFont" if chosen else "Helvetica", "candidates": found}


def system_integrity_report(conn):
    """Lightweight consistency checks. Returns a report, does not modify data."""
    p = ph()
    checks = []

    def add(code, title, severity, rows):
        checks.append({"code": code, "title": title, "severity": severity, "count": len(rows), "items": rows[:50]})

    add("bad_unit_quantity", "Одиниці складу з кількістю менше 1", "critical",
        all_rows(conn, "SELECT id,unit_code,component_type,quantity,status FROM stock_units WHERE quantity < 1 LIMIT 50"))
    add("issued_without_request", "Видані/використані одиниці без прив’язаної вимоги", "critical",
        all_rows(conn, "SELECT id,unit_code,status,request_id FROM stock_units WHERE status IN ('issued','used') AND request_id IS NULL LIMIT 50"))
    add("request_overdelivered", "Вимоги, де видано більше одиниць, ніж замовлено", "critical",
        all_rows(conn, """SELECT r.id,r.patient_name,r.component_type,r.quantity,COUNT(su.id) AS delivered
            FROM requests r LEFT JOIN stock_units su ON su.request_id=r.id AND su.status IN ('issued','used')
            WHERE r.deleted_at IS NULL
            GROUP BY r.id,r.patient_name,r.component_type,r.quantity
            HAVING COUNT(su.id) > r.quantity LIMIT 50"""))
    add("expired_still_in_stock", "Протерміновані одиниці, які ще мають статус на складі", "warning",
        all_rows(conn, f"SELECT id,unit_code,component_type,expiry_date,status FROM stock_units WHERE status='in_stock' AND expiry_date<>'' AND expiry_date < {p} LIMIT 50", (today_iso(),)))
    add("orphan_unit_request", "Одиниці прив’язані до неіснуючої вимоги", "critical",
        all_rows(conn, """SELECT su.id,su.unit_code,su.request_id,su.status FROM stock_units su
            LEFT JOIN requests r ON r.id=su.request_id
            WHERE su.request_id IS NOT NULL AND r.id IS NULL LIMIT 50"""))
    add("orphan_movement_unit", "Рухи з неіснуючою одиницею складу", "warning",
        all_rows(conn, """SELECT m.id,m.unit_id,m.action,m.created_at FROM movements m
            LEFT JOIN stock_units su ON su.id=m.unit_id
            WHERE m.unit_id IS NOT NULL AND su.id IS NULL LIMIT 50"""))
    add("orphan_movement_request", "Рухи з неіснуючою вимогою", "warning",
        all_rows(conn, """SELECT m.id,m.request_id,m.action,m.created_at FROM movements m
            LEFT JOIN requests r ON r.id=m.request_id
            WHERE m.request_id IS NOT NULL AND r.id IS NULL LIMIT 50"""))
    add("orphan_temperature_device", "Температурні записи з неіснуючим пристроєм", "warning",
        all_rows(conn, """SELECT tr.id,tr.device_id,tr.temperature,tr.measured_at FROM temperature_readings tr
            LEFT JOIN temperature_devices td ON td.id=tr.device_id
            WHERE tr.device_id IS NOT NULL AND td.id IS NULL LIMIT 50"""))
    add("requests_without_patient_card", "Вимоги без прив’язаної картки пацієнта", "info",
        all_rows(conn, "SELECT id,patient_name,birth_date,created_at FROM requests WHERE patient_id IS NULL AND deleted_at IS NULL LIMIT 50"))

    totals = {
        "users": count_rows(conn, "users"),
        "active_users": count_rows(conn, "users", "active=1"),
        "patients": count_rows(conn, "patients", "deleted_at IS NULL"),
        "requests": count_rows(conn, "requests", "deleted_at IS NULL"),
        "stock_in": count_rows(conn, "stock_units", "status='in_stock'"),
        "stock_issued": count_rows(conn, "stock_units", "status='issued'"),
        "stock_used": count_rows(conn, "stock_units", "status='used'"),
        "movements": count_rows(conn, "movements", "deleted_at IS NULL"),
        "temperature_devices": count_rows(conn, "temperature_devices", "deleted_at IS NULL"),
        "temperature_readings": count_rows(conn, "temperature_readings", "deleted_at IS NULL"),
    }
    critical = sum(c["count"] for c in checks if c["severity"] == "critical")
    warning = sum(c["count"] for c in checks if c["severity"] == "warning")
    status = "ok" if critical == 0 and warning == 0 else ("critical" if critical else "warning")
    return {"status": status, "critical": critical, "warning": warning, "generated_at": datetime.now().isoformat(timespec="seconds"), "totals": totals, "checks": checks}


def permissions_matrix():
    return {
        "admin": ["all", "users", "requests_all", "approve", "issue", "stock", "reports", "trash", "backup_restore", "migration", "telegram_config", "temperature", "audit", "integrity"],
        "transfusion": ["requests_all", "approve", "reject", "issue", "return", "writeoff", "stock", "reports", "trash", "telegram_personal", "temperature", "audit", "integrity"],
        "doctor": ["create_request", "own_requests", "patients", "confirm_used", "telegram_personal", "reports_limited"],
        "nurse": ["patients", "own_requests_view", "confirm_used", "telegram_personal", "temperature_reading"],
    }



def is_staff(user=None):
    user = user or current_user()
    return bool(user and user.get("role") in ("admin", "transfusion"))


def user_can_access_patient(conn, user, patient_id):
    if not user:
        return False
    if user.get("role") in ("admin", "transfusion"):
        return True
    p = ph()
    # Non-staff may access patients they created directly or patients tied to accessible requests.
    if one(conn, f"SELECT id FROM patients WHERE id={p} AND created_by={p} AND deleted_at IS NULL LIMIT 1", (patient_id, user["id"])):
        return True
    if user.get("role") == "nurse":
        return bool(one(conn, f"""SELECT id FROM requests
            WHERE patient_id={p} AND status<>'deleted'
            AND (requested_by={p} OR nurse_user_id={p} OR LOWER(TRIM(COALESCE(nurse_name,'')))=LOWER(TRIM({p})))
            LIMIT 1""", (patient_id, user["id"], user["id"], user.get("full_name", ""))))
    return bool(one(conn, f"SELECT id FROM requests WHERE patient_id={p} AND requested_by={p} AND status<>'deleted' LIMIT 1", (patient_id, user["id"])))


def user_can_access_request(conn, user, request_id):
    if not user:
        return False
    if user.get("role") in ("admin", "transfusion"):
        return True
    p = ph()
    if user.get("role") == "nurse":
        return bool(one(conn, f"""SELECT id FROM requests
            WHERE id={p} AND status<>'deleted'
            AND (requested_by={p} OR nurse_user_id={p} OR LOWER(TRIM(COALESCE(nurse_name,'')))=LOWER(TRIM({p})))""",
            (request_id, user["id"], user["id"], user.get("full_name", ""))))
    return bool(one(conn, f"SELECT id FROM requests WHERE id={p} AND requested_by={p} AND status<>'deleted'", (request_id, user["id"])))


def request_access_sql(user, alias="requests"):
    """Return SQL condition and params for non-staff request visibility."""
    p = ph()
    if is_staff(user):
        return "1=1", []
    prefix = f"{alias}." if alias else ""
    if user.get("role") == "nurse":
        return f"({prefix}requested_by={p} OR {prefix}nurse_user_id={p} OR LOWER(TRIM(COALESCE({prefix}nurse_name,'')))=LOWER(TRIM({p})))", [user["id"], user["id"], user.get("full_name", "")]
    return f"{prefix}requested_by={p}", [user["id"]]


def limited_reports_allowed(user):
    return bool(user and user.get("role") in ("admin", "transfusion"))


def filter_report_rows_for_user(conn, user, rows):
    if is_staff(user):
        return rows
    # For non-staff, only details tied to their own requests. Summary exports are staff-only.
    own_ids = {r["id"] for r in all_rows(conn, f"SELECT id FROM requests WHERE requested_by={ph()}", (user["id"],))}
    return [r for r in rows if r.get("request_id") in own_ids]

def table_export(conn, name):
    return all_rows(conn, f"SELECT * FROM {name} ORDER BY id")


def sanitize_backup_tables(tables, include_secrets=False):
    """Remove secrets from default backups. Medical data remains sensitive and must be protected."""
    if include_secrets:
        return tables
    cleaned = json.loads(json.dumps(tables, ensure_ascii=False, default=str))
    for row in cleaned.get("telegram_config", []) or []:
        if "bot_token" in row:
            row["bot_token"] = ""
    return cleaned

def backup_dir_path():
    path = current_app.config["BACKUP_DIR"]
    if not os.path.isabs(path):
        path = os.path.abspath(path)
    return path


def backup_tables():
    return [
        "users", "patients", "requests", "stock_units", "movements", "report_adjustments", "component_catalog",
        "telegram_config", "telegram_subscribers", "telegram_log",
        "temperature_devices", "temperature_readings", "audit_log", "backup_log", "migration_log", "login_events", "backup_policy"
    ]


def table_columns(conn, name):
    if is_pg():
        p = ph()
        rows = all_rows(conn, f"""
            SELECT column_name AS name FROM information_schema.columns
            WHERE table_schema='public' AND table_name={p}
            ORDER BY ordinal_position
        """, (name,))
        return [r["name"] for r in rows]
    return [r["name"] for r in all_rows(conn, f"PRAGMA table_info({name})")]


def backup_checksum(payload):
    raw = json.dumps(payload.get("tables", {}), ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def backup_log(conn, action, file_name="", ok_value=0, error="", details=""):
    p = ph()
    exec_sql(conn, f"INSERT INTO backup_log(action,file_name,ok,error,details,user_id) VALUES ({p},{p},{p},{p},{p},{p})",
             (action, file_name, 1 if ok_value else 0, str(error)[:1000], str(details)[:2000], session.get("uid")))


def build_backup_payload(conn, include_secrets=False):
    raw_tables = {t: table_export(conn, t) for t in backup_tables()}
    tables = sanitize_backup_tables(raw_tables, include_secrets=include_secrets)
    payload = {
        "kind": "blood_bank_v7_backup",
        "version": current_app.config["VERSION"],
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "db": "postgresql" if is_pg() else "sqlite",
        "tables": tables,
        "manifest": {
            "tables": {t: len(rows) for t, rows in tables.items()},
            "format": 3,
            "secrets_included": bool(include_secrets),
            "note": "V7.1.9 backup. Contains confidential medical data. Telegram bot token is excluded by default."
        }
    }
    payload["manifest"]["sha256"] = backup_checksum(payload)
    return payload


def validate_backup_payload(payload):
    if not isinstance(payload, dict) or payload.get("kind") != "blood_bank_v7_backup":
        return False, "Це не резервна копія Blood Bank V7"
    tables = payload.get("tables")
    if not isinstance(tables, dict):
        return False, "У файлі немає секції tables"
    required = ["users", "requests", "stock_units", "movements"]
    missing = [t for t in required if t not in tables]
    if missing:
        return False, "У копії відсутні таблиці: " + ", ".join(missing)
    expected = (payload.get("manifest") or {}).get("sha256")
    if expected and expected != backup_checksum(payload):
        return False, "Контрольна сума backup не збігається. Файл міг бути змінений або пошкоджений"
    return True, "OK"


def restore_backup_payload(conn, payload):
    okp, msg = validate_backup_payload(payload)
    if not okp:
        raise ValueError(msg)
    tables = [t for t in backup_tables() if t in payload.get("tables", {})]
    pmark = ph()
    manifest = payload.get("manifest") or {}
    secrets_included = bool(manifest.get("secrets_included"))
    preserved_telegram_token = ""
    if not secrets_included and "telegram_config" in tables:
        current_cfg = one(conn, "SELECT bot_token FROM telegram_config WHERE id=1") or {}
        preserved_telegram_token = current_cfg.get("bot_token") or current_app.config.get("TELEGRAM_BOT_TOKEN", "") or ""
    # clear child-ish tables first, then reference tables. audit/backup_log last is intentionally restored too.
    for t in reversed(tables):
        exec_sql(conn, f"DELETE FROM {t}")
    for t in tables:
        cols = table_columns(conn, t)
        rows = payload["tables"].get(t, []) or []
        for row in rows:
            use_cols = [c for c in cols if c in row]
            if not use_cols:
                continue
            placeholders = ",".join([pmark] * len(use_cols))
            sql = f"INSERT INTO {t} ({','.join(use_cols)}) VALUES ({placeholders})"
            exec_sql(conn, sql, tuple(row.get(c) for c in use_cols))
    if preserved_telegram_token:
        exec_sql(conn, f"UPDATE telegram_config SET bot_token={pmark} WHERE id=1", (preserved_telegram_token,))
    reset_postgres_sequences(conn, tables)
    result = {t: len(payload["tables"].get(t, []) or []) for t in tables}
    if preserved_telegram_token:
        result["telegram_token_preserved"] = True
    return result


def audit(conn, action, entity="", entity_id=None, details=""):
    p = ph()
    exec_sql(conn, f"INSERT INTO audit_log(user_id,action,entity,entity_id,details) VALUES ({p},{p},{p},{p},{p})",
             (session.get("uid"), action, entity, entity_id, details))


class DateFilterError(ValueError):
    pass

class IntParamError(ValueError):
    pass

class FloatParamError(ValueError):
    pass

def parse_int_param(value, name, *, required=False, default=None, min_value=None, max_value=None):
    """Validate integer query/body parameters and return a safe int or default.

    Prevents accidental 500 errors and PostgreSQL type-cast failures when
    callers send values like ?device_id=abc or ?limit=abc.
    """
    if value is None or str(value).strip() == "":
        if required:
            raise IntParamError(f"{name} має бути числом")
        return default
    s = str(value).strip()
    if not re.fullmatch(r"-?\d+", s):
        raise IntParamError(f"{name} має бути числом")
    n = int(s)
    if min_value is not None and n < min_value:
        raise IntParamError(f"{name} має бути не менше {min_value}")
    if max_value is not None and n > max_value:
        raise IntParamError(f"{name} має бути не більше {max_value}")
    return n

def parse_float_param(value, name, *, required=False, default=None, min_value=None, max_value=None):
    """Validate numeric query/body parameters and return a safe float or default."""
    if value is None or str(value).strip() == "":
        if required:
            raise FloatParamError(f"{name} має бути числом")
        return default
    s = str(value).strip().replace(",", ".")
    try:
        n = float(s)
    except Exception:
        raise FloatParamError(f"{name} має бути числом")
    if min_value is not None and n < min_value:
        raise FloatParamError(f"{name} має бути не менше {min_value}")
    if max_value is not None and n > max_value:
        raise FloatParamError(f"{name} має бути не більше {max_value}")
    return n


def validate_date_param(value, name="date"):
    value = (value or "").strip()
    if not value:
        return ""
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        raise DateFilterError(f"Невірний формат дати {name}. Потрібно YYYY-MM-DD")
    try:
        date.fromisoformat(value)
    except Exception:
        raise DateFilterError(f"Невірна дата {name}. Потрібно YYYY-MM-DD")
    return value

def sql_date(field):
    # PostgreSQL does not accept DATE(text) reliably when timestamps are stored as TEXT.
    # SQLite supports DATE(text). Keep all date filters portable.
    return f"CAST({field} AS DATE)" if is_pg() else f"DATE({field})"


def date_filter_sql(field, start, end, params):
    start = validate_date_param(start, "start")
    end = validate_date_param(end, "end")
    p = ph()
    field_date = sql_date(field)
    parts = []
    if start:
        parts.append(f"{field_date} >= {p}")
        params.append(start)
    if end:
        parts.append(f"{field_date} <= {p}")
        params.append(end)
    return (" AND " + " AND ".join(parts)) if parts else ""

@bp.errorhandler(DateFilterError)
def _date_filter_error(e):
    return err(str(e), 400)

@bp.errorhandler(IntParamError)
def _int_param_error(e):
    return err(str(e), 400)

@bp.errorhandler(FloatParamError)
def _float_param_error(e):
    return err(str(e), 400)


def boolint(v):
    if isinstance(v, bool):
        return 1 if v else 0
    return 1 if str(v).lower() in ("1", "true", "on", "yes", "так") else 0


def telegram_config(conn):
    cfg = one(conn, "SELECT * FROM telegram_config WHERE id=1") or {}
    token = (cfg.get("bot_token") or current_app.config.get("TELEGRAM_BOT_TOKEN") or "").strip()
    enabled = bool(cfg.get("enabled")) or bool(current_app.config.get("TELEGRAM_BOT_TOKEN"))
    return {"bot_token": token, "enabled": enabled, "critical_threshold": int(cfg.get("critical_threshold") or 2)}


def telegram_send_raw(conn, event, chat_id, text):
    cfg = telegram_config(conn)
    token = cfg.get("bot_token")
    if not cfg.get("enabled") or not token or not chat_id:
        return False
    ok_send = 0
    error = ""
    try:
        safe_text = html.escape(str(text or ""), quote=False)
        data = urllib.parse.urlencode({"chat_id": chat_id, "text": safe_text, "parse_mode": "HTML"}).encode()
        req = urllib.request.Request(f"https://api.telegram.org/bot{token}/sendMessage", data=data, method="POST")
        with urllib.request.urlopen(req, timeout=8) as resp:
            ok_send = 1 if resp.status == 200 else 0
            if not ok_send:
                error = f"HTTP {resp.status}"
    except Exception as e:
        error = str(e)[:500]
    p = ph()
    exec_sql(conn, f"INSERT INTO telegram_log(event,chat_id,ok,error,text) VALUES ({p},{p},{p},{p},{p})", (event, str(chat_id), ok_send, error, str(text or "")[:1000]))
    return bool(ok_send)


def telegram_notify(conn, event, text, role_filter=None):
    """Best-effort notification. Never blocks main clinical workflow."""
    try:
        flag = {
            "new_request": "notify_new_request", "approve": "notify_approve", "reject": "notify_reject",
            "issue": "notify_issue", "used": "notify_used", "critical": "notify_critical",
            "expired": "notify_expired", "system": "notify_system"
        }.get(event, "notify_system")
        where = f"s.enabled=1 AND COALESCE(s.chat_id,'')<>'' AND s.{flag}=1 AND u.active=1"
        params = []
        if role_filter:
            placeholders = ",".join([ph()] * len(role_filter))
            where += f" AND u.role IN ({placeholders})"
            params.extend(role_filter)
        rows = all_rows(conn, f"SELECT s.chat_id FROM telegram_subscribers s JOIN users u ON u.id=s.user_id WHERE {where}", tuple(params))
        for r in rows:
            telegram_send_raw(conn, event, r.get("chat_id"), text)
    except Exception as e:
        try:
            p = ph()
            exec_sql(conn, f"INSERT INTO telegram_log(event,chat_id,ok,error,text) VALUES ({p},{p},0,{p},{p})", (event, "", str(e)[:500], str(text or "")[:1000]))
        except Exception:
            pass


def check_critical_stock(conn):
    cfg = telegram_config(conn)
    threshold = int(cfg.get("critical_threshold") or 0)
    if threshold <= 0:
        return
    rows = all_rows(conn, "SELECT component_type, abo, rh, COUNT(*) qty FROM stock_units WHERE status='in_stock' GROUP BY component_type,abo,rh HAVING COUNT(*) <= %s ORDER BY qty" if is_pg() else "SELECT component_type, abo, rh, COUNT(*) qty FROM stock_units WHERE status='in_stock' GROUP BY component_type,abo,rh HAVING COUNT(*) <= ? ORDER BY qty", (threshold,))
    if rows:
        lines = ["⚠️ Критичний залишок компонентів:"]
        for r in rows[:20]:
            lines.append(f"{r['component_type']} {r['abo']} {r['rh']}: {r['qty']}")
        telegram_notify(conn, "critical", "\n".join(lines), role_filter=["admin", "transfusion"])


@bp.after_app_request
def headers(resp):
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    resp.headers.setdefault("Referrer-Policy", "same-origin")
    return resp


@bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        key = login_client_key() + ":" + (username.lower() or "-")
        wait = login_blocked(key)
        if wait:
            record_login_event(username, False, f"blocked {wait}s")
            return render_template("login.html", error=f"Забагато невдалих спроб. Спробуйте через {wait//60 + 1} хв.", version=current_app.config["VERSION"]), 429
        u = login_user(username, request.form.get("password", ""))
        if not u:
            login_failure(key)
            record_login_event(username, False, "bad_credentials")
            return render_template("login.html", error="Невірний логін або пароль", version=current_app.config["VERSION"])
        login_success(key)
        record_login_event(username, True, "first_login" if u.get("first_login") else "ok", u)
        if u.get("first_login"):
            return redirect("/change-password")
        return redirect("/")
    return render_template("login.html", error="", version=current_app.config["VERSION"])


@bp.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


@bp.route("/change-password", methods=["GET", "POST"])
@require_login
def change_password():
    u = current_user()
    if request.method == "POST":
        if request.form.get("csrf") != session.get("csrf"):
            return render_template("change_password.html", error="CSRF token invalid", csrf=session.get("csrf"), version=current_app.config["VERSION"], first_login=u.get("first_login"))
        old_password = request.form.get("old_password", "")
        new = request.form.get("new_password", "")
        if not u.get("first_login") and not check_password_hash(u.get("password_hash", ""), old_password):
            return render_template("change_password.html", error="Старий пароль невірний", csrf=session.get("csrf"), version=current_app.config["VERSION"], first_login=u.get("first_login"))
        if len(new) < 6:
            return render_template("change_password.html", error="Пароль має бути мінімум 6 символів", csrf=session.get("csrf"), version=current_app.config["VERSION"], first_login=u.get("first_login"))
        p = ph()
        with db() as conn:
            exec_sql(conn, f"UPDATE users SET password_hash={p}, first_login=0 WHERE id={p}", (generate_password_hash(new), u["id"]))
            audit(conn, "change_password", "users", u["id"])
        return redirect("/")
    return render_template("change_password.html", error="", csrf=session.get("csrf"), version=current_app.config["VERSION"], first_login=u.get("first_login"))


@bp.route("/")
@require_login
def index():
    return render_template("app.html", user=current_user(), csrf=session.get("csrf"), version=current_app.config["VERSION"])


@bp.route("/offline")
def offline():
    return render_template("offline.html", version=current_app.config["VERSION"])


@bp.route("/manifest.json")
def manifest():
    return send_from_directory("static", "manifest.json")


@bp.route("/api/public-health")
def public_health():
    try:
        with db() as conn:
            scalar(conn, "SELECT 1")
        return ok(version=current_app.config["VERSION"], db="postgresql" if is_pg() else "sqlite")
    except Exception as e:
        return err(str(e), 500)


@bp.route("/api/render-readiness")
def readiness():
    token = current_app.config.get("API_TOKEN")
    production_like = is_pg() or current_app.config.get("POSTGRES_ONLY") or current_app.config.get("REQUIRE_HTTPS") or current_app.config.get("REQUIRE_API_TOKEN")
    if production_like and not token:
        return err("API_TOKEN не налаштований", 503)
    if token and request.args.get("token") != token:
        return err("API_TOKEN неправильний", 403)
    with db() as conn:
        counts = {t: scalar(conn, f"SELECT COUNT(*) FROM {t}") for t in ["users", "patients", "requests", "stock_units", "movements", "report_adjustments", "component_catalog", "telegram_config", "telegram_subscribers", "telegram_log", "temperature_devices", "temperature_readings", "audit_log", "backup_log", "migration_log"]}
        integrity = system_integrity_report(conn)
    warnings = []
    if current_app.config.get("SECRET_KEY_RANDOM"):
        warnings.append("SECRET_KEY не заданий у ENV. На production це призведе до скидання сесій після рестарту.")
    if production_like and not current_app.config.get("API_TOKEN"):
        warnings.append("API_TOKEN не налаштований для production-like режиму.")
    return ok(version=current_app.config["VERSION"], db="postgresql" if is_pg() else "sqlite", counts=counts, integrity={"status": integrity["status"], "critical": integrity["critical"], "warning": integrity["warning"]}, security_warnings=warnings)


@bp.route("/api/me")
@require_login
def me():
    return ok(user=current_user(), csrf=session.get("csrf"), version=current_app.config["VERSION"])


@bp.route("/api/components")
@require_login
def components_api():
    with db() as conn:
        rows = all_rows(conn, "SELECT id,name,category,active,sort_order FROM component_catalog WHERE active=1 ORDER BY sort_order,name")
    if not rows:
        rows = [{"id": i+1, **c, "active": 1, "sort_order": i+1} for i,c in enumerate(DEFAULT_COMPONENTS)]
    return ok(components=rows)

@bp.route("/api/nurses")
@require_login
def nurses_api():
    """Active nurses for request assignment.

    Accessible to roles that can create or work with requests. Returns only
    minimal public staff data needed for selecting nurse_user_id in the form.
    """
    u = current_user()
    if u["role"] not in ["admin", "transfusion", "doctor"]:
        return err("Недостатньо прав", 403)
    with db() as conn:
        rows = all_rows(conn, "SELECT id, full_name, position FROM users WHERE role='nurse' AND active=1 ORDER BY full_name, id")
    return ok(nurses=rows)

@bp.route("/api/users", methods=["GET", "POST"])
@require_login
@require_csrf
def users():
    u = current_user()
    p = ph()
    if u["role"] != "admin":
        return err("Недостатньо прав", 403)
    if request.method == "GET":
        with db() as conn:
            return ok(users=all_rows(conn, "SELECT id,username,full_name,position,role,active,first_login,created_at FROM users ORDER BY id"))
    j = getj()
    role = j.get("role", "doctor")
    if role not in ["admin", "transfusion", "doctor", "nurse"]:
        return err("Невірна роль")
    if not j.get("username") or not j.get("full_name"):
        return err("Логін і ПІБ обов’язкові")
    with db() as conn:
        user_id = insert_and_get_id(conn, "users", ["username","full_name","position","role","password_hash","first_login"],
                 [j["username"].strip(), j["full_name"].strip(), j.get("position", ""), role, generate_password_hash(j.get("password", "123456")), 1])
        audit(conn, "create_user", "users", user_id, j["username"])
    return ok(user={"id": user_id})


@bp.route("/api/users/<int:uid>", methods=["PUT", "DELETE"])
@require_login
@require_csrf
def user_update(uid):
    u = current_user()
    if u["role"] != "admin":
        return err("Недостатньо прав", 403)
    p = ph(); j = getj()
    with db() as conn:
        target = one(conn, f"SELECT * FROM users WHERE id={p}", (uid,))
        if not target:
            return err("Користувача не знайдено", 404)
        if request.method == "DELETE":
            if uid == u["id"]:
                return err("Не можна вимкнути самого себе")
            exec_sql(conn, f"UPDATE users SET active=0 WHERE id={p}", (uid,))
            audit(conn, "deactivate_user", "users", uid)
            return ok()
        role = j.get("role", target["role"])
        if role not in ["admin", "transfusion", "doctor", "nurse"]:
            return err("Невірна роль")
        exec_sql(conn, f"UPDATE users SET full_name={p}, position={p}, role={p}, active={p} WHERE id={p}",
                 (j.get("full_name", target["full_name"]), j.get("position", target.get("position", "")), role, int(j.get("active", target["active"])), uid))
        if j.get("password"):
            exec_sql(conn, f"UPDATE users SET password_hash={p}, first_login=1 WHERE id={p}", (generate_password_hash(j["password"]), uid))
        audit(conn, "update_user", "users", uid)
    return ok()



@bp.route("/api/patients", methods=["GET", "POST"])
@require_login
@require_csrf
def patients_api():
    u = current_user(); p = ph()
    if request.method == "GET":
        qtext = (request.args.get("q") or "").strip()
        with db() as conn:
            params = []
            if is_staff(u):
                base = "SELECT DISTINCT p.* FROM patients p WHERE p.deleted_at IS NULL"
            else:
                # Doctor/nurse see patients they created directly and patients tied to their own/assigned requests.
                if u.get("role") == "nurse":
                    base = f"""SELECT DISTINCT p.* FROM patients p
                        LEFT JOIN requests r ON r.patient_id=p.id AND r.status<>'deleted'
                        WHERE p.deleted_at IS NULL AND (p.created_by={p} OR r.requested_by={p} OR r.nurse_user_id={p} OR LOWER(TRIM(COALESCE(r.nurse_name,'')))=LOWER(TRIM({p})))"""
                    params.extend([u["id"], u["id"], u["id"], u.get("full_name", "")])
                else:
                    base = f"""SELECT DISTINCT p.* FROM patients p
                        LEFT JOIN requests r ON r.patient_id=p.id AND r.status<>'deleted'
                        WHERE p.deleted_at IS NULL AND (p.created_by={p} OR r.requested_by={p})"""
                    params.extend([u["id"], u["id"]])
            if qtext:
                like = f"%{qtext}%"
                base += f" AND (p.full_name LIKE {p} OR p.birth_date LIKE {p} OR p.address LIKE {p} OR p.department LIKE {p} OR p.diagnosis LIKE {p})"
                params.extend([like, like, like, like, like])
            rows = all_rows(conn, base + " ORDER BY p.updated_at DESC, p.id DESC LIMIT 200", tuple(params))
        return ok(patients=rows)
    if u["role"] not in ["admin", "transfusion", "doctor", "nurse"]:
        return err("Недостатньо прав", 403)
    j = getj()
    if not (j.get("full_name") or j.get("patient_name")):
        return err("ПІБ пацієнта обов’язкове")
    with db() as conn:
        pid_new = insert_and_get_id(conn, "patients",
            ["full_name","birth_date","address","department","diagnosis","patient_status","abo","rh","note","created_by"],
            [(j.get("full_name") or j.get("patient_name") or "").strip(), j.get("birth_date", ""), j.get("address", ""),
             j.get("department", ""), j.get("diagnosis", ""), j.get("patient_status", ""), normalize_abo(j.get("abo", "")) if j.get("abo") else "", normalize_rh(j.get("rh", "")) if j.get("rh") else "", j.get("note", ""), u["id"]])
        audit(conn, "create_patient", "patients", pid_new, j.get("full_name") or j.get("patient_name") or "")
    return ok(patient={"id": pid_new})


@bp.route("/api/patients/<int:pid>", methods=["GET", "PUT", "DELETE"])
@require_login
@require_csrf
def patient_detail(pid):
    u = current_user(); p = ph(); j = getj()
    with db() as conn:
        patient = one(conn, f"SELECT * FROM patients WHERE id={p}", (pid,))
        if not patient:
            return err("Пацієнта не знайдено", 404)
        if not user_can_access_patient(conn, u, pid):
            return err("Недостатньо прав", 403)
        if request.method == "GET":
            if is_staff(u):
                reqs = all_rows(conn, f"SELECT * FROM requests WHERE patient_id={p} ORDER BY id DESC", (pid,))
                units = all_rows(conn, f"""SELECT su.* FROM stock_units su JOIN requests r ON r.id=su.request_id
                    WHERE r.patient_id={p} ORDER BY su.id DESC""", (pid,))
            else:
                cond, cond_params = request_access_sql(u, "requests")
                reqs = all_rows(conn, f"SELECT * FROM requests WHERE patient_id={p} AND ({cond}) ORDER BY id DESC", tuple([pid] + cond_params))
                cond_r, cond_r_params = request_access_sql(u, "r")
                units = all_rows(conn, f"""SELECT su.* FROM stock_units su JOIN requests r ON r.id=su.request_id
                    WHERE r.patient_id={p} AND ({cond_r}) ORDER BY su.id DESC""", tuple([pid] + cond_r_params))
            return ok(patient=patient, requests=reqs, units=units)
        if u["role"] not in ["admin", "transfusion", "doctor"]:
            return err("Недостатньо прав", 403)
        if not is_staff(u):
            # Doctor may edit patients they created or patients tied to their own requests. Nurse is view/confirm-only.
            if u["role"] != "doctor" or not user_can_access_patient(conn, u, pid):
                return err("Недостатньо прав", 403)
        if request.method == "DELETE":
            if u["role"] not in ["admin", "transfusion"]:
                return err("Недостатньо прав", 403)
            exec_sql(conn, f"UPDATE patients SET deleted_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP WHERE id={p}", (pid,))
            audit(conn, "delete_patient", "patients", pid, j.get("reason", ""))
            return ok()
        new_patient = {
            "full_name": j.get("full_name", patient["full_name"]), "birth_date": j.get("birth_date", patient.get("birth_date", "")),
            "address": j.get("address", patient.get("address", "")), "department": j.get("department", patient.get("department", "")),
            "diagnosis": j.get("diagnosis", patient.get("diagnosis", "")), "patient_status": j.get("patient_status", patient.get("patient_status", "")),
            "abo": normalize_abo(j.get("abo", patient.get("abo", ""))), "rh": normalize_rh(j.get("rh", patient.get("rh", ""))), "note": j.get("note", patient.get("note", ""))
        }
        exec_sql(conn, f"""UPDATE patients SET full_name={p}, birth_date={p}, address={p}, department={p}, diagnosis={p},
            patient_status={p}, abo={p}, rh={p}, note={p}, updated_at=CURRENT_TIMESTAMP WHERE id={p}""",
            (new_patient["full_name"], new_patient["birth_date"], new_patient["address"], new_patient["department"], new_patient["diagnosis"], new_patient["patient_status"],
             new_patient["abo"], new_patient["rh"], new_patient["note"], pid))
        audit(conn, "update_patient", "patients", pid, audit_diff(patient, new_patient, ["full_name","birth_date","address","department","diagnosis","patient_status","abo","rh","note"]))
    return ok()


def upsert_patient_from_request(conn, j):
    p = ph()
    name = (j.get("patient_name") or j.get("full_name") or "").strip()
    birth = j.get("birth_date", "")
    if not name:
        return None
    found = one(conn, f"SELECT * FROM patients WHERE deleted_at IS NULL AND full_name={p} AND COALESCE(birth_date,'')={p} ORDER BY id DESC LIMIT 1", (name, birth or ""))
    if found:
        exec_sql(conn, f"""UPDATE patients SET address={p}, department={p}, diagnosis={p}, patient_status={p}, abo={p}, rh={p}, updated_at=CURRENT_TIMESTAMP WHERE id={p}""",
                 (j.get("address", found.get("address", "")), j.get("department", found.get("department", "")), j.get("diagnosis", found.get("diagnosis", "")),
                  j.get("patient_status", found.get("patient_status", "")), j.get("abo", found.get("abo", "")), j.get("rh", found.get("rh", "")), found["id"]))
        return found["id"]
    return insert_and_get_id(conn, "patients",
        ["full_name","birth_date","address","department","diagnosis","patient_status","abo","rh","note","created_by"],
        [name, birth, j.get("address", ""), j.get("department", ""), j.get("diagnosis", ""), j.get("patient_status", ""), j.get("abo", ""), j.get("rh", ""), "Створено автоматично з вимоги", j.get("requested_by")])



def request_progress(conn, rid):
    p = ph()
    total = scalar(conn, f"SELECT quantity FROM requests WHERE id={p}", (rid,)) or 0
    issued = scalar(conn, f"SELECT COUNT(*) FROM stock_units WHERE request_id={p} AND status='issued'", (rid,)) or 0
    used = scalar(conn, f"SELECT COUNT(*) FROM stock_units WHERE request_id={p} AND status='used'", (rid,)) or 0
    delivered = issued + used
    remaining = max(int(total) - int(delivered), 0)
    return {"issued_count": int(issued), "used_count": int(used), "delivered_count": int(delivered), "remaining_quantity": int(remaining)}

def enrich_requests(conn, rows):
    out = []
    for r in rows:
        d = dict(r)
        d.update(request_progress(conn, d["id"]))
        out.append(d)
    return out

def resolve_nurse_user_id(conn, nurse_user_id=None, nurse_name=""):
    p = ph()
    nurse_id = parse_int_param(nurse_user_id, "nurse_user_id", min_value=1)
    if nurse_id:
        row = one(conn, f"SELECT id, full_name, position FROM users WHERE id={p} AND role='nurse' AND active=1", (nurse_id,))
        if row:
            return row["id"], row.get("full_name") or nurse_name, row.get("position") or ""
    name = (nurse_name or "").strip()
    if name:
        row = one(conn, f"SELECT id, full_name, position FROM users WHERE role='nurse' AND active=1 AND LOWER(TRIM(full_name))=LOWER(TRIM({p})) ORDER BY id DESC LIMIT 1", (name,))
        if row:
            return row["id"], row.get("full_name") or name, row.get("position") or ""
    return None, name, ""

def update_request_status(conn, rid):
    p = ph()
    r = one(conn, f"SELECT * FROM requests WHERE id={p}", (rid,))
    if not r or r.get("status") in ["created", "rejected", "deleted"]:
        return
    pr = request_progress(conn, rid)
    total = int(r.get("quantity") or 0)
    delivered = pr["delivered_count"]
    issued = pr["issued_count"]
    used = pr["used_count"]
    if delivered <= 0:
        new_status = "approved"
    elif delivered < total:
        new_status = "partial_issued"
    elif issued > 0:
        new_status = "issued"
    elif used >= total:
        new_status = "used"
    else:
        new_status = "issued"
    exec_sql(conn, f"UPDATE requests SET status={p}, updated_at=CURRENT_TIMESTAMP WHERE id={p}", (new_status, rid))

@bp.route("/api/requests", methods=["GET", "POST"])
@require_login
@require_csrf
def requests_api():
    u = current_user()
    p = ph()
    if request.method == "GET":
        with db() as conn:
            if u["role"] == "nurse":
                rows = all_rows(conn, f"""SELECT * FROM requests WHERE status<>'deleted'
                    AND (requested_by={p} OR nurse_user_id={p} OR LOWER(TRIM(COALESCE(nurse_name,'')))=LOWER(TRIM({p}))) ORDER BY id DESC""", (u["id"], u["id"], u.get("full_name", "")))
            elif u["role"] == "doctor":
                rows = all_rows(conn, f"SELECT * FROM requests WHERE requested_by={p} AND status<>'deleted' ORDER BY id DESC", (u["id"],))
            else:
                rows = all_rows(conn, "SELECT * FROM requests WHERE status<>'deleted' ORDER BY id DESC")
            rows = enrich_requests(conn, rows)
        return ok(requests=rows)
    if u["role"] not in ["admin", "transfusion", "doctor"]:
        return err("Медсестра не створює вимоги у V7 Clean Core", 403)
    j = getj()
    qty = parse_int_param(j.get("quantity"), "Кількість", required=True, min_value=1, max_value=100)
    if not j.get("patient_name") or not j.get("component_type"):
        return err("Не заповнені обов’язкові поля")
    with db() as conn:
        try:
            j["component_type"] = require_catalog_component(conn, j.get("component_type"))
        except ValueError as e:
            return err(str(e))
        j["abo"] = normalize_abo(j.get("abo", "O(I)"))
        j["rh"] = normalize_rh(j.get("rh", "Rh+"))
        j["requested_by"] = u["id"]
        nurse_uid, nurse_name, nurse_pos = resolve_nurse_user_id(conn, j.get("nurse_user_id"), j.get("nurse_name", ""))
        j["nurse_user_id"] = nurse_uid
        if nurse_name: j["nurse_name"] = nurse_name
        if nurse_pos and not j.get("nurse_position"): j["nurse_position"] = nurse_pos
        patient_id = upsert_patient_from_request(conn, j)
        rid_new = insert_and_get_id(conn, "requests",
            ["patient_id","patient_name","birth_date","address","department","diagnosis","patient_status","component_type","abo","rh","quantity","needed_date","urgency","indication","request_note","transfusion_history","reaction_note","doctor_name","doctor_position","nurse_name","nurse_user_id","nurse_position","requested_by"],
            [patient_id, j.get("patient_name", ""), j.get("birth_date", ""), j.get("address", ""), j.get("department", ""), j.get("diagnosis", ""),
             j.get("patient_status", ""), j.get("component_type", ""), j.get("abo", "O(I)"), j.get("rh", "Rh+"), qty,
             j.get("needed_date", ""), j.get("urgency", "routine"), j.get("indication", ""), j.get("request_note", ""),
             j.get("transfusion_history", ""), j.get("reaction_note", ""), j.get("doctor_name") or u["full_name"],
             j.get("doctor_position") or u.get("position", ""), j.get("nurse_name", ""), j.get("nurse_user_id"), j.get("nurse_position", ""), u["id"]])
        audit(conn, "create_request", "requests", rid_new, j.get("patient_name", ""))
        telegram_notify(conn, "new_request", f"🩸 Нова вимога №{rid_new}\nПацієнт: {j.get('patient_name','')}\nКомпонент: {j.get('component_type','')} {j.get('abo','')} {j.get('rh','')} x{qty}\nПотрібно на: {j.get('needed_date','')}", role_filter=["admin", "transfusion"])
    return ok(request={"id": rid_new})


@bp.route("/api/requests/<int:rid>/<action>", methods=["POST"])
@require_login
@require_csrf
def request_action(rid, action):
    u = current_user()
    p = ph()
    j = getj()
    with db() as conn:
        r = one(conn, f"SELECT * FROM requests WHERE id={p}", (rid,))
        if not r:
            return err("Вимогу не знайдено", 404)
        if action == "approve":
            if u["role"] not in ["admin", "transfusion"]:
                return err("Недостатньо прав", 403)
            if r["status"] != "created":
                return err("Погодити можна тільки створену вимогу")
            exec_sql(conn, f"UPDATE requests SET status='approved', updated_at=CURRENT_TIMESTAMP WHERE id={p}", (rid,))
            audit(conn, "approve_request", "requests", rid)
            telegram_notify(conn, "approve", f"✅ Вимогу №{rid} погоджено\nПацієнт: {r.get('patient_name','')}\nКомпонент: {r.get('component_type','')} {r.get('abo','')} {r.get('rh','')} x{r.get('quantity','')}")
            return ok()
        if action == "reject":
            if u["role"] not in ["admin", "transfusion"]:
                return err("Недостатньо прав", 403)
            exec_sql(conn, f"UPDATE requests SET status='rejected', reject_reason={p}, updated_at=CURRENT_TIMESTAMP WHERE id={p}", (j.get("reason", ""), rid))
            audit(conn, "reject_request", "requests", rid, j.get("reason", ""))
            telegram_notify(conn, "reject", f"❌ Вимогу №{rid} відхилено\nПацієнт: {r.get('patient_name','')}\nПричина: {j.get('reason','')}")
            return ok()
        if action == "delete":
            if u["role"] not in ["admin", "transfusion"] and r["requested_by"] != u["id"]:
                return err("Недостатньо прав", 403)
            if r["status"] != "created":
                return err("Погоджені/видані вимоги не видаляються напряму")
            exec_sql(conn, f"UPDATE requests SET status='deleted', deleted_at=CURRENT_TIMESTAMP, deleted_by={p}, delete_reason={p}, updated_at=CURRENT_TIMESTAMP WHERE id={p}", (u["id"], j.get("reason", ""), rid))
            audit(conn, "delete_request", "requests", rid, j.get("reason", ""))
            return ok()
        if action == "restore":
            if u["role"] not in ["admin", "transfusion"]:
                return err("Недостатньо прав", 403)
            if r["status"] != "deleted":
                return err("Вимога не в кошику")
            exec_sql(conn, f"UPDATE requests SET status='created', restored_at=CURRENT_TIMESTAMP, restored_by={p}, updated_at=CURRENT_TIMESTAMP WHERE id={p}", (u["id"], rid))
            audit(conn, "restore_request", "requests", rid)
            return ok()
    return err("Невідома дія")


@bp.route("/api/stock", methods=["GET", "POST"])
@require_login
@require_csrf
def stock_api():
    u = current_user()
    p = ph()
    if request.method == "GET":
        if u["role"] not in ["admin", "transfusion"]:
            return err("Недостатньо прав", 403)
        with db() as conn:
            return ok(units=all_rows(conn, "SELECT * FROM stock_units WHERE status<>'deleted' ORDER BY id DESC"))
    if u["role"] not in ["admin", "transfusion"]:
        return err("Недостатньо прав", 403)
    j = getj()
    qty = parse_int_param(j.get("quantity"), "Кількість", default=1, min_value=1, max_value=1000)
    with db() as conn:
        try:
            j["component_type"] = require_catalog_component(conn, j.get("component_type"))
        except ValueError as e:
            return err(str(e))
        j["abo"] = normalize_abo(j.get("abo", "O(I)"))
        j["rh"] = normalize_rh(j.get("rh", "Rh+"))
        base_code = (j.get("unit_code") or "").strip()
        for i in range(qty):
            code = (f"{base_code}-{i+1}" if base_code and qty > 1 else base_code) or f"UNIT-{int(time.time()*1000)}-{i}"
            new_unit_id = insert_and_get_id(conn, "stock_units",
                ["component_type","abo","rh","quantity","unit_code","series","source","received_date","expiry_date","note"],
                [j.get("component_type", ""), j.get("abo", "O(I)"), j.get("rh", "Rh+"), 1, code, j.get("series", ""), j.get("source", ""), j.get("received_date", ""), j.get("expiry_date", ""), j.get("note", "")])
            exec_sql(conn, f"INSERT INTO movements(unit_id,action,quantity,user_id,details) VALUES ({p},'income',1,{p},{p})", (new_unit_id, u["id"], code))
        audit(conn, "add_stock", "stock_units", None, f"{j.get('component_type','')} x{qty}")
        telegram_notify(conn, "system", f"➕ Надходження на склад: {j.get('component_type','')} {j.get('abo','')} {j.get('rh','')} x{qty}", role_filter=["admin", "transfusion"])
    return ok()


@bp.route("/api/issue", methods=["POST"])
@require_login
@require_csrf
def issue():
    u = current_user()
    p = ph(); j = getj()
    rid = parse_int_param(j.get("request_id"), "request_id", required=True, min_value=1)
    raw_unit_ids = j.get("unit_ids") or []
    if not isinstance(raw_unit_ids, list):
        return err("unit_ids має бути списком", 400)
    unit_ids = [parse_int_param(x, "unit_id", required=True, min_value=1) for x in raw_unit_ids]
    if u["role"] not in ["admin", "transfusion"]:
        return err("Недостатньо прав", 403)
    if not rid or not unit_ids:
        return err("Вкажіть вимогу і одиниці")
    with db() as conn:
        r = one(conn, f"SELECT * FROM requests WHERE id={p}", (rid,))
        if not r or r["status"] not in ["approved", "reserved", "partial_issued", "issued"]:
            return err("Вимога має бути погоджена")
        pr = request_progress(conn, rid)
        remaining = int(pr["remaining_quantity"])
        if remaining <= 0:
            return err("Вимога вже повністю видана")
        if len(unit_ids) <= 0:
            return err("Оберіть хоча б одну одиницю")
        if len(unit_ids) > remaining:
            return err(f"Не можна видати більше залишку по вимозі. Залишок: {remaining}")
        if len(set(unit_ids)) != len(unit_ids):
            return err("Одиниці повторюються у списку")
        for uid in unit_ids:
            unit = one(conn, f"SELECT * FROM stock_units WHERE id={p}", (uid,))
            if not unit or unit["status"] != "in_stock":
                return err(f"Одиниця {uid} недоступна")
            compatible, reason = unit_matches_request(r, unit)
            if not compatible:
                return err(f"Одиниця {uid} не відповідає вимозі: {reason}")
        for uid in unit_ids:
            exec_sql(conn, f"UPDATE stock_units SET status='issued', request_id={p}, issued_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP WHERE id={p}", (rid, uid))
            exec_sql(conn, f"INSERT INTO movements(unit_id,request_id,action,quantity,user_id) VALUES ({p},{p},'issue',1,{p})", (uid, rid, u["id"]))
        update_request_status(conn, rid)
        pr_after = request_progress(conn, rid)
        audit(conn, "issue", "requests", rid, ",".join(map(str, unit_ids)))
        telegram_notify(conn, "issue", f"📦 Видано по вимозі №{rid}\nПацієнт: {r.get('patient_name','')}\nОдиниці: {','.join(map(str, unit_ids))}\nЗалишок до видачі: {pr_after['remaining_quantity']}")
        check_critical_stock(conn)
    return ok()


@bp.route("/api/units/<int:uid>/<action>", methods=["POST"])
@require_login
@require_csrf
def unit_action(uid, action):
    u = current_user(); p = ph(); j = getj()
    with db() as conn:
        unit = one(conn, f"SELECT * FROM stock_units WHERE id={p}", (uid,))
        if not unit:
            return err("Одиницю не знайдено", 404)
        if action == "used":
            if unit["status"] != "issued":
                return err("Використати можна тільки видану одиницю")
            r = one(conn, f"SELECT * FROM requests WHERE id={p}", (unit["request_id"],))
            if u["role"] == "doctor" and r and r["requested_by"] != u["id"]:
                return err("Недостатньо прав", 403)
            if u["role"] == "nurse" and r and not (r["requested_by"] == u["id"] or r.get("nurse_user_id") == u["id"] or (r.get("nurse_name") or "").strip().lower() == (u.get("full_name") or "").strip().lower()):
                return err("Недостатньо прав", 403)
            reaction_type, reaction_severity, reaction_note = normalize_reaction_payload(j)
            reaction_text = reaction_label(reaction_type, reaction_severity, reaction_note)
            exec_sql(conn, f"""UPDATE stock_units SET status='used', used_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP,
                reaction_type={p}, reaction_severity={p}, reaction_note={p}, reaction_recorded_at=CURRENT_TIMESTAMP WHERE id={p}""",
                (reaction_type, reaction_severity, reaction_note, uid))
            exec_sql(conn, f"""INSERT INTO movements(unit_id,request_id,action,quantity,user_id,reason,reaction_type,reaction_severity)
                VALUES ({p},{p},'used',1,{p},{p},{p},{p})""", (uid, unit["request_id"], u["id"], reaction_text, reaction_type, reaction_severity))
            update_request_status(conn, unit["request_id"])
            audit(conn, "used", "stock_units", uid, reaction_text)
            if reaction_type != "none":
                telegram_notify(conn, "used", f"⚠️ Використано компонент із реакцією, одиниця №{uid}\nВимога: {unit.get('request_id') or ''}\n{reaction_text}", role_filter=["admin", "transfusion"])
            else:
                telegram_notify(conn, "used", f"✅ Використано компонент, одиниця №{uid}\nВимога: {unit.get('request_id') or ''}\nПримітка: {reaction_text}")
            return ok(reaction={"type": reaction_type, "severity": reaction_severity, "note": reaction_note, "label": reaction_text})
        if action == "return":
            if u["role"] not in ["admin", "transfusion"]:
                return err("Недостатньо прав", 403)
            if unit["status"] != "issued":
                return err("Повернути можна тільки видану, але ще не використану одиницю")
            old_request_id = unit.get("request_id")
            reason = j.get("reason", "Повернення на склад")
            exec_sql(conn, f"UPDATE stock_units SET status='in_stock', request_id=NULL, returned_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP WHERE id={p}", (uid,))
            exec_sql(conn, f"INSERT INTO movements(unit_id,request_id,action,quantity,user_id,reason) VALUES ({p},{p},'return',1,{p},{p})", (uid, old_request_id, u["id"], reason))
            if old_request_id:
                update_request_status(conn, old_request_id)
            audit(conn, "return_unit", "stock_units", uid, reason)
            telegram_notify(conn, "system", f"↩️ Повернено на склад одиницю №{uid}\nВимога: {old_request_id or ''}\nПричина: {reason}", role_filter=["admin", "transfusion"])
            return ok()
        if action == "writeoff":
            if u["role"] not in ["admin", "transfusion"]:
                return err("Недостатньо прав", 403)
            if unit["status"] not in ["in_stock", "reserved"]:
                return err("Списати можна тільки на складі/резерв")
            exec_sql(conn, f"UPDATE stock_units SET status='written_off', written_off_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP WHERE id={p}", (uid,))
            exec_sql(conn, f"INSERT INTO movements(unit_id,action,quantity,user_id,reason) VALUES ({p},'writeoff',1,{p},{p})", (uid, u["id"], j.get("reason", "")))
            audit(conn, "writeoff", "stock_units", uid, j.get("reason", ""))
            telegram_notify(conn, "system", f"🗑️ Списано одиницю №{uid}\nПричина: {j.get('reason','')}", role_filter=["admin", "transfusion"])
            check_critical_stock(conn)
            return ok()
        if action == "delete":
            if u["role"] not in ["admin", "transfusion"]:
                return err("Недостатньо прав", 403)
            if unit["status"] not in ["in_stock", "reserved", "expired", "written_off"]:
                return err("Видану/використану одиницю не можна перемістити в кошик")
            exec_sql(conn, f"UPDATE stock_units SET status='deleted', deleted_at=CURRENT_TIMESTAMP, deleted_by={p}, delete_reason={p}, updated_at=CURRENT_TIMESTAMP WHERE id={p}", (u["id"], j.get("reason", ""), uid))
            exec_sql(conn, f"INSERT INTO movements(unit_id,action,quantity,user_id,reason) VALUES ({p},'delete',1,{p},{p})", (uid, u["id"], j.get("reason", "")))
            audit(conn, "delete_unit", "stock_units", uid, j.get("reason", ""))
            return ok()
        if action == "restore":
            if u["role"] not in ["admin", "transfusion"]:
                return err("Недостатньо прав", 403)
            if unit["status"] not in ["deleted", "written_off", "expired"]:
                return err("Одиниця не в кошику/не списана")
            exec_sql(conn, f"UPDATE stock_units SET status='in_stock', request_id=NULL, restored_at=CURRENT_TIMESTAMP, restored_by={p}, updated_at=CURRENT_TIMESTAMP WHERE id={p}", (u["id"], uid))
            exec_sql(conn, f"INSERT INTO movements(unit_id,action,quantity,user_id,reason) VALUES ({p},'restore',1,{p},{p})", (uid, u["id"], j.get("reason", "Відновлення")))
            audit(conn, "restore_unit", "stock_units", uid, j.get("reason", ""))
            return ok()
    return err("Невідома дія")


@bp.route("/api/stock/summary")
@require_login
def stock_summary():
    u = current_user()
    if u["role"] not in ["admin", "transfusion"]:
        return ok(summary=[])
    with db() as conn:
        rows = all_rows(conn, "SELECT component_type, abo, rh, COUNT(*) qty FROM stock_units WHERE status='in_stock' GROUP BY component_type,abo,rh ORDER BY component_type,abo,rh")
    return ok(summary=rows)


@bp.route("/api/dashboard/transfusion")
@require_login
@require_role("admin", "transfusion")
def transfusion_dashboard():
    p = ph()
    today = today_iso()
    with db() as conn:
        cfg = telegram_config(conn)
        threshold = int(cfg.get("critical_threshold") or 2)
        counts = {
            "new_requests": count_rows(conn, "requests", "status='created' AND deleted_at IS NULL"),
            "today_needed": count_rows(conn, "requests", f"needed_date={p} AND status NOT IN ('used','rejected','deleted')", (today,)),
            "in_stock": count_rows(conn, "stock_units", "status='in_stock'"),
            "issued": count_rows(conn, "stock_units", "status='issued'"),
            "expired": count_rows(conn, "stock_units", "status='expired'"),
            "temp_alarm": count_rows(conn, "temperature_readings", "deleted_at IS NULL AND status IN ('low','high','alarm')"),
        }
        new_requests = enrich_requests(conn, all_rows(conn, "SELECT * FROM requests WHERE status='created' AND deleted_at IS NULL ORDER BY created_at DESC, id DESC LIMIT 10"))
        today_requests = enrich_requests(conn, all_rows(conn, f"SELECT * FROM requests WHERE needed_date={p} AND status NOT IN ('used','rejected','deleted') ORDER BY urgency DESC, created_at DESC, id DESC LIMIT 10", (today,)))
        critical_stock = all_rows(conn, f"""SELECT component_type, abo, rh, COUNT(*) qty
            FROM stock_units
            WHERE status='in_stock'
            GROUP BY component_type, abo, rh
            HAVING COUNT(*) <= {p}
            ORDER BY qty ASC, component_type, abo, rh
            LIMIT 20""", (threshold,))
        expired_units = all_rows(conn, f"""SELECT id, component_type, abo, rh, unit_code, series, expiry_date, status
            FROM stock_units
            WHERE status IN ('expired') OR (status='in_stock' AND expiry_date<>'' AND expiry_date < {p})
            ORDER BY expiry_date, id
            LIMIT 20""", (today,))
        temp_alerts = all_rows(conn, """SELECT tr.id, tr.measured_at, tr.temperature, tr.humidity, tr.status, tr.note,
                td.name device_name, td.location, td.min_temp, td.max_temp
            FROM temperature_readings tr
            LEFT JOIN temperature_devices td ON td.id=tr.device_id
            WHERE tr.deleted_at IS NULL AND tr.status IN ('low','high','alarm')
            ORDER BY tr.measured_at DESC, tr.id DESC
            LIMIT 10""")
        recent_movements = all_rows(conn, """SELECT m.id, m.created_at, m.action, m.quantity, m.reason,
                su.component_type, su.abo, su.rh, su.unit_code, su.series, r.patient_name, r.id request_id, u.full_name user_name
            FROM movements m
            LEFT JOIN stock_units su ON su.id=m.unit_id
            LEFT JOIN requests r ON r.id=COALESCE(m.request_id, su.request_id)
            LEFT JOIN users u ON u.id=m.user_id
            WHERE m.deleted_at IS NULL
            ORDER BY m.created_at DESC, m.id DESC
            LIMIT 15""")
    return ok(today=today, threshold=threshold, counts=counts, new_requests=new_requests,
              today_requests=today_requests, critical_stock=critical_stock,
              expired_units=expired_units, temp_alerts=temp_alerts, recent_movements=recent_movements)


@bp.route("/api/stock/expire", methods=["POST"])
@require_login
@require_csrf
def expire_stock():
    u = current_user()
    if u["role"] not in ["admin", "transfusion"]:
        return err("Недостатньо прав", 403)
    p = ph(); today = today_iso()
    with db() as conn:
        rows = all_rows(conn, f"SELECT id FROM stock_units WHERE status='in_stock' AND expiry_date<>'' AND expiry_date < {p}", (today,))
        for r in rows:
            exec_sql(conn, f"UPDATE stock_units SET status='expired', updated_at=CURRENT_TIMESTAMP WHERE id={p}", (r["id"],))
            exec_sql(conn, f"INSERT INTO movements(unit_id,action,quantity,user_id,reason) VALUES ({p},'expired',1,{p},{p})", (r["id"], u["id"], "Автоматичне блокування протермінованих"))
        audit(conn, "expire_stock", "stock_units", None, str(len(rows)))
        if rows:
            telegram_notify(conn, "expired", f"⏰ Заблоковано протерміновані компоненти: {len(rows)}", role_filter=["admin", "transfusion"])
            check_critical_stock(conn)
    return ok(expired=len(rows))


@bp.route("/api/requests/<int:rid>/available-units")
@require_login
def available_units(rid):
    p = ph()
    with db() as conn:
        r = one(conn, f"SELECT * FROM requests WHERE id={p}", (rid,))
        if not r:
            return err("Вимогу не знайдено", 404)
        if u := current_user():
            if not user_can_access_request(conn, u, rid):
                return err("Недостатньо прав", 403)
        candidate_rows = all_rows(conn, f"""SELECT * FROM stock_units
            WHERE status='in_stock' AND component_type={p} AND rh={p}
            ORDER BY CASE WHEN expiry_date='' THEN 1 ELSE 0 END, expiry_date, id""", (r["component_type"], normalize_rh(r["rh"])))
        rows = [unit for unit in candidate_rows if unit_matches_request(r, unit)[0]]
        rd = dict(r); rd.update(request_progress(conn, rid))
    return ok(request=rd, units=rows)


@bp.route("/api/traceability/<code>")
@require_login
def traceability(code):
    p = ph(); code = (code or "").strip()
    if not code:
        return err("Порожній код")
    with db() as conn:
        unit = one(conn, f"SELECT * FROM stock_units WHERE unit_code={p} OR series={p} OR CAST(id AS TEXT)={p} ORDER BY id DESC", (code, code, code))
        if not unit:
            return err("Одиницю не знайдено", 404)
        req = one(conn, f"SELECT * FROM requests WHERE id={p}", (unit.get("request_id"),)) if unit.get("request_id") else None
        u = current_user()
        if not is_staff(u):
            if not req or not user_can_access_request(conn, u, req["id"]):
                return err("Недостатньо прав", 403)
        moves = all_rows(conn, f"SELECT * FROM movements WHERE unit_id={p} ORDER BY id", (unit["id"],))
    return ok(unit=unit, movements=moves, request=req)




@bp.route("/api/units/<int:uid>/trace.pdf")
@require_login
def unit_trace_pdf(uid):
    """Printable traceability card for one component unit."""
    p = ph()
    u = current_user()
    with db() as conn:
        unit = one(conn, f"SELECT * FROM stock_units WHERE id={p}", (uid,))
        if not unit:
            return err("Одиницю не знайдено", 404)
        req = one(conn, f"SELECT * FROM requests WHERE id={p}", (unit.get("request_id"),)) if unit.get("request_id") else None
        patient = one(conn, f"SELECT * FROM patients WHERE id={p}", (req.get("patient_id"),)) if req and req.get("patient_id") else None
        if not is_staff(u):
            if not req or not user_can_access_request(conn, u, req["id"]):
                return err("Недостатньо прав", 403)
        moves = all_rows(conn, f"""SELECT m.*, usr.full_name user_name, usr.role user_role
            FROM movements m
            LEFT JOIN users usr ON usr.id=m.user_id
            WHERE m.unit_id={p}
            ORDER BY m.created_at, m.id""", (uid,))

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    st = pdf_font_status()
    font = st.get("font_name") or "Helvetica"
    if st.get("chosen"):
        pdfmetrics.registerFont(TTFont("AppFont", st["chosen"]))

    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="UA", fontName=font, fontSize=9, leading=11))
    styles.add(ParagraphStyle(name="UAHead", fontName=font, fontSize=14, leading=16, spaceAfter=8, alignment=1))
    styles.add(ParagraphStyle(name="UASmall", fontName=font, fontSize=8, leading=10))

    def P(v, small=False):
        return Paragraph(html.escape(str(v if v is not None else "")), styles["UASmall" if small else "UA"])
    def kv_rows(items):
        return [[P(k), P(v)] for k, v in items]
    def add_table(story, title, headers, rows, widths=None):
        story.append(Paragraph(title, styles["UA"]))
        story.append(Spacer(1, 2*mm))
        body = [[P(h) for h in headers]]
        for row in rows:
            body.append([P(x, small=True) for x in row])
        t = Table(body, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('FONTNAME',(0,0),(-1,-1),font),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#D9EAF7')),
            ('GRID',(0,0),(-1,-1),0.25,colors.grey),
            ('VALIGN',(0,0),(-1,-1),'TOP')
        ]))
        story.append(t); story.append(Spacer(1, 4*mm))

    story = [Paragraph(f"Картка простежуваності одиниці №{uid}", styles["UAHead"])]
    story.append(Paragraph(f"Сформовано: {datetime.now().isoformat(timespec='seconds')}", styles["UASmall"]))
    story.append(Spacer(1, 4*mm))

    unit_items = [
        ("ID одиниці", unit.get("id")),
        ("Код", unit.get("unit_code")),
        ("Серія", unit.get("series")),
        ("Компонент", unit.get("component_type")),
        ("Група / Rh", f"{unit.get('abo','')} {unit.get('rh','')}"),
        ("Джерело", unit.get("source")),
        ("Надійшло", unit.get("received_date")),
        ("Термін придатності", unit.get("expiry_date")),
        ("Статус", ua_status(unit.get("status"))),
        ("Прив’язана вимога", unit.get("request_id")),
        ("Реакція / примітка", reaction_label(unit.get("reaction_type"), unit.get("reaction_severity"), unit.get("reaction_note"))),
        ("Створено", unit.get("created_at")),
        ("Оновлено", unit.get("updated_at")),
    ]
    add_table(story, "Дані одиниці компонента", ["Поле", "Значення"], kv_rows(unit_items), [55*mm, 120*mm])

    if req:
        req_items = [
            ("Вимога №", req.get("id")),
            ("Статус вимоги", ua_status(req.get("status"))),
            ("Пацієнт", req.get("patient_name")),
            ("Дата народження", req.get("birth_date")),
            ("Проживання / адреса", req.get("address")),
            ("Відділення", req.get("department")),
            ("Діагноз", req.get("diagnosis")),
            ("Потрібно на дату", req.get("needed_date")),
            ("Лікар", f"{req.get('doctor_name','')} / {req.get('doctor_position','')}"),
            ("Медсестра", f"{req.get('nurse_name','')} / {req.get('nurse_position','')}"),
            ("Показання", req.get("indication")),
            ("Анамнез", req.get("transfusion_history")),
            ("Попередні реакції", req.get("reaction_note")),
        ]
        add_table(story, "Пов’язана вимога", ["Поле", "Значення"], kv_rows(req_items), [55*mm, 120*mm])
    if patient:
        pat_items = [
            ("ID пацієнта", patient.get("id")),
            ("ПІБ", patient.get("full_name")),
            ("Дата народження", patient.get("birth_date")),
            ("Адреса", patient.get("address")),
            ("Відділення", patient.get("department")),
            ("Діагноз", patient.get("diagnosis")),
            ("Група / Rh", f"{patient.get('abo','')} {patient.get('rh','')}"),
        ]
        add_table(story, "Картка пацієнта", ["Поле", "Значення"], kv_rows(pat_items), [55*mm, 120*mm])

    add_table(story, "Повний журнал рухів одиниці", ["Дата","Дія","К-сть","Користувач","Роль","Причина","Деталі"],
              [[m.get("created_at"), ua_action(m.get("action")), m.get("quantity"), m.get("user_name"), m.get("user_role"), m.get("reason"), m.get("details")] for m in moves],
              [30*mm,25*mm,12*mm,35*mm,20*mm,35*mm,35*mm])

    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("Підпис трансфузіолога: ______________________    Підпис відповідальної особи: ______________________", styles["UA"]))
    doc.build(story)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"trace_unit_{uid}.pdf", mimetype="application/pdf")


@bp.route("/api/trash")
@require_login
@require_role("admin", "transfusion")
def trash():
    with db() as conn:
        reqs = all_rows(conn, "SELECT * FROM requests WHERE status='deleted' ORDER BY id DESC LIMIT 300")
        units = all_rows(conn, "SELECT * FROM stock_units WHERE status IN ('deleted','written_off','expired') ORDER BY id DESC LIMIT 300")
        moves = all_rows(conn, "SELECT * FROM movements WHERE deleted_at IS NOT NULL ORDER BY id DESC LIMIT 300")
    return ok(requests=reqs, units=units, movements=moves)



def report_dataset(conn, start="", end=""):
    params = []
    filt = "WHERE m.deleted_at IS NULL" + date_filter_sql("m.created_at", start, end, params)
    movements = all_rows(conn, f"SELECT m.action, COUNT(*) qty FROM movements m {filt} GROUP BY m.action ORDER BY m.action", tuple(params))
    params2 = []
    filt2 = "WHERE m.deleted_at IS NULL" + date_filter_sql("m.created_at", start, end, params2)
    by_component = all_rows(conn, f"""SELECT COALESCE(s.component_type,'') component_type, COALESCE(s.abo,'') abo, COALESCE(s.rh,'') rh, m.action, COUNT(*) qty
        FROM movements m LEFT JOIN stock_units s ON s.id=m.unit_id {filt2}
        GROUP BY COALESCE(s.component_type,''), COALESCE(s.abo,''), COALESCE(s.rh,''), m.action
        ORDER BY component_type, abo, rh, m.action""", tuple(params2))
    # Net views make returns clear: journal facts remain unchanged, while net issue subtracts returns.
    def net_rows(rows, key_fields):
        acc = {}
        for row in rows:
            key = tuple(row.get(k, "") for k in key_fields)
            item = acc.setdefault(key, {k: row.get(k, "") for k in key_fields} | {"income":0,"issue":0,"return":0,"used":0,"writeoff":0,"expired":0,"delete":0,"restore":0})
            a = row.get("action", "")
            if a in item:
                item[a] += int(row.get("qty") or 0)
        out = []
        for item in acc.values():
            item["issue_net"] = max(int(item.get("issue",0)) - int(item.get("return",0)), 0)
            out.append(item)
        return out
    net_movements = net_rows(movements, [])
    net_by_component = net_rows(by_component, ["component_type","abo","rh"])
    params3 = []
    filt3 = "WHERE m.deleted_at IS NULL" + date_filter_sql("m.created_at", start, end, params3)
    day_expr = sql_date("m.created_at")
    daily = all_rows(conn, f"SELECT {day_expr} day, m.action, COUNT(*) qty FROM movements m {filt3} GROUP BY {day_expr}, m.action ORDER BY day DESC, m.action", tuple(params3))
    params4 = []
    filt4 = "WHERE m.deleted_at IS NULL" + date_filter_sql("m.created_at", start, end, params4)
    details = all_rows(conn, f"""SELECT m.id,m.created_at,m.action,m.quantity,m.unit_id,m.request_id,m.user_id,m.reason,m.details,
        COALESCE(s.component_type,'') component_type, COALESCE(s.abo,'') abo, COALESCE(s.rh,'') rh, COALESCE(s.unit_code,'') unit_code, COALESCE(s.series,'') series
        FROM movements m LEFT JOIN stock_units s ON s.id=m.unit_id {filt4} ORDER BY m.id""", tuple(params4))
    stock = all_rows(conn, "SELECT status, COUNT(*) qty FROM stock_units GROUP BY status ORDER BY status")
    reqs = all_rows(conn, "SELECT status, COUNT(*) qty FROM requests GROUP BY status ORDER BY status")
    adj_params = []
    adj_where = "WHERE deleted_at IS NULL"
    # Adjustment belongs to report when its correction period overlaps the selected report period.
    if start:
        adj_where += f" AND (period_end='' OR period_end IS NULL OR period_end >= {ph()})"; adj_params.append(start)
    if end:
        adj_where += f" AND (period_start='' OR period_start IS NULL OR period_start <= {ph()})"; adj_params.append(end)
    adjustments = all_rows(conn, f"SELECT * FROM report_adjustments {adj_where} ORDER BY id DESC LIMIT 500", tuple(adj_params))

    def adjusted_rows(net_rows_in, key_fields):
        acc = {}
        for row in net_rows_in:
            key = tuple(row.get(k, "") for k in key_fields)
            acc[key] = dict(row)
            acc[key]["adjustment"] = 0
        for adj in adjustments:
            key = tuple(adj.get(k, "") for k in key_fields)
            item = acc.setdefault(key, {k: adj.get(k, "") for k in key_fields} | {"income":0,"issue":0,"return":0,"issue_net":0,"used":0,"writeoff":0,"expired":0,"adjustment":0})
            delta = int(adj.get("quantity_delta") or 0)
            action = adj.get("action") or "adjustment"
            if action in item:
                item[action] = int(item.get(action) or 0) + delta
            item["adjustment"] = int(item.get("adjustment") or 0) + delta
        out = []
        for item in acc.values():
            item["issue_net"] = max(int(item.get("issue",0)) - int(item.get("return",0)), 0)
            out.append(item)
        return out

    adjusted_net_movements = adjusted_rows(net_movements, [])
    adjusted_net_by_component = adjusted_rows(net_by_component, ["component_type","abo","rh"])
    return {"period":{"start":start,"end":end}, "movements":movements, "net_movements":net_movements, "adjusted_net_movements": adjusted_net_movements, "by_component":by_component, "net_by_component":net_by_component, "adjusted_net_by_component": adjusted_net_by_component, "daily":daily, "details":details, "stock":stock, "requests":reqs, "adjustments":adjustments}


def ua_action(a):
    return {"income":"Надходження", "issue":"Видача", "issue_net":"Видача нетто", "used":"Використано", "writeoff":"Списання", "expired":"Протерміновано", "delete":"Видалення", "restore":"Відновлення", "return":"Повернення"}.get(a, a or "")


def ua_status(a):
    return {"created":"Створена", "approved":"Погоджена", "reserved":"Зарезервована", "issued":"Видана", "partial_issued":"Частково видана", "used":"Використана", "rejected":"Відмовлена", "deleted":"Видалена", "in_stock":"На складі", "written_off":"Списана", "expired":"Протермінована"}.get(a, a or "")


def report_filename(ext, start="", end=""):
    period = (start or "all") + "_" + (end or "today")
    return f"blood_bank_report_{period}.{ext}"



@bp.route("/api/requests/<int:rid>/print.pdf")
@require_login
def request_print_pdf(rid):
    """Single request printable PDF with patient/request/issue trace data."""
    p = ph()
    u = current_user()
    with db() as conn:
        r = one(conn, f"SELECT * FROM requests WHERE id={p}", (rid,))
        if not r:
            return err("Вимогу не знайдено", 404)
        if not user_can_access_request(conn, u, rid):
            return err("Недостатньо прав", 403)
        patient = one(conn, f"SELECT * FROM patients WHERE id={p}", (r.get("patient_id"),)) if r.get("patient_id") else None
        units = all_rows(conn, f"""SELECT su.*, 
                (SELECT GROUP_CONCAT(m.action || ' ' || COALESCE(m.created_at,''), '; ') FROM movements m WHERE m.unit_id=su.id) movement_summary
            FROM stock_units su
            WHERE su.request_id={p}
            ORDER BY su.id""", (rid,))
        moves = all_rows(conn, f"""SELECT m.*, su.unit_code, su.series, su.component_type, su.abo, su.rh, u.full_name user_name
            FROM movements m
            LEFT JOIN stock_units su ON su.id=m.unit_id
            LEFT JOIN users u ON u.id=m.user_id
            WHERE m.request_id={p} OR su.request_id={p}
            ORDER BY m.created_at, m.id""", (rid, rid))
        progress = request_progress(conn, rid)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    st = pdf_font_status()
    font = st.get("font_name") or "Helvetica"
    if st.get("chosen"):
        pdfmetrics.registerFont(TTFont("AppFont", st["chosen"]))

    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="UA", fontName=font, fontSize=9, leading=11))
    styles.add(ParagraphStyle(name="UAHead", fontName=font, fontSize=14, leading=16, spaceAfter=8, alignment=1))
    styles.add(ParagraphStyle(name="UASmall", fontName=font, fontSize=8, leading=10))
    def P(v, small=False):
        return Paragraph(html.escape(str(v if v is not None else "")), styles["UASmall" if small else "UA"])
    def kv_rows(items):
        return [[P(k), P(v)] for k, v in items]
    def add_table(story, title, headers, rows, widths=None):
        story.append(Paragraph(title, styles["UA"]))
        story.append(Spacer(1, 2*mm))
        body = [[P(h) for h in headers]]
        for row in rows:
            body.append([P(x, small=True) for x in row])
        t = Table(body, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('FONTNAME',(0,0),(-1,-1),font),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#D9EAF7')),
            ('GRID',(0,0),(-1,-1),0.25,colors.grey),
            ('VALIGN',(0,0),(-1,-1),'TOP')
        ]))
        story.append(t); story.append(Spacer(1, 4*mm))

    story = [Paragraph(f"Вимога на компоненти крові №{rid}", styles["UAHead"])]
    story.append(Paragraph(f"Сформовано: {datetime.now().isoformat(timespec='seconds')}", styles["UASmall"]))
    story.append(Spacer(1, 4*mm))

    req_items = [
        ("Статус", ua_status(r.get("status"))),
        ("Пацієнт", r.get("patient_name")),
        ("Дата народження", r.get("birth_date")),
        ("Проживання / адреса", r.get("address")),
        ("Відділення", r.get("department")),
        ("Діагноз", r.get("diagnosis")),
        ("Статус хворого", r.get("patient_status")),
        ("Терміновість", ua_status(r.get("urgency"))),
        ("Показання", r.get("indication")),
        ("Трансфузійний анамнез", r.get("transfusion_history")),
        ("Попередні реакції / алергії", r.get("reaction_note")),
        ("Компонент", r.get("component_type")),
        ("Група / Rh", f"{r.get('abo','')} {r.get('rh','')}"),
        ("Кількість", r.get("quantity")),
        ("Потрібно на дату", r.get("needed_date")),
        ("Лікар", f"{r.get('doctor_name','')} / {r.get('doctor_position','')}"),
        ("Медсестра", f"{r.get('nurse_name','')} / {r.get('nurse_position','')}"),
        ("Примітка", r.get("request_note")),
        ("Створено", r.get("created_at")),
        ("Оновлено", r.get("updated_at")),
    ]
    add_table(story, "Дані вимоги", ["Поле", "Значення"], kv_rows(req_items), [55*mm, 120*mm])

    if patient:
        pat_items = [
            ("ID пацієнта", patient.get("id")),
            ("ПІБ", patient.get("full_name")),
            ("Дата народження", patient.get("birth_date")),
            ("Проживання / адреса", patient.get("address")),
            ("Відділення", patient.get("department")),
            ("Діагноз", patient.get("diagnosis")),
            ("Група / Rh", f"{patient.get('abo','')} {patient.get('rh','')}"),
            ("Примітка", patient.get("note")),
        ]
        add_table(story, "Картка пацієнта", ["Поле", "Значення"], kv_rows(pat_items), [55*mm, 120*mm])

    prog_items = [
        ("Потрібно", r.get("quantity")),
        ("Видано зараз", progress.get("issued_count")),
        ("Використано", progress.get("used_count")),
        ("Видано разом", progress.get("delivered_count")),
        ("Залишок до видачі", progress.get("remaining_quantity")),
    ]
    add_table(story, "Прогрес виконання", ["Показник", "Значення"], kv_rows(prog_items), [55*mm, 120*mm])

    add_table(story, "Одиниці, прив’язані до вимоги", ["ID","Код","Серія","Компонент","Група","Rh","Статус","Термін","Реакція"],
              [[x.get("id"), x.get("unit_code"), x.get("series"), x.get("component_type"), x.get("abo"), x.get("rh"), ua_status(x.get("status")), x.get("expiry_date"), x.get("reaction_note")] for x in units],
              [12*mm,25*mm,24*mm,35*mm,18*mm,14*mm,24*mm,22*mm,35*mm])

    add_table(story, "Журнал рухів по вимозі", ["Дата","Дія","Одиниця","Серія","К-сть","Користувач","Деталі/причина"],
              [[m.get("created_at"), ua_action(m.get("action")), m.get("unit_code") or m.get("unit_id"), m.get("series"), m.get("quantity"), m.get("user_name"), m.get("reason") or m.get("details")] for m in moves],
              [30*mm,25*mm,24*mm,22*mm,12*mm,32*mm,45*mm])

    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("Підпис лікаря: ______________________    Підпис трансфузіолога: ______________________", styles["UA"]))
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph("Підпис медсестри: ___________________    Дата/час: ________________________________", styles["UA"]))

    doc.build(story)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"blood_request_{rid}.pdf", mimetype="application/pdf")


@bp.route("/api/reports/preview")
@require_login
@require_role("admin", "transfusion")
def reports_preview():
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    with db() as conn:
        data = report_dataset(conn, start, end)
    return ok(**data)


@bp.route("/api/reports/adjustments", methods=["GET", "POST"])
@require_login
@require_role("admin", "transfusion")
@require_csrf
def report_adjustments():
    u = current_user(); p = ph(); j = getj()
    if request.method == "GET":
        with db() as conn:
            return ok(adjustments=all_rows(conn, "SELECT * FROM report_adjustments WHERE deleted_at IS NULL ORDER BY id DESC"))
    qty = parse_int_param(j.get("quantity_delta"), "Корекція кількості", required=True, min_value=-100000, max_value=100000)
    if not j.get("action") or qty == 0:
        return err("Вкажіть дію і корекцію кількості")
    with db() as conn:
        try:
            j["component_type"] = require_catalog_component(conn, j.get("component_type", "")) if j.get("component_type") else ""
        except ValueError as e:
            return err(str(e))
        j["abo"] = normalize_abo(j.get("abo", "")) if j.get("abo") else ""
        j["rh"] = normalize_rh(j.get("rh", "")) if j.get("rh") else ""
        exec_sql(conn, f"""INSERT INTO report_adjustments(period_start,period_end,component_type,abo,rh,action,quantity_delta,reason,user_id)
            VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p})""",
            (j.get("period_start", ""), j.get("period_end", ""), j.get("component_type", ""), j.get("abo", ""), j.get("rh", ""), j.get("action", ""), qty, j.get("reason", ""), u["id"]))
        audit(conn, "report_adjustment_add", "report_adjustments", None, f"{j.get('action')} {qty}: {j.get('reason','')}")
    return ok()


@bp.route("/api/reports/adjustments/<int:aid>", methods=["DELETE"])
@require_login
@require_role("admin", "transfusion")
@require_csrf
def report_adjustment_delete(aid):
    u = current_user(); p = ph()
    with db() as conn:
        exec_sql(conn, f"UPDATE report_adjustments SET deleted_at=CURRENT_TIMESTAMP, deleted_by={p} WHERE id={p}", (u["id"], aid))
        audit(conn, "report_adjustment_delete", "report_adjustments", aid)
    return ok()


@bp.route("/api/movements/<int:mid>", methods=["PUT", "DELETE"])
@require_login
@require_role("admin", "transfusion")
@require_csrf
def movement_update(mid):
    u = current_user(); p = ph(); j = getj()
    with db() as conn:
        m = one(conn, f"SELECT * FROM movements WHERE id={p}", (mid,))
        if not m:
            return err("Рух не знайдено", 404)
        if request.method == "DELETE":
            exec_sql(conn, f"UPDATE movements SET deleted_at=CURRENT_TIMESTAMP, deleted_by={p}, delete_reason={p} WHERE id={p}", (u["id"], j.get("reason", ""), mid))
            audit(conn, "movement_soft_delete", "movements", mid, j.get("reason", ""))
            return ok()
        new_m = {"reason": j.get("reason", m.get("reason", "")), "details": j.get("details", m.get("details", ""))}
        exec_sql(conn, f"UPDATE movements SET reason={p}, details={p} WHERE id={p}", (new_m["reason"], new_m["details"], mid))
        audit(conn, "movement_edit", "movements", mid, audit_diff(m, new_m, ["reason","details"]))
    return ok()


@bp.route("/api/reports/history/<action>", methods=["POST"])
@require_login
@require_role("admin")
@require_csrf
def report_history(action):
    u = current_user(); p = ph(); j = getj()
    start = j.get("start", ""); end = j.get("end", "")
    if not start or not end:
        return err("Вкажіть початок і кінець періоду")
    with db() as conn:
        created_day = sql_date("created_at")
        if action == "clear":
            exec_sql(conn, f"UPDATE movements SET deleted_at=CURRENT_TIMESTAMP, deleted_by={p}, delete_reason={p} WHERE deleted_at IS NULL AND {created_day}>={p} AND {created_day}<={p}", (u["id"], j.get("reason", "Очищення періоду"), start, end))
            audit(conn, "report_history_clear", "movements", None, f"{start}..{end}: {j.get('reason','')}")
            return ok()
        if action == "restore":
            exec_sql(conn, f"UPDATE movements SET deleted_at=NULL, restored_at=CURRENT_TIMESTAMP, restored_by={p} WHERE deleted_at IS NOT NULL AND {created_day}>={p} AND {created_day}<={p}", (u["id"], start, end))
            audit(conn, "report_history_restore", "movements", None, f"{start}..{end}")
            return ok()
    return err("Невідома дія")



@bp.route("/api/telegram/config", methods=["GET", "POST"])
@require_login
@require_role("admin", "transfusion")
@require_csrf
def telegram_config_api():
    u = current_user(); p = ph(); j = getj()
    with db() as conn:
        if request.method == "POST":
            if u["role"] != "admin":
                return err("Налаштування бота змінює тільки адміністратор", 403)
            current = one(conn, "SELECT * FROM telegram_config WHERE id=1") or {}
            token = (j.get("bot_token") or "").strip() or current.get("bot_token") or current_app.config.get("TELEGRAM_BOT_TOKEN", "")
            enabled = boolint(j.get("enabled", current.get("enabled", 0)))
            threshold = parse_int_param(j.get("critical_threshold"), "Поріг критичного залишку", default=int(current.get("critical_threshold") or 2), min_value=0, max_value=100000)
            exec_sql(conn, f"UPDATE telegram_config SET bot_token={p}, enabled={p}, critical_threshold={p}, updated_at=CURRENT_TIMESTAMP WHERE id=1", (token, enabled, threshold))
            audit(conn, "telegram_config", "telegram_config", 1, f"enabled={enabled}; threshold={threshold}")
        cfg = telegram_config(conn)
    return ok(config={"enabled": 1 if cfg.get("enabled") else 0, "has_bot_token": bool(cfg.get("bot_token")), "critical_threshold": cfg.get("critical_threshold")})


@bp.route("/api/telegram/me", methods=["GET", "POST"])
@require_login
@require_csrf
def telegram_me():
    u = current_user(); p = ph(); j = getj()
    fields = ["enabled","notify_new_request","notify_approve","notify_reject","notify_issue","notify_used","notify_critical","notify_expired","notify_system"]
    with db() as conn:
        sub = one(conn, f"SELECT * FROM telegram_subscribers WHERE user_id={p} ORDER BY id DESC LIMIT 1", (u["id"],))
        if request.method == "POST":
            chat_id = (j.get("chat_id") or (sub or {}).get("chat_id") or "").strip()
            vals = {k: boolint(j.get(k, (sub or {}).get(k, 1 if k!='enabled' else 0))) for k in fields}
            if sub:
                exec_sql(conn, f"""UPDATE telegram_subscribers SET chat_id={p}, enabled={p}, notify_new_request={p}, notify_approve={p}, notify_reject={p}, notify_issue={p}, notify_used={p}, notify_critical={p}, notify_expired={p}, notify_system={p}, updated_at=CURRENT_TIMESTAMP WHERE id={p}""",
                         (chat_id, vals['enabled'], vals['notify_new_request'], vals['notify_approve'], vals['notify_reject'], vals['notify_issue'], vals['notify_used'], vals['notify_critical'], vals['notify_expired'], vals['notify_system'], sub['id']))
            else:
                exec_sql(conn, f"""INSERT INTO telegram_subscribers(user_id,chat_id,enabled,notify_new_request,notify_approve,notify_reject,notify_issue,notify_used,notify_critical,notify_expired,notify_system) VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p})""",
                         (u['id'], chat_id, vals['enabled'], vals['notify_new_request'], vals['notify_approve'], vals['notify_reject'], vals['notify_issue'], vals['notify_used'], vals['notify_critical'], vals['notify_expired'], vals['notify_system']))
            audit(conn, "telegram_me", "telegram_subscribers", u["id"], chat_id)
        sub = one(conn, f"SELECT * FROM telegram_subscribers WHERE user_id={p} ORDER BY id DESC LIMIT 1", (u["id"],)) or {"user_id": u["id"], "chat_id": "", "enabled": 0, **{k:1 for k in fields if k!='enabled'}}
    return ok(subscriber=sub)


@bp.route("/api/telegram/test", methods=["POST"])
@require_login
@require_csrf
def telegram_test():
    u = current_user(); p = ph(); j = getj()
    with db() as conn:
        sub = one(conn, f"SELECT * FROM telegram_subscribers WHERE user_id={p} ORDER BY id DESC LIMIT 1", (u["id"],))
        chat_id = (j.get("chat_id") or (sub or {}).get("chat_id") or "").strip()
        if not chat_id:
            return err("Вкажіть Telegram chat_id")
        ok_send = telegram_send_raw(conn, "test", chat_id, f"🧪 Тестове повідомлення банку крові\\nКористувач: {u['full_name']}\\nВерсія: {current_app.config['VERSION']}")
        audit(conn, "telegram_test", "telegram_subscribers", u["id"], chat_id)
    return ok(sent=ok_send)


@bp.route("/api/telegram/log")
@require_login
@require_role("admin", "transfusion")
def telegram_log_api():
    with db() as conn:
        rows = all_rows(conn, "SELECT * FROM telegram_log ORDER BY id DESC LIMIT 200")
    return ok(log=rows)


def temperature_status(temp, min_temp, max_temp):
    try:
        t = float(temp); mn = float(min_temp); mx = float(max_temp)
    except Exception:
        return "alarm"
    if t < mn:
        return "low"
    if t > mx:
        return "high"
    return "ok"


@bp.route("/api/temperature/devices", methods=["GET", "POST"])
@require_login
@require_csrf
def temperature_devices_api():
    u = current_user(); p = ph()
    if request.method == "GET":
        if u["role"] not in ["admin", "transfusion", "nurse"]:
            return err("Недостатньо прав", 403)
        with db() as conn:
            rows = all_rows(conn, "SELECT * FROM temperature_devices WHERE deleted_at IS NULL ORDER BY active DESC, name")
        return ok(devices=rows)
    if u["role"] not in ["admin", "transfusion"]:
        return err("Недостатньо прав", 403)
    j = getj()
    name = (j.get("name") or "").strip()
    if not name:
        return err("Назва холодильника/морозильника обов’язкова")
    min_temp = parse_float_param(j.get("min_temp"), "Мінімальна температура", default=2)
    max_temp = parse_float_param(j.get("max_temp"), "Максимальна температура", default=6)
    if min_temp >= max_temp:
        return err("Мінімальна температура має бути меншою за максимальну")
    with db() as conn:
        device_id = insert_and_get_id(conn, "temperature_devices",
            ["name", "device_type", "location", "min_temp", "max_temp", "active"],
            [name, j.get("device_type", "fridge"), j.get("location", ""), min_temp, max_temp, boolint(j.get("active", 1))])
        audit(conn, "temperature_device_create", "temperature_devices", device_id, name)
    return ok(device={"id": device_id})


@bp.route("/api/temperature/devices/<int:did>", methods=["PUT", "DELETE"])
@require_login
@require_role("admin", "transfusion")
@require_csrf
def temperature_device_item(did):
    u = current_user(); p = ph(); j = getj()
    with db() as conn:
        dev = one(conn, f"SELECT * FROM temperature_devices WHERE id={p}", (did,))
        if not dev:
            return err("Пристрій не знайдено", 404)
        if request.method == "DELETE":
            exec_sql(conn, f"UPDATE temperature_devices SET deleted_at=CURRENT_TIMESTAMP, deleted_by={p}, delete_reason={p}, active=0 WHERE id={p}", (u["id"], j.get("reason", ""), did))
            audit(conn, "temperature_device_delete", "temperature_devices", did, j.get("reason", ""))
            return ok()
        min_temp = parse_float_param(j.get("min_temp"), "Мінімальна температура", default=float(dev.get("min_temp") or 2))
        max_temp = parse_float_param(j.get("max_temp"), "Максимальна температура", default=float(dev.get("max_temp") or 6))
        if min_temp >= max_temp:
            return err("Мінімальна температура має бути меншою за максимальну")
        exec_sql(conn, f"""UPDATE temperature_devices SET name={p}, device_type={p}, location={p}, min_temp={p}, max_temp={p}, active={p}, updated_at=CURRENT_TIMESTAMP WHERE id={p}""",
                 (j.get("name", dev.get("name")), j.get("device_type", dev.get("device_type")), j.get("location", dev.get("location")), min_temp, max_temp, boolint(j.get("active", dev.get("active", 1))), did))
        audit(conn, "temperature_device_update", "temperature_devices", did)
    return ok()


@bp.route("/api/temperature/readings", methods=["GET", "POST"])
@require_login
@require_csrf
def temperature_readings_api():
    u = current_user(); p = ph()
    if request.method == "GET":
        if u["role"] not in ["admin", "transfusion", "nurse"]:
            return err("Недостатньо прав", 403)
        start = request.args.get("start", ""); end = request.args.get("end", ""); device_id = parse_int_param(request.args.get("device_id"), "device_id", min_value=1)
        params=[]; where="WHERE r.deleted_at IS NULL"
        if device_id:
            where += f" AND r.device_id={p}"; params.append(device_id)
        where += date_filter_sql("r.measured_at", start, end, params)
        with db() as conn:
            rows = all_rows(conn, f"""SELECT r.*, d.name device_name, d.device_type, d.location, d.min_temp, d.max_temp, u.full_name user_name
                FROM temperature_readings r
                JOIN temperature_devices d ON d.id=r.device_id
                LEFT JOIN users u ON u.id=r.user_id
                {where} ORDER BY r.measured_at DESC, r.id DESC LIMIT 500""", tuple(params))
            summary = all_rows(conn, """SELECT d.name device_name, r.status, COUNT(*) qty
                FROM temperature_readings r JOIN temperature_devices d ON d.id=r.device_id
                WHERE r.deleted_at IS NULL GROUP BY d.name, r.status ORDER BY d.name, r.status""")
        return ok(readings=rows, summary=summary)
    if u["role"] not in ["admin", "transfusion", "nurse"]:
        return err("Недостатньо прав для внесення температури", 403)
    j = getj()
    device_id = parse_int_param(j.get("device_id"), "device_id", required=True, min_value=1)
    temp = parse_float_param(j.get("temperature"), "Температура", required=True)
    humidity = parse_float_param(j.get("humidity"), "Вологість", default=None, min_value=0, max_value=100) if j.get("humidity") not in (None, "") else None
    measured_at = j.get("measured_at") or datetime.now().isoformat(timespec="seconds")
    with db() as conn:
        dev = one(conn, f"SELECT * FROM temperature_devices WHERE id={p} AND deleted_at IS NULL", (device_id,))
        if not dev:
            return err("Пристрій не знайдено", 404)
        status = temperature_status(temp, dev.get("min_temp"), dev.get("max_temp"))
        exec_sql(conn, f"""INSERT INTO temperature_readings(device_id,measured_at,temperature,humidity,status,note,user_id)
            VALUES ({p},{p},{p},{p},{p},{p},{p})""", (device_id, measured_at, temp, humidity, status, j.get("note", ""), u["id"]))
        audit(conn, "temperature_reading_create", "temperature_readings", None, f"{dev.get('name')}: {temp} °C ({status})")
        if status != "ok":
            telegram_notify(conn, "temperature", f"🌡️ Порушення температури\nПристрій: {dev.get('name')}\nМісце: {dev.get('location') or '—'}\nТемпература: {temp} °C\nНорма: {dev.get('min_temp')}…{dev.get('max_temp')} °C\nСтатус: {'нижче норми' if status=='low' else 'вище норми'}", role_filter=["admin", "transfusion"])
    return ok(status=status)


@bp.route("/api/temperature/readings/<int:rid>", methods=["DELETE"])
@require_login
@require_role("admin", "transfusion")
@require_csrf
def temperature_reading_delete(rid):
    u = current_user(); p = ph(); j = getj()
    with db() as conn:
        exec_sql(conn, f"UPDATE temperature_readings SET deleted_at=CURRENT_TIMESTAMP, deleted_by={p}, delete_reason={p} WHERE id={p}", (u["id"], j.get("reason", ""), rid))
        audit(conn, "temperature_reading_delete", "temperature_readings", rid, j.get("reason", ""))
    return ok()


@bp.route("/api/temperature/export.csv")
@require_login
@require_role("admin", "transfusion", "nurse")
def temperature_export_csv():
    start = request.args.get("start", ""); end = request.args.get("end", ""); device_id = parse_int_param(request.args.get("device_id"), "device_id", min_value=1)
    params=[]; where="WHERE r.deleted_at IS NULL"
    if device_id:
        where += f" AND r.device_id={ph()}"; params.append(device_id)
    where += date_filter_sql("r.measured_at", start, end, params)
    with db() as conn:
        rows = all_rows(conn, f"""SELECT r.measured_at,d.name device_name,d.location,r.temperature,r.humidity,r.status,r.note,u.full_name user_name
            FROM temperature_readings r JOIN temperature_devices d ON d.id=r.device_id LEFT JOIN users u ON u.id=r.user_id {where}
            ORDER BY r.measured_at DESC, r.id DESC""", tuple(params))
    def clean(v): return '"'+str(v if v is not None else '').replace('"','""')+'"'
    header=["measured_at","device_name","location","temperature","humidity","status","note","user_name"]
    body=[','.join(header)] + [','.join(clean(r.get(h,'')) for h in header) for r in rows]
    return Response('\ufeff'+'\n'.join(body), mimetype='text/csv; charset=utf-8', headers={"Content-Disposition":"attachment; filename=temperature_journal.csv"})


@bp.route("/api/temperature/export.xlsx")
@require_login
@require_role("admin", "transfusion", "nurse")
def temperature_export_xlsx():
    start = request.args.get("start", ""); end = request.args.get("end", ""); device_id = parse_int_param(request.args.get("device_id"), "device_id", min_value=1)
    params=[]; where="WHERE r.deleted_at IS NULL"
    if device_id:
        where += f" AND r.device_id={ph()}"; params.append(device_id)
    where += date_filter_sql("r.measured_at", start, end, params)
    with db() as conn:
        rows = all_rows(conn, f"""SELECT r.measured_at,d.name device_name,d.device_type,d.location,d.min_temp,d.max_temp,r.temperature,r.humidity,r.status,r.note,u.full_name user_name
            FROM temperature_readings r JOIN temperature_devices d ON d.id=r.device_id LEFT JOIN users u ON u.id=r.user_id {where}
            ORDER BY r.measured_at DESC, r.id DESC""", tuple(params))
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title="Температурний журнал"
    ws.append(["Температурний журнал банку крові", "", "", "", "Версія", current_app.config["VERSION"]])
    ws.append(["Період", start or "початок", end or "сьогодні"])
    ws.append([])
    headers=["Дата/час", "Пристрій", "Тип", "Місце", "Мін °C", "Макс °C", "Темп. °C", "Вологість", "Статус", "Примітка", "Користувач"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get("measured_at"), r.get("device_name"), r.get("device_type"), r.get("location"), r.get("min_temp"), r.get("max_temp"), r.get("temperature"), r.get("humidity"), r.get("status"), r.get("note"), r.get("user_name")])
    thin=Side(style="thin", color="9CA3AF"); border=Border(left=thin,right=thin,top=thin,bottom=thin); fill=PatternFill("solid", fgColor="D9EAF7")
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border=border; cell.alignment=Alignment(wrap_text=True, vertical="top")
            if cell.row==4: cell.font=Font(bold=True); cell.fill=fill; cell.alignment=Alignment(horizontal="center")
    ws["A1"].font=Font(bold=True, size=14); ws.merge_cells("A1:D1")
    for c in range(1, len(headers)+1): ws.column_dimensions[get_column_letter(c)].width=18
    ws.freeze_panes="A5"
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="temperature_journal.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



@bp.route("/api/login-events")
@require_login
@require_role("admin", "transfusion")
def login_events_api():
    p = ph()
    limit = parse_int_param(request.args.get("limit"), "limit", default=100, min_value=1, max_value=500)
    success = (request.args.get("success") or "").strip()
    username = (request.args.get("username") or "").strip()
    where = ["1=1"]
    params = []
    if success in ("0", "1"):
        where.append(f"success={p}"); params.append(int(success))
    if username:
        where.append(f"LOWER(username) LIKE LOWER({p})"); params.append(f"%{username}%")
    with db() as conn:
        rows = all_rows(conn, f"""SELECT id, username, user_id, role, full_name, ip, user_agent, success, reason, created_at
            FROM login_events WHERE {' AND '.join(where)}
            ORDER BY created_at DESC, id DESC LIMIT {p}""", tuple(params + [limit]))
    return ok(events=rows)


@bp.route("/api/audit")
@require_login
@require_role("admin", "transfusion")
def audit_api():
    p = ph()
    limit = parse_int_param(request.args.get("limit"), "limit", default=200, min_value=1, max_value=500)
    action = (request.args.get("action") or "").strip()
    entity = (request.args.get("entity") or "").strip()
    user_id = parse_int_param(request.args.get("user_id"), "user_id", min_value=1)
    where = ["1=1"]
    params = []
    if action:
        where.append(f"a.action LIKE {p}"); params.append(f"%{action}%")
    if entity:
        where.append(f"a.entity LIKE {p}"); params.append(f"%{entity}%")
    if user_id:
        where.append(f"a.user_id={p}"); params.append(user_id)
    with db() as conn:
        sql = f"""SELECT a.*, u.full_name AS user_name, u.username
            FROM audit_log a LEFT JOIN users u ON u.id=a.user_id
            WHERE {' AND '.join(where)} ORDER BY a.id DESC LIMIT {limit}"""
        rows = all_rows(conn, sql, tuple(params))
    return ok(audit=rows)



def list_backup_files():
    os.makedirs(backup_dir_path(), exist_ok=True)
    files = []
    for name in os.listdir(backup_dir_path()):
        if not name.endswith(".json"):
            continue
        path = os.path.join(backup_dir_path(), name)
        try:
            st = os.stat(path)
            files.append({"name": name, "path": path, "mtime": st.st_mtime, "size": st.st_size})
        except OSError:
            pass
    return sorted(files, key=lambda x: x["mtime"], reverse=True)

def prune_backups(conn, keep_last=14):
    keep = max(int(keep_last or 14), 1)
    removed = []
    files = list_backup_files()
    for f in files[keep:]:
        try:
            os.remove(f["path"])
            removed.append(f["name"])
        except Exception:
            pass
    if removed:
        backup_log(conn, "prune", "", 1, "", json.dumps({"removed": removed, "keep_last": keep}, ensure_ascii=False))
    return removed

def create_backup_file(conn, prefix="blood_bank_v7_backup", include_secrets=False, action="create"):
    os.makedirs(backup_dir_path(), exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    fname = f"{prefix}_{ts}.json"
    path = os.path.join(backup_dir_path(), fname)
    payload = build_backup_payload(conn, include_secrets=include_secrets)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    audit(conn, action, "backup", None, fname)
    backup_log(conn, action, fname, 1, "", json.dumps(payload.get("manifest", {}), ensure_ascii=False))
    return fname, payload

def api_token_valid():
    token = current_app.config.get("API_TOKEN")
    if not token:
        return False, "API_TOKEN не налаштований"
    supplied = request.args.get("token") or request.headers.get("X-API-Token") or ""
    if supplied != token:
        return False, "API_TOKEN неправильний"
    return True, ""

def perform_auto_backup(conn, *, force=False, source="auto_backup"):
    pol = one(conn, "SELECT * FROM backup_policy WHERE id=1")
    if not pol:
        exec_sql(conn, "INSERT INTO backup_policy(id, enabled, keep_last) VALUES (1,0,14)")
        pol = one(conn, "SELECT * FROM backup_policy WHERE id=1")
    if not force and not boolint(pol.get("enabled")):
        backup_log(conn, source, "", 0, "auto-backup disabled", "")
        return None, None, [], pol, "disabled"
    fname, payload = create_backup_file(conn, prefix=source, include_secrets=False, action=source)
    removed = prune_backups(conn, int(pol.get("keep_last") or 14))
    p = ph()
    exec_sql(conn, f"UPDATE backup_policy SET last_run_at=CURRENT_TIMESTAMP, last_file={p}, last_ok=1, last_error='', updated_at=CURRENT_TIMESTAMP WHERE id=1", (fname,))
    pol = one(conn, "SELECT * FROM backup_policy WHERE id=1")
    return fname, payload, removed, pol, ""


@bp.route("/api/backup/create", methods=["POST"])
@require_login
@require_role("admin")
@require_csrf
def backup_create():
    fname = ""
    try:
        with db() as conn:
            fname, payload = create_backup_file(conn, prefix="blood_bank_v7_backup", include_secrets=(request.args.get("include_secrets") == "1"), action="backup_create")
            pol = one(conn, "SELECT * FROM backup_policy WHERE id=1") or {"keep_last": 14}
            removed = prune_backups(conn, int(pol.get("keep_last") or 14))
        return ok(file=fname, manifest=payload.get("manifest", {}), pruned=removed)
    except Exception as e:
        with db() as conn:
            backup_log(conn, "backup_create", fname, 0, str(e), "")
        return err("Backup error: " + str(e), 500)



@bp.route("/api/backup/policy", methods=["GET", "PUT"])
@require_login
@require_role("admin")
@require_csrf
def backup_policy_api():
    if request.method == "GET":
        with db() as conn:
            pol = one(conn, "SELECT * FROM backup_policy WHERE id=1")
            if not pol:
                exec_sql(conn, "INSERT INTO backup_policy(id, enabled, keep_last) VALUES (1,0,14)")
                pol = one(conn, "SELECT * FROM backup_policy WHERE id=1")
        return ok(policy=pol)
    j = getj()
    keep_last = parse_int_param(j.get("keep_last"), "keep_last", default=14, min_value=1, max_value=100)
    enabled = 1 if boolint(j.get("enabled")) else 0
    with db() as conn:
        p = ph()
        exec_sql(conn, f"""UPDATE backup_policy SET enabled={p}, keep_last={p}, updated_at=CURRENT_TIMESTAMP WHERE id=1""", (enabled, keep_last))
        if not one(conn, "SELECT * FROM backup_policy WHERE id=1"):
            exec_sql(conn, "INSERT INTO backup_policy(id, enabled, keep_last) VALUES (1,0,14)")
        audit(conn, "backup_policy_update", "backup_policy", 1, json.dumps({"enabled": enabled, "keep_last": keep_last}, ensure_ascii=False))
        pol = one(conn, "SELECT * FROM backup_policy WHERE id=1")
    return ok(policy=pol)

@bp.route("/api/backup/auto-run", methods=["POST"])
@require_login
@require_role("admin")
@require_csrf
def backup_auto_run():
    fname = ""
    try:
        with db() as conn:
            fname, payload, removed, pol, skipped = perform_auto_backup(conn, force=True, source="auto_backup")
        return ok(file=fname, manifest=payload.get("manifest", {}) if payload else {}, pruned=removed, policy=pol, skipped=skipped)
    except Exception as e:
        with db() as conn:
            p = ph()
            exec_sql(conn, f"UPDATE backup_policy SET last_run_at=CURRENT_TIMESTAMP, last_file={p}, last_ok=0, last_error={p}, updated_at=CURRENT_TIMESTAMP WHERE id=1", (fname, str(e)[:500]))
            backup_log(conn, "auto_backup", fname, 0, str(e), "")
        return err("Auto backup error: " + str(e), 500)



@bp.route("/api/cron/auto-backup", methods=["GET", "POST"])
def cron_auto_backup():
    valid, msg = api_token_valid()
    if not valid:
        return err(msg, 403 if "неправильний" in msg else 503)
    fname = ""
    try:
        with db() as conn:
            fname, payload, removed, pol, skipped = perform_auto_backup(conn, force=False, source="cron_auto_backup")
        if skipped == "disabled":
            return ok(skipped="disabled", message="Auto-backup вимкнений у backup_policy", policy=pol)
        return ok(file=fname, manifest=payload.get("manifest", {}) if payload else {}, pruned=removed, policy=pol)
    except Exception as e:
        with db() as conn:
            p = ph()
            exec_sql(conn, f"UPDATE backup_policy SET last_run_at=CURRENT_TIMESTAMP, last_file={p}, last_ok=0, last_error={p}, updated_at=CURRENT_TIMESTAMP WHERE id=1", (fname, str(e)[:500]))
            backup_log(conn, "cron_auto_backup", fname, 0, str(e), "")
        return err("Cron auto backup error: " + str(e), 500)


@bp.route("/api/backups")
@require_login
@require_role("admin")
def backups_list():
    os.makedirs(backup_dir_path(), exist_ok=True)
    rows=[]
    for name in sorted(os.listdir(backup_dir_path()), reverse=True):
        if not name.endswith(".json"):
            continue
        path=os.path.join(backup_dir_path(), name)
        st=os.stat(path)
        meta={}
        try:
            with open(path, "r", encoding="utf-8") as f:
                meta=json.load(f).get("manifest", {})
        except Exception:
            meta={"warning":"Не вдалося прочитати manifest"}
        rows.append({"file":name,"size":st.st_size,"created_at":datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),"manifest":meta})
    return ok(backups=rows)


@bp.route("/api/backups/<path:name>")
@require_login
@require_role("admin")
def backup_download(name):
    if "/" in name or "\\" in name or not name.endswith(".json"):
        return err("Некоректна назва файлу", 400)
    return send_file(os.path.join(backup_dir_path(), name), as_attachment=True, download_name=name)


@bp.route("/api/backup/verify", methods=["POST"])
@require_login
@require_role("admin")
@require_csrf
def backup_verify():
    file = request.files.get("file")
    if not file:
        return err("Файл backup не передано")
    try:
        payload = json.loads(file.read().decode("utf-8"))
        valid, msg = validate_backup_payload(payload)
        manifest = payload.get("manifest", {}) if isinstance(payload, dict) else {}
        return ok(valid=valid, message=msg, manifest=manifest)
    except Exception as e:
        return err("Не вдалося прочитати backup: " + str(e), 400)


@bp.route("/api/backup/restore", methods=["POST"])
@require_login
@require_role("admin")
@require_csrf
def backup_restore():
    file = request.files.get("file")
    confirm = request.form.get("confirm", "")
    if confirm != "RESTORE":
        return err("Для відновлення введіть RESTORE")
    if not file:
        return err("Файл backup не передано")
    name = file.filename or "uploaded_backup.json"
    try:
        payload = json.loads(file.read().decode("utf-8"))
        os.makedirs(backup_dir_path(), exist_ok=True)
        emergency_name = f"emergency_before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        emergency_path = os.path.join(backup_dir_path(), emergency_name)
        with db() as conn:
            emergency_payload = build_backup_payload(conn)
            with open(emergency_path, "w", encoding="utf-8") as f:
                json.dump(emergency_payload, f, ensure_ascii=False, indent=2, default=str)
            backup_log(conn, "emergency_before_restore", emergency_name, 1, "", json.dumps(emergency_payload.get("manifest", {}), ensure_ascii=False))
            counts = restore_backup_payload(conn, payload)
            backup_log(conn, "restore", name, 1, "", json.dumps({"restored": counts, "emergency_backup": emergency_name}, ensure_ascii=False))
            audit(conn, "backup_restore", "backup", None, f"{name}; emergency={emergency_name}")
        return ok(restored=counts, emergency_backup=emergency_name)
    except Exception as e:
        try:
            with db() as conn:
                backup_log(conn, "restore", name, 0, str(e), "")
        except Exception:
            pass
        return err("Restore error: " + str(e), 500)


@bp.route("/api/backup/log")
@require_login
@require_role("admin")
def backup_log_api():
    with db() as conn:
        rows = all_rows(conn, "SELECT * FROM backup_log ORDER BY id DESC LIMIT 200")
    return ok(log=rows)






# ---------------- V7.1.0: QA / release candidate control ----------------

@bp.route("/api/system/integrity")
@require_login
def system_integrity_api():
    u = current_user()
    if u["role"] not in ["admin", "transfusion"]:
        return err("Недостатньо прав", 403)
    with db() as conn:
        return ok(report=system_integrity_report(conn))


@bp.route("/api/system/permissions")
@require_login
def system_permissions_api():
    u = current_user()
    if u["role"] not in ["admin", "transfusion"]:
        return err("Недостатньо прав", 403)
    return ok(matrix=permissions_matrix())


@bp.post("/api/system/run-expire-check")
@require_login
@require_csrf
@require_role("admin", "transfusion")
def system_run_expire_check():
    # Same logic as /api/stock/expire, exposed from QA panel.
    return expire_stock()


# ---------------- V7.0.9: V6 -> V7 migration ----------------

LEGACY_STATUS_MAP = {
    "нова": "created", "new": "created", "created": "created", "незатверджена": "created",
    "погоджено": "approved", "погоджена": "approved", "approved": "approved", "активна": "approved",
    "видано": "issued", "видана": "issued", "issued": "issued",
    "частково видано": "partial_issued", "partial_issued": "partial_issued",
    "використано": "used", "використана": "used", "used": "used",
    "відмовлено": "rejected", "відмовлена": "rejected", "rejected": "rejected",
    "списано": "used", "списана": "used",
}
LEGACY_UNIT_STATUS_MAP = {
    "available": "in_stock", "in_stock": "in_stock", "на складі": "in_stock", "доступний": "in_stock",
    "reserved": "reserved", "резерв": "reserved", "зарезервовано": "reserved",
    "issued": "issued", "видано": "issued", "видана": "issued",
    "used": "used", "використано": "used",
    "written_off": "written_off", "writeoff": "written_off", "списано": "written_off",
    "expired": "expired", "протерміновано": "expired",
    "deleted": "deleted", "видалено": "deleted",
}
LEGACY_ROLE_MAP = {"admin":"admin", "administrator":"admin", "адмін":"admin", "адміністратор":"admin", "transfusion":"transfusion", "трансфузіолог":"transfusion", "doctor":"doctor", "лікар":"doctor", "nurse":"nurse", "медсестра":"nurse"}

# V7.1.5: legacy V6 component aliases are normalized before import.
# The mapping is intentionally conservative: unknown names are reported in analysis and skipped during import.
LEGACY_COMPONENT_ALIASES = {
    "цільна кров": "Цільна кров", "whole blood": "Цільна кров",
    "еритроцити": "Еритроцити", "ер маса": "Еритроцити", "ер.маса": "Еритроцити", "еритроцитарна маса": "Еритроцити", "rbc": "Еритроцити",
    "еритроцити в додатковому розчині": "Еритроцити в додатковому розчині",
    "еритроцити з видаленим лейкоцитарно-тромбоцитарним шаром": "Еритроцити з видаленим лейкоцитарно-тромбоцитарним шаром",
    "еритроцити збіднені на лейкоцити": "Еритроцити збіднені на лейкоцити",
    "еритроцити відмиті": "Еритроцити відмиті",
    "плазма": "Плазма свіжозаморожена", "плазма сз": "Плазма свіжозаморожена", "сзп": "Плазма свіжозаморожена", "ffp": "Плазма свіжозаморожена",
    "плазма свіжозаморожена": "Плазма свіжозаморожена",
    "плазма збіднена на кріопреципітат": "Плазма свіжозаморожена збіднена на кріопреципітат",
    "плазма свіжозаморожена збіднена на кріопреципітат": "Плазма свіжозаморожена збіднена на кріопреципітат",
    "плазма після патогенредукції": "Плазма свіжозаморожена оброблена методом патогенредукції",
    "плазма свіжозаморожена оброблена методом патогенредукції": "Плазма свіжозаморожена оброблена методом патогенредукції",
    "тромбоцити": "Тромбоцити відновлені", "тромбоцитарна маса": "Тромбоцити відновлені", "тромбоконцентрат": "Тромбоцити відновлені",
    "тромбоцити відновлені": "Тромбоцити відновлені", "тромбоцити аферез": "Тромбоцити аферез",
    "тромбоцити аферез оброблені методом патогенредукції": "Тромбоцити аферез оброблені методом патогенредукції",
    "кріопреципітат": "Кріопреципітат", "криопреципитат": "Кріопреципітат", "cryo": "Кріопреципітат",
}


def legacy_component_candidates(tables):
    items = []
    for table in ("requests", "stock_units", "blood_units", "stock_entries"):
        for row in tables.get(table, []) or []:
            if isinstance(row, dict):
                val = vget(row, "component_type", "component", "component_name", "blood_component", default="")
                if val:
                    items.append((table, str(val).strip()))
    return items


def normalize_legacy_component(conn, value):
    raw = clean_text(value, 180)
    if not raw:
        return "", "empty"
    exact = normalize_component(conn, raw)
    if exact:
        return exact, "exact"
    key = " ".join(raw.lower().replace("\u00a0", " ").replace(".", " ").replace("-", " ").split())
    if key in LEGACY_COMPONENT_ALIASES:
        mapped = LEGACY_COMPONENT_ALIASES[key]
        if normalize_component(conn, mapped):
            return mapped, "mapped"
    # keyword fallback, conservative order
    if "патоген" in key and "тромбо" in key:
        mapped = "Тромбоцити аферез оброблені методом патогенредукції"
    elif "афер" in key and "тромбо" in key:
        mapped = "Тромбоцити аферез"
    elif "тромбо" in key:
        mapped = "Тромбоцити відновлені"
    elif "кріо" in key or "крио" in key or "cryo" in key:
        mapped = "Кріопреципітат"
    elif "патоген" in key and "плаз" in key:
        mapped = "Плазма свіжозаморожена оброблена методом патогенредукції"
    elif "збід" in key and "кріо" in key and "плаз" in key:
        mapped = "Плазма свіжозаморожена збіднена на кріопреципітат"
    elif "плаз" in key or "сзп" in key or "ffp" in key:
        mapped = "Плазма свіжозаморожена"
    elif "відмит" in key and "ерит" in key:
        mapped = "Еритроцити відмиті"
    elif "лейко" in key and "ерит" in key:
        mapped = "Еритроцити збіднені на лейкоцити"
    elif "ерит" in key or "ер мас" in key or "rbc" in key:
        mapped = "Еритроцити"
    elif "ціль" in key and "кров" in key:
        mapped = "Цільна кров"
    else:
        return "", "unknown"
    if normalize_component(conn, mapped):
        return mapped, "mapped"
    return "", "unknown"


def vget(row, *names, default=""):
    if not isinstance(row, dict):
        return default
    for n in names:
        if n in row and row.get(n) not in (None, ""):
            return row.get(n)
    return default


def norm_status(v, mapping, default):
    key = str(v or "").strip().lower()
    return mapping.get(key, default)


def to_int(v, default=1):
    try:
        return max(1, int(float(str(v).replace(",", "."))))
    except Exception:
        return default


def read_migration_source(upload):
    if not upload:
        raise ValueError("Файл не передано")
    name = upload.filename or "uploaded"
    raw = upload.read()
    if not raw:
        raise ValueError("Файл порожній")
    # JSON backup / export
    if name.lower().endswith(".json") or raw[:1] in (b"{", b"["):
        payload = json.loads(raw.decode("utf-8-sig"))
        if isinstance(payload, dict) and isinstance(payload.get("tables"), dict):
            return "json", name, payload.get("tables") or {}
        if isinstance(payload, dict):
            return "json", name, payload
        raise ValueError("JSON має містити об’єкт з таблицями")
    # SQLite database file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".sqlite")
    tmp.write(raw); tmp.close()
    try:
        con = sqlite3.connect(tmp.name)
        con.row_factory = sqlite3.Row
        tables = [r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name")]
        data = {t: [dict(x) for x in con.execute(f"SELECT * FROM {t}").fetchall()] for t in tables}
        con.close()
        return "sqlite", name, data
    finally:
        try: os.unlink(tmp.name)
        except Exception: pass


def detect_legacy_tables(tables, conn=None):
    names = set(tables.keys())
    component_preview = {"exact": {}, "mapped": {}, "unknown": {}}
    if conn is not None:
        for table, raw in legacy_component_candidates(tables):
            normalized, status = normalize_legacy_component(conn, raw)
            bucket = component_preview.setdefault(status, {})
            key = raw if status == "unknown" else f"{raw} → {normalized}"
            bucket[key] = bucket.get(key, 0) + 1
    return {
        "source_version": "V7 backup" if "stock_units" in names else ("V6 legacy" if ("stock_entries" in names or "blood_units" in names) else "unknown"),
        "tables": sorted(names),
        "counts": {k: len(v or []) for k, v in tables.items()},
        "component_preview": component_preview,
        "unknown_components": sorted((component_preview.get("unknown") or {}).keys()),
        "importable": {
            "users": len(tables.get("users", []) or []),
            "patients": len(tables.get("patients", []) or []),
            "requests": len(tables.get("requests", []) or []),
            "stock_entries": len(tables.get("stock_entries", []) or []),
            "blood_units": len(tables.get("blood_units", []) or []),
            "stock_units": len(tables.get("stock_units", []) or []),
            "movements_or_events": len(tables.get("movements", []) or []) + len(tables.get("unit_events", []) or []),
            "temperature": len(tables.get("fridge_temperature_log", []) or []) + len(tables.get("temperature_readings", []) or []),
            "telegram_logs": len(tables.get("telegram_logs", []) or []) + len(tables.get("telegram_log", []) or []),
        }
    }


def migration_log(conn, action, source_kind="", file_name="", ok_value=0, stats=None, error="", details=""):
    stats = stats or {}
    p = ph()
    exec_sql(conn, f"""INSERT INTO migration_log(action,source_kind,file_name,ok,imported_users,imported_patients,imported_requests,imported_units,imported_movements,skipped,error,details,user_id)
        VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p})""",
        (action, source_kind, file_name, 1 if ok_value else 0, int(stats.get("users",0)), int(stats.get("patients",0)), int(stats.get("requests",0)), int(stats.get("units",0)), int(stats.get("movements",0)), int(stats.get("skipped",0)), str(error)[:1000], str(details)[:4000], session.get("uid")))


def find_or_create_patient(conn, row, stats):
    p = ph()
    name = str(vget(row, "full_name", "patient_name", "name", default="")).strip()
    birth = str(vget(row, "birth_date", "dob", default="") or "").strip()
    if not name:
        return None
    found = one(conn, f"SELECT id FROM patients WHERE full_name={p} AND COALESCE(birth_date,'')={p} AND deleted_at IS NULL", (name, birth))
    if found:
        return found["id"]
    new_id = insert_and_get_id(conn, "patients",
        ["full_name","birth_date","address","department","diagnosis","patient_status","abo","rh","note"],
        [name, birth, vget(row,"address","living_address"), vget(row,"department"), vget(row,"diagnosis"), vget(row,"patient_status"), vget(row,"abo","patient_group"), vget(row,"rh","patient_rh"), vget(row,"note","request_note")])
    stats["patients"] = stats.get("patients",0) + 1
    return new_id


def import_legacy_data(conn, tables, source_kind, file_name, mode="merge"):
    p = ph(); stats = {"users":0,"patients":0,"requests":0,"units":0,"movements":0,"skipped":0}
    details = {"request_id_map":{}, "unit_id_map":{}}
    # Users: import as inactive-safe reset users with temp password if username missing locally.
    for u in tables.get("users", []) or []:
        username = str(vget(u, "username", default="")).strip()
        if not username or username.lower() == "sepsis":
            stats["skipped"] += 1; continue
        if one(conn, f"SELECT id FROM users WHERE username={p}", (username,)):
            stats["skipped"] += 1; continue
        role = LEGACY_ROLE_MAP.get(str(vget(u,"role",default="doctor")).strip().lower(), "doctor")
        exec_sql(conn, f"""INSERT INTO users(username,full_name,position,role,password_hash,active,first_login)
            VALUES ({p},{p},{p},{p},{p},1,1)""", (username, vget(u,"full_name", default=username), vget(u,"position"), role, generate_password_hash("ChangeMe123!")))
        stats["users"] += 1
    # Patients explicit table first.
    for pat in tables.get("patients", []) or []:
        find_or_create_patient(conn, pat, stats)
    # Requests.
    for r in tables.get("requests", []) or []:
        patient_id = find_or_create_patient(conn, r, stats)
        component_raw = vget(r, "component_type", "component", "component_name", "blood_component", default="")
        component, comp_status = normalize_legacy_component(conn, component_raw)
        if not vget(r,"patient_name","full_name") or not component:
            stats["skipped"] += 1
            details.setdefault("unknown_components", []).append({"table":"requests", "value": component_raw})
            continue
        status = norm_status(vget(r,"status", default="created"), LEGACY_STATUS_MAP, "created")
        qty = to_int(vget(r,"quantity","amount", default=1), 1)
        new_id = insert_and_get_id(conn, "requests",
            ["patient_id","patient_name","birth_date","address","department","diagnosis","patient_status","component_type","abo","rh","quantity","needed_date","urgency","indication","request_note","transfusion_history","reaction_note","doctor_name","doctor_position","nurse_name","nurse_user_id","nurse_position","status","reject_reason","created_at"],
            [patient_id, vget(r,"patient_name","full_name"), vget(r,"birth_date"), vget(r,"address"), vget(r,"department"), vget(r,"diagnosis"), vget(r,"patient_status"), component, vget(r,"abo","patient_group", default="O(I)"), vget(r,"rh","patient_rh", default="Rh+"), qty, vget(r,"needed_date"), vget(r,"urgency", default="routine"), vget(r,"indication"), vget(r,"note","request_note"), vget(r,"transfusion_history"), vget(r,"reaction_description","reaction_note"), vget(r,"doctor_name"), vget(r,"doctor_position"), vget(r,"nurse_name"), None, vget(r,"nurse_position"), status, vget(r,"reject_reason","compatibility_warning"), vget(r,"created_at") or datetime.now().isoformat(timespec="seconds")])
        details["request_id_map"][str(vget(r,"id", default=new_id))] = new_id
        stats["requests"] += 1
    # Stock units from V7 stock_units, V6 blood_units and V6 stock_entries.
    stock_sources = []
    for u in tables.get("stock_units", []) or []:
        stock_sources.append(("stock_units", u))
    for u in tables.get("blood_units", []) or []:
        stock_sources.append(("blood_units", u))
    for se in tables.get("stock_entries", []) or []:
        stock_sources.append(("stock_entries", se))
    seen_codes = set([x.get("unit_code") for x in all_rows(conn, "SELECT unit_code FROM stock_units WHERE unit_code IS NOT NULL AND unit_code<>''")])
    for source, u in stock_sources:
        component_raw = vget(u, "component_type", "component", "component_name", "blood_component", default="")
        component, comp_status = normalize_legacy_component(conn, component_raw)
        if not component:
            stats["skipped"] += 1
            details.setdefault("unknown_components", []).append({"table": source, "value": component_raw})
            continue
        qty = to_int(vget(u,"quantity","amount", default=1), 1)
        # stock_entries with amount >1 becomes one unit per count; blood_units is one physical unit.
        copies = qty if source == "stock_entries" else 1
        for i in range(copies):
            raw_code = str(vget(u,"unit_code","pack_no","qr_code", default="")).strip()
            base = raw_code or f"MIG-{source}-{vget(u,'id', default=int(time.time()*1000))}"
            code = base if copies == 1 else f"{base}-{i+1}"
            suffix=1
            while code in seen_codes:
                suffix += 1; code = f"{base}-M{suffix}"
            seen_codes.add(code)
            status = norm_status(vget(u,"status", default="in_stock"), LEGACY_UNIT_STATUS_MAP, "in_stock")
            if source == "stock_entries":
                typ = str(vget(u,"type", default="")).lower()
                if "спис" in typ or "write" in typ:
                    status = "written_off"
                else:
                    status = "in_stock"
            old_rid = str(vget(u,"request_id", default=""))
            new_rid = details["request_id_map"].get(old_rid)
            new_uid = insert_and_get_id(conn, "stock_units",
                ["component_type","abo","rh","quantity","unit_code","series","source","received_date","expiry_date","status","request_id","issued_at","used_at","written_off_at","note","created_at"],
                [component, vget(u,"abo","donor_group", default="O(I)"), vget(u,"rh","donor_rh", default="Rh+"), 1, code, vget(u,"series"), vget(u,"source","location"), vget(u,"received_date","created_at"), vget(u,"expiry_date","expiry"), status, new_rid, vget(u,"issued_at"), vget(u,"used_at"), vget(u,"written_off_at"), vget(u,"note","writeoff_reason"), vget(u,"created_at") or datetime.now().isoformat(timespec="seconds")])
            details["unit_id_map"][str(vget(u,"id", default=new_uid))] = new_uid
            stats["units"] += 1
            exec_sql(conn, f"INSERT INTO movements(unit_id,request_id,action,quantity,user_id,reason,details,created_at) VALUES ({p},{p},'income',1,{p},{p},{p},{p})", (new_uid, new_rid, session.get("uid"), "Міграція V6→V7", f"source={source}; old_id={vget(u,'id',default='')}", vget(u,"created_at") or datetime.now().isoformat(timespec="seconds")))
            stats["movements"] += 1
            if status in ("issued","used","written_off","expired"):
                action = {"issued":"issue","used":"used","written_off":"writeoff","expired":"expired"}.get(status, "issue")
                exec_sql(conn, f"INSERT INTO movements(unit_id,request_id,action,quantity,user_id,reason,details,created_at) VALUES ({p},{p},{p},1,{p},{p},{p},{p})", (new_uid, new_rid, action, session.get("uid"), vget(u,"writeoff_reason"), "Автостворено з імпортованого статусу", vget(u,"updated_at","issued_at","used_at","written_off_at") or datetime.now().isoformat(timespec="seconds")))
                stats["movements"] += 1
    reset_postgres_sequences(conn)
    audit(conn, "migration_import", "migration", None, json.dumps(stats, ensure_ascii=False))
    migration_log(conn, "import", source_kind, file_name, 1, stats, "", json.dumps(details, ensure_ascii=False)[:3500])
    return stats


@bp.route("/api/migration/analyze", methods=["POST"])
@require_login
@require_role("admin")
@require_csrf
def migration_analyze():
    try:
        source_kind, file_name, tables = read_migration_source(request.files.get("file"))
        with db() as conn:
            summary = detect_legacy_tables(tables, conn)
            migration_log(conn, "analyze", source_kind, file_name, 1, {"skipped":0}, "", json.dumps(summary, ensure_ascii=False)[:3500])
        return ok(source_kind=source_kind, file=file_name, summary=summary)
    except Exception as e:
        with db() as conn:
            migration_log(conn, "analyze", "upload", getattr(request.files.get("file"), "filename", ""), 0, {}, str(e), "")
        return err("Не вдалося проаналізувати джерело: " + str(e), 400)


@bp.route("/api/migration/import", methods=["POST"])
@require_login
@require_role("admin")
@require_csrf
def migration_import():
    if (request.form.get("confirm") or "").strip().upper() != "MIGRATE":
        return err("Для підтвердження введіть MIGRATE")
    try:
        source_kind, file_name, tables = read_migration_source(request.files.get("file"))
        with db() as conn:
            stats = import_legacy_data(conn, tables, source_kind, file_name)
        return ok(imported=stats, source_kind=source_kind, file=file_name)
    except Exception as e:
        with db() as conn:
            migration_log(conn, "import", "upload", getattr(request.files.get("file"), "filename", ""), 0, {}, str(e), "")
        return err("Міграція не виконана: " + str(e), 400)


@bp.route("/api/migration/log")
@require_login
@require_role("admin")
def migration_log_api():
    with db() as conn:
        rows = all_rows(conn, "SELECT * FROM migration_log ORDER BY id DESC LIMIT 200")
    return ok(log=rows)

@bp.route("/api/reports/export.xlsx")
@require_login
@require_role("admin", "transfusion")
def reports_export_xlsx():
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    with db() as conn:
        data = report_dataset(conn, start, end)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Підсумок"
    title = f"Звіт банку крові за період: {start or 'початок'} - {end or 'сьогодні'}"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin,right=thin,top=thin,bottom=thin)
    def add_table(sheet, title, headers, rows, fields, start_row):
        sheet.cell(start_row,1,title).font = Font(bold=True, size=12)
        r=start_row+1
        for c,h in enumerate(headers,1):
            cell=sheet.cell(r,c,h); cell.font=Font(bold=True); cell.fill=header_fill; cell.border=border; cell.alignment=Alignment(horizontal="center")
        for row in rows:
            r += 1
            for c,f in enumerate(fields,1):
                v = row.get(f, "")
                if f in ("action",): v = ua_action(v)
                if f in ("status",): v = ua_status(v)
                cell=sheet.cell(r,c,v); cell.border=border; cell.alignment=Alignment(wrap_text=True, vertical="top")
        for c in range(1, len(headers)+1):
            sheet.column_dimensions[get_column_letter(c)].width = min(max(len(str(headers[c-1]))+4, 12), 32)
        return r + 2
    r=3
    r=add_table(ws,"Журнал рухів за діями",["Дія","Кількість"],data["movements"],["action","qty"],r)
    r=add_table(ws,"Нетто-підсумок видачі",["Надходження","Видача факт","Повернення","Видача нетто","Використано","Списано","Протерміновано"],data["net_movements"],["income","issue","return","issue_net","used","writeoff","expired"],r)
    r=add_table(ws,"Скоригований підсумок",["Надходження","Видача факт","Повернення","Видача нетто","Використано","Списано","Протерміновано","Корекції"],data["adjusted_net_movements"],["income","issue","return","issue_net","used","writeoff","expired","adjustment"],r)
    r=add_table(ws,"Поточний стан складу",["Статус","Кількість"],data["stock"],["status","qty"],r)
    r=add_table(ws,"Поточний стан вимог",["Статус","Кількість"],data["requests"],["status","qty"],r)
    ws2=wb.create_sheet("По компонентах")
    r2=add_table(ws2,"Журнал рухів по компонентах",["Компонент","Група","Rh","Дія","Кількість"],data["by_component"],["component_type","abo","rh","action","qty"],1)
    r2=add_table(ws2,"Нетто по компонентах",["Компонент","Група","Rh","Надходження","Видача факт","Повернення","Видача нетто","Використано","Списано","Протерміновано"],data["net_by_component"],["component_type","abo","rh","income","issue","return","issue_net","used","writeoff","expired"],r2)
    add_table(ws2,"Скориговано по компонентах",["Компонент","Група","Rh","Надходження","Видача факт","Повернення","Видача нетто","Використано","Списано","Протерміновано","Корекції"],data["adjusted_net_by_component"],["component_type","abo","rh","income","issue","return","issue_net","used","writeoff","expired","adjustment"],r2)
    ws3=wb.create_sheet("По днях")
    add_table(ws3,"Рухи по днях",["Дата","Дія","Кількість"],data["daily"],["day","action","qty"],1)
    ws4=wb.create_sheet("Детально")
    add_table(ws4,"Детальний журнал рухів",["Дата","Дія","К-сть","Компонент","Група","Rh","Код","Серія","Вимога","Причина","Деталі"],data["details"],["created_at","action","quantity","component_type","abo","rh","unit_code","series","request_id","reason","details"],1)
    ws5=wb.create_sheet("Корекції")
    add_table(ws5,"Корекції звітів",["Дата","Період з","Період по","Компонент","Група","Rh","Дія","+/-","Причина"],data["adjustments"],["created_at","period_start","period_end","component_type","abo","rh","action","quantity_delta","reason"],1)
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=report_filename("xlsx", start, end), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



@bp.route("/api/system/pdf-fonts")
@require_login
@require_role("admin", "transfusion")
def pdf_fonts_api():
    return ok(pdf=pdf_font_status())

@bp.route("/api/system/pdf-test.pdf")
@require_login
@require_role("admin", "transfusion")
def pdf_test_api():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    st = pdf_font_status()
    font = st.get("font_name") or "Helvetica"
    if st.get("chosen"):
        pdfmetrics.registerFont(TTFont("AppFont", st["chosen"]))
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="UA", fontName=font, fontSize=11, leading=14))
    styles.add(ParagraphStyle(name="UAHead", fontName=font, fontSize=16, leading=18, spaceAfter=10, alignment=1))
    story = [
        Paragraph("Тест PDF кирилиці / Ukrainian PDF test", styles["UAHead"]),
        Paragraph("Якщо цей текст читається українською без квадратів — PDF-шрифт на Render працює.", styles["UA"]),
        Spacer(1, 5*mm),
        Paragraph("Банк крові: вимога, пацієнт, проживання, відділення, діагноз, трансфузіолог, медсестра, еритроцити, плазма, тромбоцити, кріопреципітат.", styles["UA"]),
        Spacer(1, 5*mm),
        Paragraph("Обраний шрифт: " + html.escape(st.get("chosen") or "Helvetica fallback"), styles["UA"]),
    ]
    doc.build(story)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="pdf_cyrillic_test.pdf", mimetype="application/pdf")


@bp.route("/api/reports/export.pdf")
@require_login
@require_role("admin", "transfusion")
def reports_export_pdf():
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    with db() as conn:
        data = report_dataset(conn, start, end)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    st = pdf_font_status()
    font = st.get("font_name") or "Helvetica"
    if st.get("chosen"):
        pdfmetrics.registerFont(TTFont("AppFont", st["chosen"]))
    bio=BytesIO()
    doc=SimpleDocTemplate(bio, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    styles=getSampleStyleSheet()
    styles.add(ParagraphStyle(name="UA", fontName=font, fontSize=9, leading=11))
    styles.add(ParagraphStyle(name="UAHead", fontName=font, fontSize=14, leading=16, spaceAfter=8, alignment=1))
    story=[Paragraph(f"Звіт банку крові за період: {start or 'початок'} - {end or 'сьогодні'}", styles["UAHead"])]
    def pdf_table(title, headers, rows, fields, max_rows=80):
        story.append(Paragraph(title, styles["UA"])); story.append(Spacer(1,2*mm))
        body=[headers]
        for row in rows[:max_rows]:
            vals=[]
            for f in fields:
                v=row.get(f,"")
                if f=="action": v=ua_action(v)
                if f=="status": v=ua_status(v)
                vals.append(Paragraph(str(v if v is not None else ""), styles["UA"]))
            body.append(vals)
        t=Table(body, repeatRows=1)
        t.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),font),('FONTSIZE',(0,0),(-1,-1),8),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#D9EAF7')),('GRID',(0,0),(-1,-1),0.25,colors.grey),('VALIGN',(0,0),(-1,-1),'TOP')]))
        story.append(t); story.append(Spacer(1,5*mm))
    pdf_table("Журнал рухів за діями", ["Дія","Кількість"], data["movements"], ["action","qty"])
    pdf_table("Нетто-підсумок видачі", ["Надходження","Видача факт","Повернення","Видача нетто","Використано","Списано","Протерміновано"], data["net_movements"], ["income","issue","return","issue_net","used","writeoff","expired"])
    pdf_table("Скоригований підсумок", ["Надходження","Видача факт","Повернення","Видача нетто","Використано","Корекції"], data["adjusted_net_movements"], ["income","issue","return","issue_net","used","adjustment"])
    pdf_table("Журнал по компонентах", ["Компонент","Група","Rh","Дія","К-сть"], data["by_component"], ["component_type","abo","rh","action","qty"])
    pdf_table("Нетто по компонентах", ["Компонент","Група","Rh","Видача факт","Повернення","Видача нетто","Використано"], data["net_by_component"], ["component_type","abo","rh","issue","return","issue_net","used"])
    pdf_table("Скориговано по компонентах", ["Компонент","Група","Rh","Видача нетто","Використано","Корекції"], data["adjusted_net_by_component"], ["component_type","abo","rh","issue_net","used","adjustment"])
    pdf_table("По днях", ["Дата","Дія","К-сть"], data["daily"], ["day","action","qty"])
    pdf_table("Поточний стан складу", ["Статус","К-сть"], data["stock"], ["status","qty"])
    story.append(PageBreak())
    pdf_table("Детальний журнал рухів", ["Дата","Дія","К-сть","Компонент","Група","Rh","Код","Серія","Вимога","Причина"], data["details"], ["created_at","action","quantity","component_type","abo","rh","unit_code","series","request_id","reason"], max_rows=250)
    doc.build(story)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=report_filename("pdf", start, end), mimetype="application/pdf")


@bp.route("/api/reports/moz-template.xlsx")
@require_login
@require_role("admin", "transfusion")
def reports_moz_template_xlsx():
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    with db() as conn:
        data = report_dataset(conn, start, end)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "МОЗ шаблон"
    ws.append(["Внутрішній шаблон звіту банку крові", "", "", "", ""])
    ws.append(["Період", start or "початок", end or "сьогодні", "Версія", current_app.config["VERSION"]])
    ws.append([])
    headers=["Компонент", "Група", "Rh", "Надходження", "Видача факт", "Повернення", "Видача нетто", "Використано", "Списано", "Протерміновано", "Корекції"]
    ws.append(headers)
    matrix={}
    for r in data["by_component"]:
        key=(r.get("component_type",""), r.get("abo",""), r.get("rh",""))
        matrix.setdefault(key, {"income":0,"issue":0,"return":0,"issue_net":0,"used":0,"writeoff":0,"expired":0,"adj":0})
        matrix[key][r.get("action","")] = matrix[key].get(r.get("action",""),0) + int(r.get("qty") or 0)
    for a in data["adjustments"]:
        key=(a.get("component_type",""), a.get("abo",""), a.get("rh",""))
        matrix.setdefault(key, {"income":0,"issue":0,"return":0,"issue_net":0,"used":0,"writeoff":0,"expired":0,"adj":0})
        delta = int(a.get("quantity_delta") or 0)
        matrix[key]["adj"] += delta
        action = a.get("action") or ""
        if action in matrix[key]:
            matrix[key][action] += delta
    for key, vals in sorted(matrix.items()):
        vals["issue_net"] = max(int(vals.get("issue",0)) - int(vals.get("return",0)), 0)
        ws.append([key[0], key[1], key[2], vals.get("income",0), vals.get("issue",0), vals.get("return",0), vals.get("issue_net",0), vals.get("used",0), vals.get("writeoff",0), vals.get("expired",0), vals.get("adj",0)])
    thin=Side(style="thin", color="9CA3AF"); border=Border(left=thin,right=thin,top=thin,bottom=thin); fill=PatternFill("solid", fgColor="D9EAF7")
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border=border; cell.alignment=Alignment(wrap_text=True, vertical="top")
            if cell.row==4: cell.font=Font(bold=True); cell.fill=fill; cell.alignment=Alignment(horizontal="center")
    ws["A1"].font=Font(bold=True, size=14); ws.merge_cells("A1:E1")
    for c in range(1, len(headers)+1): ws.column_dimensions[get_column_letter(c)].width=18
    ws.freeze_panes="A5"
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"moz_blood_bank_template_{start or 'all'}_{end or 'today'}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@bp.route("/api/reports/export.csv")
@require_login
@require_role("admin", "transfusion")
def reports_export_csv():
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    params = []
    filt = "WHERE m.deleted_at IS NULL" + date_filter_sql("m.created_at", start, end, params)
    with db() as conn:
        rows = all_rows(conn, f"SELECT m.created_at,m.action,m.quantity,m.unit_id,m.request_id,m.user_id,m.reason,m.details FROM movements m {filt} ORDER BY m.id", tuple(params))
    def clean(v):
        return '"'+str(v if v is not None else '').replace('"','""')+'"'
    header = ["created_at","action","quantity","unit_id","request_id","user_id","reason","details"]
    body = [','.join(header)] + [','.join(clean(r.get(h,'')) for h in header) for r in rows]
    return Response('\ufeff'+'\n'.join(body), mimetype='text/csv; charset=utf-8', headers={"Content-Disposition":"attachment; filename=blood_bank_v7_movements.csv"})
