
import os, sqlite3, hashlib, secrets, json, time, shutil, zipfile, base64, hmac, struct, math
from datetime import datetime, timedelta
from datetime import datetime
from functools import wraps
from flask import Flask, request, session, redirect, url_for, render_template, jsonify, send_file, g
from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4, landscape
from io import BytesIO
try:
    import psycopg2
    import psycopg2.extras
except Exception:
    psycopg2 = None

APP_TITLE = "Банк крові V6.4.70"
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
IS_POSTGRES = DATABASE_URL.startswith("postgres")
DB_PATH = os.environ.get("DB_PATH", "blood_bank_v4.db")
BACKUP_DIR = os.environ.get("BACKUP_DIR", "backups")
SESSION_TIMEOUT_MINUTES = int(os.environ.get("SESSION_TIMEOUT_MINUTES","30"))
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN","")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID","")
TELEGRAM_BOT_USERNAME = os.environ.get("TELEGRAM_BOT_USERNAME","")
TELEGRAM_ENABLED = os.environ.get("TELEGRAM_ENABLED","1") == "1"
TELEGRAM_SILENT_START = int(os.environ.get("TELEGRAM_SILENT_START","22"))
TELEGRAM_SILENT_END = int(os.environ.get("TELEGRAM_SILENT_END","7"))
TELEGRAM_ANTI_SPAM_MINUTES = int(os.environ.get("TELEGRAM_ANTI_SPAM_MINUTES","10"))
API_TOKEN = os.environ.get("API_TOKEN","")
BACKUP_ENCRYPTION_KEY = os.environ.get("BACKUP_ENCRYPTION_KEY","")
AUTO_BACKUP_HOURS = int(os.environ.get("AUTO_BACKUP_HOURS","24"))
AUTO_BACKUP_ENABLED = os.environ.get("AUTO_BACKUP_ENABLED","1") == "1"
AUTO_BACKUP_HOUR = int(os.environ.get("AUTO_BACKUP_HOUR","3"))
BACKUP_KEEP_DAYS = int(os.environ.get("BACKUP_KEEP_DAYS","14"))
POSTGRES_ONLY = os.environ.get("POSTGRES_ONLY","0") == "1"
SECRET_KEY = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
DEBUG = os.environ.get("DEBUG","0") == "1"
COOKIE_SECURE = os.environ.get("COOKIE_SECURE","0") == "1"
REQUIRE_HTTPS = os.environ.get("REQUIRE_HTTPS","0") == "1"

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE="Lax", SESSION_COOKIE_SECURE=COOKIE_SECURE)

RATE = {}
BANS = {}

def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def log_suppressed_error(context, exc):
    """Small production-safe logger for places where older builds silently ignored errors.
    It prints only class/message, not secrets or payload dumps. Toggle with LOG_SUPPRESSED_ERRORS=0.
    """
    if os.environ.get("LOG_SUPPRESSED_ERRORS", "1") != "1":
        return
    try:
        print(f"[suppressed-error] {context}: {exc.__class__.__name__}: {exc}", flush=True)
    except Exception:
        pass

def ensure_dirs():
    if os.path.dirname(DB_PATH):
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)


def db():
    if "db" not in g:
        ensure_dirs()
        if IS_POSTGRES:
            if psycopg2 is None:
                raise RuntimeError("psycopg2-binary is required for PostgreSQL")
            g.db = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
            g.db.autocommit = True  # V6.0: prevent InFailedSqlTransaction after migration errors
        else:
            g.db = sqlite3.connect(DB_PATH)
            g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    conn = g.pop("db", None)
    if conn:
        conn.close()

def sql_convert(sql):
    if IS_POSTGRES:
        return sql.replace("?", "%s")
    return sql


# ================= V6.0 REAL POSTGRES TRANSACTION RECOVERY =================
def db_rollback_safe():
    try:
        db().rollback()
    except Exception as e:
        log_suppressed_error('suppressed', e)

def db_commit_safe():
    try:
        if not IS_POSTGRES:
            db().commit()
    except Exception as e:
        log_suppressed_error('suppressed', e)
# ================= END V6.0 REAL POSTGRES TRANSACTION RECOVERY =================

def execute(sql, params=()):
    conn = db()
    try:
        cur = conn.cursor()
        cur.execute(sql_convert(sql), params)
        if not IS_POSTGRES:
            conn.commit()
        return cur
    except Exception:
        try:
            conn.rollback()
        except Exception as e:
            log_suppressed_error('suppressed', e)
        raise



def insert_returning_id(sql, params=()):
    """DB-agnostic INSERT helper.
    SQLite uses cursor.lastrowid. PostgreSQL uses RETURNING id to avoid unsafe
    SELECT last id fallbacks under concurrent users.
    """
    if IS_POSTGRES:
        q = sql.strip().rstrip(';')
        if ' RETURNING ' not in q.upper():
            q += ' RETURNING id'
        cur = execute(q, params)
        try:
            rec = cur.fetchone()
            if isinstance(rec, dict):
                return rec.get('id')
            if rec is not None:
                return rec[0]
        except Exception:
            db_rollback_safe()
        return None
    cur = execute(sql, params)
    return getattr(cur, 'lastrowid', None)

def rows(sql, params=()):
    conn = db()
    try:
        cur = conn.cursor()
        cur.execute(sql_convert(sql), params)
        return [dict(x) for x in cur.fetchall()]
    except Exception:
        try:
            conn.rollback()
        except Exception as e:
            log_suppressed_error('suppressed', e)
        raise

def row(sql, params=()):
    conn = db()
    try:
        cur = conn.cursor()
        cur.execute(sql_convert(sql), params)
        r = cur.fetchone()
        return dict(r) if r else None
    except Exception:
        try:
            conn.rollback()
        except Exception as e:
            log_suppressed_error('suppressed', e)
        raise



def table_columns(table):
    """Return existing column names for SQLite/PostgreSQL without noisy ALTER failures."""
    try:
        if IS_POSTGRES:
            recs = rows("SELECT column_name FROM information_schema.columns WHERE table_name=?", (table,))
            return {str(r.get("column_name", "")).lower() for r in recs}
        recs = rows(f"PRAGMA table_info({table})")
        return {str(r.get("name", "")).lower() for r in recs}
    except Exception as e:
        log_suppressed_error(f"table_columns:{table}", e)
        return set()

def column_exists(table, column):
    return str(column).lower() in table_columns(table)

def add_column_if_missing(table, column, definition):
    """Idempotent ALTER TABLE helper. Avoids duplicate-column noise in Render logs."""
    try:
        if column_exists(table, column):
            return False
        execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
        return True
    except Exception as e:
        # Keep safe fallback for races/old DBs, but log non-duplicate errors only.
        msg = str(e).lower()
        if "duplicate" not in msg and "already exists" not in msg:
            log_suppressed_error(f"add_column_if_missing:{table}.{column}", e)
        db_rollback_safe()
        return False

def hash_password(password, salt=None):
    salt = salt or secrets.token_hex(16)
    ph = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 200000).hex()
    return salt, ph

def password_ok(password, salt, ph):
    return hash_password(password, salt)[1] == ph

def password_policy_ok(p):
    if not p or len(p) < 6:
        return False, "Пароль мінімум 6 символів"
    return True, ""

def safe_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(value)
    except Exception:
        return default

def require_id_from_json(d, key="id"):
    rid = safe_int((d or {}).get(key), 0)
    if rid <= 0:
        return None, (jsonify(ok=False, error="ID запису обов’язковий"), 400)
    return rid, None

def current_user():
    uid = session.get("user_id")
    if not uid: return None
    return row("SELECT * FROM users WHERE id=?", (uid,))

def audit(action, details=""):
    u = current_user()
    try:
        execute("INSERT INTO audit(created_at,user_id,username,role,action,details,ip,user_agent) VALUES(?,?,?,?,?,?,?,?)",
                (now(), u.get("id") if u else None, u.get("username") if u else "", u.get("role") if u else "", action, details,
                 request.headers.get("X-Forwarded-For", request.remote_addr or ""), request.headers.get("User-Agent","")))
    except Exception:
        execute("INSERT INTO audit(created_at,user_id,username,role,action,details) VALUES(?,?,?,?,?,?)",
                (now(), u.get("id") if u else None, u.get("username") if u else "", u.get("role") if u else "", action, details))

def notify(user_id, title, body):
    execute("INSERT INTO notifications(created_at,user_id,title,body,read_at) VALUES(?,?,?,?,NULL)", (now(), user_id, title, body))

def create_user(username, password, role, full_name="", position="", active=1, first_login=1):
    salt, ph = hash_password(password)
    if row("SELECT id FROM users WHERE username=?", (username,)):
        return
    execute("""INSERT INTO users(username,password_hash,salt,role,full_name,position,active,created_at,first_login,spellcheck_enabled,failed_logins,locked_until)
               VALUES(?,?,?,?,?,?,?,?,?,1,0,NULL)""", (username, ph, salt, role, full_name, position, active, now(), first_login))

def init_db():
    ensure_dirs()
    if IS_POSTGRES:
        ddl = [
            """CREATE TABLE IF NOT EXISTS users(
                id SERIAL PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                salt TEXT NOT NULL,
                role TEXT NOT NULL,
                full_name TEXT,
                position TEXT,
                active INTEGER DEFAULT 1,
                created_at TEXT,
                first_login INTEGER DEFAULT 1,
                spellcheck_enabled INTEGER DEFAULT 1,
                failed_logins INTEGER DEFAULT 0,
                locked_until TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS requests(
                id SERIAL PRIMARY KEY,
                created_at TEXT,
                created_by INTEGER,
                doctor_name TEXT,
                doctor_position TEXT,
                patient_name TEXT,
                birth_date TEXT,
                address TEXT,
                patient_status TEXT,
                department TEXT,
                component TEXT,
                patient_group TEXT,
                patient_rh TEXT,
                amount REAL,
                urgency TEXT,
                needed_date TEXT,
                diagnosis TEXT,
                note TEXT,
                status TEXT DEFAULT 'Нова',
                compatibility_ok INTEGER DEFAULT 1,
                compatibility_warning TEXT,
                approved_by TEXT,
                issued_by TEXT,
                issued_at TEXT,
                pack_no TEXT,
                series TEXT,
                expiry TEXT,
                donor_group TEXT,
                donor_rh TEXT,
                used_at TEXT,
                used_by TEXT,
                use_confirm TEXT,
                writeoff_at TEXT,
                written_by TEXT,
                writeoff_reason TEXT,
                reaction_present TEXT DEFAULT 'Ні',
                reaction_type TEXT,
                reaction_severity TEXT,
                reaction_description TEXT,
                reaction_result TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS stock_entries(
                id SERIAL PRIMARY KEY,
                created_at TEXT,
                user_id INTEGER,
                type TEXT,
                component TEXT,
                donor_group TEXT,
                donor_rh TEXT,
                amount REAL,
                pack_no TEXT,
                series TEXT,
                expiry TEXT,
                patient_name TEXT,
                note TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS audit(
                id SERIAL PRIMARY KEY,
                created_at TEXT,
                user_id INTEGER,
                username TEXT,
                role TEXT,
                action TEXT,
                details TEXT,
                ip TEXT,
                user_agent TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS notifications(
                id SERIAL PRIMARY KEY,
                created_at TEXT,
                user_id INTEGER,
                title TEXT,
                body TEXT,
                read_at TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS trash(
                id SERIAL PRIMARY KEY,
                created_at TEXT,
                source_table TEXT,
                source_id INTEGER,
                data TEXT,
                deleted_by TEXT,
                reason TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS login_attempts(
                id SERIAL PRIMARY KEY,
                created_at TEXT,
                username TEXT,
                ip TEXT,
                user_agent TEXT,
                ok INTEGER
            )""",
            """CREATE TABLE IF NOT EXISTS backups(
                id SERIAL PRIMARY KEY,
                created_at TEXT,
                filename TEXT,
                size_bytes INTEGER,
                created_by TEXT
            )"""
        ]
        for q in ddl:
            execute(q)
    else:
        execute("""CREATE TABLE IF NOT EXISTS users(
            id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL, salt TEXT NOT NULL,
            role TEXT NOT NULL, full_name TEXT, position TEXT, active INTEGER DEFAULT 1, created_at TEXT, first_login INTEGER DEFAULT 1,
            spellcheck_enabled INTEGER DEFAULT 1, failed_logins INTEGER DEFAULT 0, locked_until TEXT
        )""")
        execute("""CREATE TABLE IF NOT EXISTS requests(
            id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT, created_by INTEGER, doctor_name TEXT, doctor_position TEXT,
            patient_name TEXT, birth_date TEXT, address TEXT, patient_status TEXT, department TEXT, component TEXT,
            patient_group TEXT, patient_rh TEXT, amount REAL, urgency TEXT, needed_date TEXT, diagnosis TEXT, note TEXT, status TEXT DEFAULT 'Нова',
            compatibility_ok INTEGER DEFAULT 1, compatibility_warning TEXT,
            approved_by TEXT, issued_by TEXT, issued_at TEXT, pack_no TEXT, series TEXT, expiry TEXT, donor_group TEXT, donor_rh TEXT,
            used_at TEXT, used_by TEXT, use_confirm TEXT, writeoff_at TEXT, written_by TEXT, writeoff_reason TEXT,
            reaction_present TEXT DEFAULT 'Ні', reaction_type TEXT, reaction_severity TEXT, reaction_description TEXT, reaction_result TEXT
        )""")
        execute("""CREATE TABLE IF NOT EXISTS stock_entries(
            id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT, user_id INTEGER, type TEXT, component TEXT, donor_group TEXT, donor_rh TEXT,
            amount REAL, pack_no TEXT, series TEXT, expiry TEXT, patient_name TEXT, note TEXT
        )""")
        execute("""CREATE TABLE IF NOT EXISTS audit(id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT,user_id INTEGER,username TEXT,role TEXT,action TEXT,details TEXT,ip TEXT,user_agent TEXT)""")
        execute("""CREATE TABLE IF NOT EXISTS notifications(id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT,user_id INTEGER,title TEXT,body TEXT,read_at TEXT)""")
        execute("""CREATE TABLE IF NOT EXISTS trash(id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT,source_table TEXT,source_id INTEGER,data TEXT,deleted_by TEXT,reason TEXT)""")
        execute("""CREATE TABLE IF NOT EXISTS login_attempts(id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT,username TEXT,ip TEXT,user_agent TEXT,ok INTEGER)""")
        execute("""CREATE TABLE IF NOT EXISTS backups(id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT,filename TEXT,size_bytes INTEGER,created_by TEXT)""")
    
    # Legacy auxiliary tables use SQLite AUTOINCREMENT syntax. Do not execute
    # them on PostgreSQL; PostgreSQL receives *_pg tables below.
    if not IS_POSTGRES:
        for sql in [
            "CREATE TABLE IF NOT EXISTS transfusion_events(id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT, request_id INTEGER, patient_name TEXT, component TEXT, pack_no TEXT, nurse_name TEXT, doctor_name TEXT, started_at TEXT, finished_at TEXT, result TEXT, signature TEXT)",
            "CREATE TABLE IF NOT EXISTS reaction_registry(id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT, request_id INTEGER, patient_name TEXT, reaction_type TEXT, severity TEXT, description TEXT, action_taken TEXT, result TEXT, reported_by TEXT)",
            "CREATE TABLE IF NOT EXISTS api_events(id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT, source TEXT, event_type TEXT, payload TEXT)"
        ]:
            try: execute(sql)
            except Exception as e: log_suppressed_error('legacy_aux_sqlite_ddl', e)

    if IS_POSTGRES:
        for sql in [
            "CREATE TABLE IF NOT EXISTS transfusion_events_pg(id SERIAL PRIMARY KEY, created_at TEXT, request_id INTEGER, patient_name TEXT, component TEXT, pack_no TEXT, nurse_name TEXT, doctor_name TEXT, started_at TEXT, finished_at TEXT, result TEXT, signature TEXT)",
            "CREATE TABLE IF NOT EXISTS reaction_registry_pg(id SERIAL PRIMARY KEY, created_at TEXT, request_id INTEGER, patient_name TEXT, reaction_type TEXT, severity TEXT, description TEXT, action_taken TEXT, result TEXT, reported_by TEXT)",
            "CREATE TABLE IF NOT EXISTS api_events_pg(id SERIAL PRIMARY KEY, created_at TEXT, source TEXT, event_type TEXT, payload TEXT)"
        ]:
            try: execute(sql)
            except Exception as e: log_suppressed_error('suppressed', e)

    
    # V5_4_1_QR_SCHEMA_FIX: ensure qr_code exists after upgrades
    try:
        add_column_if_missing("stock_entries", "qr_code", "TEXT")
    except Exception as e:
        log_suppressed_error('suppressed', e)

    
    # V5.5 Telegram Edition tables. Use native table names on SQLite and *_pg
    # tables on PostgreSQL to avoid AUTOINCREMENT errors in Render logs.
    if not IS_POSTGRES:
        for sql in [
            "CREATE TABLE IF NOT EXISTS telegram_logs(id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT, event_type TEXT, chat_id TEXT, message TEXT, ok INTEGER, response TEXT)",
            "CREATE TABLE IF NOT EXISTS telegram_queue(id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT, event_type TEXT, chat_id TEXT, message TEXT, attempts INTEGER DEFAULT 0, last_error TEXT, sent INTEGER DEFAULT 0)"
        ]:
            try: execute(sql)
            except Exception as e: log_suppressed_error('telegram_sqlite_ddl', e)
    if IS_POSTGRES:
        for sql in [
            "CREATE TABLE IF NOT EXISTS telegram_logs_pg(id SERIAL PRIMARY KEY, created_at TEXT, event_type TEXT, chat_id TEXT, message TEXT, ok INTEGER, response TEXT)",
            "CREATE TABLE IF NOT EXISTS telegram_queue_pg(id SERIAL PRIMARY KEY, created_at TEXT, event_type TEXT, chat_id TEXT, message TEXT, attempts INTEGER DEFAULT 0, last_error TEXT, sent INTEGER DEFAULT 0)"
        ]:
            try: execute(sql)
            except Exception as e: log_suppressed_error('suppressed', e)

        # V5_5_1_STOCK_SCHEMA_FIX
    try:
        add_column_if_missing("stock_entries", "qr_code", "TEXT")
    except Exception as e:
        log_suppressed_error('suppressed', e)

    

    # V6.4.67: дата, на яку потрібен компонент у вимозі/резервуванні
    try:
        add_column_if_missing("requests", "needed_date", "TEXT")
    except Exception as e:
        db_rollback_safe(); log_suppressed_error('suppressed', e)

    # V5.7 Telegram PRO user columns
    for name, definition in [
        ("telegram_chat_id", "TEXT"),
        ("telegram_username", "TEXT"),
        ("telegram_enabled", "INTEGER DEFAULT 0"),
        ("telegram_notify_new_requests", "INTEGER DEFAULT 1"),
        ("telegram_notify_critical", "INTEGER DEFAULT 1"),
        ("telegram_notify_expiring", "INTEGER DEFAULT 1"),
        ("telegram_notify_reactions", "INTEGER DEFAULT 1"),
        ("telegram_notify_backups", "INTEGER DEFAULT 0"),
    ]:
        add_column_if_missing("users", name, definition)

    # V6.4.0: create improved unit/traceability tables during startup migrations
    try:
        ensure_v630_schema()
    except Exception:
        db_rollback_safe()

    try:
        ensure_postgres_compat_tables()
    except Exception:
        db_rollback_safe()
    ensure_default_admin()

def make_backup(created_by="system"):
    ensure_dirs()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if IS_POSTGRES:
        name = f"blood_bank_pg_backup_{ts}.json.zip"
        path = os.path.join(BACKUP_DIR, name)
        tables=["users","requests","stock_entries","audit","notifications","trash","login_attempts","backups"]
        data={}
        for t in tables:
            try: data[t]=rows(f"SELECT * FROM {t}")
            except Exception: data[t]=[]
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr("blood_bank_v4_backup.json", json.dumps(data, ensure_ascii=False))
    else:
        if not os.path.exists(DB_PATH):
            with app.app_context():
                init_db()
        name = f"blood_bank_backup_{ts}.db.zip"
        path = os.path.join(BACKUP_DIR, name)
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
            z.write(DB_PATH, "blood_bank_v4.db")
    size = os.path.getsize(path)
    try:
        execute("INSERT INTO backups(created_at,filename,size_bytes,created_by) VALUES(?,?,?,?)", (now(), name, size, created_by))
    except Exception as e:
        log_suppressed_error('suppressed', e)
    return path

# ================= V6.4.0 BLOOD UNIT / FEFO / TRACEABILITY HELPERS =================
def table_exists(name):
    try:
        if IS_POSTGRES:
            return bool(row("SELECT to_regclass(?) AS name", (name,)).get("name"))
        return bool(rows("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (name,)))
    except Exception:
        db_rollback_safe()
        return False

def parse_date_safe(value):
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(str(value)[:19], fmt).date()
        except Exception as e:
            log_suppressed_error('suppressed', e)
    return None

def expiry_status(expiry):
    d = parse_date_safe(expiry)
    if not d:
        return "available"
    today = datetime.now().date()
    if d < today:
        return "expired"
    return "available"

def ensure_v630_schema():
    iddef = "SERIAL PRIMARY KEY" if IS_POSTGRES else "INTEGER PRIMARY KEY AUTOINCREMENT"
    ddl = [
        f"""CREATE TABLE IF NOT EXISTS blood_units(
            id {iddef}, created_at TEXT, updated_at TEXT, component TEXT, donor_group TEXT, donor_rh TEXT,
            amount REAL, pack_no TEXT, series TEXT, expiry TEXT, status TEXT DEFAULT 'available', location TEXT,
            patient_name TEXT, request_id INTEGER, received_by TEXT, issued_by TEXT, issued_at TEXT,
            used_at TEXT, written_off_at TEXT, writeoff_reason TEXT, qr_code TEXT, note TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS unit_events(
            id {iddef}, created_at TEXT, unit_id INTEGER, request_id INTEGER, actor_id INTEGER, actor_name TEXT,
            action TEXT, details TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS temperature_alerts(
            id {iddef}, created_at TEXT, fridge TEXT, temperature REAL, level TEXT, text TEXT, active INTEGER DEFAULT 1
        )"""
    ]
    for q in ddl:
        try: execute(q)
        except Exception: db_rollback_safe()
    for table, name, definition in [
        ("stock_entries", "unit_id", "INTEGER"),
        ("requests", "unit_id", "INTEGER"),
        ("requests", "reserved_at", "TEXT"),
        ("requests", "reserved_by", "TEXT"),
        ("requests", "needed_date", "TEXT"),
    ]:
        add_column_if_missing(table, name, definition)
    try:
        if IS_POSTGRES:
            execute("CREATE INDEX IF NOT EXISTS idx_blood_units_status_expiry ON blood_units(status, expiry)")
            execute("CREATE INDEX IF NOT EXISTS idx_blood_units_pack ON blood_units(pack_no, series, qr_code)")
        else:
            execute("CREATE INDEX IF NOT EXISTS idx_blood_units_status_expiry ON blood_units(status, expiry)")
            execute("CREATE INDEX IF NOT EXISTS idx_blood_units_pack ON blood_units(pack_no, series, qr_code)")
    except Exception:
        db_rollback_safe()

def unit_event(unit_id, action, details='', request_id=None):
    try:
        u=current_user() or {}
        execute("INSERT INTO unit_events(created_at,unit_id,request_id,actor_id,actor_name,action,details) VALUES(?,?,?,?,?,?,?)",
                (now(), unit_id, request_id, u.get('id'), u.get('full_name') or u.get('username') or '', action, details))
    except Exception:
        db_rollback_safe()

def normalize_unit_statuses():
    try:
        ensure_v630_schema()
        for x in rows("SELECT id, expiry, status FROM blood_units WHERE status IN ('available','reserved')"):
            if expiry_status(x.get('expiry')) == 'expired':
                execute("UPDATE blood_units SET status='expired', updated_at=? WHERE id=?", (now(), x['id']))
                unit_event(x['id'], 'auto_expired', 'Автоматично заблоковано через завершення терміну придатності')
    except Exception:
        db_rollback_safe()


def _validate_lot_qr_conflict(component, donor_group, donor_rh, pack_no='', series='', qr_code=''):
    """V6.4.67: захист QR/партії.

    Один pack_no+series не може мати різні QR-коди, і один QR-код не може
    вказувати на іншу партію. Інакше stock_entries і blood_units починають
    показувати різну простежуваність.
    """
    pack_no=(pack_no or '').strip(); series=(series or '').strip(); qr_code=(qr_code or '').strip()
    component=(component or '').strip(); donor_group=(donor_group or '').strip(); donor_rh=(donor_rh or '').strip()
    if qr_code:
        ex=row("SELECT * FROM blood_units WHERE qr_code<>'' AND qr_code=? ORDER BY id DESC LIMIT 1", (qr_code,))
        if ex:
            # той самий QR можна використати тільки для тієї самої партії/компонента
            if component and (ex.get('component') or '') and (ex.get('component') or '') != component:
                return False, f"QR {qr_code} вже прив’язаний до іншого компонента: {ex.get('component')}"
            if donor_group and (ex.get('donor_group') or '') and (ex.get('donor_group') or '') != donor_group:
                return False, f"QR {qr_code} вже прив’язаний до іншої групи крові: {ex.get('donor_group')}"
            if donor_rh and (ex.get('donor_rh') or '') and (ex.get('donor_rh') or '') != donor_rh:
                return False, f"QR {qr_code} вже прив’язаний до іншого Rh: {ex.get('donor_rh')}"
            if pack_no and (ex.get('pack_no') or '') and (ex.get('pack_no') or '') != pack_no:
                return False, f"QR {qr_code} вже прив’язаний до іншого пакета: {ex.get('pack_no')}"
            if series and (ex.get('series') or '') and (ex.get('series') or '') != series:
                return False, f"QR {qr_code} вже прив’язаний до іншої серії: {ex.get('series')}"
    if pack_no or series:
        clauses=["component=?", "donor_group=?", "donor_rh=?"]
        params=[component, donor_group, donor_rh]
        if pack_no:
            clauses.append("pack_no=?"); params.append(pack_no)
        else:
            clauses.append("COALESCE(pack_no,'')=''")
        if series:
            clauses.append("series=?"); params.append(series)
        else:
            clauses.append("COALESCE(series,'')=''")
        ex=row("SELECT * FROM blood_units WHERE "+" AND ".join(clauses)+" ORDER BY id DESC LIMIT 1", tuple(params))
        if ex and qr_code and (ex.get('qr_code') or '') and (ex.get('qr_code') or '') != qr_code:
            return False, f"Партія {pack_no or '—'}/{series or '—'} вже має інший QR: {ex.get('qr_code')}"
    return True, ""


def _request_is_locked_to_other_unit(req, unit):
    """V6.4.67: якщо вимога вже зарезервована під unit_id/pack/series,
    її не можна видати або списати з іншої партії без контрольованого скасування/зміни резерву.
    """
    if not req or not unit:
        return False, ""
    st=(req.get('status') or '').strip()
    if st not in ('Зарезервовано','Резерв','reserved'):
        return False, ""
    try:
        rid_unit=int(req.get('unit_id') or 0)
        uid=int(unit.get('id') or 0)
        if rid_unit and uid and rid_unit != uid:
            return True, f"Вимога вже зарезервована під іншу партію/одиницю. Спочатку скасуйте резерв або змініть резерв контрольованою дією."
    except Exception as e:
        log_suppressed_error('suppressed', e)
    r_pack=(req.get('pack_no') or '').strip(); r_series=(req.get('series') or '').strip()
    u_pack=(unit.get('pack_no') or '').strip(); u_series=(unit.get('series') or '').strip()
    if r_pack and u_pack and r_pack != u_pack:
        return True, f"Вимога вже зарезервована під пакет {r_pack}, а обрано {u_pack}."
    if r_series and u_series and r_series != u_series:
        return True, f"Вимога вже зарезервована під серію {r_series}, а обрано {u_series}."
    return False, ""

def create_or_update_unit_from_stock(d, amount, stock_type, user):
    """Створення/пошук blood_units для руху складу.

    V6.4.61: для надходження одна партія (component/group/Rh/pack/series) має
    один blood_unit-представник. Повторне надходження тієї ж партії не
    перезаписує amount/status вручну — після запису stock_entries статус і
    кількість перераховує _recalc_blood_unit_status_for_id().

    Для списання/видачі функція тільки знаходить відповідний blood_unit і НЕ
    ставить status='written_off/issued' напряму, бо при частковому списанні
    партія ще може мати залишок.
    """
    ensure_v630_schema()
    pack_no=(d.get('pack_no') or '').strip()
    series=(d.get('series') or '').strip()
    qr=(d.get('qr_code') or '').strip()
    component=d.get('component') or d.get('stock_component') or ''
    donor_group=d.get('donor_group') or d.get('stock_group') or d.get('group') or ''
    donor_rh=d.get('donor_rh') or d.get('stock_rh') or d.get('rh') or ''
    expiry=d.get('expiry') or ''

    if stock_type == 'Надходження':
        ok_qr, msg_qr = _validate_lot_qr_conflict(component, donor_group, donor_rh, pack_no, series, qr)
        if not ok_qr:
            raise ValueError(msg_qr)

    existing=None
    if qr:
        existing=row("SELECT * FROM blood_units WHERE qr_code<>'' AND qr_code=? ORDER BY id DESC LIMIT 1", (qr,))
    if not existing and pack_no and series:
        existing=row("""SELECT * FROM blood_units
                        WHERE component=? AND donor_group=? AND donor_rh=?
                          AND pack_no<>'' AND series<>'' AND pack_no=? AND series=?
                        ORDER BY id DESC LIMIT 1""", (component, donor_group, donor_rh, pack_no, series))
    if not existing and pack_no and not series:
        existing=row("""SELECT * FROM blood_units
                        WHERE component=? AND donor_group=? AND donor_rh=?
                          AND pack_no<>'' AND COALESCE(series,'')='' AND pack_no=?
                        ORDER BY id DESC LIMIT 1""", (component, donor_group, donor_rh, pack_no))
    if not existing and series and not pack_no:
        existing=row("""SELECT * FROM blood_units
                        WHERE component=? AND donor_group=? AND donor_rh=?
                          AND series<>'' AND COALESCE(pack_no,'')='' AND series=?
                        ORDER BY id DESC LIMIT 1""", (component, donor_group, donor_rh, series))

    if stock_type == 'Надходження':
        if existing:
            # Не перезаписуємо amount/status при дубльованому надходженні тієї ж партії.
            # Після INSERT у stock_entries буде викликаний перерахунок фактичного залишку.
            execute("""UPDATE blood_units SET updated_at=?, component=?, donor_group=?, donor_rh=?,
                      pack_no=?, series=?, expiry=COALESCE(NULLIF(?,''), expiry), location=?, received_by=?,
                      qr_code=COALESCE(NULLIF(?,''), qr_code), note=? WHERE id=?""",
                    (now(), component, donor_group, donor_rh, pack_no, series, expiry,
                     d.get('location','Склад'), user.get('username',''), qr, d.get('note',''), existing['id']))
            unit_event(existing['id'], 'stock_received_existing_lot', 'Додано рух надходження до існуючої партії')
            return existing['id']
        uid=insert_returning_id("""INSERT INTO blood_units(created_at,updated_at,component,donor_group,donor_rh,amount,pack_no,series,expiry,status,location,received_by,qr_code,note)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (now(),now(),component,donor_group,donor_rh,amount,pack_no,series,expiry,expiry_status(expiry),d.get('location','Склад'),user.get('username',''),qr,d.get('note','')))
        unit_event(uid, 'stock_received', 'Надходження на склад')
        return uid

    # Outgoing: тільки знаходимо точну одиницю партії, не міняємо status напряму.
    return existing.get('id') if existing else None

def find_fefo_unit_for_request(req, donor_group='', donor_rh='', pack_no='', series='', qr_code=''):
    normalize_unit_statuses()
    params=[]
    # V6.4.61: partial reserve means the blood_unit can still be status='reserved'
    # but have a non-reserved available remainder. FEFO must evaluate quantities,
    # not only the single status field.
    where=["COALESCE(status,'available') IN ('available','reserved')", "component=?"]
    params.append(req.get('component') or '')
    if pack_no or series or qr_code:
        where.append("((pack_no<>'' AND pack_no=?) OR (series<>'' AND series=?) OR (qr_code<>'' AND qr_code=?))")
        params += [pack_no, series, qr_code]
    if donor_group:
        where.append("donor_group=?"); params.append(donor_group)
    if donor_rh:
        where.append("donor_rh=?"); params.append(donor_rh)
    candidates=rows("SELECT * FROM blood_units WHERE "+" AND ".join(where)+" ORDER BY CASE WHEN expiry IS NULL OR expiry='' THEN 1 ELSE 0 END, expiry ASC, id ASC", tuple(params))
    needed=_request_amount_float(req, 1.0)
    fallback=None
    for u in candidates:
        if _available_qty_for_unit(u, exclude_request_id=req.get('id')) + 0.000001 < needed:
            continue
        ok1,_=abo_compatible(req.get('patient_group'), u.get('donor_group'), req.get('component'))
        ok2,_=rh_compatible(req.get('patient_rh'), u.get('donor_rh'))
        if ok1 and ok2:
            return u
        if fallback is None:
            fallback=u
    return fallback

def reserve_unit_for_request(req, unit, actor_name=''):
    if not req or not unit:
        return False
    if expiry_status(unit.get('expiry')) == 'expired':
        return False
    ok, msg, bal, qty = _can_reserve_unit_for_request(req, unit) if '_can_reserve_unit_for_request' in globals() else (True, '', 0, 0)
    if not ok:
        return False
    execute("UPDATE requests SET status='Зарезервовано', unit_id=?, reserved_at=?, reserved_by=?, pack_no=?, series=?, expiry=?, donor_group=?, donor_rh=? WHERE id=?",
            (int(unit['id']), now(), actor_name, unit.get('pack_no',''), unit.get('series',''), unit.get('expiry',''), unit.get('donor_group',''), unit.get('donor_rh',''), int(req['id'])))
    # V6.4.61: не блокуємо всю партію при резерві частини. Статус/amount
    # blood_units перераховується з урахуванням reserved_qty.
    _recalc_blood_unit_status_for_id(unit['id'], f"Резерв під вимогу №{req['id']}")
    unit_event(unit['id'], 'reserved', f"Резерв під вимогу №{req['id']}; кількість {qty:g} із доступних {bal:g}", int(req['id']))
    return True
# ================= END V6.4.0 BLOOD UNIT / FEFO / TRACEABILITY HELPERS =================

def abo_compatible(patient, donor, component):
    patient=normalize_group(patient); donor=normalize_group(donor); component=(component or "").lower()
    if not patient or not donor: return True, ""
    if "плаз" in component:
        allowed={"0":["0","A","B","AB"],"A":["A","AB"],"B":["B","AB"],"AB":["AB"]}
    else:
        allowed={"0":["0"],"A":["0","A"],"B":["0","B"],"AB":["0","A","B","AB"]}
    ok = donor in allowed.get(patient, [])
    return ok, "" if ok else f"ABO несумісність: пацієнт {patient}, донор {donor}"

def rh_compatible(patient_rh, donor_rh):
    patient_rh=normalize_rh(patient_rh); donor_rh=normalize_rh(donor_rh)
    if not patient_rh or not donor_rh: return True, ""
    ok = not (patient_rh == "-" and donor_rh == "+")
    return ok, "" if ok else "Rh несумісність"


def normalize_group(x):
    x=(x or "").upper().strip()
    if x in ["O(I)","0(I)","O","0","I"]: return "0"
    if x in ["A(II)","A","II"]: return "A"
    if x in ["B(III)","B","III"]: return "B"
    if x in ["AB(IV)","AB","IV"]: return "AB"
    return x

def normalize_rh(x):
    x=(x or "").upper().strip()
    if x in ["RH+","PLUS","+","POSITIVE"]: return "+"
    if x in ["RH-","MINUS","-","NEGATIVE"]: return "-"
    return x


def maybe_auto_backup():
    try:
        last = row("SELECT created_at FROM backups ORDER BY id DESC LIMIT 1")
        if not last:
            make_backup("auto")
            return True
        dt = datetime.strptime(last["created_at"], "%Y-%m-%d %H:%M:%S")
        if datetime.now() - dt > timedelta(hours=AUTO_BACKUP_HOURS):
            make_backup("auto")
            telegram_alert("Банк крові: автоматичну резервну копію створено")
            return True
    except Exception:
        return False
    return False


def signature_hash(text):
    salt = SECRET_KEY[:16]
    return hashlib.sha256((salt + "|" + str(text or "")).encode()).hexdigest()

def api_token_required(f):
    @wraps(f)
    def w(*a, **kw):
        if not API_TOKEN:
            return jsonify(ok=False,error="API_TOKEN не налаштований"), 403
        # Prefer header for production, but allow ?token= for quick Render browser checks.
        token = request.headers.get("X-API-Token","") or request.args.get("token", "")
        if token != API_TOKEN:
            return jsonify(ok=False,error="API token invalid"), 403
        return f(*a, **kw)
    return w

def v54_table(base_name):
    if IS_POSTGRES and base_name in ["transfusion_events","reaction_registry","api_events"]:
        return base_name + "_pg"
    return base_name


FIELD_LABELS_UA = {
    "patient_name": "ПІБ пацієнта",
    "birth_date": "Дата народження",
    "department": "Відділення",
    "component": "Компонент крові",
    "patient_group": "Група крові",
    "patient_rh": "Резус",
    "amount": "Кількість",
    "needed_date": "Дата, на яку потрібен компонент",
    "diagnosis": "Діагноз",
    "stock_component": "Компонент",
    "stock_group": "Група крові",
    "stock_rh": "Резус",
}

def validate_required_ua(data, fields):
    for f in fields:
        v = data.get(f, "")
        if v is None or str(v).strip() == "":
            return False, f'Заповніть поле: {FIELD_LABELS_UA.get(f, f)}', f
    return True, "", ""


def telegram_table(name):
    if IS_POSTGRES and name in ["telegram_logs","telegram_queue"]:
        return name + "_pg"
    return name



def ensure_postgres_compat_tables():
    """Create PostgreSQL-only legacy/auxiliary tables that previously relied on
    SQLite AUTOINCREMENT DDL. Safe to call on every startup/health check.
    """
    if not IS_POSTGRES:
        return []
    errors=[]
    ddl=[
        "CREATE TABLE IF NOT EXISTS transfusion_events_pg(id SERIAL PRIMARY KEY, created_at TEXT, request_id INTEGER, patient_name TEXT, component TEXT, pack_no TEXT, nurse_name TEXT, doctor_name TEXT, started_at TEXT, finished_at TEXT, result TEXT, signature TEXT)",
        "CREATE TABLE IF NOT EXISTS reaction_registry_pg(id SERIAL PRIMARY KEY, created_at TEXT, request_id INTEGER, patient_name TEXT, reaction_type TEXT, severity TEXT, description TEXT, action_taken TEXT, result TEXT, reported_by TEXT)",
        "CREATE TABLE IF NOT EXISTS api_events_pg(id SERIAL PRIMARY KEY, created_at TEXT, source TEXT, event_type TEXT, payload TEXT)",
        "CREATE TABLE IF NOT EXISTS telegram_logs_pg(id SERIAL PRIMARY KEY, created_at TEXT, event_type TEXT, chat_id TEXT, message TEXT, ok INTEGER, response TEXT)",
        "CREATE TABLE IF NOT EXISTS telegram_queue_pg(id SERIAL PRIMARY KEY, created_at TEXT, event_type TEXT, chat_id TEXT, message TEXT, attempts INTEGER DEFAULT 0, last_error TEXT, sent INTEGER DEFAULT 0)",
    ]
    for q in ddl:
        try:
            execute(q)
        except Exception as e:
            db_rollback_safe(); errors.append(str(e))
    return errors

def pg_table_health():
    """Return table availability for health-debug. Works for SQLite and PG."""
    names=['users','requests','stock_entries','blood_units','unit_events','trash','audit','notifications','login_attempts','backups','telegram_logs_pg' if IS_POSTGRES else 'telegram_logs','telegram_queue_pg' if IS_POSTGRES else 'telegram_queue','transfusion_protocols','reaction_acts','component_writeoffs','daily_reports','report_overrides','history_period_actions']
    out={}
    for name in names:
        try:
            if IS_POSTGRES:
                r=row('SELECT to_regclass(?) AS name',(name,))
                out[name]=bool(r and r.get('name'))
            else:
                r=row("SELECT name FROM sqlite_master WHERE type='table' AND name=?",(name,))
                out[name]=bool(r)
        except Exception as e:
            db_rollback_safe(); out[name]='ERR: '+str(e)[:120]
    return out

def schema_column_health():
    """Verify important columns used by current UI/workflow are present.
    This is especially useful after Render/PostgreSQL migrations.
    """
    checks = {
        "requests": ["needed_date", "address", "unit_id", "pack_no", "series", "expiry", "donor_group", "donor_rh", "reserved_at", "reserved_by", "compatibility_ok", "compatibility_warning"],
        "stock_entries": ["qr_code", "unit_id", "pack_no", "series", "expiry", "type", "amount"],
        "blood_units": ["qr_code", "pack_no", "series", "amount", "status", "request_id", "patient_name", "writeoff_reason", "written_off_at"],
        "users": ["telegram_chat_id", "telegram_username", "telegram_enabled"],
        "login_attempts": ["ip_address"],
    }
    out = {}
    for table, cols in checks.items():
        existing = table_columns(table)
        out[table] = {c: (c.lower() in existing) for c in cols}
    return out

def telegram_in_silent_time():
    try:
        h = datetime.now().hour
        if TELEGRAM_SILENT_START == TELEGRAM_SILENT_END:
            return False
        if TELEGRAM_SILENT_START > TELEGRAM_SILENT_END:
            return h >= TELEGRAM_SILENT_START or h < TELEGRAM_SILENT_END
        return TELEGRAM_SILENT_START <= h < TELEGRAM_SILENT_END
    except Exception:
        return False

def telegram_recent_duplicate(event_type, message):
    try:
        table = telegram_table("telegram_logs")
        recent = rows(f"SELECT created_at,message FROM {table} WHERE event_type=? AND ok=1 ORDER BY id DESC LIMIT 20", (event_type,))
        cutoff = datetime.now() - timedelta(minutes=TELEGRAM_ANTI_SPAM_MINUTES)
        for r in recent:
            try:
                dt = datetime.strptime(r["created_at"], "%Y-%m-%d %H:%M:%S")
                if dt >= cutoff and r.get("message") == message:
                    return True
            except Exception as e:
                log_suppressed_error('suppressed', e)
    except Exception as e:
        log_suppressed_error('suppressed', e)
    return False

def telegram_log(event_type, chat_id, message, ok, response):
    try:
        table = telegram_table("telegram_logs")
        execute(f"INSERT INTO {table}(created_at,event_type,chat_id,message,ok,response) VALUES(?,?,?,?,?,?)",
                (now(), event_type, str(chat_id or ""), str(message or "")[:3900], 1 if ok else 0, str(response or "")[:1000]))
    except Exception as e:
        log_suppressed_error('suppressed', e)

def telegram_queue_message(event_type, chat_id, message, error=""):
    try:
        table = telegram_table("telegram_queue")
        execute(f"INSERT INTO {table}(created_at,event_type,chat_id,message,attempts,last_error,sent) VALUES(?,?,?,?,?,?,0)",
                (now(), event_type, str(chat_id or ""), str(message or "")[:3900], 0, str(error or "")[:1000]))
    except Exception as e:
        log_suppressed_error('suppressed', e)

def telegram_send_message(message, event_type="system", chat_id=None, force=False):
    if not TELEGRAM_ENABLED:
        return False, "Telegram disabled"
    if not TELEGRAM_BOT_TOKEN:
        return False, "TELEGRAM_BOT_TOKEN missing"
    chat_id = chat_id or TELEGRAM_CHAT_ID
    if not chat_id:
        return False, "TELEGRAM_CHAT_ID missing"
    if telegram_in_silent_time() and not force and event_type not in ["critical","reaction","security"]:
        return False, "silent time"
    if telegram_recent_duplicate(event_type, message) and not force:
        return False, "duplicate suppressed"
    try:
        import urllib.request, urllib.parse
        data = urllib.parse.urlencode({
            "chat_id": chat_id,
            "text": str(message or "")[:3900],
            "parse_mode": "HTML",
            "disable_web_page_preview": "true"
        }).encode()
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        with urllib.request.urlopen(url, data=data, timeout=8) as resp:
            body = resp.read().decode("utf-8", "ignore")
        telegram_log(event_type, chat_id, message, True, body)
        return True, body
    except Exception as e:
        telegram_log(event_type, chat_id, message, False, e)
        telegram_queue_message(event_type, chat_id, message, e)
        return False, str(e)

def telegram_alert(text, event_type="system", force=False):
    ok, resp = telegram_send_message(text, event_type=event_type, force=force)
    return ok

def telegram_event_new_request(req):
    msg = (
        "🩸 <b>Нова вимога</b>\\n"
        f"Пацієнт: {req.get('patient_name','')}\\n"
        f"Відділення: {req.get('department','')}\\n"
        f"Компонент: {req.get('component','')}\\n"
        f"Група/Rh: {req.get('patient_group','')} {req.get('patient_rh','')}\\n"
        f"Кількість: {req.get('amount','')}"
    )
    return telegram_broadcast_roles(msg, ("admin","transfusion"), "new_request")

def telegram_event_reaction(req_id, reaction_type, severity, patient=""):
    msg = (
        "⚠️ <b>Трансфузійна реакція</b>\\n"
        f"Вимога №{req_id}\\n"
        f"Пацієнт: {patient}\\n"
        f"Тип: {reaction_type}\\n"
        f"Тяжкість: {severity}"
    )
    return telegram_broadcast_roles(msg, ("admin","transfusion"), "reaction", force=True)

def telegram_event_backup(ok=True):
    return telegram_broadcast_roles("✅ Резервну копію створено" if ok else "❌ Помилка резервної копії", ("admin","transfusion"), "backup")

def telegram_retry_queue(limit=20):
    table = telegram_table("telegram_queue")
    q = rows(f"SELECT * FROM {table} WHERE sent=0 ORDER BY id ASC LIMIT {int(limit)}")
    sent = 0
    for item in q:
        ok, resp = telegram_send_message(item.get("message",""), item.get("event_type","system"), item.get("chat_id") or TELEGRAM_CHAT_ID, force=True)
        if ok:
            execute(f"UPDATE {table} SET sent=1, attempts=attempts+1, last_error='' WHERE id=?", (item["id"],))
            sent += 1
        else:
            execute(f"UPDATE {table} SET attempts=attempts+1, last_error=? WHERE id=?", (str(resp)[:1000], item["id"]))
    return sent


def run_db_indexes():
    for name, table, col in [
        ("idx_requests_status","requests","status"),
        ("idx_requests_created_at","requests","created_at"),
        ("idx_requests_patient_name","requests","patient_name"),
        ("idx_stock_component","stock_entries","component"),
        ("idx_stock_expiry","stock_entries","expiry"),
        ("idx_audit_created_at","audit","created_at"),
    ]:
        try: execute(f"CREATE INDEX IF NOT EXISTS {name} ON {table}({col})")
        except Exception as e: log_suppressed_error('suppressed', e)

def cleanup_old_backups():
    try:
        cutoff = datetime.now() - timedelta(days=BACKUP_KEEP_DAYS)
        for b in rows("SELECT * FROM backups ORDER BY id"):
            try:
                dt = datetime.strptime(b["created_at"], "%Y-%m-%d %H:%M:%S")
                if dt < cutoff:
                    fp=os.path.join(BACKUP_DIR,b["filename"])
                    if os.path.exists(fp): os.remove(fp)
                    execute("DELETE FROM backups WHERE id=?", (b["id"],))
            except Exception as e: log_suppressed_error('suppressed', e)
    except Exception as e: log_suppressed_error('suppressed', e)

def make_rollback_snapshot(label="manual"):
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        fn=f"rollback_{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        p=os.path.join(BACKUP_DIR,fn)
        with zipfile.ZipFile(p,"w",zipfile.ZIP_DEFLATED) as z:
            if os.path.exists(DB_PATH): z.write(DB_PATH, os.path.basename(DB_PATH))
            for folder in ["templates","static"]:
                if os.path.exists(folder):
                    for root,dirs,files in os.walk(folder):
                        for f in files:
                            full=os.path.join(root,f); z.write(full, full)
        execute("INSERT INTO backups(created_at,filename,created_by) VALUES(?,?,?)", (now(),fn,"rollback"))
        return p
    except Exception:
        return None

def nightly_backup_if_due():
    if not AUTO_BACKUP_ENABLED: return False
    try:
        if datetime.now().hour != AUTO_BACKUP_HOUR: return False
        last=row("SELECT created_at FROM backups WHERE created_by='nightly' ORDER BY id DESC LIMIT 1")
        if last and datetime.strptime(last["created_at"], "%Y-%m-%d %H:%M:%S").date()==datetime.now().date(): return False
        make_backup("nightly"); cleanup_old_backups()
        try: telegram_event_backup(True)
        except Exception as e: log_suppressed_error('suppressed', e)
        return True
    except Exception:
        try: telegram_event_backup(False)
        except Exception as e: log_suppressed_error('suppressed', e)
        return False

def _human_age_from_hours(age):
    if age is None:
        return "немає резервної копії"
    try:
        age=float(age)
    except Exception:
        return "невідомо"
    if age < 1:
        return f"{round(age*60)} хв"
    if age < 48:
        return f"{round(age,1)} год"
    return f"{round(age/24,1)} діб"


def _parse_backup_dt(value):
    """V6.4.40: parse backup timestamps from DB rows or file mtime fallback."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    txt=str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f"):
        try:
            return datetime.strptime(txt.replace("Z", ""), fmt)
        except Exception as e:
            log_suppressed_error('suppressed', e)
    try:
        return datetime.fromisoformat(txt.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def _latest_backup_info():
    """V6.4.40: show last backup from DB; if record is missing, detect newest ZIP/DB backup file in BACKUP_DIR."""
    latest_backup_at=None; latest_backup_filename=None; latest_backup_by=None; latest_dt=None
    try:
        b=row("SELECT created_at, filename, created_by FROM backups ORDER BY id DESC LIMIT 1")
        if b:
            latest_backup_at=b.get("created_at")
            latest_backup_filename=b.get("filename")
            latest_backup_by=b.get("created_by")
            latest_dt=_parse_backup_dt(latest_backup_at)
    except Exception as e:
        log_suppressed_error('suppressed', e)
    if latest_dt is None:
        try:
            if os.path.isdir(BACKUP_DIR):
                candidates=[]
                for name in os.listdir(BACKUP_DIR):
                    lower=name.lower()
                    if lower.endswith((".zip", ".db", ".sqlite", ".sqlite3", ".json")):
                        full=os.path.join(BACKUP_DIR, name)
                        if os.path.isfile(full):
                            candidates.append((os.path.getmtime(full), name))
                if candidates:
                    mtime, name=max(candidates)
                    latest_dt=datetime.fromtimestamp(mtime)
                    latest_backup_at=latest_dt.strftime("%Y-%m-%d %H:%M:%S")
                    latest_backup_filename=name
                    latest_backup_by=latest_backup_by or "file"
        except Exception as e:
            log_suppressed_error('suppressed', e)
    age=None
    if latest_dt:
        age=round(max(0, (datetime.now()-latest_dt).total_seconds())/3600, 2)
    return age, latest_backup_at, latest_backup_filename, latest_backup_by


def health_payload():
    ok=True; err=""
    try:
        row("SELECT 1 AS ok")
    except Exception as e:
        ok=False; err=str(e)
    age, latest_backup_at, latest_backup_filename, latest_backup_by = _latest_backup_info()
    telegram_ok=bool(TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID)
    backup_ok=age is not None
    try:
        admin_count=len(rows("SELECT id FROM users WHERE role='admin' AND active=1"))
    except Exception:
        admin_count=0
    return {
        "ok": ok,
        "version": APP_VERSION if 'APP_VERSION' in globals() else "V6.4.70",
        "database": "ok" if ok else "error",
        "database_status_label": "✅ Працює" if ok else "❌ Помилка",
        "database_error": err,
        "postgres": IS_POSTGRES,
        "database_type": "PostgreSQL" if IS_POSTGRES else "SQLite",
        "telegram_configured": telegram_ok,
        "telegram_status_label": "✅ Налаштовано" if telegram_ok else "⚠️ Не налаштовано",
        "backup_age_hours": age,
        "backup_age_label": _human_age_from_hours(age),
        "latest_backup_at": latest_backup_at,
        "latest_backup_filename": latest_backup_filename,
        "latest_backup_by": latest_backup_by,
        "backup_status_label": "✅ " + _human_age_from_hours(age) if backup_ok else "⚠️ немає резервної копії",
        "auto_backup_enabled": AUTO_BACKUP_ENABLED,
        "admin_count": admin_count,
        "time": now()
    }


def ensure_default_admin():
    """V5.6.4: створює першого admin, якщо в БД немає активного admin."""
    try:
        admin = row("SELECT * FROM users WHERE role='admin' AND active=1 LIMIT 1")
        if admin:
            return False
        existing = row("SELECT * FROM users WHERE username=?", ("Sepsis",))
        if existing:
            execute("UPDATE users SET role='admin', active=1, must_change_password=1 WHERE username=?", ("Sepsis",))
            return True
        create_user("Sepsis","1986","admin","Адміністратор","Завідувач",1,1)
        try:
            audit("admin_bootstrap", "Default admin Sepsis created")
        except Exception:
            db_rollback_safe()
            pass
        return True
    except Exception as e:
        try:
            print("ADMIN_BOOTSTRAP_ERROR:", e)
        except Exception:
            db_rollback_safe()
            pass
        return False


def telegram_user_enabled(user_id=None, event_type=None):
    try:
        if user_id is None:
            u = current_user()
            user_id = u.get("id")
        u = row("SELECT * FROM users WHERE id=?", (user_id,))
        if not u:
            return False
        if not u.get("telegram_enabled") or not u.get("telegram_chat_id"):
            return False
        if event_type == "new_request" and int(u.get("telegram_notify_new_requests") or 0) != 1:
            return False
        if event_type == "critical" and int(u.get("telegram_notify_critical") or 0) != 1:
            return False
        if event_type == "expiring" and int(u.get("telegram_notify_expiring") or 0) != 1:
            return False
        if event_type == "reaction" and int(u.get("telegram_notify_reactions") or 0) != 1:
            return False
        if event_type == "backup" and int(u.get("telegram_notify_backups") or 0) != 1:
            return False
        return True
    except Exception:
        return False

def telegram_send_to_user(user_id, message, event_type="system", force=False):
    try:
        u = row("SELECT * FROM users WHERE id=?", (user_id,))
        if not u or not u.get("telegram_chat_id"):
            return False, "user telegram not configured"
        if not telegram_user_enabled(user_id, event_type) and not force:
            return False, "user telegram disabled"
        return telegram_send_message(message, event_type=event_type, chat_id=u.get("telegram_chat_id"), force=force)
    except Exception as e:
        return False, str(e)

def telegram_broadcast_roles(message, roles=("admin","transfusion"), event_type="system", force=False):
    sent = 0
    try:
        placeholders = ",".join(["?"]*len(roles))
        users = rows(f"SELECT * FROM users WHERE active=1 AND role IN ({placeholders})", tuple(roles))
        for u in users:
            ok, _ = telegram_send_to_user(u["id"], message, event_type=event_type, force=force)
            if ok:
                sent += 1
    except Exception as e:
        log_suppressed_error('suppressed', e)
    # fallback to global chat id
    if sent == 0 and TELEGRAM_CHAT_ID:
        ok, _ = telegram_send_message(message, event_type=event_type, force=force)
        if ok:
            sent += 1
    return sent

def telegram_link_url(user=None):
    try:
        u = user or current_user()
        username = TELEGRAM_BOT_USERNAME.strip().lstrip("@")
        if not username:
            return ""
        return f"https://t.me/{username}?start=link_{u.get('id')}_{u.get('username')}"
    except Exception:
        return ""

def telegram_process_update(update):
    try:
        msg = update.get("message") or update.get("edited_message") or {}
        text = (msg.get("text") or "").strip()
        chat = msg.get("chat") or {}
        from_user = msg.get("from") or {}
        chat_id = str(chat.get("id") or "")
        tg_username = from_user.get("username") or chat.get("username") or ""
        if not chat_id:
            return "no chat"

        tg_user = telegram_user_by_chat_id(chat_id)
        tg_role = (tg_user.get("role") if tg_user else "") or "guest"
        is_staff = tg_role in ("admin", "transfusion")

        if text.startswith("/start"):
            parts = text.split(maxsplit=1)
            payload = parts[1] if len(parts) > 1 else ""
            if payload.startswith("link_"):
                bits = payload.split("_")
                uid = None
                try:
                    uid = int(bits[1])
                except Exception:
                    uid = None
                if uid:
                    execute("""UPDATE users SET telegram_chat_id=?, telegram_username=?, telegram_enabled=1
                               WHERE id=?""", (chat_id, tg_username, uid))
                    telegram_send_message("✅ Telegram підключено до користувача в системі Банк крові.", "link", chat_id=chat_id, force=True)
                    return "linked"
            telegram_send_message("👋 Бот Банку крові активний. Команди: /requests /help" + (" /stock /critical /expiring" if is_staff else ""), "start", chat_id=chat_id, force=True)
            return "start"

        if text.startswith("/help"):
            msg = "Команди: /requests, /help"
            if is_staff:
                msg = "Команди: /stock, /critical, /requests, /expiring, /help"
            telegram_send_message(msg, "command", chat_id=chat_id, force=True)
            return "help"

        if text.startswith("/stock") or text.startswith("/critical") or text.startswith("/expiring"):
            if not is_staff:
                telegram_send_message("⛔ Ця команда доступна тільки адміну або трансфузіологу.", "command", chat_id=chat_id, force=True)
                return "forbidden"

        if text.startswith("/stock"):
            stock = rows("SELECT component, donor_group, donor_rh, SUM(amount) AS total FROM stock_entries GROUP BY component, donor_group, donor_rh ORDER BY component LIMIT 25")
            lines = ["📦 <b>Склад крові</b>"]
            for s in stock:
                lines.append(f"{s.get('component','')} {s.get('donor_group','')} {s.get('donor_rh','')}: {s.get('total',0)}")
            telegram_send_message("\\n".join(lines) if len(lines)>1 else "Склад порожній", "command", chat_id=chat_id, force=True)
            return "stock"

        if text.startswith("/critical"):
            alerts = get_alerts_data() if "get_alerts_data" in globals() else {}
            low = alerts.get("low", []) if isinstance(alerts, dict) else []
            lines = ["🔴 <b>Критичні залишки</b>"]
            for x in low[:20]:
                lines.append(f"{x.get('component','')} {x.get('donor_group','')} {x.get('donor_rh','')}: {x.get('amount','')}")
            telegram_send_message("\\n".join(lines) if len(lines)>1 else "✅ Критичних залишків немає", "command", chat_id=chat_id, force=True)
            return "critical"

        if text.startswith("/requests"):
            req = rows("SELECT id,patient_name,department,component,status FROM requests ORDER BY id DESC LIMIT 10")
            lines = ["📋 <b>Останні вимоги</b>"]
            for r in req:
                lines.append(f"№{r.get('id')} {r.get('patient_name','')} — {r.get('component','')} — {r.get('status','')}")
            telegram_send_message("\\n".join(lines) if len(lines)>1 else "Вимог немає", "command", chat_id=chat_id, force=True)
            return "requests"

        if text.startswith("/expiring"):
            alerts = get_alerts_data() if "get_alerts_data" in globals() else {}
            exp = alerts.get("expiry", []) if isinstance(alerts, dict) else []
            lines = ["⏰ <b>Термін придатності</b>"]
            for x in exp[:20]:
                lines.append(f"{x.get('component','')} {x.get('donor_group','')} {x.get('donor_rh','')} до {x.get('expiry','')}: {x.get('amount','')}")
            telegram_send_message("\\n".join(lines) if len(lines)>1 else "✅ Немає близьких термінів", "command", chat_id=chat_id, force=True)
            return "expiring"

        telegram_send_message("Команди: /requests /help" + (" /stock /critical /expiring" if is_staff else ""), "command", chat_id=chat_id, force=True)
        return "unknown"
    except Exception as e:
        return f"error: {e}"



@app.before_request
def before():
    # V593_PUBLIC_EMERGENCY_PATHS
    if request.path in ["/api/health-debug","/api/emergency-db-fix","/api/public-health"]:
        return None
    # V592_CSRF_EXEMPT_SECURITY_LOGIN_ATTEMPT
    if request.path == "/api/security/login-attempt":
        return None
    if REQUIRE_HTTPS and request.headers.get("X-Forwarded-Proto","https") != "https":
        return redirect(request.url.replace("http://","https://"), code=301)
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "local").split(",")[0]
    t = time.time()
    if BANS.get(ip,0) > t:
        return jsonify(ok=False,error="Забагато запитів"), 429
    key = (ip, int(t//60))
    RATE[key] = RATE.get(key,0)+1
    if RATE[key] > 240:
        BANS[ip] = t + 120
        return jsonify(ok=False,error="Rate limit"), 429
    if session.get("user_id"):
        last = session.get("last_seen", t)
        if t - last > SESSION_TIMEOUT_MINUTES*60:
            session.clear()
            if request.path.startswith("/api/"):
                return jsonify(ok=False,error="Сесія завершена"), 401
            return redirect(url_for("index"))
        session["last_seen"] = t
    if request.method in ["POST","PUT","DELETE"] and request.path.startswith("/api/"):
        if request.path.startswith("/api/external/") and API_TOKEN and request.headers.get("X-API-Token","") == API_TOKEN:
            pass
        elif session.get("csrf") != request.headers.get("X-CSRF-Token"):
            return jsonify(ok=False,error="CSRF token invalid"), 403

@app.after_request
def headers(resp):
    resp.headers["X-Frame-Options"]="SAMEORIGIN"
    resp.headers["X-Content-Type-Options"]="nosniff"
    resp.headers["Referrer-Policy"]="same-origin"
    resp.headers["Content-Security-Policy"]="default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data:; connect-src 'self'"
    return resp

def login_required(f):
    @wraps(f)
    def w(*a, **kw):
        if not current_user(): return redirect(url_for("index"))
        return f(*a, **kw)
    return w

def role_required(*roles):
    def deco(f):
        @wraps(f)
        def w(*a, **kw):
            u = current_user()
            if not u or u["role"] not in roles:
                return jsonify(ok=False,error="Недостатньо прав"), 403
            return f(*a, **kw)
        return w
    return deco

# ================= V6.4.0 PERMISSION MATRIX =================
PERMISSIONS = {
    "admin": {"*"},
    "transfusion": {
        "view_stock","edit_stock","issue_blood","reserve_units","writeoff_units",
        "view_all_requests","approve_requests","view_all_patients","view_reports",
        "view_temperature","edit_temperature","view_audit","view_traceability",
        "export_reports","telegram_manage","pwa_manage"
    },
    "doctor": {
        "create_request","view_own_requests","view_own_patients",
        "mark_used_own","record_reaction_own","view_own_dashboard"
    },
    "nurse": {
        "view_own_requests","mark_used_own","record_reaction_own",
        "view_own_patients","view_own_dashboard"
    }
}

def has_perm(user, perm):
    if not user:
        return False
    perms = PERMISSIONS.get(user.get("role"), set())
    return "*" in perms or perm in perms

def permission_required(perm):
    def deco(f):
        @wraps(f)
        def w(*a, **kw):
            u = current_user()
            if not has_perm(u, perm):
                return jsonify(ok=False, error="Недостатньо прав", permission=perm), 403
            return f(*a, **kw)
        return w
    return deco

def owns_request(user, request_id):
    if not user:
        return False
    if user.get("role") in ("admin", "transfusion"):
        return True
    r = row("SELECT created_by FROM requests WHERE id=?", (int(request_id),))
    return bool(r and int(r.get("created_by") or 0) == int(user.get("id") or 0))

def can_access_request_for_current_user(request_id):
    u = current_user()
    if not u:
        return False
    if u.get("role") in ("admin", "transfusion"):
        return True
    if not request_id:
        return False
    return owns_request(u, int(request_id))

def forbid_json():
    return jsonify(ok=False, error="Недостатньо прав"), 403
# ================= END V6.4.40 PERMISSION MATRIX =================

@app.route("/")
def index():
    # V6.4.40 FAST LOGIN: do not run schema migrations on every page open.
    # Full DB checks run once at Gunicorn startup; this keeps login/opening fast on Render/mobile.
    if os.environ.get("RUN_STARTUP_CHECKS_EACH_REQUEST") == "1":
        try:
            safe_startup_check()
            ensure_default_admin()
        except Exception as e:
            log_suppressed_error('suppressed', e)
    if "csrf" not in session: session["csrf"] = secrets.token_hex(24)
    session["last_seen"] = time.time()
    u = current_user()
    if not u: return render_template("login.html", title=APP_TITLE, error=None)
    if u["first_login"]: return render_template("change_password.html", title=APP_TITLE, user=u, csrf=session["csrf"])
    return render_template("app.html", title=APP_TITLE, user=u, csrf=session["csrf"])

@app.route("/login", methods=["GET","POST"])
def login():
    # V603_LOGIN_GET_FIX
    if request.method == "GET":
        try:
            v60_db_safe_reset()
        except Exception as e:
            log_suppressed_error('suppressed', e)
        return render_template("login.html", title=APP_TITLE, error=None)
    # V595_LOGIN_ROLLBACK_BEFORE_SELECT
    db_rollback_safe()
    # V594_LOGIN_ROLLBACK_BEFORE_SELECT
    db_rollback_safe()
    # V6.4.40 FAST LOGIN: migrations are not executed during password submit.
    # They are expensive on Render/PostgreSQL and caused long login delay.
    # Enable RUN_STARTUP_CHECKS_EACH_REQUEST=1 only for emergency repair.
    if os.environ.get("RUN_STARTUP_CHECKS_EACH_REQUEST") == "1":
        try:
            safe_startup_check()
            ensure_telegram_user_columns_safe()
            ensure_default_admin()
        except Exception as e:
            try: print("LOGIN_OPTIONAL_STARTUP_CHECK_ERROR:", e)
            except Exception as e: log_suppressed_error('suppressed', e)
    username = request.form.get("username","").strip()
    password = request.form.get("password","")
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "")
    u = row("SELECT * FROM users WHERE username=?", (username,))
    ok = False
    msg = "Невірний логін або пароль"
    if u and u.get("locked_until"):
        try:
            if datetime.strptime(u["locked_until"], "%Y-%m-%d %H:%M:%S") > datetime.now():
                msg = "Акаунт тимчасово заблоковано"
            else:
                execute("UPDATE users SET locked_until=NULL,failed_logins=0 WHERE id=?", (u["id"],))
                u = row("SELECT * FROM users WHERE username=?", (username,))
        except Exception as e: log_suppressed_error('suppressed', e)
    if u and u["active"] and not u.get("locked_until") and password_ok(password, u["salt"], u["password_hash"]):
        ok = True
    execute("INSERT INTO login_attempts(created_at,username,ip,ok) VALUES(?,?,?,?)", (now(), username, ip, 1 if ok else 0))
    if not ok:
        if u:
            failed=(u.get("failed_logins") or 0)+1
            lock=None
            if failed>=5: lock=(datetime.now()+timedelta(minutes=15)).strftime("%Y-%m-%d %H:%M:%S")
            execute("UPDATE users SET failed_logins=?,locked_until=? WHERE id=?", (failed,lock,u["id"]))
            if failed >= 10:
                try: telegram_broadcast_roles(f"🚨 10 невдалих спроб входу для {username} з IP {ip}", ("admin","transfusion"), "security", True)
                except Exception as e: log_suppressed_error('suppressed', e)
        return render_template("login.html", title=APP_TITLE, error=msg)
    # V6.4.0 optional 2FA for admin/transfusion. Disabled by default until enabled in Security UI.
    try:
        if u.get("role") in ("admin", "transfusion"):
            ensure_v630_schema()
            mode = get_setting_v640("2fa_mode", "off")
            enabled = get_setting_v640("2fa_enabled", "0") == "1"
            if enabled and mode in ("telegram", "totp"):
                session.clear()
                session["pending_2fa_user_id"] = u["id"]
                session["csrf"] = secrets.token_hex(24)
                session["last_seen"] = time.time()
                if mode == "telegram":
                    code = str(secrets.randbelow(900000) + 100000)
                    execute("INSERT INTO two_factor_codes(created_at,user_id,code,expires_at,used) VALUES(?,?,?,?,0)",
                            (now(), u["id"], code, (datetime.now()+timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S")))
                    telegram_broadcast_roles(f"🔐 Код входу в Банк крові: <b>{code}</b>. Діє 5 хв.", ("admin","transfusion"), "security", True)
                return redirect(url_for("two_factor_page_v640"))
    except Exception:
        db_rollback_safe()
    session.clear(); session["user_id"] = u["id"]; session["csrf"] = secrets.token_hex(24)
    session["last_seen"] = time.time()
    execute("UPDATE users SET failed_logins=0,locked_until=NULL WHERE id=?", (u["id"],))
    audit("login","Вхід у систему")
    return redirect(url_for("index"))

@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))

@app.post("/change-password")
@login_required
def change_password():
    u = current_user()
    old = request.form.get("old_password","")
    p1 = request.form.get("new_password","")
    p2 = request.form.get("new_password2","")
    stored = row("SELECT * FROM users WHERE id=?", (u["id"],))
    if not password_ok(old, stored["salt"], stored["password_hash"]):
        return render_template("change_password.html", title=APP_TITLE, user=u, csrf=session["csrf"], error="Старий пароль невірний")
    if p1 != p2:
        return render_template("change_password.html", title=APP_TITLE, user=u, csrf=session["csrf"], error="Паролі не співпадають")
    ok,msg = password_policy_ok(p1)
    if not ok:
        return render_template("change_password.html", title=APP_TITLE, user=u, csrf=session["csrf"], error=msg)
    salt, ph = hash_password(p1)
    execute("UPDATE users SET salt=?,password_hash=?,first_login=0 WHERE id=?", (salt,ph,u["id"]))
    audit("change_password","Користувач змінив пароль")
    return redirect(url_for("index"))

@app.get("/api/me")
@login_required
def api_me(): return jsonify(current_user())

@app.get("/api/users")
@login_required
def api_users():
    u=current_user()
    if u["role"]=="admin":
        return jsonify(rows("SELECT id,username,role,full_name,position,active,created_at FROM users ORDER BY id"))
    if u["role"]=="transfusion":
        return jsonify(rows("SELECT id,username,role,full_name,position,active,created_at FROM users WHERE role IN ('doctor','nurse') ORDER BY id"))
    return jsonify(ok=False,error="Недостатньо прав"),403

@app.post("/api/users/create")
@role_required("admin","transfusion")
def api_user_create():
    u=current_user(); d=request.json or {}
    role=d.get("role")
    if u["role"]=="transfusion" and role not in ["doctor","nurse"]:
        return jsonify(ok=False,error="Трансфузіолог може створювати тільки лікарів і медсестер")
    if role not in ["admin","transfusion","doctor","nurse"]:
        return jsonify(ok=False,error="Невірна роль")
    username=(d.get("username") or "").strip()
    password=d.get("password") or ""
    ok,msg=password_policy_ok(password)
    if not username or not ok: return jsonify(ok=False,error=msg or "Логін обов’язковий")
    try:
        create_user(username,password,role,d.get("full_name",""),d.get("position",""),1,1)
        audit("create_user", username)
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False,error=str(e))

@app.post("/api/users/update")
@role_required("admin","transfusion")
def api_user_update():
    cu=current_user(); d=request.json or {}; uid,err=require_id_from_json(d)
    if err: return err
    target=row("SELECT * FROM users WHERE id=?", (uid,))
    if not target: return jsonify(ok=False,error="Користувача не знайдено")
    if target["username"]=="Sepsis" and cu["username"]!="Sepsis": return jsonify(ok=False,error="Sepsis може редагувати тільки Sepsis")
    if cu["role"]=="transfusion" and target["role"] not in ["doctor","nurse"]: return jsonify(ok=False,error="Трансфузіолог керує тільки лікарями/медсестрами")
    allowed={}
    for k in ["full_name","position","role","active"]:
        if k in d: allowed[k]=d[k]
    if "role" in allowed:
        if cu["role"]=="transfusion" and allowed["role"] not in ["doctor","nurse"]: return jsonify(ok=False,error="Трансфузіолог може призначати тільки doctor/nurse")
        if allowed["role"] not in ["admin","transfusion","doctor","nurse"]: return jsonify(ok=False,error="Невірна роль")
    if "username" in d:
        username=(d.get("username") or "").strip()
        if len(username)<3: return jsonify(ok=False,error="Логін мінімум 3 символи")
        if row("SELECT id FROM users WHERE username=? AND id<>?",(username,uid)): return jsonify(ok=False,error="Такий логін вже існує")
        allowed["username"]=username
    if "password" in d and d.get("password"):
        ok,msg=password_policy_ok(d["password"])
        if not ok: return jsonify(ok=False,error=msg)
        salt,ph=hash_password(d["password"])
        allowed["salt"]=salt; allowed["password_hash"]=ph; allowed["first_login"]=1
    if "active" in allowed: allowed["active"]=1 if allowed["active"] else 0
    if not allowed: return jsonify(ok=False,error="Немає змін")
    execute("UPDATE users SET "+",".join([f"{k}=?" for k in allowed])+" WHERE id=?", tuple(list(allowed.values())+[uid]))
    audit("update_user", "user updated")
    notify(uid,"Профіль змінено","Ваш профіль змінено.")
    return jsonify(ok=True, password_changed=("password_hash" in allowed))

@app.get("/api/requests")
@login_required
def api_requests():
    u=current_user()
    if u["role"] in ("doctor", "nurse"):
        return jsonify(rows("SELECT * FROM requests WHERE created_by=? ORDER BY id DESC", (u["id"],)))
    return jsonify(rows("SELECT * FROM requests ORDER BY id DESC"))

@app.get("/api/requests/mine")
@login_required
def api_requests_mine():
    return jsonify(rows("SELECT * FROM requests WHERE created_by=? ORDER BY id DESC",(current_user()["id"],)))

@app.post("/api/request/create")
@role_required("doctor","admin","transfusion")
def api_request_create():
    u=current_user(); d=request.json or {}
    for k in ["patient_name","birth_date","department","component","patient_group","patient_rh","amount","diagnosis"]:
        if not d.get(k): return jsonify(ok=False,error=f"Заповніть поле: {FIELD_LABELS_UA.get(k,k)}", field=k), 400
    execute("""INSERT INTO requests(created_at,created_by,doctor_name,doctor_position,patient_name,birth_date,address,patient_status,department,component,patient_group,patient_rh,amount,urgency,needed_date,diagnosis,note,status)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (now(),u["id"],u["full_name"],u["position"],d["patient_name"],d["birth_date"],d.get("address",""),d.get("patient_status",""),d["department"],d["component"],d["patient_group"],d["patient_rh"],float(d["amount"]),d.get("urgency","Планово"),d.get("needed_date",""),d["diagnosis"],d.get("note",""),"Нова"))
    audit("create_request", d["patient_name"])
    telegram_event_new_request(d)
    return jsonify(ok=True)

@app.post("/api/request/update")
@role_required("admin","transfusion","doctor","nurse")
def api_request_update_v6414():
    """V6.4.61: безпечне редагування вимоги.

    Критичні поля обробленої вимоги не можна міняти напряму, бо це обходить
    складські рухи, blood_units і аудит. Для погодження/видачі/скасування
    використовуються окремі контрольовані дії.
    """
    u=current_user(); d=request.json or {}; rid,err=require_id_from_json(d)
    if err: return err
    req=row("SELECT * FROM requests WHERE id=?", (rid,))
    if not req: return jsonify(ok=False,error="Вимогу не знайдено"),404
    st=(req.get("status") or "").strip()
    if u.get("role") in ("doctor","nurse"):
        if not owns_request(u, rid): return jsonify(ok=False,error="Можна редагувати тільки свої вимоги"),403
        if st not in ("Нова", "Чернетка", ""):
            return jsonify(ok=False,error="Після обробки вимогу редагує тільки трансфузіолог або адміністратор через контрольовані дії"),403

    allowed={}
    all_fields=["patient_name","birth_date","address","patient_status","department","component","patient_group","patient_rh","amount","urgency","needed_date","diagnosis","note","status","donor_group","donor_rh","pack_no","series","expiry"]
    for k in all_fields:
        if k in d: allowed[k]=d.get(k)

    critical={"status","component","amount","patient_group","patient_rh","needed_date","donor_group","donor_rh","pack_no","series","expiry","unit_id"}
    locked_statuses={"Погоджено","Зарезервовано","Резерв","Видано","Використано","Списано","Відмовлено","Потребує перегляду"}
    attempted_critical=[k for k in allowed if k in critical]
    if st in locked_statuses and attempted_critical:
        return jsonify(ok=False,error="Критичні поля обробленої вимоги не можна редагувати напряму. Використайте погодження, резервування, видачу, скасування видачі або списання."),409
    # Навіть для нової вимоги admin/transfusion не повинні вручну ставити фінальний статус через update.
    if "status" in allowed:
        return jsonify(ok=False,error="Статус вимоги змінюється тільки через контрольовані дії: погодити, відмовити, резервувати, видати, скасувати видачу, використано або списати."),409

    if "amount" in allowed:
        try: allowed["amount"]=float(str(allowed["amount"]).replace(",","."))
        except Exception: return jsonify(ok=False,error="Кількість має бути числом"),400
        if allowed["amount"] <= 0:
            return jsonify(ok=False,error="Кількість має бути більше 0"),400
    if u.get("role") in ("doctor","nurse"):
        for k in ("donor_group","donor_rh","pack_no","series","expiry"):
            allowed.pop(k, None)
    if not allowed: return jsonify(ok=False,error="Немає змін")
    execute("UPDATE requests SET "+",".join([f"{k}=?" for k in allowed])+" WHERE id=?", tuple(list(allowed.values())+[rid]))
    audit("request_update", f"requests:{rid}")
    return jsonify(ok=True)

@app.post("/api/request/delete")
@role_required("admin","transfusion","doctor","nurse")
def api_request_delete_v6414():
    """V6.4.61: заборона видалення оброблених вимог.

    Видаляти напряму можна лише чернетки/нові/відмовлені без активної видачі.
    Видано/Використано/Списано та зарезервовані/погоджені вимоги не видаляються
    простим delete, щоб не залишити stock_entries/blood_units із посиланням на
    відсутню вимогу.
    """
    u=current_user(); d=request.json or {}; rid,err=require_id_from_json(d)
    if err: return err
    req=row("SELECT * FROM requests WHERE id=?", (rid,))
    if not req: return jsonify(ok=False,error="Вимогу не знайдено"),404
    st=(req.get("status") or "").strip()
    blocked={"Погоджено","Зарезервовано","Резерв","Видано","Використано","Списано","Потребує перегляду"}
    if st in blocked:
        return jsonify(ok=False,error=f"Вимогу зі статусом «{st}» не можна видалити напряму. Спочатку виконайте контрольовану дію: скасувати видачу/анулювати/архівувати."),409
    if u.get("role") in ("doctor","nurse"):
        if not owns_request(u, rid): return jsonify(ok=False,error="Можна видаляти тільки свої вимоги"),403
        if st not in ("Нова", "Чернетка", ""):
            return jsonify(ok=False,error="Оброблені вимоги видаляє тільки трансфузіолог або адміністратор"),403
    # Якщо це відмовлена вимога з випадково зарезервованою одиницею — спершу звільнимо резерв.
    if req.get("unit_id"):
        unit=row("SELECT * FROM blood_units WHERE id=?", (int(req.get("unit_id")),))
        if unit and (unit.get("status") or "") == "reserved":
            execute("UPDATE blood_units SET status='available', request_id=NULL, patient_name='', updated_at=? WHERE id=?", (now(), int(req.get("unit_id"))))
            _recalc_blood_unit_status_for_id(req.get("unit_id"), f"Видалено необроблену вимогу №{rid}")
            unit_event(req.get("unit_id"), "request_deleted_release", f"Видалено вимогу №{rid}", rid)
    execute("INSERT INTO trash(created_at,source_table,source_id,data,deleted_by,reason) VALUES(?,?,?,?,?,?)", (now(),"requests",rid,json.dumps(dict(req),ensure_ascii=False),u.get("username"),d.get("reason") or "Видалено вимогу"))
    execute("DELETE FROM requests WHERE id=?", (rid,))
    audit("request_delete", f"requests:{rid}")
    return jsonify(ok=True)

@app.post("/api/request/action")
@role_required("admin","transfusion")
def api_request_action():
    u=current_user(); d=request.json or {}; rid,err=require_id_from_json(d)
    if err: return err
    action=d.get("action")
    req = row("SELECT * FROM requests WHERE id=?", (rid,))
    if not req:
        return jsonify(ok=False,error="Вимогу не знайдено"), 404
    actor = u.get("full_name") or u.get("username") or ""
    if action=="approve":
        if (req.get("status") or "").strip() not in ("Нова", "Чернетка", "", "Відмовлено", "Потребує перегляду"):
            return jsonify(ok=False,error="Погодити можна тільки нову/чернеткову або відмовлену вимогу"),409
        execute("UPDATE requests SET status='Погоджено',approved_by=? WHERE id=?", (actor, rid))
        audit("request_approve", str(rid))
        return jsonify(ok=True)
    elif action=="reject":
        if (req.get("status") or "").strip() in ("Видано", "Використано", "Списано"):
            return jsonify(ok=False,error="Видану/використану/списану вимогу не можна відмовити напряму"),409
        if req.get("unit_id"):
            unit=row("SELECT * FROM blood_units WHERE id=?", (int(req.get("unit_id")),))
            if unit and (unit.get("status") or "") == "reserved":
                execute("UPDATE blood_units SET status='available', request_id=NULL, patient_name='', updated_at=? WHERE id=?", (now(), int(req.get("unit_id"))))
                _recalc_blood_unit_status_for_id(req.get("unit_id"), f"Відмова від вимоги №{rid}")
                unit_event(req.get("unit_id"), "reserve_released", "Вимогу відмовлено", rid)
        execute("UPDATE requests SET status='Відмовлено', unit_id=NULL WHERE id=?", (rid,))
        audit("request_reject", str(rid))
        return jsonify(ok=True)
    elif action=="reserve":
        if (req.get("status") or "").strip() not in ("Погоджено", "Нова", "Чернетка", ""):
            return jsonify(ok=False,error="Резервувати можна тільки нову або погоджену вимогу"),409
        unit = find_fefo_unit_for_request(req, d.get("donor_group",""), d.get("donor_rh",""), d.get("pack_no",""), d.get("series",""), d.get("qr_code",""))
        if not unit:
            return jsonify(ok=False,error="Немає доступної сумісної одиниці компонента для резервування"), 409
        ok1,w1 = abo_compatible(req.get("patient_group"), unit.get("donor_group"), req.get("component"))
        ok2,w2 = rh_compatible(req.get("patient_rh"), unit.get("donor_rh"))
        if not (ok1 and ok2):
            return jsonify(ok=False,error="; ".join([x for x in [w1,w2] if x]) or "Несумісність"), 409
        ok_res, msg_res, bal_res, qty_res = _can_reserve_unit_for_request(req, unit)
        if not ok_res:
            return jsonify(ok=False,error=msg_res),409
        if d.get("needed_date"):
            execute("UPDATE requests SET needed_date=? WHERE id=?", (d.get("needed_date"), rid))
            req = row("SELECT * FROM requests WHERE id=?", (rid,)) or req
        if not reserve_unit_for_request(req, unit, actor):
            return jsonify(ok=False,error=msg_res or "Не вдалося зарезервувати одиницю"),409
        audit("request_reserve", f"request={rid}; unit={unit.get('id')}; qty={qty_res:g}; lot_balance={bal_res:g}; needed_date={req.get('needed_date','')}")
        return jsonify(ok=True, unit=unit, message="Одиницю зарезервовано за FEFO")
    elif action=="issue":
        unit = None
        if req.get("unit_id"):
            unit = row("SELECT * FROM blood_units WHERE id=?", (int(req.get("unit_id")),))
        if not unit:
            unit = find_fefo_unit_for_request(req, d.get("donor_group",""), d.get("donor_rh",""), d.get("pack_no",""), d.get("series",""), d.get("qr_code",""))
        if not unit:
            return jsonify(ok=False,error="Немає доступної сумісної одиниці компонента для видачі"), 409
        if (req.get("status") or "").strip() not in ("Погоджено", "Зарезервовано", "Резерв"):
            return jsonify(ok=False,error="Видача дозволена тільки після погодження або резервування вимоги"), 409
        if unit.get("status") not in ("available","reserved"):
            return jsonify(ok=False,error=f"Одиниця недоступна: {unit.get('status')}"), 409
        if expiry_status(unit.get("expiry")) == "expired":
            execute("UPDATE blood_units SET status='expired', updated_at=? WHERE id=?", (now(), unit.get("id")))
            return jsonify(ok=False,error="Термін придатності компонента завершився. Видачу заблоковано."), 409
        try:
            issue_qty = float(str(req.get("amount") or 1).replace(",", "."))
        except Exception:
            return jsonify(ok=False,error="Кількість у вимозі має бути числом"), 400
        if issue_qty <= 0:
            return jsonify(ok=False,error="Кількість у вимозі має бути більше 0"), 400
        # V6.4.61: видача не може забирати кількість, зарезервовану під інші вимоги.
        ok_issue_res, msg_issue_res, avail_issue_res, qty_issue_res = _can_issue_unit_for_request(req, unit)
        if not ok_issue_res:
            return jsonify(ok=False,error=msg_issue_res),409
        # V6.4.61: видача по вимозі має перевіряти фактичний залишок конкретної партії.
        # Раніше можна було видати 2 з P1/S1, де на складі була лише 1 одиниця.
        receipt, lot_qty, lot_err = _stock_find_receipt_for_outgoing(
            unit.get("component"), unit.get("donor_group"), unit.get("donor_rh"), issue_qty,
            unit.get("pack_no") or "", unit.get("series") or "", None
        )
        if not receipt:
            return jsonify(ok=False,error=f"Недостатньо в партії {unit.get('pack_no') or '—'}/{unit.get('series') or '—'}: доступно {lot_qty:g}, спроба видати {issue_qty:g}"), 409
        real_unit = _find_unit_for_stock_receipt(receipt)
        if not real_unit or int(real_unit.get('id') or 0) != int(unit.get('id') or 0):
            return jsonify(ok=False,error="Видачу заблоковано: партія складу не відповідає blood_units цієї одиниці"), 409
        ok1,w1 = abo_compatible(req["patient_group"], unit.get("donor_group"), req["component"])
        ok2,w2 = rh_compatible(req["patient_rh"], unit.get("donor_rh"))
        warning = "; ".join([x for x in [w1,w2] if x])
        comp_ok = 1 if (ok1 and ok2) else 0
        if not comp_ok and not d.get("override"):
            return jsonify(ok=False,error=warning or "Несумісність крові"), 409
        # Дані видачі беремо з підтвердженого надходження/партії, а не з довільного JSON.
        unit = real_unit
        execute("""UPDATE requests SET status='Видано',issued_by=?,issued_at=?,donor_group=?,donor_rh=?,pack_no=?,series=?,expiry=?,compatibility_ok=?,compatibility_warning=?,unit_id=? WHERE id=?""",
                (actor,now(),unit.get("donor_group",""),unit.get("donor_rh",""),unit.get("pack_no",""),unit.get("series",""),unit.get("expiry",""),comp_ok,warning,int(unit.get("id")),rid))
        execute("UPDATE blood_units SET updated_at=?, request_id=?, patient_name=?, issued_by=?, issued_at=? WHERE id=?",
                (now(), rid, req.get("patient_name",""), actor, now(), int(unit.get("id"))))
        try:
            execute("""INSERT INTO stock_entries(created_at,user_id,type,component,donor_group,donor_rh,amount,pack_no,series,expiry,patient_name,note,qr_code,unit_id)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (now(), u["id"], "Видача", unit.get("component"), unit.get("donor_group"), unit.get("donor_rh"), issue_qty, unit.get("pack_no",""), unit.get("series",""), unit.get("expiry",""), req.get("patient_name",""), f"Видача по вимозі №{rid}", unit.get("qr_code",""), int(unit.get("id"))))
        except Exception as e:
            db_rollback_safe()
            return jsonify(ok=False,error=f"Помилка запису видачі: {str(e)}"),500
        after_issue=_recalc_blood_unit_status_for_id(unit.get("id"), f"Видача по вимозі №{rid}")
        try:
            if float((after_issue or {}).get('amount') or 0) < -0.000001:
                return jsonify(ok=False,error="Видачу заблоковано: після операції залишок партії став би від’ємним"),409
        except Exception as e:
            log_suppressed_error('suppressed', e)
        for _code in [unit.get("qr_code"), unit.get("pack_no"), unit.get("series")]:
            if _code:
                log_traceability(_code, "issued", req.get("patient_name",""), rid, req.get("department",""), f"Видано по вимозі №{rid}")
        unit_event(unit.get("id"), "issued", f"Видано пацієнту {req.get('patient_name','')}", rid)
        audit("request_issue", f"request={rid}; unit={unit.get('id')}; comp_ok={comp_ok}")
        return jsonify(ok=True, unit=unit, compatibility_ok=comp_ok, warning=warning)
    else:
        return jsonify(ok=False,error="Невідома дія")

@app.post("/api/request/used")
@role_required("admin","transfusion","doctor","nurse")
def api_request_used():
    """V6.4.61: підтвердження використання вимоги без ламання залишку партії.

    Раніше цей endpoint напряму ставив blood_units.status='used'. Якщо з
    партії P1/S1 було видано 1 з 2, то у stock_entries лишався залишок 1,
    але blood_units уже ставав used. Тепер статус blood_units завжди
    перераховується по фактичному залишку партії. Якщо залишок >0 — партія
    лишається available і очищається від прив'язки до пацієнта цієї вимоги.
    Якщо залишок 0 — після перерахунку можна позначити одиницю як used.
    """
    u=current_user(); d=request.json or {}; rid,err=require_id_from_json(d)
    if err: return err
    if u.get("role") in ("doctor","nurse") and not owns_request(u, rid):
        return jsonify(ok=False,error="Можна підтверджувати тільки свої записи"), 403
    if not d.get("use_date") or not d.get("used_by") or not d.get("use_confirm"):
        return jsonify(ok=False,error="Заповніть дату, хто підтвердив і підтвердження"),400
    req=row("SELECT * FROM requests WHERE id=?", (rid,))
    if not req:
        return jsonify(ok=False,error="Вимогу не знайдено"),404
    if (req.get("status") or "").strip() != "Видано":
        return jsonify(ok=False,error="Використання можна підтвердити тільки після видачі компонента"),409
    if not req.get("unit_id"):
        return jsonify(ok=False,error="Використання неможливе: вимога не має прив’язаної одиниці компонента"),409

    execute("UPDATE requests SET status='Використано',used_at=?,used_by=?,use_confirm=? WHERE id=?",
            (d["use_date"],d["used_by"],d["use_confirm"],rid))

    unit_after=None
    if req.get("unit_id"):
        uid=int(req.get("unit_id"))
        unit_event(uid, "used", d.get("use_confirm",""), rid)
        unit_after=_recalc_blood_unit_status_for_id(uid, f"Підтверджено використання вимоги №{rid}")
        try:
            left=float((unit_after or {}).get('amount') or 0)
        except Exception:
            left=0.0
        # Статус used допустимий лише коли за цією партією фактичний залишок 0.
        # Для часткової видачі залишок >0, тому партія лишається available.
        if left <= 0.000001:
            execute("UPDATE blood_units SET status='used', updated_at=?, used_at=? WHERE id=?",
                    (now(), d["use_date"], uid))
            unit_after=row("SELECT * FROM blood_units WHERE id=?", (uid,))

    audit("request_used", str(rid))
    return jsonify(ok=True, unit=unit_after)


@app.post("/api/request/cancel-issue")
@role_required("admin","transfusion")
def api_request_cancel_issue_v6451():
    """V6.4.61: контрольоване скасування видачі замість прямого видалення руху складу.

    Дозволяємо скасувати лише вимогу зі статусом «Видано». Для «Використано»
    потрібне окреме адміністративне рішення/акт, тому тут блокуємо, щоб не
    зламати фінальний медичний запис.
    """
    u=current_user(); d=request.json or {}; rid,err=require_id_from_json(d)
    if err: return err
    req=row("SELECT * FROM requests WHERE id=?", (rid,))
    if not req:
        return jsonify(ok=False,error="Вимогу не знайдено"),404
    st=(req.get('status') or '').strip()
    if st != 'Видано':
        return jsonify(ok=False,error="Скасувати видачу можна тільки для вимоги зі статусом «Видано». Для «Використано» потрібне окреме анулювання актом."),409
    uid=req.get('unit_id')
    if not uid:
        return jsonify(ok=False,error="Вимога не має прив’язаної одиниці/партії"),409
    reason=d.get('reason') or 'Скасовано видачу по вимозі'
    issue_rows=rows("""SELECT * FROM stock_entries
                      WHERE unit_id=? AND type='Видача'
                        AND (note LIKE ? OR patient_name=?)
                      ORDER BY id DESC""", (int(uid), f"%№{rid}%", req.get('patient_name') or ''))
    if not issue_rows:
        # fallback за партією, якщо старий запис не мав unit_id, але має note з № вимоги
        issue_rows=rows("""SELECT * FROM stock_entries
                          WHERE type='Видача' AND note LIKE ?
                          ORDER BY id DESC""", (f"%№{rid}%",))
    if not issue_rows:
        return jsonify(ok=False,error="Не знайдено рух видачі по цій вимозі"),404
    rec=issue_rows[0]
    execute("INSERT INTO trash(created_at,source_table,source_id,data,deleted_by,reason) VALUES(?,?,?,?,?,?)",
            (now(),'stock_entries',int(rec.get('id')),json.dumps(dict(rec),ensure_ascii=False),u.get('username'),reason))
    execute("DELETE FROM stock_entries WHERE id=?", (int(rec.get('id')),))
    execute("""UPDATE requests SET status='Погоджено', issued_by='', issued_at=NULL,
              used_at=NULL, used_by='', use_confirm='', unit_id=NULL,
              pack_no='', series='', expiry='', donor_group='', donor_rh=''
              WHERE id=?""", (rid,))
    try:
        _recalc_blood_unit_status_for_id(uid, f"Скасовано видачу по вимозі №{rid}")
        unit_event(uid, 'issue_cancelled', reason, rid)
    except Exception:
        db_rollback_safe()
    audit('request_cancel_issue', f"request={rid}; stock_entry={rec.get('id')}; unit={uid}")
    return jsonify(ok=True, message="Видачу скасовано, вимогу повернено в статус «Погоджено»")

@app.post("/api/request/writeoff")
@role_required("admin","transfusion","doctor","nurse")
def api_request_writeoff():
    """V6.4.67: не плутати анулювання вимоги зі списанням компонента.

    Якщо компонент ще не видано, цей endpoint НЕ створює статус «Списано» без
    руху складу. Для резервованої/погодженої вимоги треба використати відмову
    або скасування резерву. «Списано» тут дозволено тільки після статусу
    «Видано», тобто коли рух «Видача» вже існує і компонент вийшов зі складу.
    """
    u=current_user(); d=request.json or {}; rid,err=require_id_from_json(d)
    if err: return err
    req=row("SELECT * FROM requests WHERE id=?", (rid,))
    if not req:
        return jsonify(ok=False,error="Вимогу не знайдено"),404
    role=(u.get("role") or "").lower()
    if role in ("doctor", "nurse") and not owns_request(u, rid):
        return jsonify(ok=False,error="Можна обробляти тільки свої вимоги"),403
    st=(req.get("status") or "").strip()
    if st in ("", "Нова", "Чернетка", "Очікує", "Очікує розгляду", "Погоджено", "Зарезервовано", "Резерв", "Потребує перегляду"):
        return jsonify(ok=False,error="Цей endpoint не анулює/не відмовляє вимогу. Списання можливе тільки після видачі компонента; для скасування використайте відмову або скасування резерву."),409
    if st in ("Використано", "Списано", "Відмовлено"):
        return jsonify(ok=False,error=f"Вимога вже має фінальний статус: {st}"),409
    if st != "Видано":
        return jsonify(ok=False,error="Списання вимоги можливе тільки зі статусу «Видано»"),409
    if not d.get("writeoff_date") or not d.get("written_by") or not d.get("writeoff_reason"):
        return jsonify(ok=False,error="Заповніть дату, хто списав і причину"),400
    uid=req.get("unit_id")
    if not uid:
        return jsonify(ok=False,error="Вимога не має прив’язаної одиниці компонента"),409
    # Переконуємось, що була реальна видача по цій вимозі.
    issue=row("""SELECT * FROM stock_entries WHERE unit_id=? AND type='Видача' AND (note LIKE ? OR patient_name=?) ORDER BY id DESC LIMIT 1""",
              (int(uid), f"%№{rid}%", req.get('patient_name') or ''))
    if not issue:
        return jsonify(ok=False,error="Не знайдено рух видачі по цій вимозі. Списання без руху складу заборонене."),409
    execute("UPDATE requests SET status='Списано',writeoff_at=?,written_by=?,writeoff_reason=? WHERE id=?", (d["writeoff_date"],d["written_by"],d["writeoff_reason"],rid))
    # V6.4.67: після списання виданого компонента статус blood_units не має
    # повертатися в issued після перерахунку. Якщо фактичний залишок партії 0 —
    # одиниця/партія має фінальний статус written_off. Якщо залишок >0, це
    # означає, що списана лише видана частина агрегованої партії, а активний
    # залишок партії лишається available/reserved і не прив'язується до цієї вимоги.
    execute("UPDATE blood_units SET updated_at=?, written_off_at=?, writeoff_reason=? WHERE id=?",
            (now(), d["writeoff_date"], d["writeoff_reason"], int(uid)))
    unit_after=_recalc_blood_unit_status_for_id(uid, f"Списання виданої вимоги №{rid}")
    try:
        left=float((unit_after or {}).get('amount') or 0)
    except Exception:
        left=0.0
    if left <= 0.000001:
        execute("""UPDATE blood_units SET status='written_off', amount=0, updated_at=?,
                  written_off_at=?, writeoff_reason=?, request_id=NULL, patient_name='',
                  issued_by='', issued_at=NULL, used_at=NULL WHERE id=?""",
                (now(), d["writeoff_date"], d["writeoff_reason"], int(uid)))
        unit_after=row("SELECT * FROM blood_units WHERE id=?", (int(uid),))
    unit_event(uid, "request_issued_component_written_off", d.get("writeoff_reason",""), rid)
    for _code in [req.get("pack_no"), req.get("series")]:
        if _code:
            log_traceability(_code, "request_writeoff_after_issue", req.get("patient_name",""), rid, req.get("department",""), d.get("writeoff_reason",""))
    audit("request_writeoff", str(rid))
    return jsonify(ok=True)

@app.post("/api/request/reaction")
@role_required("admin","transfusion","doctor","nurse")
def api_reaction():
    u=current_user(); d=request.json or {}; rid=int(d.get("id") or 0)
    if u.get("role") in ("doctor","nurse") and not owns_request(u, rid):
        return forbid_json()
    execute("""UPDATE requests SET reaction_present=?,reaction_type=?,reaction_severity=?,reaction_description=?,reaction_result=? WHERE id=?""",
            (d.get("reaction_present","Так"),d.get("reaction_type",""),d.get("reaction_severity",""),d.get("reaction_description",""),d.get("reaction_result",""),rid))
    audit("reaction", str(rid))
    return jsonify(ok=True)

@app.get("/api/doctor/reminders")
@role_required("doctor","admin","transfusion")
def api_doc_reminders():
    u=current_user()
    q="SELECT * FROM requests WHERE status='Видано' AND (used_at IS NULL OR writeoff_at IS NULL)"
    params=()
    if u["role"]=="doctor":
        q+=" AND created_by=?"; params=(u["id"],)
    return jsonify(rows(q+" ORDER BY id DESC", params))

@app.post("/api/stock/add")
@role_required("admin","transfusion")
def api_stock_add():
    try:
        u=current_user()
        d=request.json or {}
        component = d.get("component") or d.get("stock_component") or ""
        donor_group = d.get("donor_group") or d.get("stock_group") or d.get("group") or ""
        donor_rh = d.get("donor_rh") or d.get("stock_rh") or d.get("rh") or ""
        amount_raw = d.get("amount") or d.get("qty") or d.get("quantity") or ""
        stock_type = d.get("type") or "Надходження"
        stock_type_raw = str(stock_type).strip().lower()
        if stock_type_raw in ["in", "income", "add", "plus", "+", "надходження", "прихід"]:
            stock_type = "Надходження"
        elif stock_type_raw in ["out", "writeoff", "minus", "-", "списання", "видано", "видача"]:
            stock_type = "Списання"
        if not component:
            return jsonify(ok=False,error="Заповніть поле: Компонент"), 400
        if not amount_raw:
            return jsonify(ok=False,error="Заповніть поле: Кількість"), 400
        try:
            amount = float(str(amount_raw).replace(",", "."))
        except Exception:
            return jsonify(ok=False,error="Кількість має бути числом"), 400
        if amount <= 0:
            return jsonify(ok=False,error="Кількість має бути більше 0"), 400

        # V6.4.61: ручне списання/видача має перевіряти не тільки загальний залишок,
        # а конкретну партію/серію. Інакше можна було списати P2/S2, хоча на складі є лише P1/S1.
        pack_no = (d.get("pack_no") or "").strip()
        series = (d.get("series") or "").strip()
        expiry = (d.get("expiry") or "").strip()
        if stock_type != "Надходження":
            current_qty = _component_current_qty(component, donor_group, donor_rh)
            if current_qty < amount:
                return jsonify(ok=False,error=f"Недостатньо на складі: доступно {current_qty:g}, спроба списати {amount:g}"), 409
            if not (pack_no or series):
                return jsonify(ok=False,error="Для списання/видачі потрібно вказати пакет або серію"),400
            receipt, lot_qty, lot_err = _stock_find_receipt_for_outgoing(component, donor_group, donor_rh, amount, pack_no, series, d.get('stock_entry_id') or None)
            if not receipt:
                return jsonify(ok=False,error=f"Недостатньо в партії {pack_no or '—'}/{series or '—'}: доступно {lot_qty:g}, спроба списати {amount:g}"),409
            pack_no = (receipt.get('pack_no') or pack_no or '').strip()
            series = (receipt.get('series') or series or '').strip()
            expiry = (receipt.get('expiry') or expiry or '').strip()
            # V6.4.61: ручне списання/видача не може забирати активний резерв.
            available_qty, lot_balance, reserved_qty = _available_qty_for_stock_record(receipt)
            if amount > available_qty + 0.000001:
                return jsonify(ok=False,error=_reserved_guard_message(pack_no, series, available_qty, reserved_qty, lot_balance, 'списати/видати')),409

        d2={**d, "component": component, "donor_group": donor_group, "donor_rh": donor_rh, "pack_no": pack_no, "series": series, "expiry": expiry}
        unit_id = create_or_update_unit_from_stock(d2, amount, stock_type, u)
        cur = execute("""INSERT INTO stock_entries(created_at,user_id,type,component,donor_group,donor_rh,amount,pack_no,series,expiry,patient_name,note,qr_code,unit_id)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (now(),u["id"],stock_type,component,donor_group,donor_rh,amount,pack_no,series,expiry,d.get("patient_name",""),d.get("note",""),d.get("qr_code",""),unit_id))
        if unit_id:
            _recalc_blood_unit_status_for_id(unit_id, 'Рух складу: ' + stock_type)
        audit("stock_add", f"{component}; unit_id={unit_id}")
        return jsonify(ok=True, unit_id=unit_id, status="Одиницю компонента оновлено")
    except ValueError as e:
        audit("stock_add_conflict", str(e))
        return jsonify(ok=False,error=str(e)), 409
    except Exception as e:
        audit("stock_add_error", str(e))
        return jsonify(ok=False,error=f"Помилка складу: {str(e)}"), 500

@app.get("/api/stock")
@role_required("admin","transfusion")
def api_stock():
    delta = _stock_delta_sql()
    return jsonify(rows(f"""SELECT component,donor_group AS donor_group, donor_rh AS donor_rh,
                                  COALESCE(SUM({delta}),0) qty
                           FROM stock_entries
                           GROUP BY component,donor_group,donor_rh
                           HAVING COALESCE(SUM({delta}),0) <> 0
                           ORDER BY component"""))

def _stock_entry_lot_key(x):
    """Поштучний ключ партії/пакета для зв'язку надходження зі списанням/видачею.
    Якщо є серія або номер пакета — працюємо саме по них, щоб пакет, який списали,
    не продовжував виглядати як активне надходження.
    """
    return (
        _txt(x.get('component')).strip().lower(),
        moz_norm_group(x.get('donor_group') or x.get('group')).strip().lower(),
        moz_norm_rh(x.get('donor_rh') or x.get('rh')).strip().lower(),
        _txt(x.get('pack_no')).strip().lower(),
        _txt(x.get('series')).strip().lower(),
        _txt(x.get('expiry')).strip().lower(),
    )

def _annotate_stock_entry_reconciliation(items):
    """V6.4.40: позначає закриті надходження.
    Раніше після списання пакет залишався окремим рядком 'Надходження' і користувач
    сприймав його як активний залишок. Тепер такі надходження приховуються у звичайному
    списку або позначаються як 'закрито списанням/видачею'.
    """
    outs={}
    out_kind={}
    for x in items:
        kind=_moz_type_kind(x.get('type'))
        if kind not in ('issue','writeoff'):
            continue
        key=_stock_entry_lot_key(x)
        # Без серії/пакета не закриваємо конкретне надходження, щоб не списати випадковий рядок.
        if not (key[3] or key[4]):
            continue
        amt=_moz_amount(x.get('amount'))
        outs[key]=outs.get(key,0.0)+amt
        if kind=='writeoff': out_kind[key]='Списано'
        elif key not in out_kind: out_kind[key]='Видано'
    # розподіл списання/видачі по старіших надходженнях у межах одного ключа
    remaining_out=dict(outs)
    annotated=[]
    for x in sorted(items, key=lambda z: (str(z.get('created_at') or ''), int(z.get('id') or 0))):
        y=dict(x)
        kind=_moz_type_kind(y.get('type'))
        y['movement_kind']=kind
        y['is_closed_receipt']=False
        y['active_amount']=_moz_amount(y.get('amount'))
        y['closed_reason']=''
        if kind=='in':
            key=_stock_entry_lot_key(y)
            if (key[3] or key[4]) and remaining_out.get(key,0)>0:
                amt=_moz_amount(y.get('amount'))
                used=min(amt, remaining_out.get(key,0.0))
                remaining_out[key]=max(0.0, remaining_out.get(key,0.0)-used)
                active=max(0.0, amt-used)
                y['active_amount']=active
                if active <= 0.000001:
                    y['is_closed_receipt']=True
                    y['closed_reason']=out_kind.get(key,'Закрито')
                elif used>0:
                    y['closed_reason']=f'Частково закрито: {used:g}'
        annotated.append(y)
    return sorted(annotated, key=lambda z: int(z.get('id') or 0), reverse=True)

@app.get("/api/stock/entries")
@role_required("admin","transfusion")
def api_stock_entries_v6414():
    items=rows("""SELECT id,created_at,type,component,donor_group,donor_rh,amount,pack_no,series,expiry,patient_name,note,qr_code,unit_id
                                         FROM stock_entries ORDER BY id DESC LIMIT 1000""")
    items=_annotate_stock_entry_reconciliation(items)
    include_closed=(request.args.get('include_closed') or '').lower() in ('1','true','yes','так')
    if not include_closed:
        # У робочому списку не показуємо надходження, які вже повністю списані/видані.
        # Сам рух не видаляється: він лишається у простежуваності, звіті та при include_closed=1.
        items=[x for x in items if not (x.get('movement_kind')=='in' and x.get('is_closed_receipt'))]
    return jsonify(ok=True, items=items, count=len(items), include_closed=include_closed)

def _stock_normalize_type(t):
    """Єдина нормалізація типу руху складу для ручного редагування."""
    x=str(t or '').strip().lower()
    if x in ["in", "income", "add", "plus", "+", "надходження", "прихід"]:
        return "Надходження"
    if x in ["out", "issue", "issued", "writeoff", "written_off", "minus", "-", "списання", "списано", "видано", "видача"]:
        return "Списання"
    return str(t or '').strip() or "Надходження"


def _stock_entry_delta_py(x):
    """Python-еквівалент _stock_delta_sql() для перевірки редагування до UPDATE."""
    try:
        amount=abs(float(str((x or {}).get('amount') or 0).replace(',', '.')))
    except Exception:
        amount=0.0
    kind=_moz_type_kind((x or {}).get('type'))
    if kind in ('issue','writeoff'):
        return -amount
    return amount


def _stock_entry_group_key(x):
    return (
        (x or {}).get('component') or '—',
        (x or {}).get('donor_group') or '—',
        (x or {}).get('donor_rh') or '—',
    )


def _stock_entry_lot_identity(x):
    """Ключ партії без терміну придатності: використовується для заборони
    відриву списання/видачі від уже прийнятого пакета/серії.
    """
    return (
        _txt((x or {}).get('component')).strip().lower(),
        moz_norm_group((x or {}).get('donor_group') or (x or {}).get('group')).strip().lower(),
        moz_norm_rh((x or {}).get('donor_rh') or (x or {}).get('rh')).strip().lower(),
        _txt((x or {}).get('pack_no')).strip().lower(),
        _txt((x or {}).get('series')).strip().lower(),
    )


def _stock_lot_has_identifier(x):
    key=_stock_entry_lot_identity(x or {})
    return bool(key[3] or key[4])


def _stock_lot_balance_py(key, exclude_id=None):
    """Баланс конкретної партії/пакета. Порівняння робимо Python-нормалізацією,
    бо в БД група може бути записана як A, а в ключі нормалізується як A(II).
    """
    if not key or not (key[3] or key[4]):
        return None
    total=0.0
    for it in rows("SELECT * FROM stock_entries"):
        try:
            if exclude_id is not None and int(it.get('id') or 0)==int(exclude_id):
                continue
        except Exception as e:
            log_suppressed_error('suppressed', e)
        if _stock_entry_lot_identity(it) == key:
            total += _stock_entry_delta_py(it)
    return total



def _stock_lot_matches_filters(it, component, donor_group, donor_rh, pack_no='', series=''):
    """V6.4.61: matching руху складу з конкретним пакетом/серією."""
    if _txt(it.get('component')).strip().lower() != _txt(component).strip().lower():
        return False
    if moz_norm_group(it.get('donor_group') or it.get('group')).strip().lower() != moz_norm_group(donor_group).strip().lower():
        return False
    if moz_norm_rh(it.get('donor_rh') or it.get('rh')).strip().lower() != moz_norm_rh(donor_rh).strip().lower():
        return False
    if pack_no and _txt(it.get('pack_no')).strip().lower() != _txt(pack_no).strip().lower():
        return False
    if series and _txt(it.get('series')).strip().lower() != _txt(series).strip().lower():
        return False
    return True


def _stock_lot_balance_by_filters(component, donor_group, donor_rh, pack_no='', series='', exclude_id=None):
    if not (_txt(pack_no).strip() or _txt(series).strip()):
        return None
    total=0.0
    for it in rows("SELECT * FROM stock_entries"):
        try:
            if exclude_id is not None and int(it.get('id') or 0)==int(exclude_id):
                continue
        except Exception as e:
            log_suppressed_error('suppressed', e)
        if _stock_lot_matches_filters(it, component, donor_group, donor_rh, pack_no, series):
            total += _stock_entry_delta_py(it)
    return total




def _request_amount_float(req, default=1.0):
    """Безпечне числове значення кількості з вимоги."""
    try:
        val=float(str((req or {}).get('amount') or default).replace(',', '.'))
    except Exception:
        val=float(default)
    return val


def _lot_balance_for_unit(unit, exclude_id=None):
    """Фактичний залишок конкретної партії blood_units за stock_entries."""
    if not unit:
        return 0.0
    key=_stock_entry_lot_identity(unit)
    bal=_stock_lot_balance_py(key, exclude_id=exclude_id)
    try:
        return float(bal or 0.0)
    except Exception:
        return 0.0


def _active_reserved_request_for_unit(unit_id):
    """Активний резерв зберігається тільки якщо є вимога у статусі резерву."""
    if not unit_id:
        return None
    return row("""SELECT * FROM requests
                  WHERE unit_id=? AND status IN ('Зарезервовано','Резерв')
                  ORDER BY id DESC LIMIT 1""", (int(unit_id),))


def _active_reserved_requests_for_unit(unit_id):
    """V6.4.61: усі активні резерви по цій партії/одиниці."""
    if not unit_id:
        return []
    return rows("""SELECT * FROM requests
                   WHERE unit_id=? AND status IN ('Зарезервовано','Резерв')
                   ORDER BY id ASC""", (int(unit_id),))


def _reserved_qty_for_unit(unit_id, exclude_request_id=None):
    """V6.4.61: зарезервована кількість окремо від загального залишку партії."""
    total=0.0
    for r in _active_reserved_requests_for_unit(unit_id):
        try:
            if exclude_request_id is not None and int(r.get('id') or 0) == int(exclude_request_id):
                continue
        except Exception as e:
            log_suppressed_error('suppressed', e)
        total += _request_amount_float(r, 1.0)
    return float(total or 0.0)


def _available_qty_for_unit(unit, exclude_request_id=None):
    """Фактично доступно для нового резерву/видачі: баланс партії мінус активні резерви інших вимог."""
    if not unit:
        return 0.0
    bal=_lot_balance_for_unit(unit)
    reserved=_reserved_qty_for_unit(unit.get('id'), exclude_request_id=exclude_request_id)
    avail=float(bal or 0.0)-float(reserved or 0.0)
    if avail < 0 and avail > -0.000001:
        avail=0.0
    return avail



def _reserved_qty_for_lot_key(key, exclude_request_id=None):
    """V6.4.61: активний резерв по конкретній партії/серії, незалежно від status blood_units."""
    if not key or not (key[3] or key[4]):
        return 0.0
    total=0.0
    for u in rows("SELECT * FROM blood_units WHERE COALESCE(status,'available') IN ('available','reserved')"):
        try:
            if _stock_entry_lot_identity(u) == key:
                total += _reserved_qty_for_unit(u.get('id'), exclude_request_id=exclude_request_id)
        except Exception as e:
            log_suppressed_error('suppressed', e)
    return float(total or 0.0)


def _reserved_qty_for_stock_record(rec, exclude_request_id=None):
    """V6.4.61: резерв по партії stock_entries/receipt."""
    if not rec:
        return 0.0
    unit=_find_unit_for_stock_receipt(rec) if '_find_unit_for_stock_receipt' in globals() else None
    if unit and unit.get('id'):
        return _reserved_qty_for_unit(unit.get('id'), exclude_request_id=exclude_request_id)
    return _reserved_qty_for_lot_key(_stock_entry_lot_identity(rec), exclude_request_id=exclude_request_id)


def _available_qty_for_stock_record(rec, exclude_request_id=None, exclude_stock_entry_id=None):
    """V6.4.61: доступний незарезервований залишок партії stock_entries."""
    if not rec:
        return 0.0, 0.0, 0.0
    key=_stock_entry_lot_identity(rec)
    balance=_stock_lot_balance_py(key, exclude_id=exclude_stock_entry_id) or 0.0
    reserved=_reserved_qty_for_stock_record(rec, exclude_request_id=exclude_request_id)
    available=float(balance or 0.0)-float(reserved or 0.0)
    if available < 0 and available > -0.000001:
        available=0.0
    return available, float(balance or 0.0), float(reserved or 0.0)


def _reserved_guard_message(pack_no, series, available, reserved, balance, action='списати'):
    return (f"Неможливо {action}: у партії {pack_no or '—'}/{series or '—'} доступно {available:g}, "
            f"зарезервовано {reserved:g}, фактичний залишок {balance:g}. "
            "Спочатку скасуйте резерв або обробіть відповідну вимогу.")


def _safe_lot_writeoff(component, donor_group, donor_rh, amount, pack_no='', series='', stock_entry_id=None, reason='Списання', source_label='legacy_writeoff', patient_name='', qr_code='', request_id=None, writeoff_record=True):
    """V6.4.61: єдина безпечна логіка для всіх старих/нових списань.

    Старі endpoints (/api/writeoff, /api/units/writeoff, /api/writeoff/act)
    раніше напряму додавали stock_entries і обходили резерви. Тут завжди:
    - шукаємо реальне надходження/партію;
    - перевіряємо фактичний баланс;
    - перевіряємо активний резерв;
    - прив'язуємо правильний blood_units.id;
    - перераховуємо статус партії.
    """
    try:
        qty=float(str(amount or 0).replace(',', '.'))
    except Exception:
        return False, "Кількість має бути числом", None
    if qty <= 0:
        return False, "Кількість має бути більше 0", None
    component=(component or '').strip(); donor_group=(donor_group or '').strip(); donor_rh=(donor_rh or '').strip()
    pack_no=(pack_no or '').strip(); series=(series or '').strip()
    if not component or not donor_group or not donor_rh:
        return False, "Компонент, група і Rh обов’язкові для списання", None
    if not (pack_no or series or stock_entry_id):
        return False, "Для списання потрібно вказати конкретний пакет або серію", None
    receipt, lot_qty, lot_err = _stock_find_receipt_for_outgoing(component, donor_group, donor_rh, qty, pack_no, series, stock_entry_id)
    if not receipt:
        return False, f"Недостатньо в партії {pack_no or '—'}/{series or '—'}: доступно {lot_qty:g}, спроба списати {qty:g}", None
    available_qty, lot_balance, reserved_qty = _available_qty_for_stock_record(receipt)
    real_pack=(receipt.get('pack_no') or pack_no or '').strip(); real_series=(receipt.get('series') or series or '').strip()
    if qty > available_qty + 0.000001:
        return False, _reserved_guard_message(real_pack, real_series, available_qty, reserved_qty, lot_balance, 'списати'), None
    unit=_find_unit_for_stock_receipt(receipt)
    if not unit or not unit.get('id'):
        return False, "Для цієї партії немає коректно прив’язаної одиниці blood_units", None
    unit_id=int(unit.get('id'))
    actor=current_user().get('id') if current_user() else None
    try:
        execute("""INSERT INTO stock_entries(created_at,user_id,type,component,donor_group,donor_rh,amount,pack_no,series,expiry,patient_name,note,qr_code,unit_id)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (now(), actor, 'Списання', component, donor_group, donor_rh, qty, real_pack, real_series, receipt.get('expiry',''), patient_name or '', reason, qr_code or unit.get('qr_code',''), unit_id))
        if writeoff_record:
            try:
                execute("INSERT INTO component_writeoffs(created_at,package_code,component,amount,reason,written_by,notes) VALUES(?,?,?,?,?,?,?)",
                        (now(), qr_code or real_pack or real_series or str(unit_id), component, qty, reason, _v581_user_name(), source_label))
            except Exception:
                db_rollback_safe()
        execute("UPDATE blood_units SET updated_at=?, written_off_at=?, writeoff_reason=? WHERE id=?", (now(), now(), reason, unit_id))
        after=_recalc_blood_unit_status_for_id(unit_id, source_label or reason)
        unit_event(unit_id, 'written_off', reason, request_id or unit.get('request_id'))
        for _code in [qr_code, real_pack, real_series, unit.get('qr_code')]:
            if _code:
                log_traceability(_code, 'writeoff', patient_name or '', request_id, '', reason)
        return True, "", {'unit_id':unit_id, 'stock_entry': receipt, 'unit': after or row('SELECT * FROM blood_units WHERE id=?',(unit_id,)), 'amount': qty, 'pack_no': real_pack, 'series': real_series}
    except Exception as e:
        db_rollback_safe()
        return False, f"Помилка списання: {str(e)}", None


def _mark_reserved_requests_expired_review(unit, reason='Резервована партія прострочена'):
    """V6.4.61: прострочений резерв не може лишатися нормальним резервом.

    Якщо auto-expire бачить активні резерви на простроченій партії, вимоги
    переводяться у статус «Потребує перегляду», прив'язка до unit/пакета
    очищається, а blood_unit потім можна коректно перерахувати/списати.
    """
    if not unit or not unit.get('id'):
        return 0
    changed=0
    reqs=_active_reserved_requests_for_unit(unit.get('id')) if '_active_reserved_requests_for_unit' in globals() else []
    for r in reqs:
        rid=int(r.get('id'))
        msg=(f"Автоматично знято резерв: партія {unit.get('pack_no') or '—'}/{unit.get('series') or '—'} "
             f"прострочена. Потрібен перегляд і вибір іншого компонента.")
        old_note=(r.get('note') or '').strip()
        new_note=(old_note + '\n' if old_note else '') + msg
        execute("""UPDATE requests
                   SET status='Потребує перегляду', unit_id=NULL, pack_no='', series='', expiry='',
                       donor_group='', donor_rh='', note=?
                   WHERE id=?""", (new_note, rid))
        try:
            notify(r.get('created_by'), 'Резерв прострочений', msg)
        except Exception:
            db_rollback_safe()
        try:
            unit_event(unit.get('id'), 'reserved_expired_review', msg, rid)
        except Exception:
            db_rollback_safe()
        changed += 1
    if changed:
        try:
            audit('expired_reserved_review', f"unit={unit.get('id')}; requests={changed}; lot={unit.get('pack_no') or '—'}/{unit.get('series') or '—'}")
        except Exception:
            db_rollback_safe()
    return changed

def _can_reserve_unit_for_request(req, unit):
    """V6.4.61: резервування не може перевищувати незарезервований залишок партії."""
    if not req or not unit:
        return False, "Немає вимоги або одиниці для резервування", 0.0, 0.0
    qty=_request_amount_float(req)
    if qty <= 0:
        return False, "Кількість у вимозі має бути більше 0", 0.0, qty
    locked, lock_msg = _request_is_locked_to_other_unit(req, unit) if '_request_is_locked_to_other_unit' in globals() else (False, '')
    if locked:
        return False, lock_msg, 0.0, qty
    avail=_available_qty_for_unit(unit, exclude_request_id=req.get('id'))
    if avail + 0.000001 < qty:
        return False, f"Недостатньо доступного залишку в партії {unit.get('pack_no') or '—'}/{unit.get('series') or '—'}: доступно {avail:g}, спроба зарезервувати {qty:g}", avail, qty
    return True, "", avail, qty


def _can_issue_unit_for_request(req, unit):
    """V6.4.61: видача не може забрати кількість, зарезервовану під інші вимоги."""
    if not req or not unit:
        return False, "Немає вимоги або одиниці для видачі", 0.0, 0.0
    qty=_request_amount_float(req)
    if qty <= 0:
        return False, "Кількість у вимозі має бути більше 0", 0.0, qty
    locked, lock_msg = _request_is_locked_to_other_unit(req, unit) if '_request_is_locked_to_other_unit' in globals() else (False, '')
    if locked:
        return False, lock_msg, 0.0, qty
    avail=_available_qty_for_unit(unit, exclude_request_id=req.get('id'))
    if avail + 0.000001 < qty:
        return False, f"Недостатньо доступного залишку в партії {unit.get('pack_no') or '—'}/{unit.get('series') or '—'}: доступно {avail:g}, спроба видати {qty:g}", avail, qty
    return True, "", avail, qty

def _stock_find_receipt_for_outgoing(component, donor_group, donor_rh, qty, pack_no='', series='', stock_entry_id=None, exclude_id=None):
    """Повертає реальне активне надходження/партію для списання або видачі."""
    try:
        qty=float(qty or 0)
    except Exception:
        qty=0.0
    candidates=[]
    if stock_entry_id:
        rec=row("SELECT * FROM stock_entries WHERE id=?", (int(stock_entry_id),))
        if not rec:
            return None, 0.0, "Обране надходження не знайдено"
        if _moz_type_kind(rec.get('type')) != 'in':
            return None, 0.0, "Для списання/видачі можна обрати тільки запис надходження"
        if not _stock_lot_matches_filters(rec, component, donor_group, donor_rh, '', ''):
            return None, 0.0, "Обране надходження не відповідає компоненту, групі або Rh"
        candidates=[rec]
    else:
        for rec in rows("SELECT * FROM stock_entries ORDER BY CASE WHEN expiry IS NULL OR expiry='' THEN 1 ELSE 0 END, expiry ASC, id ASC"):
            if _moz_type_kind(rec.get('type')) != 'in':
                continue
            if not _stock_lot_matches_filters(rec, component, donor_group, donor_rh, pack_no, series):
                continue
            if not _stock_lot_has_identifier(rec):
                continue
            candidates.append(rec)
    seen=set(); total_available=0.0
    for rec in candidates:
        key=_stock_entry_lot_identity(rec)
        if key in seen:
            continue
        seen.add(key)
        bal=_stock_lot_balance_py(key, exclude_id=exclude_id) or 0.0
        if bal > 0:
            total_available += bal
        if bal + 0.000001 >= qty:
            return rec, bal, ""
    return None, total_available, "Недостатньо в конкретній партії/серії"


def _stock_lot_rows_for_unit(unit):
    if not unit:
        return []
    return rows("""SELECT * FROM stock_entries
                   WHERE COALESCE(NULLIF(component,''),'—')=?
                     AND COALESCE(NULLIF(donor_group,''),'—')=?
                     AND COALESCE(NULLIF(donor_rh,''),'—')=?
                     AND COALESCE(NULLIF(pack_no,''),'')=?
                     AND COALESCE(NULLIF(series,''),'')=?
                   ORDER BY id ASC""",
                (unit.get('component') or '—', unit.get('donor_group') or '—', unit.get('donor_rh') or '—',
                 unit.get('pack_no') or '', unit.get('series') or ''))


def _recalc_blood_unit_status_for_id(unit_id, event_note='Перерахунок статусу партії'):
    """V6.4.61: синхронізація blood_units зі stock_entries.

    status не виставляється напряму після кожного руху. Ми рахуємо фактичний
    залишок конкретної партії. Якщо після часткового списання залишок >0 —
    одиниця/партія лишається available/reserved, а не written_off.
    """
    try:
        if not unit_id:
            return None
        unit=row("SELECT * FROM blood_units WHERE id=?", (int(unit_id),))
        if not unit:
            return None
        entries=_stock_lot_rows_for_unit(unit)
        balance=0.0
        last_out=None
        last_req_id=None
        last_patient=''
        for it in entries:
            delta=_stock_entry_delta_py(it)
            balance += delta
            if delta < 0:
                last_out=it
                if it.get('patient_name'):
                    last_patient=it.get('patient_name') or ''
        if balance < 0 and balance > -0.000001:
            balance=0.0
        if balance > 0.000001:
            expiry_state = expiry_status(unit.get('expiry'))
            reserved_qty = _reserved_qty_for_unit(unit_id) if '_reserved_qty_for_unit' in globals() else 0.0
            available_qty = balance - reserved_qty
            if available_qty < 0 and available_qty > -0.000001:
                available_qty = 0.0
            active_reserve = _active_reserved_request_for_unit(unit_id) if '_active_reserved_request_for_unit' in globals() else None
            # V6.4.61: reserved тільки якщо весь фактичний залишок покритий активними резервами.
            # Якщо P1/S1=2 і зарезервовано 1, статус лишається available, amount=2,
            # а доступна кількість рахується як balance-reserved_qty.
            if expiry_state == 'expired':
                new_status='expired'
                execute("""UPDATE blood_units SET status=?, amount=?, updated_at=?, written_off_at=NULL,
                          writeoff_reason='', request_id=NULL, patient_name='', issued_by='', issued_at=NULL, used_at=NULL
                          WHERE id=?""", (new_status, balance, now(), int(unit_id)))
            elif reserved_qty > 0.000001 and available_qty <= 0.000001 and active_reserve:
                new_status='reserved'
                # V6.4.61: якщо одна партія повністю покрита кількома частковими резервами,
                # blood_units.request_id не повинен показувати випадково лише одну вимогу.
                # Баланс резервів рахується з requests, а тут зберігаємо request_id тільки коли
                # активний резерв один; при кількох — очищаємо, щоб UI/аудит не вводив в оману.
                active_reserves = _active_reserved_requests_for_unit(unit_id) if '_active_reserved_requests_for_unit' in globals() else []
                if len(active_reserves) == 1:
                    rid = int(active_reserves[0].get('id'))
                    pname = active_reserves[0].get('patient_name','')
                else:
                    rid = None
                    pname = 'Кілька активних резервів' if len(active_reserves) > 1 else ''
                execute("""UPDATE blood_units SET status=?, amount=?, updated_at=?, written_off_at=NULL,
                          request_id=?, patient_name=?, writeoff_reason=''
                          WHERE id=?""", (new_status, balance, now(), rid, pname, int(unit_id)))
            else:
                new_status = 'available'
                execute("""UPDATE blood_units SET status=?, amount=?, updated_at=?, written_off_at=NULL,
                          writeoff_reason='', request_id=NULL, patient_name='', issued_by='', issued_at=NULL, used_at=NULL
                          WHERE id=?""", (new_status, balance, now(), int(unit_id)))
        else:
            kind=_moz_type_kind((last_out or {}).get('type')) if last_out else ''
            note=((last_out or {}).get('note') or '').lower() if last_out else ''
            if kind == 'issue' or 'видач' in note or 'видано' in note:
                new_status='issued'
            elif kind == 'writeoff' or last_out:
                new_status='written_off'
            else:
                new_status='available'
            
            active_any = _active_reserved_request_for_unit(unit_id) if '_active_reserved_request_for_unit' in globals() else None
            if active_any:
                execute("""UPDATE blood_units SET status=?, amount=0, updated_at=?, request_id=?, patient_name=? WHERE id=?""",
                        (new_status, now(), int(active_any.get('id')), active_any.get('patient_name',''), int(unit_id)))
            else:
                execute("""UPDATE blood_units SET status=?, amount=0, updated_at=?, request_id=NULL, patient_name='', issued_by='', issued_at=NULL WHERE id=?""",
                        (new_status, now(), int(unit_id)))
        unit_event(unit_id, 'lot_status_recalculated', event_note)
        return row("SELECT * FROM blood_units WHERE id=?", (int(unit_id),))
    except Exception:
        db_rollback_safe()
        return None


def _recalc_blood_unit_status_for_stock_record(rec, event_note='Перерахунок статусу партії'):
    """Знайти/перерахувати blood_unit за stock_entries або словником з trash."""
    try:
        if not rec:
            return None
        if rec.get('unit_id'):
            return _recalc_blood_unit_status_for_id(rec.get('unit_id'), event_note)
        unit=_find_unit_for_stock_receipt(rec) if '_find_unit_for_stock_receipt' in globals() else None
        if unit:
            return _recalc_blood_unit_status_for_id(unit.get('id'), event_note)
    except Exception:
        db_rollback_safe()
    return None


def _stock_lot_outgoing_qty_py(x, exclude_id=None):
    """Скільки вже списано/видано з цієї партії. Для надходження це визначає,
    чи можна змінювати пакет/серію/групу/Rh або зменшувати кількість.
    """
    key=_stock_entry_lot_identity(x or {})
    if not (key[3] or key[4]):
        return 0.0
    total=0.0
    for it in rows("SELECT * FROM stock_entries"):
        try:
            if exclude_id is not None and int(it.get('id') or 0)==int(exclude_id):
                continue
        except Exception as e:
            log_suppressed_error('suppressed', e)
        if _stock_entry_lot_identity(it) == key and _moz_type_kind(it.get('type')) in ('issue','writeoff'):
            try:
                total += abs(float(str(it.get('amount') or 0).replace(',', '.')))
            except Exception as e:
                log_suppressed_error('suppressed', e)
    return total


def _stock_same_value(a,b):
    return _txt(a).strip().lower() == _txt(b).strip().lower()


@app.post("/api/stock/update")
@role_required("admin","transfusion")
def api_stock_update_v6414():
    """V6.4.61: безпечне редагування руху складу без можливості створити мінусовий залишок."""
    u=current_user(); d=request.json or {}; sid=int(d.get("id") or 0)
    rec=row("SELECT * FROM stock_entries WHERE id=?", (sid,))
    if not rec: return jsonify(ok=False,error="Запис складу не знайдено"),404
    allowed={}
    for k in ["type","component","donor_group","donor_rh","amount","pack_no","series","expiry","patient_name","note","qr_code"]:
        if k in d: allowed[k]=d.get(k)
    if "type" in allowed:
        allowed["type"]=_stock_normalize_type(allowed.get("type"))
    for required in ["component"]:
        if required in allowed and not str(allowed.get(required) or '').strip():
            return jsonify(ok=False,error="Компонент не може бути порожнім"),400
    if "amount" in allowed:
        try:
            allowed["amount"]=float(str(allowed["amount"]).replace(",","."))
        except Exception:
            return jsonify(ok=False,error="Кількість має бути числом"),400
        if allowed["amount"] <= 0:
            return jsonify(ok=False,error="Кількість має бути більше 0"),400
    if not allowed: return jsonify(ok=False,error="Немає змін")

    old_rec=dict(rec)
    new_rec=dict(rec)
    new_rec.update(allowed)
    old_key=_stock_entry_group_key(old_rec)
    new_key=_stock_entry_group_key(new_rec)
    old_delta=_stock_entry_delta_py(old_rec)
    new_delta=_stock_entry_delta_py(new_rec)

    # V6.4.61: партійний контроль. Якщо надходження вже має видачу/списання
    # по цьому пакету/серії, забороняємо зміну ключових ідентифікаторів партії.
    # Інакше списання P1/S1 може залишитися без надходження P1/S1 після редагування.
    old_kind=_moz_type_kind(old_rec.get('type'))
    new_kind=_moz_type_kind(new_rec.get('type'))
    immutable_lot_fields=["component","donor_group","donor_rh","pack_no","series","expiry"]
    if old_kind == 'in' and _stock_lot_has_identifier(old_rec):
        linked_out=_stock_lot_outgoing_qty_py(old_rec, exclude_id=sid)
        if linked_out > 0.000001:
            changed=[k for k in immutable_lot_fields if k in allowed and not _stock_same_value(old_rec.get(k), new_rec.get(k))]
            if changed:
                return jsonify(ok=False,error="Неможливо змінити пакет/серію або ключові поля: по цій партії вже є списання/видача"),409
            if "type" in allowed and new_kind != 'in':
                return jsonify(ok=False,error="Неможливо змінити тип надходження: по цій партії вже є списання/видача"),409
            if "amount" in allowed and float(allowed.get("amount") or 0) + 0.000001 < linked_out:
                return jsonify(ok=False,error=f"Неможливо зменшити надходження нижче вже списаної/виданої кількості: {linked_out:g}"),409

    # Для редагування списання/видачі з конкретним пакетом/серією додатково
    # перевіряємо баланс саме цієї партії, а не тільки загальний залишок компонента.
    old_lot_key=_stock_entry_lot_identity(old_rec)
    new_lot_key=_stock_entry_lot_identity(new_rec)
    if old_kind in ('issue','writeoff') and _stock_lot_has_identifier(old_rec):
        if not _stock_lot_has_identifier(new_rec):
            return jsonify(ok=False,error="Неможливо очистити пакет/серію у списанні або видачі: запис уже прив’язаний до партії"),409
    impacted_lots=set()
    if _stock_lot_has_identifier(old_rec): impacted_lots.add(old_lot_key)
    if _stock_lot_has_identifier(new_rec): impacted_lots.add(new_lot_key)
    if impacted_lots:
        lot_sim={}
        for key in impacted_lots:
            lot_sim[key]=_stock_lot_balance_py(key, exclude_id=sid) or 0.0
        # Баланс рахується без поточного запису, тому додаємо лише новий вплив.
        # Старий вплив віднімати не треба — він уже виключений через exclude_id.
        if new_lot_key in lot_sim:
            lot_sim[new_lot_key]=lot_sim.get(new_lot_key,0.0)+new_delta
        for key, qty in lot_sim.items():
            if qty < -0.000001:
                return jsonify(ok=False,error=f"Редагування неможливе: залишок партії/серії стане від’ємним для {key[0]} {key[1]} {key[2]} пакет {key[3] or '—'} серія {key[4] or '—'} ({qty:g})"),409
            # V6.4.61: після редагування фактичний залишок партії не може бути меншим за активні резерви.
            reserved_qty = _reserved_qty_for_lot_key(key)
            if qty + 0.000001 < reserved_qty:
                return jsonify(ok=False,error=f"Редагування неможливе: у партії {key[3] or '—'}/{key[4] or '—'} зарезервовано {reserved_qty:g}, а після редагування залишиться {qty:g}"),409

    # Симуляція залишків після UPDATE. Це закриває випадки:
    # - зменшили старе надходження, яке вже було видано/списано;
    # - змінили тип/кількість/компонент/групу/Rh так, що стара або нова група йде в мінус;
    # - ввели від’ємну або нульову кількість.
    impacted={old_key, new_key}
    simulated={}
    for key in impacted:
        simulated[key]=_component_current_qty(*key)
    simulated[old_key]=simulated.get(old_key,0.0)-old_delta
    simulated[new_key]=simulated.get(new_key,0.0)+new_delta
    for key, qty in simulated.items():
        if qty < -0.000001:
            return jsonify(ok=False,error=f"Редагування неможливе: залишок стане від’ємним для {key[0]} {key[1]} {key[2]} ({qty:g})"),409

    execute("UPDATE stock_entries SET "+",".join([f"{k}=?" for k in allowed])+" WHERE id=?", tuple(list(allowed.values())+[sid]))
    if rec.get("unit_id"):
        unit_fields={}
        mapping={"component":"component","donor_group":"donor_group","donor_rh":"donor_rh","amount":"amount","pack_no":"pack_no","series":"series","expiry":"expiry","patient_name":"patient_name","note":"note","qr_code":"qr_code"}
        for src,dst in mapping.items():
            if src in allowed: unit_fields[dst]=allowed[src]
        if unit_fields:
            unit_fields["updated_at"]=now()
            execute("UPDATE blood_units SET "+",".join([f"{k}=?" for k in unit_fields])+" WHERE id=?", tuple(list(unit_fields.values())+[int(rec.get("unit_id"))]))
            _recalc_blood_unit_status_for_id(rec.get("unit_id"), f"Редаговано запис складу №{sid}")
            unit_event(rec.get("unit_id"), "stock_entry_updated", f"Редаговано запис складу №{sid}")
    audit("stock_update", f"stock_entries:{sid}")
    return jsonify(ok=True)


def _request_lock_for_stock_records(stock_records):
    """V6.4.61: перевірка, чи рух складу пов'язаний із активною або фінальною вимогою.

    Блокує видалення партій/рухів, якщо по них є вимога у статусі:
    Зарезервовано, Резерв, Видано, Використано, Списано.
    """
    lock_statuses=('Зарезервовано','Резерв','Видано','Використано','Списано')
    ph=','.join(['?']*len(lock_statuses))
    seen=[]
    for rec in stock_records or []:
        candidates=[]
        try:
            uid=rec.get('unit_id')
            if uid:
                candidates += rows(f"SELECT * FROM requests WHERE unit_id=? AND status IN ({ph})", (int(uid),)+lock_statuses)
        except Exception:
            db_rollback_safe()
        try:
            note=rec.get('note') or ''
            import re
            for m in re.findall(r'№\s*(\d+)', note):
                rr=row(f"SELECT * FROM requests WHERE id=? AND status IN ({ph})", (int(m),)+lock_statuses)
                if rr:
                    candidates.append(rr)
        except Exception:
            db_rollback_safe()
        # Додатковий захист: якщо реквізити партії збігаються з активною/фінальною вимогою.
        try:
            candidates += rows(f"""SELECT * FROM requests WHERE status IN ({ph})
                           AND COALESCE(NULLIF(component,''),'—')=?
                           AND COALESCE(NULLIF(donor_group,''),'—')=?
                           AND COALESCE(NULLIF(donor_rh,''),'—')=?
                           AND COALESCE(NULLIF(pack_no,''),'')=?
                           AND COALESCE(NULLIF(series,''),'')=?""",
                           lock_statuses + (rec.get('component') or '—', rec.get('donor_group') or '—', rec.get('donor_rh') or '—', rec.get('pack_no') or '', rec.get('series') or ''))
        except Exception:
            db_rollback_safe()
        for rr in candidates:
            rid=rr.get('id')
            if rid not in seen:
                seen.append(rid)
                return rr
    return None

@app.post("/api/stock/delete")
@role_required("admin","transfusion")
def api_stock_delete_v6414():
    """V6.4.61: безпечне видалення рухів складу з коректним перерахунком blood_units.

    - Якщо видаляється списання/видача — blood_unit НЕ переводиться в deleted;
      після видалення руху статус/залишок партії перераховується.
    - Якщо видаляється надходження — видаляється весь рух тієї ж партії/серії,
      щоб не лишити списання без надходження; після цього blood_unit цієї партії
      переводиться в deleted, бо активних рухів партії не лишається.
    """
    u=current_user(); d=request.json or {}; sid=int(d.get("id") or 0)
    rec=row("SELECT * FROM stock_entries WHERE id=?", (sid,))
    if not rec:
        return jsonify(ok=False,error="Запис складу не знайдено"),404
    reason=d.get("reason") or "Видалено зі складу"
    kind=_moz_type_kind(rec.get('type'))
    recs=[rec]
    deleting_whole_lot = False

    # Для надходження видаляємо весь рух тієї ж партії, щоб не залишити
    # списання/видачу без надходження. Для списання/видачі видаляємо тільки
    # цей рух і потім перераховуємо статус партії.
    if kind == 'in' and _stock_lot_has_identifier(rec):
        try:
            recs=rows("""SELECT * FROM stock_entries
                       WHERE COALESCE(NULLIF(component,''),'—')=?
                         AND COALESCE(NULLIF(donor_group,''),'—')=?
                         AND COALESCE(NULLIF(donor_rh,''),'—')=?
                         AND COALESCE(NULLIF(pack_no,''),'')=?
                         AND COALESCE(NULLIF(series,''),'')=?
                       ORDER BY id ASC""",
                      (rec.get('component') or '—', rec.get('donor_group') or '—', rec.get('donor_rh') or '—', rec.get('pack_no') or '', rec.get('series') or ''))
            if not recs:
                recs=[rec]
            deleting_whole_lot = True
        except Exception:
            db_rollback_safe(); recs=[rec]
    elif kind == 'in':
        deleting_whole_lot = True

    locked_req=_request_lock_for_stock_records(recs)
    if locked_req:
        return jsonify(ok=False,error=f"Неможливо видалити рух складу: він пов’язаний із вимогою №{locked_req.get('id')}, статус «{locked_req.get('status')}». Спочатку скасуйте резерв/видачу або переведіть вимогу в безпечний статус."),409

    affected_unit_ids=set()
    affected_lot_records=[]
    ids=[]
    for r0 in recs:
        rid=int(r0.get('id') or 0); ids.append(rid)
        if r0.get('unit_id'):
            try:
                affected_unit_ids.add(int(r0.get('unit_id')))
            except Exception as e:
                log_suppressed_error('suppressed', e)
        affected_lot_records.append(dict(r0))
        execute("INSERT INTO trash(created_at,source_table,source_id,data,deleted_by,reason) VALUES(?,?,?,?,?,?)",
                (now(),"stock_entries",rid,json.dumps(dict(r0),ensure_ascii=False),u.get("username"),reason))
        if r0.get("unit_id"):
            try:
                unit_event(r0.get("unit_id"), "stock_entry_deleted", f"Видалено запис складу №{rid}")
            except Exception:
                db_rollback_safe()

    if ids:
        ph=','.join(['?']*len(ids))
        execute(f"DELETE FROM stock_entries WHERE id IN ({ph})", tuple(ids))

    # Після видалення — синхронізуємо blood_units. Важливо: видалення
    # списання/видачі має збільшити залишок і повернути статус available,
    # а не залишити unit у deleted.
    if deleting_whole_lot:
        for uid in affected_unit_ids:
            try:
                execute("UPDATE blood_units SET status='deleted', amount=0, updated_at=?, note=? WHERE id=?",
                        (now(), reason, int(uid)))
                unit_event(uid, 'lot_deleted', 'Видалено всю партію/рух партії зі складу')
            except Exception:
                db_rollback_safe()
        # Якщо у частини рухів не було unit_id, пробуємо знайти unit за партією.
        for r0 in affected_lot_records:
            try:
                unit=_find_unit_for_stock_receipt(r0) if '_find_unit_for_stock_receipt' in globals() else None
                if unit and int(unit.get('id')) not in affected_unit_ids:
                    execute("UPDATE blood_units SET status='deleted', amount=0, updated_at=?, note=? WHERE id=?",
                            (now(), reason, int(unit.get('id'))))
            except Exception:
                db_rollback_safe()
    else:
        for uid in affected_unit_ids:
            _recalc_blood_unit_status_for_id(uid, f"Перерахунок після видалення руху складу №{sid}")
        for r0 in affected_lot_records:
            if not r0.get('unit_id'):
                _recalc_blood_unit_status_for_stock_record(r0, f"Перерахунок після видалення руху складу №{sid}")

    audit("stock_delete", f"stock_entries:{ids}")
    return jsonify(ok=True, deleted=len(ids), linked_deleted=max(0,len(ids)-1), message=("Видалено всю партію/пов’язаний рух" if len(ids)>1 else "Запис видалено"))

@app.get("/api/alerts")
@role_required("admin","transfusion")
def api_alerts():
    low=rows("""SELECT component,donor_group as "group",donor_rh as rh,SUM(CASE WHEN type='Надходження' THEN amount ELSE -amount END) qty
                FROM stock_entries GROUP BY component,donor_group,donor_rh HAVING SUM(CASE WHEN type='Надходження' THEN amount ELSE -amount END)<5""")
    all_exp=rows("SELECT * FROM stock_entries WHERE expiry IS NOT NULL AND expiry<>'' ORDER BY expiry")
    exp=[]
    try:
        from datetime import date
        today=datetime.now().date()
        for x in all_exp:
            try:
                if datetime.strptime(x.get("expiry",""), "%Y-%m-%d").date() <= today + timedelta(days=7):
                    exp.append(x)
            except Exception as e:
                log_suppressed_error('suppressed', e)
    except Exception:
        exp=[]
    return jsonify(low=low, expiry=exp)

@app.get("/api/audit")
@role_required("admin","transfusion")
def api_audit(): return jsonify(rows("SELECT * FROM audit ORDER BY id DESC LIMIT 200"))

@app.get("/api/notifications")
@login_required
def api_notifications():
    return jsonify(rows("SELECT * FROM notifications WHERE user_id=? ORDER BY id DESC LIMIT 50",(current_user()["id"],)))

@app.get("/api/login-history")
@role_required("admin")
def api_login_history(): return jsonify(rows("SELECT * FROM login_attempts ORDER BY id DESC LIMIT 200"))

@app.get("/api/reports/preview")
@role_required("admin","transfusion")
def api_reports_preview():
    return jsonify(rows=rows("SELECT id,created_at,patient_name,component,amount,status,department,doctor_name FROM requests ORDER BY id DESC LIMIT 500"))

@app.get("/api/reports/reactions")
@role_required("admin","transfusion")
def api_reports_reactions(): return jsonify(rows=rows("SELECT * FROM requests WHERE reaction_present='Так' ORDER BY id DESC"))


def setup_pdf_font(c, size=12):
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        candidates = ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed.ttf", "DejaVuSans.ttf"]
        for fp in candidates:
            if os.path.exists(fp):
                pdfmetrics.registerFont(TTFont("DejaVu", fp))
                c.setFont("DejaVu", size)
                return "DejaVu"
    except Exception as e:
        log_suppressed_error('suppressed', e)
    c.setFont("Helvetica", size)
    return "Helvetica"

def pdf_text(c, x, y, text):
    try:
        c.drawString(x, y, str(text or ""))
    except Exception:
        safe = str(text or "").encode("latin-1", "replace").decode("latin-1")
        c.drawString(x, y, safe)


@app.get("/reports/request/<int:rid>.pdf")
@login_required
def request_pdf(rid):
    req = row("SELECT * FROM requests WHERE id=?", (rid,))
    if not req:
        return "Not found", 404
    u = current_user()
    if u and u.get("role") in ("doctor", "nurse") and int(req.get("created_by") or 0) != int(u.get("id") or 0):
        return "Недостатньо прав", 403
    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    try:
        setup_pdf_font(c, 12)
    except Exception:
        try: c.setFont("Helvetica", 12)
        except Exception as e: log_suppressed_error('suppressed', e)
    y = 800
    try:
        pdf_text(c, 40, y, f"Вимога №{rid}")
    except Exception:
        c.drawString(40, y, f"Request #{rid}")
    y -= 30
    fields = [
        ("Пацієнт", req.get("patient_name","")),
        ("Дата народження", req.get("birth_date","")),
        ("Відділення", req.get("department","")),
        ("Компонент", req.get("component","")),
        ("Група/Rh", f"{req.get('patient_group','')} {req.get('patient_rh','')}"),
        ("Кількість", req.get("amount","")),
        ("Статус", req.get("status","")),
        ("Діагноз", req.get("diagnosis","")),
        ("Лікар", req.get("doctor_name","")),
        ("Пакет", req.get("pack_no","")),
        ("Серія", req.get("series","")),
        ("Термін", req.get("expiry","")),
        ("Сумісність", req.get("compatibility_warning","") or "OK"),
    ]
    for k, v in fields:
        try:
            pdf_text(c, 40, y, f"{k}: {v}")
        except Exception:
            c.drawString(40, y, f"{k}: {v}")
        y -= 20
        if y < 60:
            c.showPage()
            try: setup_pdf_font(c, 12)
            except Exception as e: log_suppressed_error('suppressed', e)
            y = 800
    c.save()
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"request_{rid}.pdf")

@app.get("/reports/export/<fmt>")
@role_required("admin","transfusion")
def report_export(fmt):
    data=rows("SELECT id,created_at,patient_name,component,amount,status FROM requests ORDER BY id DESC")
    if fmt=="xlsx":
        wb=Workbook(); ws=wb.active; ws.append(["ID","Дата","Пацієнт","Компонент","К-сть","Статус"])
        for x in data: ws.append([x["id"],x["created_at"],x["patient_name"],x["component"],x["amount"],x["status"]])
        bio=BytesIO(); wb.save(bio); bio.seek(0); return send_file(bio, as_attachment=True, download_name="requests.xlsx")
    bio=BytesIO(); c=canvas.Canvas(bio,pagesize=A4); setup_pdf_font(c,11); y=800; pdf_text(c,40,y,"Звіт по вимогах"); y-=30
    for x in data[:40]: pdf_text(c,40,y,f'{x["id"]} {x["patient_name"]} {x["component"]} {x["status"]}'); y-=18
    c.save(); bio.seek(0); return send_file(bio, as_attachment=True, download_name="requests.pdf")

@app.get("/backup")
@role_required("admin")
def backup():
    init_db()
    return send_file(DB_PATH, as_attachment=True, download_name="blood_bank_backup.db")

@app.post("/api/admin/delete-record")
@role_required("admin","transfusion")
def delete_record():
    d=request.json or {}; table=d.get("table"); rid,err=require_id_from_json(d)
    if err: return err
    if table not in ["requests","stock_entries","users","audit"]:
        return jsonify(ok=False,error="Недозволена таблиця")
    if table == "stock_entries":
        return jsonify(ok=False,error="Пряме видалення записів складу через admin/delete-record заборонене. Використовуйте безпечне видалення складу /api/stock/delete, щоб перерахувати партії і blood_units."),409
    u = current_user()
    if table == "users":
        target = row("SELECT * FROM users WHERE id=?", (rid,))
        if not target:
            return jsonify(ok=False,error="Не знайдено")
        if target["username"] == "Sepsis":
            return jsonify(ok=False,error="Sepsis видаляти не можна")
        if u["role"] == "transfusion" and target["role"] not in ["doctor","nurse"]:
            return jsonify(ok=False,error="Трансфузіолог може видаляти тільки лікарів/медсестер")
    rec=row(f"SELECT * FROM {table} WHERE id=?", (rid,))
    if not rec:
        return jsonify(ok=False,error="Не знайдено")
    execute("INSERT INTO trash(created_at,source_table,source_id,data,deleted_by,reason) VALUES(?,?,?,?,?,?)", (now(),table,rid,json.dumps(dict(rec),ensure_ascii=False),u["username"],d.get("reason","")))
    execute(f"DELETE FROM {table} WHERE id=?", (rid,))
    audit("delete_record", f"{table}:{rid}")
    return jsonify(ok=True)




@app.post("/api/trash/restore")
@role_required("admin","transfusion")
def restore_record():
    """V6.4.61: безпечне відновлення.

    Для stock_entries з конкретною партією відновлюємо всі записи цієї партії,
    які лежать у кошику. Це не дає відновити списання без надходження або лише
    частину руху партії після видалення надходження.
    """
    d = request.json or {}
    tid,err=require_id_from_json(d)
    if err: return err
    tr = row("SELECT * FROM trash WHERE id=?", (tid,))
    if not tr:
        return jsonify(ok=False,error="Запис у кошику не знайдено")
    table = tr["source_table"]
    if table not in ["requests","stock_entries","users","audit"]:
        return jsonify(ok=False,error="Відновлення для цієї таблиці не дозволено")

    def _insert_restored(table, data, tid_local):
        old_id = data.get("id")
        exists = row(f"SELECT id FROM {table} WHERE id=?", (old_id,)) if old_id is not None else None
        restore_data = dict(data)
        if exists and "id" in restore_data:
            restore_data.pop("id", None)
        if table == "users" and restore_data.get("username"):
            existing_user = row("SELECT id FROM users WHERE username=?", (restore_data["username"],))
            if existing_user:
                restore_data["username"] = restore_data["username"] + "_restored_" + str(tid_local)
        cols = list(restore_data.keys())
        vals = [restore_data[c] for c in cols]
        placeholders = ",".join(["?"] * len(cols))
        execute(f"INSERT INTO {table} ({','.join(cols)}) VALUES ({placeholders})", tuple(vals))
        return restore_data

    restored=[]

    if table == "stock_entries":
        data = json.loads(tr["data"] or "{}")
        # Якщо це партійний запис — відновлюємо весь комплект рухів партії з кошика.
        group_trash=[]
        if _stock_lot_has_identifier(data):
            target_key=_stock_entry_lot_identity(data)
            for tx in rows("SELECT * FROM trash WHERE source_table='stock_entries' ORDER BY id ASC"):
                try:
                    dx=json.loads(tx.get('data') or '{}')
                except Exception:
                    continue
                if _stock_entry_lot_identity(dx) == target_key:
                    group_trash.append(tx)
        if not group_trash:
            group_trash=[tr]

        # Якщо відновлюється вихідний рух без активного/відновлюваного надходження — блокуємо.
        has_receipt=False
        for tx in group_trash:
            try:
                dx=json.loads(tx.get('data') or '{}')
                if _moz_type_kind(dx.get('type')) == 'in':
                    has_receipt=True
                    break
            except Exception as e:
                log_suppressed_error('suppressed', e)
        if not has_receipt and _stock_lot_has_identifier(data):
            key=_stock_entry_lot_identity(data)
            bal_in=0.0
            for it in rows("SELECT * FROM stock_entries"):
                if _stock_entry_lot_identity(it)==key and _moz_type_kind(it.get('type'))=='in':
                    bal_in += abs(float(it.get('amount') or 0))
            if bal_in <= 0.000001:
                return jsonify(ok=False,error="Неможливо відновити списання/видачу без активного надходження цієї партії. Відновіть усю партію разом."),409

        restored_tids=[]
        affected_unit_ids=set()
        affected_records=[]
        for tx in group_trash:
            dx=json.loads(tx.get('data') or '{}')
            rd=_insert_restored('stock_entries', dx, tx.get('id'))
            restored.append(rd); affected_records.append(rd); restored_tids.append(int(tx.get('id')))
            if rd.get('unit_id'):
                try: affected_unit_ids.add(int(rd.get('unit_id')))
                except Exception as e: log_suppressed_error('suppressed', e)
        if restored_tids:
            ph=','.join(['?']*len(restored_tids))
            execute(f"DELETE FROM trash WHERE id IN ({ph})", tuple(restored_tids))
        for uid in affected_unit_ids:
            _recalc_blood_unit_status_for_id(uid, 'Відновлено рух/партію складу з кошика')
        for rd in affected_records:
            if not rd.get('unit_id'):
                _recalc_blood_unit_status_for_stock_record(rd, 'Відновлено рух/партію складу з кошика')
        audit("restore_record", f"stock_entries:{[x.get('id') for x in restored]}")
        return jsonify(ok=True, restored=len(restored), message=("Відновлено весь рух партії" if len(restored)>1 else "Запис відновлено"))

    # Інші таблиці — старий безпечний режим одного запису.
    data = json.loads(tr["data"] or "{}")
    old_id = data.get("id")
    restore_data=_insert_restored(table, data, tid)
    execute("DELETE FROM trash WHERE id=?", (tid,))
    audit("restore_record", f"{table}:{old_id}")
    return jsonify(ok=True)



@app.post("/api/trash/delete")
@role_required("admin","transfusion")
def delete_trash_record_permanent():
    """V6.4.40: остаточно стерти запис із кошика.
    Це не відновлює робочий запис, а тільки прибирає його із trash.
    """
    d = request.json or {}
    tid, err = require_id_from_json(d)
    if err:
        return err
    tr = row("SELECT * FROM trash WHERE id=?", (tid,))
    if not tr:
        return jsonify(ok=False, error="Запис у кошику не знайдено")
    execute("DELETE FROM trash WHERE id=?", (tid,))
    audit("trash_delete_permanent", f"trash:{tid} table:{tr.get('source_table') if isinstance(tr, dict) else ''}")
    return jsonify(ok=True)

@app.get("/api/security/integrity")
@role_required("admin")
def integrity(): return jsonify(ok=True, time=now(), db_exists=os.path.exists(DB_PATH))

@app.post("/api/spellcheck")
@login_required
def spellcheck():
    text=(request.json or {}).get("text","")
    issues=[]
    for bad in ["темммпература","пацієнтт","гемоглобінн"]:
        if bad in text: issues.append({"word":bad,"suggestion":bad.replace("ммм","м").replace("тт","т").replace("нн","н")})
    return jsonify(ok=True, issues=issues)

# ===== RENDER / GUNICORN STARTUP DB INIT =====
try:
    with app.app_context():
        init_db()
except Exception as e:
    print("DB startup init error:", e)


@app.get("/api/dashboard")
@login_required
def api_dashboard():
    u=current_user()
    if u["role"] in ("doctor", "nurse"):
        uid = u["id"]
        return jsonify(
            requests=rows("SELECT status, COUNT(*) count FROM requests WHERE created_by=? GROUP BY status", (uid,)),
            components=rows("SELECT component, SUM(amount) amount FROM requests WHERE created_by=? GROUP BY component", (uid,)),
            departments=rows("SELECT department, COUNT(*) count FROM requests WHERE created_by=? GROUP BY department", (uid,)),
            reactions=rows("SELECT reaction_type, COUNT(*) count FROM requests WHERE created_by=? AND reaction_present='Так' GROUP BY reaction_type", (uid,)),
            daily=rows("SELECT substr(created_at,1,10) day, COUNT(*) count FROM requests WHERE created_by=? GROUP BY substr(created_at,1,10) ORDER BY day DESC LIMIT 14", (uid,))
        )
    return jsonify(
        requests=rows("SELECT status, COUNT(*) count FROM requests GROUP BY status"),
        components=rows("SELECT component, SUM(amount) amount FROM requests GROUP BY component"),
        departments=rows("SELECT department, COUNT(*) count FROM requests GROUP BY department"),
        reactions=rows("SELECT reaction_type, COUNT(*) count FROM requests WHERE reaction_present='Так' GROUP BY reaction_type"),
        daily=rows("SELECT substr(created_at,1,10) day, COUNT(*) count FROM requests GROUP BY substr(created_at,1,10) ORDER BY day DESC LIMIT 14")
    )

@app.get("/api/patients/history")
@role_required("admin","transfusion","doctor","nurse")
def api_patient_history():
    # V6.4.40: doctor/nurse бачать історію тільки власних вимог, включно зі списаними/відмовленими/використаними.
    u=current_user(); name=request.args.get("name","").strip()
    if not name:
        if u.get("role") in ("doctor","nurse"):
            return jsonify(rows=rows("SELECT * FROM requests WHERE created_by=? ORDER BY id DESC LIMIT 50", (u.get("id"),)))
        return jsonify(rows=rows("SELECT * FROM requests ORDER BY id DESC LIMIT 50"))
    if u.get("role") in ("doctor","nurse"):
        return jsonify(rows=rows("SELECT * FROM requests WHERE created_by=? AND patient_name LIKE ? ORDER BY id DESC", (u.get("id"), f"%{name}%")))
    return jsonify(rows=rows("SELECT * FROM requests WHERE patient_name LIKE ? ORDER BY id DESC", (f"%{name}%",)))

@app.get("/api/backups")
@role_required("admin")
def api_backups():
    return jsonify(rows("SELECT * FROM backups ORDER BY id DESC"))

@app.post("/api/backups/create")
@role_required("admin")
def api_backup_create():
    path=make_backup(current_user()["username"])
    audit("backup_create", os.path.basename(path))
    telegram_event_backup(True)
    return jsonify(ok=True, filename=os.path.basename(path))

@app.get("/api/backups/download/<int:bid>")
@role_required("admin")
def api_backup_download(bid):
    b=row("SELECT * FROM backups WHERE id=?", (bid,))
    if not b:
        return jsonify(ok=False, error="Резервну копію не знайдено"), 404
    path = os.path.join(BACKUP_DIR, b["filename"])
    if not os.path.exists(path):
        return jsonify(ok=False, error="Файл резервної копії не знайдено"), 404
    return send_file(path, as_attachment=True, download_name=b["filename"])

@app.post("/api/backups/restore")
@role_required("admin")
def api_backup_restore():
    d=request.json or {}; bid,err=require_id_from_json(d)
    if err: return err
    b=row("SELECT * FROM backups WHERE id=?", (bid,))
    if not b: return jsonify(ok=False,error="Резервну копію не знайдено")
    if IS_POSTGRES:
        return jsonify(ok=False,error="Для PostgreSQL restore виконуйте через Render PostgreSQL Backups або імпорт JSON вручну. Завантаження резервної копії доступне.")
    make_backup("before_restore")
    with zipfile.ZipFile(os.path.join(BACKUP_DIR,b["filename"])) as z:
        z.extract("blood_bank_v4.db", BACKUP_DIR)
    shutil.copy(os.path.join(BACKUP_DIR,"blood_bank_v4.db"), DB_PATH)
    audit("backup_restore", b["filename"])
    return jsonify(ok=True)


@app.get("/api/system/db")
@role_required("admin")
def api_system_db():
    try:
        u = row("SELECT COUNT(*) AS count FROM users")
        return jsonify(ok=True, postgres=IS_POSTGRES, users=u.get("count", 0))
    except Exception as e:
        return jsonify(ok=False, postgres=IS_POSTGRES, error=str(e)), 500


@app.post("/api/stock/auto-expire")
@role_required("admin","transfusion")
def api_stock_auto_expire():
    """V6.4.61: списує тільки фактичний залишок простроченої партії.

    Старий варіант списував повну кількість первинного надходження, навіть якщо
    частина вже була видана/списана. Тепер працюємо по конкретній партії і
    перераховуємо blood_units після руху.
    """
    expired=[]
    all_items=rows("SELECT * FROM stock_entries WHERE type='Надходження' AND expiry IS NOT NULL AND expiry<>'' ORDER BY expiry ASC, id ASC")
    today=datetime.now().date()
    seen=set()
    for x in all_items:
        try:
            exp=datetime.strptime(x.get("expiry",""), "%Y-%m-%d").date()
            if exp >= today:
                continue
            key=_stock_entry_lot_identity(x)
            if key in seen:
                continue
            seen.add(key)
            if not _stock_lot_has_identifier(x):
                continue
            remaining=_stock_lot_balance_py(key) or 0.0
            if remaining <= 0.000001:
                _recalc_blood_unit_status_for_stock_record(x, 'Прострочена партія без залишку')
                continue
            unit=_find_unit_for_stock_receipt(x)
            if not unit:
                continue
            # V6.4.61: якщо активний резерв прострочився, вимога переходить
            # у «Потребує перегляду», резерв знімається, після чого прострочений
            # фактичний залишок партії списується як непридатний.
            reserved_qty=_reserved_qty_for_unit(unit.get('id'))
            review_count=0
            if reserved_qty > 0.000001:
                review_count=_mark_reserved_requests_expired_review(unit, 'Прострочення зарезервованої партії')
                _recalc_blood_unit_status_for_id(unit.get('id'), 'Знято прострочений резерв перед автосписанням')
                unit=row("SELECT * FROM blood_units WHERE id=?", (int(unit.get('id')),)) or unit
                reserved_qty=_reserved_qty_for_unit(unit.get('id'))
                remaining=_stock_lot_balance_py(key) or 0.0
            writeoff_qty=max(0.0, float(remaining or 0.0)-float(reserved_qty or 0.0))
            if writeoff_qty <= 0.000001:
                unit_event(unit.get('id'), 'auto_expire_blocked_reserved', f"Партія прострочена, але зарезервована кількість {reserved_qty:g}; автосписання пропущене")
                expired.append({**dict(x), 'expired_amount': 0, 'reserved_skipped': reserved_qty, 'review_requests': review_count})
                continue
            execute("""INSERT INTO stock_entries(created_at,user_id,type,component,donor_group,donor_rh,amount,pack_no,series,expiry,patient_name,note,qr_code,unit_id)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (now(), current_user()["id"], "Списання", x.get("component",""), x.get("donor_group",""), x.get("donor_rh",""),
                     writeoff_qty, x.get("pack_no",""), x.get("series",""), x.get("expiry",""), "", "Автосписання: закінчився термін",
                     x.get("qr_code","") or unit.get('qr_code',''), unit.get('id')))
            _recalc_blood_unit_status_for_id(unit.get('id'), 'Автосписання простроченої партії')
            expired.append({**dict(x), 'expired_amount': writeoff_qty, 'reserved_skipped': reserved_qty, 'review_requests': review_count})
        except Exception:
            db_rollback_safe()
            pass
    audit("auto_expire", f"expired={len(expired)}")
    if expired:
        telegram_alert(f"Банк крові: автосписано прострочені компоненти: {len(expired)}")
    return jsonify(ok=True, expired=len(expired))

@app.get("/api/transfusions")
@role_required("admin","transfusion")
def api_transfusions():
    return jsonify(rows("""SELECT id,created_at,patient_name,component,amount,status,doctor_name,issued_at,used_at,pack_no,series,donor_group,donor_rh,reaction_present,reaction_type
                           FROM requests
                           WHERE status IN ('Видано','Використано','Списано') OR used_at IS NOT NULL OR issued_at IS NOT NULL
                           ORDER BY id DESC LIMIT 500"""))

@app.get("/api/barcode/<code>")
@role_required("admin","transfusion")
def api_barcode(code):
    return jsonify(ok=True, code=code, label=f"Пакет/QR: {code}")


@app.post("/api/notifications/test")
@login_required
def api_notifications_test():
    u=current_user()
    notify(u["id"], "Тестове повідомлення", "Browser notification / PWA перевірка")
    telegram_alert(f"🔔 Тестове повідомлення для {u['username']}", "test", force=True)
    return jsonify(ok=True)

@app.get("/api/audit/security")
@role_required("admin")
def api_audit_security():
    return jsonify(
        failed_logins=rows("SELECT username, ip, user_agent, created_at FROM login_attempts WHERE ok=0 ORDER BY id DESC LIMIT 100"),
        recent_actions=rows("SELECT created_at,username,role,action,details,ip,user_agent FROM audit ORDER BY id DESC LIMIT 100")
    )


@app.post("/api/transfusions/event")
@role_required("admin","transfusion","doctor","nurse")
def api_transfusion_event():
    d=request.json or {}
    req_id=int(d.get("request_id") or d.get("id") or 0)
    req=row("SELECT * FROM requests WHERE id=?", (req_id,))
    if not req:
        return jsonify(ok=False,error="Вимогу не знайдено")
    u=current_user()
    if u.get("role") in ("doctor","nurse") and not owns_request(u, req_id):
        return forbid_json()
    sig = signature_hash(f"{req_id}|{current_user()['username']}|{now()}|{d.get('result','')}")
    table=v54_table("transfusion_events")
    execute(f"""INSERT INTO {table}(created_at,request_id,patient_name,component,pack_no,nurse_name,doctor_name,started_at,finished_at,result,signature)
                VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
            (now(),req_id,req.get("patient_name",""),req.get("component",""),req.get("pack_no",""),d.get("nurse_name",current_user().get("full_name","")),req.get("doctor_name",""),d.get("started_at",""),d.get("finished_at",""),d.get("result",""),sig))
    audit("transfusion_event", f"request={req_id}")
    return jsonify(ok=True, signature=sig)

@app.get("/api/transfusions/events")
@role_required("admin","transfusion")
def api_transfusion_events():
    table=v54_table("transfusion_events")
    return jsonify(rows(f"SELECT * FROM {table} ORDER BY id DESC LIMIT 500"))

@app.post("/api/reactions/register")
@role_required("admin","transfusion","doctor","nurse")
def api_reactions_register():
    u=current_user(); d=request.json or {}
    rid=int(d.get("request_id") or 0)
    if u.get("role") in ("doctor","nurse") and not owns_request(u, rid):
        return forbid_json()
    table=v54_table("reaction_registry")
    execute(f"""INSERT INTO {table}(created_at,request_id,patient_name,reaction_type,severity,description,action_taken,result,reported_by)
                VALUES(?,?,?,?,?,?,?,?,?)""",
            (now(),rid,d.get("patient_name",""),d.get("reaction_type",""),d.get("severity",""),d.get("description",""),d.get("action_taken",""),d.get("result",""),current_user().get("username","")))
    audit("reaction_register", d.get("patient_name",""))
    telegram_event_reaction(rid, d.get("reaction_type",""), d.get("severity",""), d.get("patient_name",""))
    return jsonify(ok=True)

@app.get("/api/reactions/registry")
@role_required("admin","transfusion")
def api_reactions_registry():
    table=v54_table("reaction_registry")
    return jsonify(rows(f"SELECT * FROM {table} ORDER BY id DESC LIMIT 500"))

@app.post("/api/external/event")
@api_token_required
def api_external_event():
    d=request.json or {}
    table=v54_table("api_events")
    execute(f"INSERT INTO {table}(created_at,source,event_type,payload) VALUES(?,?,?,?)",
            (now(),d.get("source","external"),d.get("event_type",""),json.dumps(d,ensure_ascii=False)))
    return jsonify(ok=True)

@app.get("/api/external/status")
@api_token_required
def api_external_status():
    return jsonify(ok=True, version="V6.4.70", postgres=IS_POSTGRES, time=now())

@app.get("/api/security/full-audit")
@role_required("admin")
def api_security_full_audit():
    return jsonify(
        audit=rows("SELECT * FROM audit ORDER BY id DESC LIMIT 500"),
        failed=rows("SELECT * FROM login_attempts WHERE ok=0 ORDER BY id DESC LIMIT 200"),
        users=rows("SELECT id,username,role,active,created_at,failed_logins,locked_until FROM users ORDER BY id")
    )

@app.get("/api/backup/encryption-status")
@role_required("admin")
def api_backup_encryption_status():
    return jsonify(ok=True, enabled=bool(BACKUP_ENCRYPTION_KEY), note="V5.4 placeholder: use Render PostgreSQL backups or encrypted external backup worker for full production.")



@app.get("/api/telegram/status")
@role_required("admin","transfusion")
def api_telegram_status():
    return jsonify(
        ok=True,
        enabled=TELEGRAM_ENABLED,
        bot_configured=bool(TELEGRAM_BOT_TOKEN),
        chat_configured=bool(TELEGRAM_CHAT_ID),
        silent_now=telegram_in_silent_time(),
        silent_start=TELEGRAM_SILENT_START,
        silent_end=TELEGRAM_SILENT_END,
        anti_spam_minutes=TELEGRAM_ANTI_SPAM_MINUTES
    )

@app.post("/api/telegram/test")
@role_required("admin","transfusion")
def api_telegram_test():
    u=current_user()
    ok, resp = telegram_send_message(
        f"✅ <b>Тест Telegram</b>\\nСистема: Банк крові\\nКористувач: {u.get('username','')}\\nЧас: {now()}",
        event_type="test",
        force=True
    )
    return jsonify(ok=ok, response=str(resp)[:1000])

@app.get("/api/telegram/logs")
@role_required("admin","transfusion")
def api_telegram_logs():
    table = telegram_table("telegram_logs")
    return jsonify(rows(f"SELECT * FROM {table} ORDER BY id DESC LIMIT 200"))

@app.get("/api/telegram/queue")
@role_required("admin","transfusion")
def api_telegram_queue():
    table = telegram_table("telegram_queue")
    return jsonify(rows(f"SELECT * FROM {table} ORDER BY id DESC LIMIT 200"))

@app.post("/api/telegram/retry")
@role_required("admin","transfusion")
def api_telegram_retry():
    sent = telegram_retry_queue()
    return jsonify(ok=True, sent=sent)




@app.post("/api/backups/nightly-run")
@role_required("admin")
def api_backups_nightly_run():
    try:
        p=make_backup("nightly")
        cleanup_old_backups()
        try: telegram_event_backup(True)
        except Exception as e: log_suppressed_error('suppressed', e)
        return jsonify(ok=True, filename=os.path.basename(p))
    except Exception as e:
        return jsonify(ok=False,error=str(e)),500

@app.post("/api/backups/rollback-snapshot")
@role_required("admin")
def api_rollback_snapshot():
    p=make_rollback_snapshot("manual")
    return jsonify(ok=bool(p), filename=os.path.basename(p) if p else "")

@app.get("/api/audit/export/<fmt>")
@role_required("admin")
def api_audit_export(fmt):
    data=rows("SELECT created_at,username,role,action,details,ip,user_agent FROM audit ORDER BY id DESC LIMIT 5000")
    fmt=fmt.lower()
    if fmt=="csv":
        bio=BytesIO()
        header=["created_at","username","role","action","details","ip","user_agent"]
        lines=[",".join(header)]
        for r in data:
            vals=[str(r.get(k,"")).replace('"','""') for k in header]
            lines.append(",".join([f'"{v}"' for v in vals]))
        bio.write("\n".join(lines).encode("utf-8-sig")); bio.seek(0)
        return send_file(bio,as_attachment=True,download_name="audit.csv",mimetype="text/csv")
    if fmt=="xlsx":
        wb=Workbook(); ws=wb.active; ws.title="Audit"
        header=["created_at","username","role","action","details","ip","user_agent"]; ws.append(header)
        for r in data: ws.append([r.get(k,"") for k in header])
        bio=BytesIO(); wb.save(bio); bio.seek(0)
        return send_file(bio,as_attachment=True,download_name="audit.xlsx")
    return jsonify(ok=False,error="fmt must be csv or xlsx"),400

@app.post("/api/maintenance/run")
@role_required("admin")
def api_maintenance_run():
    run_db_indexes(); cleanup_old_backups(); nightly_backup_if_due()
    try: telegram_retry_queue()
    except Exception as e: log_suppressed_error('suppressed', e)
    return jsonify(ok=True)


@app.post("/api/admin/bootstrap")
def api_admin_bootstrap():
    token = request.headers.get("X-API-Token","")
    if not API_TOKEN or token != API_TOKEN:
        return jsonify(ok=False,error="API token invalid"), 403
    created = ensure_default_admin()
    return jsonify(ok=True, created=created, login="Sepsis", password="1986")


@app.route("/api/admin/bootstrap-browser", methods=["GET","POST"])
def api_admin_bootstrap_browser():
    token = request.headers.get("X-API-Token","") or request.args.get("token","")
    if not API_TOKEN or token != API_TOKEN:
        return jsonify(ok=False,error="API token invalid"), 403
    created = ensure_default_admin()
    return jsonify(ok=True, created=created, login="Sepsis", password="1986")



def telegram_allowed_notifications_for_role(role):
    role=(role or "").lower()
    if role == "admin":
        return ["new_requests","critical","expiring","reactions","backups"]
    if role == "transfusion":
        return ["new_requests","critical","expiring","reactions"]
    return ["new_requests"]

def telegram_user_by_chat_id(chat_id):
    try:
        return row("SELECT * FROM users WHERE telegram_chat_id=? LIMIT 1", (str(chat_id),))
    except Exception:
        return None

def telegram_notify_request_owner(req, message, event_type="new_request"):
    try:
        uid = req.get("created_by") if isinstance(req, dict) else None
        if uid:
            return telegram_send_to_user(uid, message, event_type=event_type, force=False)
    except Exception as e:
        log_suppressed_error('suppressed', e)
    return False, "owner not notified"

@app.get("/api/telegram/me")
@login_required
def api_telegram_me():
    try:
        ensure_telegram_user_columns_safe()
        u=current_user()
        role=(u.get("role") or "doctor").lower()
        link=telegram_link_url(u)
        allowed=telegram_allowed_notifications_for_role(role)
        return jsonify(ok=True,
            role=role,
            allowed_notifications=allowed,
            telegram_chat_id=u.get("telegram_chat_id","") or "",
            telegram_username=u.get("telegram_username","") or "",
            telegram_enabled=bool(u.get("telegram_enabled")),
            connect_url=link,
            link_url=link,
            bot_username=TELEGRAM_BOT_USERNAME,
            settings={
                "new_requests": int(u.get("telegram_notify_new_requests") or 0),
                "critical": int(u.get("telegram_notify_critical") or 0),
                "expiring": int(u.get("telegram_notify_expiring") or 0),
                "reactions": int(u.get("telegram_notify_reactions") or 0),
                "backups": int(u.get("telegram_notify_backups") or 0),
            })
    except Exception as e:
        return jsonify(ok=False, error=f"Telegram me error: {str(e)}"), 500

@app.post("/api/telegram/me/settings")
@login_required
def api_telegram_me_settings():
    try:
        ensure_telegram_user_columns_safe()
        u=current_user()
        role=(u.get("role") or "doctor").lower()
        allowed=set(telegram_allowed_notifications_for_role(role))
        d=request.json or {}
        new_requests = 1 if ("new_requests" in allowed and d.get("new_requests")) else 0
        critical = 1 if ("critical" in allowed and d.get("critical")) else 0
        expiring = 1 if ("expiring" in allowed and d.get("expiring")) else 0
        reactions = 1 if ("reactions" in allowed and d.get("reactions")) else 0
        backups = 1 if ("backups" in allowed and d.get("backups")) else 0
        execute("""UPDATE users SET telegram_enabled=?, telegram_notify_new_requests=?, telegram_notify_critical=?,
                   telegram_notify_expiring=?, telegram_notify_reactions=?, telegram_notify_backups=? WHERE id=?""",
                (1 if d.get("telegram_enabled") else 0, new_requests, critical, expiring, reactions, backups, u["id"]))
        return jsonify(ok=True, allowed_notifications=list(allowed))
    except Exception as e:
        return jsonify(ok=False, error=f"Telegram settings error: {str(e)}"), 500

@app.post("/api/telegram/me/test")
@login_required
def api_telegram_me_test():
    u=current_user()
    ok, resp = telegram_send_to_user(u["id"], f"✅ Тест персонального Telegram для {u.get('full_name') or u.get('username')}", "test", force=True)
    return jsonify(ok=ok, response=str(resp)[:1000])

@app.route("/telegram/webhook", methods=["GET","POST"])
def telegram_webhook():
    if request.method == "GET":
        return jsonify(ok=True, info="Telegram webhook endpoint active")
    update = request.json or {}
    result = telegram_process_update(update)
    return jsonify(ok=True, result=result)

@app.post("/api/telegram/poll")
@role_required("admin","transfusion")
def api_telegram_poll():
    if not TELEGRAM_BOT_TOKEN:
        return jsonify(ok=False,error="TELEGRAM_BOT_TOKEN missing"), 400
    try:
        import urllib.request, json as _json
        url=f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates"
        with urllib.request.urlopen(url, timeout=8) as resp:
            data=_json.loads(resp.read().decode("utf-8","ignore"))
        count=0
        for upd in data.get("result",[]):
            telegram_process_update(upd)
            count += 1
        return jsonify(ok=True, processed=count)
    except Exception as e:
        return jsonify(ok=False,error=str(e)), 500


def ensure_telegram_user_columns_safe():
    try:
        cols = [
            ("telegram_chat_id","TEXT"),
            ("telegram_username","TEXT"),
            ("telegram_enabled","INTEGER"),
            ("telegram_notify_new_requests","INTEGER"),
            ("telegram_notify_critical","INTEGER"),
            ("telegram_notify_expiring","INTEGER"),
            ("telegram_notify_reactions","INTEGER"),
            ("telegram_notify_backups","INTEGER"),
        ]
        for name, typ in cols:
            add_column_if_missing("users", name, typ)
        for name, val in {
            "telegram_enabled":0,
            "telegram_notify_new_requests":1,
            "telegram_notify_critical":1,
            "telegram_notify_expiring":1,
            "telegram_notify_reactions":1,
            "telegram_notify_backups":0,
        }.items():
            try:
                execute(f"UPDATE users SET {name}=? WHERE {name} IS NULL", (val,))
            except Exception as e:
                log_suppressed_error('suppressed', e)
        return True
    except Exception as e:
        try: print("TELEGRAM_SAFE_MIGRATION_ERROR:", e)
        except Exception as e: log_suppressed_error('suppressed', e)
        return False

def safe_startup_check():
    errors=[]
    for fn_name in ["ensure_telegram_user_columns_safe","ensure_default_admin","run_db_indexes"]:
        try:
            fn=globals().get(fn_name)
            if fn: fn()
        except Exception as e:
            errors.append(fn_name + ":" + str(e))
    if errors:
        try: print("SAFE_STARTUP_ERRORS:", " | ".join(errors))
        except Exception as e: log_suppressed_error('suppressed', e)
    try:
        ensure_traceability_tables()
    except Exception as e:
        errors.append('traceability:' + str(e))
    try:
        ensure_v59_tables()
    except Exception as e:
        errors.append('v59:' + str(e))
    return errors


# V592_LOGIN_ATTEMPTS_MIGRATION_MARKER: ALTER TABLE login_attempts ADD COLUMN ip_address

def ensure_login_attempts_columns_v592():
    """Ensure login_attempts table works on SQLite and PostgreSQL."""
    if IS_POSTGRES:
        execute("""CREATE TABLE IF NOT EXISTS login_attempts(
            id SERIAL PRIMARY KEY,
            created_at TEXT,
            username TEXT,
            ip_address TEXT,
            ok INTEGER DEFAULT 0,
            user_agent TEXT
        )""")
    else:
        execute("""CREATE TABLE IF NOT EXISTS login_attempts(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT,
            username TEXT,
            ip_address TEXT,
            ok INTEGER DEFAULT 0,
            user_agent TEXT
        )""")
    try:
        add_column_if_missing("login_attempts", "ip_address", "TEXT")
    except Exception:
        db_rollback_safe()
    return True


@app.get("/api/health-debug")
@api_token_required
def api_health_debug():
    out={"ok":False,"version":APP_VERSION,"postgres":IS_POSTGRES,"database_type":"PostgreSQL" if IS_POSTGRES else "SQLite", "log_suppressed_errors": os.environ.get("LOG_SUPPRESSED_ERRORS","1") == "1", "token_modes":["X-API-Token","?token="]}
    try:
        out["db_select"]=row("SELECT 1 AS ok")
        pg_errors=ensure_postgres_compat_tables()
        out["migrations"]=v593_fix_all_known_migrations()
        try:
            ensure_v640_schema()
            ensure_v6437_report_history_tables()
        except Exception as e:
            db_rollback_safe(); out.setdefault("schema_warnings", []).append(str(e))
        if pg_errors:
            out["postgres_compat_errors"]=pg_errors
        out["tables"]=pg_table_health()
        out["columns"]=schema_column_health()
        out["admins"]=len(rows("SELECT id FROM users WHERE role='admin' AND active=1"))
        out["ok"]=True
    except Exception as e:
        db_rollback_safe()
        out["error"]=str(e)
    return jsonify(out), (200 if out.get("ok") else 500)


@app.get("/api/health-postgres")
@api_token_required
def api_health_postgres_v655():
    errors=[]
    migrations=[]
    try:
        errors.extend(ensure_postgres_compat_tables())
        try:
            migrations = v593_fix_all_known_migrations()
            ensure_login_attempts_columns_v592()
            ensure_v640_schema()
            ensure_v6437_report_history_tables()
        except Exception as e:
            db_rollback_safe(); errors.append(str(e))
        tables=pg_table_health()
        columns=schema_column_health()
        table_ok = all(v is True for v in tables.values() if not isinstance(v,str))
        column_ok = all(all(v is True for v in cols.values()) for cols in columns.values())
        ok=not errors and table_ok and column_ok
        return jsonify(ok=ok, version=APP_VERSION, postgres=IS_POSTGRES, database_url_set=bool(DATABASE_URL), tables=tables, columns=columns, migrations=migrations, table_ok=table_ok, column_ok=column_ok, errors=errors, db_select=row("SELECT 1 AS ok"), token_modes=["X-API-Token","?token="]), (200 if ok else 500)
    except Exception as e:
        db_rollback_safe()
        return jsonify(ok=False, version=APP_VERSION, postgres=IS_POSTGRES, error=str(e)), 500


@app.get("/api/render-readiness")
@api_token_required
def api_render_readiness_v670():
    """Read-only deployment readiness report for Render/PostgreSQL.
    It does not mutate business data; it verifies DB/schema/routes/critical APIs.
    """
    try:
        # Run idempotent lightweight schema preparation so this endpoint reflects
        # current readiness, not whether a lazy migration has been triggered yet.
        try:
            ensure_postgres_compat_tables()
            v593_fix_all_known_migrations()
            ensure_login_attempts_columns_v592()
            ensure_v640_schema()
            ensure_v6437_report_history_tables()
        except Exception as e:
            db_rollback_safe(); log_suppressed_error("render_readiness_schema_prepare", e)
        tables = pg_table_health()
        columns = schema_column_health()
        table_ok = all(v is True for v in tables.values() if not isinstance(v, str))
        column_ok = all(all(v is True for v in cols.values()) for cols in columns.values())
        required_routes = [
            "/api/version", "/api/health", "/api/stock", "/api/component-stock",
            "/api/component-stock/approved-requests", "/api/request/action",
            "/api/request/used", "/api/request/writeoff", "/api/request/cancel-issue",
            "/api/stock/add", "/api/stock/delete", "/api/writeoff", "/api/writeoff/act",
            "/api/units/writeoff", "/api/barcode/issue-check", "/api/barcode/issue",
            "/api/telegram/queue", "/api/health-postgres", "/api/health-debug"
        ]
        routes_by_rule = {}
        for rule in app.url_map.iter_rules():
            methods = sorted(m for m in rule.methods if m not in ("HEAD", "OPTIONS"))
            routes_by_rule.setdefault(str(rule.rule), []).append({"endpoint": rule.endpoint, "methods": methods})
        route_presence = {r: (r in routes_by_rule) for r in required_routes}
        duplicate_routes = {r: v for r, v in routes_by_rule.items() if len(v) > 1 and not (len(v) == 2 and {tuple(x["methods"]) for x in v} == {("GET",), ("POST",)})}
        counts = {}
        for t in ["users", "requests", "stock_entries", "blood_units", "unit_events", "trash"]:
            try:
                counts[t] = (row(f"SELECT COUNT(*) AS c FROM {t}") or {}).get("c", 0)
            except Exception as e:
                db_rollback_safe(); counts[t] = "ERR: " + str(e)[:120]
        ok = bool(table_ok and column_ok and all(route_presence.values()) and not duplicate_routes)
        return jsonify(
            ok=ok,
            version=APP_VERSION,
            app_title=APP_TITLE,
            postgres=IS_POSTGRES,
            database_type="PostgreSQL" if IS_POSTGRES else "SQLite",
            database_url_set=bool(DATABASE_URL),
            table_ok=table_ok,
            column_ok=column_ok,
            tables=tables,
            columns=columns,
            route_presence=route_presence,
            duplicate_routes=duplicate_routes,
            counts=counts,
            csrf_enabled=True,
            token_modes=["X-API-Token", "?token="],
            recommended_next=[
                "Create one stock receipt",
                "Create and approve one request",
                "Reserve, issue, used",
                "Generate /reports/full.xlsx and /reports/export/pdf",
                "Check Render logs for ERROR/Traceback/suppressed-error"
            ]
        ), (200 if ok else 500)
    except Exception as e:
        db_rollback_safe()
        return jsonify(ok=False, version=APP_VERSION, postgres=IS_POSTGRES, error=str(e)), 500

@app.errorhandler(500)
def v572_error_page(e):
    try: print("V572_500:", e)
    except Exception as e: log_suppressed_error('suppressed', e)
    if request.path.startswith("/api/"):
        return jsonify(ok=False, error="Помилка сервера", details=str(e) if DEBUG else ""), 500
    return render_template("login.html", error="Помилка сервера. Відкрий /api/health-debug або Render Logs."), 500


# ================= V5.8.1 TRACEABILITY CORE =================
def ensure_traceability_tables():
    """Create traceability/incompatibility tables with DB-specific primary keys."""
    pk = "SERIAL PRIMARY KEY" if IS_POSTGRES else "INTEGER PRIMARY KEY AUTOINCREMENT"
    try:
        execute(f"""CREATE TABLE IF NOT EXISTS incompatibility_log(
            id {pk},
            created_at TEXT,
            request_id INTEGER,
            patient_name TEXT,
            patient_group TEXT,
            patient_rh TEXT,
            donor_group TEXT,
            donor_rh TEXT,
            component TEXT,
            reason TEXT,
            override_used INTEGER DEFAULT 0,
            issued_by TEXT,
            workstation TEXT,
            ip_address TEXT,
            notes TEXT
        )""")
    except Exception:
        db_rollback_safe()
    try:
        execute(f"""CREATE TABLE IF NOT EXISTS package_traceability(
            id {pk},
            package_code TEXT,
            action_type TEXT,
            patient_name TEXT,
            request_id INTEGER,
            user_name TEXT,
            department TEXT,
            created_at TEXT,
            notes TEXT
        )""")
    except Exception:
        db_rollback_safe()
    return True


def _v581_user_name():
    try:
        return current_user().get("username","")
    except Exception:
        return "system"

def log_traceability(package_code, action_type, patient_name="", request_id=None, department="", notes=""):
    ensure_traceability_tables()
    try:
        execute("""INSERT INTO package_traceability(package_code,action_type,patient_name,request_id,user_name,department,created_at,notes)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (package_code or "", action_type or "", patient_name or "", request_id, _v581_user_name(), department or "", now(), notes or ""))
        return True
    except Exception as e:
        try: print("TRACEABILITY_LOG_ERROR:", e)
        except Exception as e: log_suppressed_error('suppressed', e)
        return False

def log_incompatibility(request_id, req, donor_group, donor_rh, component, reason, override_used=0, notes=""):
    ensure_traceability_tables()
    try:
        execute("""INSERT INTO incompatibility_log(created_at,request_id,patient_name,patient_group,patient_rh,donor_group,donor_rh,component,reason,override_used,issued_by,workstation,ip_address,notes)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (now(), request_id, req.get("patient_name",""), req.get("patient_group",""), req.get("patient_rh",""),
                 donor_group or "", donor_rh or "", component or req.get("component",""), reason or "", 1 if override_used else 0,
                 _v581_user_name(), request.headers.get("User-Agent","")[:500], request.remote_addr or "", notes or ""))
        try:
            telegram_broadcast_roles(
                f"🔴 <b>НЕСУМІСНІСТЬ</b>\nПацієнт: {req.get('patient_name','')}\nКомпонент: {component or req.get('component','')}\nПацієнт: {req.get('patient_group','')} {req.get('patient_rh','')}\nДонор: {donor_group} {donor_rh}\nПричина: {reason}",
                ("admin","transfusion"), "critical", True
            )
        except Exception:
            db_rollback_safe()
            pass
        return True
    except Exception as e:
        try: print("INCOMPATIBILITY_LOG_ERROR:", e)
        except Exception as e: log_suppressed_error('suppressed', e)
        return False

def barcode_find_package(code):
    code = str(code or "").strip()
    if not code:
        return None
    try:
        return row("SELECT * FROM stock_entries WHERE qr_code=? OR pack_no=? OR series=? ORDER BY id DESC LIMIT 1", (code, code, code))
    except Exception:
        return None

@app.get("/api/incompatibility")
@role_required("admin","transfusion")
def api_incompatibility_list():
    ensure_traceability_tables()
    return jsonify(rows("SELECT * FROM incompatibility_log ORDER BY id DESC LIMIT 500"))

def _traceability_rows_for_code(code):
    """V6.4.40: простежуваність має працювати не тільки з package_traceability.
    Збираємо історію з журналу простежуваності, рухів складу, вимог, blood_units і unit_events.
    """
    ensure_traceability_tables()
    code = str(code or "").strip()
    if not code:
        return []
    out=[]
    try:
        for x in rows("SELECT * FROM package_traceability WHERE package_code=? ORDER BY id DESC LIMIT 300", (code,)):
            out.append(dict(x))
    except Exception:
        db_rollback_safe()
    try:
        for x in rows("""SELECT * FROM stock_entries
                         WHERE qr_code=? OR pack_no=? OR series=?
                         ORDER BY id DESC LIMIT 300""", (code, code, code)):
            out.append({
                "created_at": x.get("created_at"),
                "action_type": x.get("type") or "Рух складу",
                "patient_name": x.get("patient_name") or "",
                "request_id": "",
                "user_name": "",
                "department": "",
                "notes": f"{x.get('component','')} {x.get('donor_group','')} {x.get('donor_rh','')} · к-сть {x.get('amount','')} · пакет {x.get('pack_no','')} · серія {x.get('series','')} · {x.get('note','')}",
                "package_code": code,
            })
    except Exception:
        db_rollback_safe()
    try:
        for r in rows("""SELECT * FROM requests
                         WHERE pack_no=? OR series=? OR CAST(unit_id AS TEXT)=?
                         ORDER BY id DESC LIMIT 200""", (code, code, code)):
            out.append({
                "created_at": r.get("created_at") or r.get("issued_at") or r.get("writeoff_at") or r.get("used_at"),
                "action_type": "Вимога: " + str(r.get("status") or ""),
                "patient_name": r.get("patient_name") or "",
                "request_id": r.get("id"),
                "user_name": r.get("issued_by") or r.get("written_by") or "",
                "department": r.get("department") or "",
                "notes": f"{r.get('component','')} · {r.get('patient_group','')} {r.get('patient_rh','')} · донор {r.get('donor_group','')} {r.get('donor_rh','')} · {r.get('writeoff_reason','')}",
                "package_code": code,
            })
    except Exception:
        db_rollback_safe()
    try:
        units=rows("""SELECT * FROM blood_units
                     WHERE qr_code=? OR pack_no=? OR series=? OR CAST(id AS TEXT)=?
                     ORDER BY id DESC LIMIT 50""", (code, code, code, code))
        for u in units:
            out.append({
                "created_at": u.get("updated_at") or u.get("received_at"),
                "action_type": "Одиниця: " + str(u.get("status") or ""),
                "patient_name": u.get("patient_name") or "",
                "request_id": u.get("request_id") or "",
                "user_name": u.get("issued_by") or "",
                "department": "",
                "notes": f"{u.get('component','')} {u.get('donor_group','')} {u.get('donor_rh','')} · пакет {u.get('pack_no','')} · серія {u.get('series','')} · термін {u.get('expiry','')} · {u.get('writeoff_reason','')}",
                "package_code": code,
            })
            try:
                for ev in rows("SELECT * FROM unit_events WHERE unit_id=? ORDER BY id DESC LIMIT 200", (u.get('id'),)):
                    out.append({
                        "created_at": ev.get("created_at"),
                        "action_type": ev.get("event_type") or "Подія одиниці",
                        "patient_name": u.get("patient_name") or "",
                        "request_id": ev.get("request_id") or u.get("request_id") or "",
                        "user_name": ev.get("user_name") or "",
                        "department": "",
                        "notes": ev.get("note") or "",
                        "package_code": code,
                    })
            except Exception:
                db_rollback_safe()
    except Exception:
        db_rollback_safe()
    # dedupe + sort desc
    seen=set(); clean=[]
    for x in out:
        key=(str(x.get('created_at')), str(x.get('action_type')), str(x.get('request_id')), str(x.get('notes')))
        if key in seen: continue
        seen.add(key); clean.append(x)
    clean.sort(key=lambda z: str(z.get('created_at') or ''), reverse=True)
    return clean

@app.get("/api/traceability/package/<code>")
@role_required("admin","transfusion")
def api_traceability_package(code):
    return jsonify(_traceability_rows_for_code(code))

@app.get("/api/traceability/<code>")
@role_required("admin","transfusion")
def api_traceability_short_v623(code):
    items = _traceability_rows_for_code(code)
    return jsonify(ok=bool(items), code=code, items=items, error=None if items else "Код не знайдено")

@app.post("/api/traceability/log")
@role_required("admin","transfusion")
def api_traceability_log():
    d = request.json or {}
    ok = log_traceability(d.get("package_code") or d.get("qr_code"), d.get("action_type"), d.get("patient_name",""), d.get("request_id"), d.get("department",""), d.get("notes",""))
    return jsonify(ok=ok)

@app.post("/api/barcode/scan")
@role_required("admin","transfusion")
def api_barcode_scan():
    d = request.json or {}
    code = (d.get("code") or d.get("barcode") or d.get("qr_code") or "").strip()
    if not code:
        return jsonify(ok=False,error="Немає QR/штрихкоду"), 400
    pkg = barcode_find_package(code)
    if not pkg:
        return jsonify(ok=False,error="Пакет не знайдено", code=code), 404
    log_traceability(code, "scan", pkg.get("patient_name",""), None, "", "barcode scan")
    return jsonify(ok=True, package=pkg, code=code)

@app.post("/api/barcode/issue-check")
@role_required("admin","transfusion")
def api_barcode_issue_check():
    d = request.json or {}
    request_id = int(d.get("request_id") or 0)
    code = (d.get("code") or d.get("barcode") or d.get("qr_code") or "").strip()
    req = row("SELECT * FROM requests WHERE id=?", (request_id,))
    pkg = barcode_find_package(code)
    if not req:
        return jsonify(ok=False,error="Вимогу не знайдено"), 404
    if not pkg:
        return jsonify(ok=False,error="Пакет не знайдено"), 404
    donor_group = pkg.get("donor_group","")
    donor_rh = pkg.get("donor_rh","")
    reasons = []
    # V6.4.61: correct compatibility call order: patient, donor, component.
    try:
        ok1, w1 = abo_compatible(req.get("patient_group"), donor_group, req.get("component") or pkg.get("component"))
        ok2, w2 = rh_compatible(req.get("patient_rh"), donor_rh)
        if not ok1:
            reasons.append(w1 or "ABO несумісність")
        if not ok2:
            reasons.append(w2 or "Rh несумісність")
    except Exception:
        # Fallback only if compatibility helper itself fails.
        if normalize_group(donor_group) != normalize_group(req.get("patient_group","")):
            reasons.append("ABO несумісність")
        if normalize_rh(donor_rh) == "+" and normalize_rh(req.get("patient_rh")) == "-":
            reasons.append("Rh несумісність")
    try:
        if pkg.get("expiry") and pkg.get("expiry") < datetime.now().strftime("%Y-%m-%d"):
            reasons.append("Пакет прострочений")
    except Exception as e:
        log_suppressed_error('suppressed', e)
    if reasons:
        reason = "; ".join(reasons)
        log_incompatibility(request_id, req, donor_group, donor_rh, pkg.get("component",""), reason, 0, "barcode issue-check")
        return jsonify(ok=False, compatible=False, red_alert=True, reason=reason, request=req, package=pkg)
    log_traceability(code, "issue_check_ok", req.get("patient_name",""), request_id, req.get("department",""), "compatible")
    return jsonify(ok=True, compatible=True, reason="Сумісно", request=req, package=pkg)

@app.post("/api/barcode/issue")
@role_required("admin","transfusion")
def api_barcode_issue():
    """V6.4.61: barcode-видача проходить ту саму безпечну логіку, що і звичайна видача.
    Не можна просто поставити requests.status='видано' без stock_entries/blood_units.
    """
    d = request.json or {}
    request_id = int(d.get("request_id") or 0)
    code = (d.get("code") or d.get("barcode") or d.get("qr_code") or "").strip()
    req = row("SELECT * FROM requests WHERE id=?", (request_id,))
    if not req:
        return jsonify(ok=False,error="Вимогу не знайдено"), 404
    if (req.get('status') or '').strip() not in ('Погоджено','Зарезервовано','Резерв'):
        return jsonify(ok=False,error="Barcode-видача дозволена тільки після погодження або резервування вимоги"),409
    unit = row("SELECT * FROM blood_units WHERE qr_code=? OR pack_no=? OR series=? ORDER BY id DESC LIMIT 1", (code,code,code)) if code else None
    if not unit:
        return jsonify(ok=False,error="Пакет/одиницю не знайдено"), 404
    if req.get('unit_id') and int(req.get('unit_id') or 0) != int(unit.get('id') or 0):
        return jsonify(ok=False,error="Вимога вже прив’язана до іншої одиниці. Скасуйте резерв або оберіть правильний пакет."),409
    if unit.get('status') not in ('available','reserved'):
        return jsonify(ok=False,error=f"Одиниця недоступна: {unit.get('status')}"),409
    if expiry_status(unit.get('expiry')) == 'expired':
        execute("UPDATE blood_units SET status='expired', updated_at=? WHERE id=?", (now(), unit.get('id')))
        return jsonify(ok=False,error="Термін придатності компонента завершився. Видачу заблоковано."),409
    if not _component_matches_request(unit.get('component',''), req.get('component','')):
        return jsonify(ok=False,error="Компонент пакета не відповідає вимозі"),409
    ok1,w1 = abo_compatible(req.get('patient_group'), unit.get('donor_group'), req.get('component'))
    ok2,w2 = rh_compatible(req.get('patient_rh'), unit.get('donor_rh'))
    warning = '; '.join([x for x in [w1,w2] if x])
    comp_ok = 1 if (ok1 and ok2) else 0
    if not comp_ok and not d.get('override'):
        log_incompatibility(request_id, req, unit.get('donor_group',''), unit.get('donor_rh',''), unit.get('component',''), warning or 'Несумісність', 0, 'barcode issue')
        return jsonify(ok=False, compatible=False, red_alert=True, reason=warning or 'Несумісність'),409
    ok_issue, msg_issue, avail_issue, reserved_issue = _can_issue_unit_for_request(req, unit)
    if not ok_issue:
        return jsonify(ok=False,error=msg_issue),409
    try:
        issue_qty=float(str(req.get('amount') or 1).replace(',', '.'))
    except Exception:
        return jsonify(ok=False,error="Кількість у вимозі має бути числом"),400
    receipt, lot_qty, lot_err = _stock_find_receipt_for_outgoing(unit.get('component'), unit.get('donor_group'), unit.get('donor_rh'), issue_qty, unit.get('pack_no') or '', unit.get('series') or '', None)
    if not receipt:
        return jsonify(ok=False,error=f"Недостатньо в партії {unit.get('pack_no') or '—'}/{unit.get('series') or '—'}: доступно {lot_qty:g}, спроба видати {issue_qty:g}"),409
    real_unit=_find_unit_for_stock_receipt(receipt)
    if not real_unit or int(real_unit.get('id') or 0) != int(unit.get('id') or 0):
        return jsonify(ok=False,error="Видачу заблоковано: партія складу не відповідає blood_units цієї одиниці"),409
    actor=current_user().get('full_name') or current_user().get('username') or ''
    execute("""UPDATE requests SET status='Видано',issued_by=?,issued_at=?,donor_group=?,donor_rh=?,pack_no=?,series=?,expiry=?,compatibility_ok=?,compatibility_warning=?,unit_id=? WHERE id=?""",
            (actor, now(), unit.get('donor_group',''), unit.get('donor_rh',''), unit.get('pack_no',''), unit.get('series',''), unit.get('expiry',''), comp_ok, warning, int(unit.get('id')), request_id))
    execute("UPDATE blood_units SET updated_at=?, request_id=?, patient_name=?, issued_by=?, issued_at=? WHERE id=?",
            (now(), request_id, req.get('patient_name',''), actor, now(), int(unit.get('id'))))
    execute("""INSERT INTO stock_entries(created_at,user_id,type,component,donor_group,donor_rh,amount,pack_no,series,expiry,patient_name,note,qr_code,unit_id)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (now(), current_user().get('id'), 'Видача', unit.get('component'), unit.get('donor_group'), unit.get('donor_rh'), issue_qty, unit.get('pack_no',''), unit.get('series',''), unit.get('expiry',''), req.get('patient_name',''), f"Barcode-видача по вимозі №{request_id}", unit.get('qr_code',''), int(unit.get('id'))))
    after=_recalc_blood_unit_status_for_id(unit.get('id'), f"Barcode-видача по вимозі №{request_id}")
    unit_event(unit.get('id'), 'barcode_issued', f"Barcode-видача по вимозі №{request_id}", request_id)
    for _code in [code, unit.get('qr_code'), unit.get('pack_no'), unit.get('series')]:
        if _code:
            log_traceability(_code, 'issued', req.get('patient_name',''), request_id, req.get('department',''), 'barcode issue safe')
    audit('barcode_issue', f"request={request_id}; barcode={code}; unit={unit.get('id')}")
    return jsonify(ok=True, issued=True, unit=after or unit, compatibility_ok=comp_ok, warning=warning)
# ================= END V5.8.1 TRACEABILITY CORE# ================= END V5.8.1 TRACEABILITY CORE =================


# ================= V5.9 PRODUCTION TRACEABILITY SUITE =================
def ensure_v59_tables():
    try:
        ensure_traceability_tables()
    except Exception:
        db_rollback_safe()
    pk = "SERIAL PRIMARY KEY" if IS_POSTGRES else "INTEGER PRIMARY KEY AUTOINCREMENT"
    for sql in [
        f"""CREATE TABLE IF NOT EXISTS fridge_temperature_log(
            id {pk},
            fridge_name TEXT,
            temperature REAL,
            entered_by TEXT,
            created_at TEXT,
            alert_triggered INTEGER DEFAULT 0,
            notes TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS component_writeoffs(
            id {pk},
            created_at TEXT,
            package_code TEXT,
            component TEXT,
            amount REAL,
            reason TEXT,
            written_by TEXT,
            notes TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS daily_reports(
            id {pk},
            created_at TEXT,
            report_type TEXT,
            report_text TEXT,
            sent_telegram INTEGER DEFAULT 0
        )""",
        f"""CREATE TABLE IF NOT EXISTS login_attempts(
            id {pk},
            created_at TEXT,
            username TEXT,
            ip_address TEXT,
            ok INTEGER DEFAULT 0,
            user_agent TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS device_sessions(
            id {pk},
            created_at TEXT,
            username TEXT,
            ip_address TEXT,
            user_agent TEXT,
            event_type TEXT
        )"""
    ]:
        try:
            execute(sql)
        except Exception:
            db_rollback_safe()
    return True


def v59_today():
    return datetime.now().strftime("%Y-%m-%d")

def v59_scalar(sql, params=()):
    try:
        r=row(sql, params)
        if not r: return 0
        return list(r.values())[0]
    except Exception:
        return 0

def _stock_delta_sql(alias=""):
    """Єдина формула залишку складу.

    V6.4.40: SQLite LOWER() не переводить кирилицю у нижній регістр.
    Через це рухи типу «Видача»/«Списання» у SQLite рахувались як +amount,
    а не як мінус. Це давало неправильні залишки, компоненти і звіти.
    Тому для українських статусів перевіряємо і точні варіанти з великої/малої
    літери, і англомовні lower-case варіанти.
    """
    prefix = (alias + ".") if alias else ""
    raw = f"COALESCE({prefix}type,'')"
    low = f"LOWER(COALESCE({prefix}type,''))"
    a = f"COALESCE(CAST({prefix}amount AS REAL),0)"
    in_terms = "'Надходження','надходження','Прихід','прихід','IN','In','in','income','add','plus','+'"
    out_terms = "'Списання','списання','Списано','списано','Видача','видача','Видано','видано','OUT','Out','out','issue','issued','writeoff','written_off','minus','-'"
    return (
        f"CASE "
        f"WHEN {raw} IN ({in_terms}) OR {low} IN ('in','income','add','plus','+') THEN {a} "
        f"WHEN {raw} IN ({out_terms}) OR {low} IN ('out','issue','issued','writeoff','written_off','minus','-') THEN -ABS({a}) "
        f"ELSE {a} END"
    )

def _active_request_where(alias=""):
    # V6.4.40: SQLite LOWER() не працює з кирилицею, тому перевіряємо обидва регістри.
    prefix = (alias + ".") if alias else ""
    st = f"COALESCE({prefix}status,'')"
    low = f"LOWER(COALESCE({prefix}status,''))"
    blocked = [
        "Відмов", "відмов", "Спис", "спис", "Використ", "використ", "Закрит", "закрит"
    ]
    parts = [f"{st} NOT LIKE '%{x}%'" for x in blocked]
    parts += [f"{low} NOT LIKE '%deleted%'", f"{low} NOT LIKE '%trash%'"]
    return "(" + " AND ".join(parts) + ")"

def v59_dashboard_stats():
    ensure_v59_tables()
    today=v59_today()
    stock_total = v59_scalar(f"SELECT COALESCE(SUM({_stock_delta_sql()}),0) FROM stock_entries")
    active_where = _active_request_where()
    return {
        "issued_today": v59_scalar("SELECT COUNT(*) FROM requests WHERE status LIKE ? AND created_at LIKE ?", ("%видано%", today+"%")),
        "used_today": v59_scalar("SELECT COUNT(*) FROM requests WHERE status LIKE ? AND created_at LIKE ?", ("%використ%", today+"%")),
        "writeoffs_today": v59_scalar("SELECT COUNT(*) FROM component_writeoffs WHERE created_at LIKE ?", (today+"%",)),
        "incompat_today": v59_scalar("SELECT COUNT(*) FROM incompatibility_log WHERE created_at LIKE ?", (today+"%",)),
        "active_requests": v59_scalar(f"SELECT COUNT(*) FROM requests WHERE {active_where}"),
        "stock_items": stock_total,
        "stock_records": v59_scalar("SELECT COUNT(*) FROM stock_entries"),
        "temperature_alerts_today": v59_scalar("SELECT COUNT(*) FROM fridge_temperature_log WHERE alert_triggered=1 AND created_at LIKE ?", (today+"%",))
    }

def v59_daily_report_text():
    s=v59_dashboard_stats()
    lines=[
        "📊 <b>Добовий звіт банку крові</b>",
        f"Дата: {v59_today()}",
        f"📦 Записів складу: {s.get('stock_items',0)}",
        f"📋 Активні вимоги: {s.get('active_requests',0)}",
        f"✅ Видано сьогодні: {s.get('issued_today',0)}",
        f"♻️ Списано сьогодні: {s.get('writeoffs_today',0)}",
        f"🔴 Несумісності сьогодні: {s.get('incompat_today',0)}",
        f"🌡️ Температурні тривоги: {s.get('temperature_alerts_today',0)}",
    ]
    try:
        al=get_alerts_data()
        lines.append(f"⚠️ Критичні залишки: {len(al.get('low',[]))}")
        lines.append(f"⏰ Близькі терміни: {len(al.get('expiry',[]))}")
    except Exception as e:
        log_suppressed_error('suppressed', e)
    return "\n".join(lines)

@app.get("/api/dashboard/pro")
@role_required("admin","transfusion")
def api_dashboard_pro():
    return jsonify(v59_dashboard_stats())

@app.post("/api/temperature/add")
@role_required("admin","transfusion")
def api_temperature_add():
    ensure_v59_tables()
    d=request.json or {}
    fridge=(d.get("fridge_name") or d.get("fridge") or "").strip()
    if not fridge:
        return jsonify(ok=False,error="Вкажіть холодильник"),400
    try:
        temp=float(str(d.get("temperature")).replace(",","."))
    except Exception:
        return jsonify(ok=False,error="Температура має бути числом"),400
    alert=1 if temp>6 or temp<2 else 0
    execute("INSERT INTO fridge_temperature_log(fridge_name,temperature,entered_by,created_at,alert_triggered,notes) VALUES(?,?,?,?,?,?)",
            (fridge,temp,_v581_user_name(),now(),alert,d.get("notes","")))
    if alert:
        try:
            telegram_broadcast_roles(f"🌡️ <b>Температурна тривога</b>\n{fridge}: {temp}°C", ("admin","transfusion"), "critical", True)
        except Exception as e: log_suppressed_error('suppressed', e)
    return jsonify(ok=True, alert=bool(alert))

@app.get("/api/temperature")
@role_required("admin","transfusion")
def api_temperature_list():
    ensure_v59_tables()
    return jsonify(rows("SELECT * FROM fridge_temperature_log ORDER BY id DESC LIMIT 500"))

@app.post("/api/writeoff")
@role_required("admin","transfusion")
def api_writeoff_component():
    """V6.4.61: legacy /api/writeoff переведено на безпечну партійну логіку.
    Більше не списує активний резерв і не створює мінусові залишки.
    """
    ensure_v59_tables()
    d=request.json or {}
    code=(d.get("package_code") or d.get("qr_code") or d.get("code") or d.get('pack_no') or d.get('series') or "").strip()
    reason=(d.get("reason") or "").strip()
    if not code and not (d.get('pack_no') or d.get('series')):
        return jsonify(ok=False,error="Вкажіть код пакета/серію"),400
    if not reason:
        return jsonify(ok=False,error="Вкажіть причину списання"),400
    unit=None
    if d.get('unit_id'):
        unit=row("SELECT * FROM blood_units WHERE id=?", (int(d.get('unit_id')),))
    if not unit and code:
        unit=row("SELECT * FROM blood_units WHERE qr_code=? OR pack_no=? OR series=? ORDER BY id DESC LIMIT 1", (code,code,code))
    component=(unit or {}).get('component') or d.get('component','')
    donor_group=(unit or {}).get('donor_group') or d.get('donor_group') or d.get('group','')
    donor_rh=(unit or {}).get('donor_rh') or d.get('donor_rh') or d.get('rh','')
    pack_no=(unit or {}).get('pack_no') or d.get('pack_no') or code
    series=(unit or {}).get('series') or d.get('series') or code
    try:
        amount=float(str(d.get('amount') or 1).replace(',','.'))
    except Exception:
        return jsonify(ok=False,error="Кількість має бути числом"),400
    ok,msg,data=_safe_lot_writeoff(component, donor_group, donor_rh, amount, pack_no, series, d.get('stock_entry_id'), reason, 'legacy /api/writeoff', '', (unit or {}).get('qr_code') or code, None, True)
    if not ok:
        return jsonify(ok=False,error=msg),409 if 'Неможливо' in msg or 'Недостатньо' in msg else 400
    audit("component_writeoff", f"{code}: {reason}; unit={data.get('unit_id')}")
    return jsonify(ok=True, unit_id=data.get('unit_id'), amount=data.get('amount'), pack_no=data.get('pack_no'), series=data.get('series'))

@app.get("/api/writeoffs")
@role_required("admin","transfusion")
def api_writeoffs():
    ensure_v59_tables()
    return jsonify(rows("SELECT * FROM component_writeoffs ORDER BY id DESC LIMIT 500"))

@app.get("/api/traceability/search")
@role_required("admin","transfusion")
def api_traceability_search():
    ensure_v59_tables()
    q=(request.args.get("q") or "").strip()
    if not q: return jsonify([])
    like="%"+q+"%"
    return jsonify(rows("""SELECT * FROM package_traceability
                           WHERE package_code LIKE ? OR patient_name LIKE ? OR notes LIKE ?
                           ORDER BY id DESC LIMIT 300""",(like,like,like)))

@app.post("/api/telegram/daily-report")
@role_required("admin","transfusion")
def api_telegram_daily_report():
    ensure_v59_tables()
    text=v59_daily_report_text()
    sent=0
    try:
        sent=telegram_broadcast_roles(text,("admin","transfusion"),"report",True)
    except Exception:
        sent=0
    execute("INSERT INTO daily_reports(created_at,report_type,report_text,sent_telegram) VALUES(?,?,?,?)",
            (now(),"daily",text,1 if sent else 0))
    return jsonify(ok=True, sent=sent, report=text)

@app.get("/api/daily-reports")
@role_required("admin","transfusion")
def api_daily_reports():
    ensure_v59_tables()
    return jsonify(rows("SELECT * FROM daily_reports ORDER BY id DESC LIMIT 100"))

@app.post("/api/security/login-attempt")
def api_security_login_attempt():
    try:
        ensure_login_attempts_columns_v592()
        d=request.json or {}
        execute("INSERT INTO login_attempts(created_at,username,ip_address,ok,user_agent) VALUES(?,?,?,?,?)",
                (now(),str(d.get("username",""))[:120],request.remote_addr or "",1 if d.get("ok") else 0,request.headers.get("User-Agent","")[:500]))
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False,error=f"login attempt log error: {str(e)}"), 500

@app.get("/api/security/login-attempts")
@role_required("admin")
def api_security_login_attempts():
    ensure_v59_tables()
    return jsonify(rows("SELECT * FROM login_attempts ORDER BY id DESC LIMIT 300"))
# ================= END V5.9 PRODUCTION TRACEABILITY SUITE =================


@app.get("/api/security/login-attempt-test")
@role_required("admin")
def api_security_login_attempt_test():
    try:
        ensure_login_attempts_columns_v592()
        return jsonify(ok=True, count=len(rows("SELECT id FROM login_attempts LIMIT 5")))
    except Exception as e:
        return jsonify(ok=False,error=str(e)),500


# ================= V6.0 EMERGENCY HEALTH / DB FIX =================
def v593_safe_exec(sql, params=()):
    try:
        execute(sql, params)
        return True, ""
    except Exception as e:
        db_rollback_safe()
        return False, str(e)

def v593_fix_all_known_migrations():
    """
    Safe public emergency migration for old Render PostgreSQL/SQLite schemas.
    Does not require login. Does not expose secrets.
    """
    errors=[]
    try:
        init_db()
    except Exception as e:
        errors.append("init_db:"+str(e))

    for fn_name in [
        "ensure_telegram_user_columns_safe",
        "ensure_traceability_tables",
        "ensure_v59_tables",
        "ensure_postgres_compat_tables",
        "ensure_login_attempts_columns_v592",
        "ensure_default_admin"
    ]:
        try:
            fn=globals().get(fn_name)
            if fn:
                fn()
        except Exception as e:
            errors.append(fn_name+":"+str(e))

    # Extra defensive columns for known tables
    table_cols = {
        "login_attempts": [
            ("created_at","TEXT"),("username","TEXT"),("ip_address","TEXT"),("ok","INTEGER"),("user_agent","TEXT")
        ],
        "users": [
            ("telegram_chat_id","TEXT"),("telegram_username","TEXT"),("telegram_enabled","INTEGER"),
            ("telegram_notify_new_requests","INTEGER"),("telegram_notify_critical","INTEGER"),
            ("telegram_notify_expiring","INTEGER"),("telegram_notify_reactions","INTEGER"),("telegram_notify_backups","INTEGER")
        ],
        "fridge_temperature_log": [
            ("fridge_name","TEXT"),("temperature","REAL"),("entered_by","TEXT"),("created_at","TEXT"),("alert_triggered","INTEGER"),("notes","TEXT")
        ],
        "component_writeoffs": [
            ("created_at","TEXT"),("package_code","TEXT"),("component","TEXT"),("amount","REAL"),("reason","TEXT"),("written_by","TEXT"),("notes","TEXT")
        ],
        "daily_reports": [
            ("created_at","TEXT"),("report_type","TEXT"),("report_text","TEXT"),("sent_telegram","INTEGER")
        ],
        "incompatibility_log": [
            ("created_at","TEXT"),("request_id","INTEGER"),("patient_name","TEXT"),("patient_group","TEXT"),("patient_rh","TEXT"),
            ("donor_group","TEXT"),("donor_rh","TEXT"),("component","TEXT"),("reason","TEXT"),("override_used","INTEGER"),
            ("issued_by","TEXT"),("workstation","TEXT"),("ip_address","TEXT"),("notes","TEXT")
        ],
        "package_traceability": [
            ("package_code","TEXT"),("action_type","TEXT"),("patient_name","TEXT"),("request_id","INTEGER"),("user_name","TEXT"),
            ("department","TEXT"),("created_at","TEXT"),("notes","TEXT")
        ]
    }
    for table, cols in table_cols.items():
        for col, typ in cols:
            add_column_if_missing(table, col, typ)

    try:
        safe_startup_check()
    except Exception as e:
        errors.append("safe_startup_check:"+str(e))

    return errors

@app.get("/api/public-health")
def api_public_health_v593():
    # V6.4.40: легка публічна перевірка без міграцій і без розкриття службових даних.
    out={"ok":False,"version":"V6.4.70"}
    try:
        row("SELECT 1 AS ok")
        out["ok"]=True
    except Exception as e:
        out["error"]="database check failed" if not DEBUG else str(e)
    return jsonify(out), (200 if out.get("ok") else 500)

@app.get("/api/emergency-db-fix")
@api_token_required
def api_emergency_db_fix_v593():
    out={"ok":False,"version":"V6.4.70"}
    try:
        out["migrations"]=v593_fix_all_known_migrations()
        out["admins"]=len(rows("SELECT id FROM users WHERE role='admin' AND active=1"))
        out["ok"]=True
    except Exception as e:
        out["error"]=str(e)
    return jsonify(out), (200 if out.get("ok") else 500)
# ================= END V6.0 EMERGENCY HEALTH / DB FIX =================


@app.get("/api/tx-reset")
@api_token_required
def api_tx_reset_v594():
    db_rollback_safe()
    return jsonify(ok=True, version="V6.4.70", message="transaction rolled back")


@app.get("/api/ui/feature-map")
@login_required
def api_ui_feature_map():
    return jsonify(ok=True, version="V6.4.70", features={
        "patient":["patientSec","patientsSec","historySec","requestSec","requestsSec"],
        "request":["requestSec","requestsSec","createRequestSec","doctorRequestSec"],
        "myRequests":["myRequestsSec","requestsSec","requestSec"],
        "history":["historySec","patientHistorySec","patientsSec"],
        "barcode":["barcodeSec","qrSec","scanSec"],
        "stock":["stockSec","warehouseSec","dashboardProSec"],
        "temperature":["temperatureSec","fridgeSec"],
        "monitor":["monitorSec","healthSec","dashboardProSec"],
        "audit":["auditSec","monitorSec","reportsSec"]
    })


# ================= V6.0.1 STABLE CLEAN ARCHITECTURE =================
APP_VERSION = "V6.4.70"

ROLE_PERMISSIONS = {
    "admin": {
        "label": "Адміністратор",
        "sections": ["dashboard","stock","requests","reports","users","telegram","telegramPersonal","pwa","monitor","audit","maintenance","barcode","traceability","incompat","dashboardPro","temperature","writeoff","dailyReport","backup"],
        "can_manage_users": True,
        "can_edit_all": True,
        "can_delete_all": True,
        "can_view_audit": True,
        "can_backup": True
    },
    "transfusion": {
        "label": "Трансфузіолог",
        "sections": ["dashboard","stock","requests","reports","users","telegram","telegramPersonal","pwa","monitor","audit","maintenance","barcode","traceability","incompat","dashboardPro","temperature","writeoff","dailyReport"],
        "can_manage_users": True,
        "can_edit_all": True,
        "can_delete_all": True,
        "can_view_audit": True,
        "can_backup": False
    },
    "doctor": {
        "label": "Лікар",
        "sections": ["dashboard","requests","patients","history","telegramPersonal","pwa","transfusionJournal","reactions","sign"],
        "can_manage_users": False,
        "can_edit_all": False,
        "can_delete_all": False,
        "can_view_audit": False,
        "can_backup": False
    },
    "nurse": {
        "label": "Медсестра",
        "sections": ["dashboard","requests","patients","history","telegramPersonal","pwa","transfusionJournal","reactions","sign"],
        "can_manage_users": False,
        "can_edit_all": False,
        "can_delete_all": False,
        "can_view_audit": False,
        "can_backup": False
    }
}

def v60_permission(role, key):
    return bool(ROLE_PERMISSIONS.get(role, {}).get(key, False))

def v60_sections_for_role(role):
    return ROLE_PERMISSIONS.get(role, ROLE_PERMISSIONS["doctor"]).get("sections", [])

def v60_db_safe_reset():
    try:
        db().rollback()
    except Exception as e:
        log_suppressed_error('suppressed', e)
    return True

def v60_safe_migrations():
    errors = []
    for fn_name in [
        "init_db",
        "ensure_default_admin",
        "ensure_telegram_user_columns_safe",
        "ensure_traceability_tables",
        "ensure_v59_tables",
        "ensure_postgres_compat_tables",
        "ensure_login_attempts_columns_v592",
        "ensure_v640_schema",
        "ensure_v6437_report_history_tables",
        "v593_fix_all_known_migrations"
    ]:
        try:
            fn = globals().get(fn_name)
            if fn:
                fn()
        except Exception as e:
            v60_db_safe_reset()
            errors.append(fn_name + ": " + str(e))
    return errors


@app.get("/api/permissions")
@login_required
def api_permissions_v60():
    u = current_user()
    role = u.get("role","doctor")
    return jsonify(ok=True, role=role, permissions=ROLE_PERMISSIONS.get(role, ROLE_PERMISSIONS["doctor"]))

@app.get("/api/health")
def api_health_v60():
    # V6.4.40: lightweight real home status. No migrations on every health call.
    out = health_payload()
    try:
        out["db_select"] = row("SELECT 1 AS ok")
    except Exception as e:
        out["db_select"] = None
        out["database_error"] = str(e)
        out["ok"] = False
    return jsonify(out), (200 if out.get("ok") else 500)

@app.get("/api/clean-architecture")
@login_required
def api_clean_architecture_v60():
    return jsonify(ok=True, version=APP_VERSION, modules=[
        "auth","roles","stock","requests","compatibility","barcode","traceability",
        "incompatibility","telegram","pwa","audit","backup","reports","health"
    ])
# ================= END V6.0.1 STABLE CLEAN ARCHITECTURE =================







@app.post("/api/users/reset-password")
@role_required("admin","transfusion")
def api_users_reset_password_v612():
    d=request.json or {}
    username=(d.get("username") or "").strip()
    password=d.get("password") or "Password123"
    if not username:
        return jsonify(ok=False,error="username required"),400
    try:
        salt, ph = hash_password(password)
        execute("UPDATE users SET password_hash=?, salt=? WHERE username=?", (ph, salt, username))
        for col in ["must_change_password","force_password_change","password_change_required"]:
            try:
                execute(f"UPDATE users SET {col}=0 WHERE username=?", (username,))
            except Exception:
                db_rollback_safe()
        try:
            execute("UPDATE users SET first_login=0 WHERE username=?", (username,))
        except Exception:
            db_rollback_safe()
        return jsonify(ok=True)
    except Exception as e:
        db_rollback_safe()
        return jsonify(ok=False,error=str(e)),500

@app.post("/api/users/clear-first-login")
@role_required("admin","transfusion")
def api_users_clear_first_login_v611():
    d=request.json or {}
    username=(d.get("username") or "").strip()
    if not username:
        return jsonify(ok=False,error="username required"),400
    try:
        for col in ["must_change_password","force_password_change","password_change_required","first_login"]:
            try:
                execute(f"UPDATE users SET {col}=0 WHERE username=?", (username,))
            except Exception:
                db_rollback_safe()
        return jsonify(ok=True)
    except Exception as e:
        db_rollback_safe()
        return jsonify(ok=False,error=str(e)),500








@app.get("/api/ui/role-config")
@login_required
def api_ui_role_config_v613():
    u=current_user()
    role=u.get("role","doctor")
    if role in ("admin","transfusion"):
        allowed=["home","patients","requests","stock","components","reports","users","telegram","telegramPersonal","pwa","monitor","maintenance","audit","barcode","traceability","incompat","dashboardPro","temperature","writeoff","dailyReport","transfusionJournal","reactions"]
        if role == "admin":
            allowed.append("backup")
        if role in ("admin","transfusion"):
            allowed.append("trash")
    elif role=="doctor":
        # V6.4.40: doctor має бачити тільки робочі персональні розділи, без складу/моніторингу/системи
        allowed=["home","patients","requests","myRequests","history","telegramPersonal","pwa","transfusionJournal","reactions","sign"]
    elif role=="nurse":
        # V6.4.40: nurse має бачити тільки робочі персональні розділи, без складу/моніторингу/системи
        allowed=["home","patients","requests","myRequests","history","telegramPersonal","pwa","transfusionJournal","reactions","sign"]
    else:
        allowed=["home","requests"]
    return jsonify(ok=True, role=role, allowed=allowed)





@app.get("/api/stock/summary")
@role_required("admin","transfusion")
def api_stock_summary_v6415():
    """V6.4.40: підсумок складу для деталей компонентів. Працює однаково в SQLite/PostgreSQL."""
    try:
        delta=_stock_delta_sql()
        items=[]
        try:
            items = rows(f"""
                SELECT
                    COALESCE(NULLIF(component,''),'Невказаний компонент') AS component,
                    COALESCE(NULLIF(donor_group,''),'—') AS donor_group,
                    COALESCE(NULLIF(donor_rh,''),'—') AS donor_rh,
                    COALESCE(SUM({delta}),0) AS total,
                    COALESCE(SUM({delta}),0) AS qty,
                    COUNT(*) AS packs,
                    MIN(NULLIF(expiry,'')) AS nearest_expiry
                FROM stock_entries
                GROUP BY COALESCE(NULLIF(component,''),'Невказаний компонент'),
                         COALESCE(NULLIF(donor_group,''),'—'),
                         COALESCE(NULLIF(donor_rh,''),'—')
                HAVING COALESCE(SUM({delta}),0) <> 0
                ORDER BY component, donor_group, donor_rh
            """)
        except Exception:
            db_rollback_safe(); items=[]
        if not items:
            try:
                items = rows("""
                    SELECT
                        COALESCE(NULLIF(component,''),'Невказаний компонент') AS component,
                        COALESCE(NULLIF(donor_group,''),'—') AS donor_group,
                        COALESCE(NULLIF(donor_rh,''),'—') AS donor_rh,
                        COALESCE(SUM(COALESCE(amount,1)),0) AS total,
                        COALESCE(SUM(COALESCE(amount,1)),0) AS qty,
                        COUNT(*) AS packs,
                        MIN(NULLIF(expiry,'')) AS nearest_expiry
                    FROM blood_units
                    WHERE COALESCE(status,'available') IN ('available','reserved')
                    GROUP BY COALESCE(NULLIF(component,''),'Невказаний компонент'),
                             COALESCE(NULLIF(donor_group,''),'—'),
                             COALESCE(NULLIF(donor_rh,''),'—')
                    HAVING COALESCE(SUM(COALESCE(amount,1)),0) <> 0
                    ORDER BY component, donor_group, donor_rh
                """)
            except Exception:
                db_rollback_safe(); items=[]
        return jsonify(ok=True, table='stock_entries_or_blood_units', items=items)
    except Exception as e:
        db_rollback_safe()
        return jsonify(ok=False, error=str(e), items=[]), 500

@app.get("/api/warnings")
@role_required("admin","transfusion")
def api_warnings_v620():
    """V6.2.0: повертає тільки активні попередження в основний список.
    Виправляє ситуацію, коли неактивні/службові записи рахувались як активні попередження.
    """
    active=[]
    inactive=[]

    def add_active(level, title, text):
        active.append({"active": True, "level": level, "title": title, "text": text})

    def add_inactive(level, title, text):
        inactive.append({"active": False, "level": level, "title": title, "text": text})

    # 1) Склад: рахуємо поточний залишок правильно, а не просто SUM(amount)
    try:
        try:
            stock_rows = rows("""
                SELECT component,
                       donor_group as donor_group,
                       donor_rh as donor_rh,
                       COALESCE(SUM(CASE WHEN type='Надходження' THEN CAST(amount AS REAL) ELSE -CAST(amount AS REAL) END),0) total
                FROM stock_entries
                GROUP BY component, donor_group, donor_rh
            """)
        except Exception:
            db_rollback_safe()
            stock_rows = rows("""
                SELECT component,
                       donor_group as donor_group,
                       donor_rh as donor_rh,
                       COALESCE(SUM(CAST(amount AS REAL)),0) total
                FROM stock
                GROUP BY component, donor_group, donor_rh
            """)

        if not stock_rows:
            add_active("warn", "Склад порожній", "Немає компонентів на складі або дані ще не внесені.")
        else:
            for x in stock_rows:
                total = float(x.get("total") or 0)
                comp = x.get("component") or "Компонент"
                gr = x.get("donor_group") or ""
                rh = x.get("donor_rh") or ""
                label = " ".join([str(comp), str(gr), str(rh)]).strip()
                if total <= 0:
                    add_active("critical", "Критичний залишок", f"{label} — 0")
                elif total < 5:
                    add_active("warn", "Низький залишок", f"{label} — {total:g}")
                else:
                    add_inactive("ok", "Залишок в нормі", f"{label} — {total:g}")
    except Exception:
        db_rollback_safe()

    # 2) Вимоги: активними є всі, крім виданих/використаних/списаних/відмовлених/закритих
    try:
        review_count_row = row("SELECT COUNT(*) c FROM requests WHERE status='Потребує перегляду'")
        review_count = int((review_count_row or {}).get('c') or 0)
        if review_count:
            add_active('critical', 'Резерв потребує перегляду', f'Вимог із простроченим/знятим резервом: {review_count}. Потрібно повторно підібрати компонент.')
    except Exception:
        db_rollback_safe()
    try:
        reqs = rows("""
            SELECT status, COUNT(*) c
            FROM requests
            GROUP BY COALESCE(status,'active')
        """)
        active_count = 0
        inactive_count = 0
        inactive_markers = [
            "видано", "використ", "списан", "відмов", "отказ", "закрит", "closed", "done", "used", "rejected", "written"
        ]
        for r in reqs:
            st = str(r.get("status") or "active").strip().lower()
            c = int(r.get("c") or 0)
            if any(m in st for m in inactive_markers):
                inactive_count += c
            else:
                active_count += c
        if active_count:
            add_active("info", "Активні вимоги", f"Активних вимог: {active_count}")
        if inactive_count:
            add_inactive("ok", "Неактивні вимоги", f"Неактивних вимог: {inactive_count}")
    except Exception:
        db_rollback_safe()

    if not active:
        active.append({"active": True, "level": "ok", "title": "Активних попереджень немає", "text": "Критичних подій не виявлено."})

    return jsonify(ok=True, warnings=active, active=active, inactive=inactive,
                   counts={"active": len(active) if not (len(active)==1 and active[0].get("level")=="ok") else 0,
                           "inactive": len(inactive),
                           "total": (0 if (len(active)==1 and active[0].get("level")=="ok") else len(active)) + len(inactive)})




@app.get("/api/trash")
@role_required("admin","transfusion")
def api_trash_v615_final():
    """V6.4.40: кошик показує реальні записи з таблиці trash + старі записи зі status deleted/trash.
    Раніше /api/request/delete переносив вимогу у trash і видаляв з requests, але /api/trash
    читав тільки status-поля у робочих таблицях, тому видалені вимоги не відображались.
    """
    items=[]
    try:
        deleted=rows("SELECT * FROM trash ORDER BY id DESC LIMIT 300")
        for tr in deleted:
            data={}
            try:
                data=json.loads(tr.get("data") or "{}")
            except Exception:
                data={}
            item=dict(data)
            item["_trash_id"]=tr.get("id")
            item["_table"]=tr.get("source_table") or "trash"
            item["_source_id"]=tr.get("source_id")
            item["_deleted_at"]=tr.get("created_at")
            item["_deleted_by"]=tr.get("deleted_by")
            item["_reason"]=tr.get("reason")
            item["trash_id"]=tr.get("id")
            items.append(item)
    except Exception:
        db_rollback_safe()
    # Сумісність зі старими версіями, де записи могли тільки позначатися status='deleted'.
    for tbl in ["requests","stock_entries"]:
        try:
            got=rows(f"SELECT * FROM {tbl} WHERE COALESCE(status,'') IN ('deleted','trash','removed') ORDER BY id DESC LIMIT 100")
            for x in got:
                x["_table"]=tbl
                x["_legacy_status_trash"]=True
                items.append(x)
        except Exception:
            db_rollback_safe()
    return jsonify(ok=True, items=items, count=len(items))







@app.get("/api/component-stock")
@role_required("admin","transfusion")
def api_component_stock_v6415():
    """V6.4.40: єдине джерело для блоку «Компоненти на складі».
    Виправлення: у PostgreSQL не можна писати AS 'group', через це endpoint падав 500,
    а фронтенд показував «Немає компонентів», хоча розширена панель бачила залишок.
    """
    try:
        delta = _stock_delta_sql()
        items = []
        # 1) Основне джерело — рух складу stock_entries.
        try:
            items = rows(f"""
                SELECT
                    COALESCE(NULLIF(component,''),'Невказаний компонент') AS component,
                    COALESCE(NULLIF(donor_group,''),'—') AS donor_group,
                    COALESCE(NULLIF(donor_rh,''),'—') AS donor_rh,
                    COALESCE(SUM({delta}),0) AS qty,
                    COALESCE(SUM({delta}),0) AS amount,
                    COUNT(*) AS packs,
                    MIN(NULLIF(expiry,'')) AS nearest_expiry
                FROM stock_entries
                GROUP BY COALESCE(NULLIF(component,''),'Невказаний компонент'),
                         COALESCE(NULLIF(donor_group,''),'—'),
                         COALESCE(NULLIF(donor_rh,''),'—')
                HAVING COALESCE(SUM({delta}),0) <> 0
                ORDER BY component, donor_group, donor_rh
            """)
        except Exception as e:
            db_rollback_safe()
            items = []

        # 2) Fallback — нова модель blood_units, якщо рух складу порожній/не дає залишків.
        if not items:
            try:
                items = rows("""
                    SELECT
                        COALESCE(NULLIF(component,''),'Невказаний компонент') AS component,
                        COALESCE(NULLIF(donor_group,''),'—') AS donor_group,
                        COALESCE(NULLIF(donor_rh,''),'—') AS donor_rh,
                        COALESCE(SUM(COALESCE(amount,1)),0) AS qty,
                        COALESCE(SUM(COALESCE(amount,1)),0) AS amount,
                        COUNT(*) AS packs,
                        MIN(NULLIF(expiry,'')) AS nearest_expiry
                    FROM blood_units
                    WHERE COALESCE(status,'available') IN ('available','reserved')
                    GROUP BY COALESCE(NULLIF(component,''),'Невказаний компонент'),
                             COALESCE(NULLIF(donor_group,''),'—'),
                             COALESCE(NULLIF(donor_rh,''),'—')
                    HAVING COALESCE(SUM(COALESCE(amount,1)),0) <> 0
                    ORDER BY component, donor_group, donor_rh
                """)
            except Exception:
                db_rollback_safe()
                items = []

        # V6.4.61: додаємо reserved_qty / available_qty, щоб UI не плутав
        # фактичний залишок із доступним після резервування.
        for x in items:
            try:
                comp=x.get('component') or 'Невказаний компонент'
                grp=x.get('donor_group') or x.get('group') or '—'
                rh=x.get('donor_rh') or x.get('rh') or '—'
                reserved_qty=0.0
                for u in rows("""SELECT id FROM blood_units
                                 WHERE COALESCE(NULLIF(component,''),'Невказаний компонент')=?
                                   AND COALESCE(NULLIF(donor_group,''),'—')=?
                                   AND COALESCE(NULLIF(donor_rh,''),'—')=?""", (comp, grp, rh)):
                    reserved_qty += _reserved_qty_for_unit(u.get('id')) if '_reserved_qty_for_unit' in globals() else 0.0
                qty=float(x.get('qty') or x.get('amount') or 0)
                x['reserved_qty']=reserved_qty
                x['available_qty']=max(0.0, qty-reserved_qty)

                # V6.4.61: packs більше не означає кількість рухів.
                # lot_count = активні фізичні партії/пакет+серія з ненульовим балансом;
                # movement_count = кількість рядків руху складу. Це прибирає плутанину
                # коли 1 надходження + 1 видача показувало «Пакетів: 2».
                try:
                    lot_rows = rows("""SELECT * FROM stock_entries
                                      WHERE COALESCE(NULLIF(component,''),'Невказаний компонент')=?
                                        AND COALESCE(NULLIF(donor_group,''),'—')=?
                                        AND COALESCE(NULLIF(donor_rh,''),'—')=?""", (comp, grp, rh))
                    x['movement_count'] = len(lot_rows)
                    lot_balances = {}
                    for lr in lot_rows:
                        k = ((lr.get('pack_no') or '').strip(), (lr.get('series') or '').strip())
                        if not (k[0] or k[1]):
                            k = ('—', '—')
                        lot_balances[k] = lot_balances.get(k, 0.0) + float(_stock_entry_delta_py(lr) or 0.0)
                    x['lot_count'] = sum(1 for v in lot_balances.values() if abs(float(v or 0.0)) > 0.000001)
                    if not lot_rows:
                        bu_count = row("""SELECT COUNT(*) c FROM blood_units
                                          WHERE COALESCE(NULLIF(component,''),'Невказаний компонент')=?
                                            AND COALESCE(NULLIF(donor_group,''),'—')=?
                                            AND COALESCE(NULLIF(donor_rh,''),'—')=?
                                            AND COALESCE(status,'available') IN ('available','reserved')""", (comp, grp, rh))
                        x['lot_count'] = int((bu_count or {}).get('c') or 0)
                    x['packs'] = x.get('lot_count', x.get('packs', 0))
                except Exception:
                    db_rollback_safe()
                    x['movement_count'] = int(x.get('packs') or 0) if str(x.get('packs') or '').isdigit() else 0
                    x['lot_count'] = int(x.get('packs') or 0) if str(x.get('packs') or '').isdigit() else 0
            except Exception:
                x['reserved_qty']=0.0
                x['available_qty']=float(x.get('qty') or x.get('amount') or 0)
                x['lot_count']=int(x.get('packs') or 0) if str(x.get('packs') or '').isdigit() else 0
                x['movement_count']=int(x.get('packs') or 0) if str(x.get('packs') or '').isdigit() else 0
        total = sum(float(x.get('qty') or x.get('amount') or 0) for x in items)
        return jsonify(ok=True, items=items, total=total, source="stock_entries_or_blood_units")
    except Exception as e:
        db_rollback_safe()
        return jsonify(ok=False, error=str(e), items=[], total=0), 500



def _component_key_where(prefix=""):
    """SQL-фільтр для групи компонента, сумісний із SQLite/PostgreSQL."""
    return "COALESCE(NULLIF(component,''),'—')=? AND COALESCE(NULLIF(donor_group,''),'—')=? AND COALESCE(NULLIF(donor_rh,''),'—')=?"

def _component_current_qty(component, donor_group, donor_rh):
    delta = _stock_delta_sql()
    r = row(f"""
        SELECT COALESCE(SUM({delta}),0) AS qty
        FROM stock_entries
        WHERE COALESCE(NULLIF(component,''),'—')=?
          AND COALESCE(NULLIF(donor_group,''),'—')=?
          AND COALESCE(NULLIF(donor_rh,''),'—')=?
    """, (component or '—', donor_group or '—', donor_rh or '—'))
    try:
        return float((r or {}).get('qty') or 0)
    except Exception:
        return 0.0

def _approved_request_statuses_for_component_actions():
    return ("Погоджено", "Зарезервовано", "approved", "reserved")

def _final_request_statuses():
    return ("Видано", "Використано", "Списано", "Відмовлено", "deleted", "trash")

def _component_matches_request(stock_component, request_component):
    """Проста сумісність назви компонента зі спрощеною вимогою.
    У вимозі лишена проста форма: Еритроцитарні компоненти / Плазма / Тромбоцити / Кріопреципітат.
    На складі — деталізований формуляр.
    """
    sc=(stock_component or '').lower()
    rc=(request_component or '').lower()
    if not rc or rc in sc or sc in rc:
        return True
    if 'еритро' in rc and 'еритро' in sc: return True
    if 'плазм' in rc and 'плазм' in sc: return True
    if 'тромбо' in rc and 'тромбо' in sc: return True
    if 'кріо' in rc and 'кріо' in sc: return True
    if 'крио' in rc and 'кріо' in sc: return True
    return False

def _approved_requests_for_component(component='', donor_group='', donor_rh=''):
    sts=_approved_request_statuses_for_component_actions()
    ph=','.join(['?']*len(sts))
    params=list(sts)
    q=f"""SELECT id,created_at,needed_date,patient_name,birth_date,patient_status,address,department,component,patient_group,patient_rh,amount,urgency,diagnosis,note,status,created_by,doctor_name,doctor_position,compatibility_ok,compatibility_warning,approved_by,issued_by,issued_at,pack_no,series,expiry
          FROM requests WHERE COALESCE(status,'') IN ({ph}) ORDER BY id DESC LIMIT 200"""
    reqs=rows(q, tuple(params))
    out=[]
    for r in reqs:
        if component and not _component_matches_request(component, r.get('component','')):
            continue
        # Відображаємо всі погоджені, а сумісність ABO/Rh перевіряємо вже при видачі/списанні.
        out.append(r)
    return out

@app.get("/api/component-stock/approved-requests")
@role_required("admin","transfusion")
def api_component_stock_approved_requests_v6422():
    component=request.args.get('component','')
    donor_group=request.args.get('donor_group','') or request.args.get('group','')
    donor_rh=request.args.get('donor_rh','') or request.args.get('rh','')
    items=_approved_requests_for_component(component, donor_group, donor_rh)
    return jsonify(ok=True, items=items, count=len(items))

def _pick_available_unit_for_component(component, donor_group, donor_rh):
    try:
        return row("""SELECT * FROM blood_units
                    WHERE COALESCE(status,'available') IN ('available','reserved')
                      AND COALESCE(NULLIF(component,''),'—')=?
                      AND COALESCE(NULLIF(donor_group,''),'—')=?
                      AND COALESCE(NULLIF(donor_rh,''),'—')=?
                    ORDER BY CASE WHEN expiry IS NULL OR expiry='' THEN 1 ELSE 0 END, expiry ASC, id ASC LIMIT 1""",
                   (component or '—', donor_group or '—', donor_rh or '—'))
    except Exception:
        db_rollback_safe()
        return None

def _unit_same_lot_as_receipt(unit, receipt):
    """V6.4.61: точна перевірка, що blood_units відповідає саме цьому надходженню.
    Не можна прив'язувати списання P3/S3 до unit_id від P2/S2.
    """
    if not unit or not receipt:
        return False
    checks = [
        _stock_same_value(unit.get('component'), receipt.get('component')),
        _stock_same_value(moz_norm_group(unit.get('donor_group')), moz_norm_group(receipt.get('donor_group'))),
        _stock_same_value(moz_norm_rh(unit.get('donor_rh')), moz_norm_rh(receipt.get('donor_rh'))),
    ]
    if not all(checks):
        return False
    r_pack=(receipt.get('pack_no') or '').strip()
    r_series=(receipt.get('series') or '').strip()
    u_pack=(unit.get('pack_no') or '').strip()
    u_series=(unit.get('series') or '').strip()
    if r_pack and not _stock_same_value(u_pack, r_pack):
        return False
    if r_series and not _stock_same_value(u_series, r_series):
        return False
    if (r_pack or r_series) and not (u_pack or u_series):
        return False
    return True

def _find_unit_for_stock_receipt(receipt):
    """V6.4.61: знайти blood_unit тільки для конкретної партії/серії надходження.
    Якщо exact-зв'язку немає — повертаємо None і дію потрібно блокувати, а не брати випадковий unit.
    """
    if not receipt:
        return None
    try:
        if receipt.get('unit_id'):
            u=row("SELECT * FROM blood_units WHERE id=?", (int(receipt.get('unit_id')),))
            if _unit_same_lot_as_receipt(u, receipt):
                return u
            return None
        component=(receipt.get('component') or '').strip()
        group=(receipt.get('donor_group') or '').strip()
        rh=(receipt.get('donor_rh') or '').strip()
        pack=(receipt.get('pack_no') or '').strip()
        series=(receipt.get('series') or '').strip()
        if pack and series:
            return row("""SELECT * FROM blood_units WHERE COALESCE(status,'available') IN ('available','reserved')
                         AND component=? AND donor_group=? AND donor_rh=? AND pack_no=? AND series=?
                         ORDER BY id DESC LIMIT 1""", (component, group, rh, pack, series))
        if pack and not series:
            return row("""SELECT * FROM blood_units WHERE COALESCE(status,'available') IN ('available','reserved')
                         AND component=? AND donor_group=? AND donor_rh=? AND pack_no=? AND COALESCE(series,'')=''
                         ORDER BY id DESC LIMIT 1""", (component, group, rh, pack))
        if series and not pack:
            return row("""SELECT * FROM blood_units WHERE COALESCE(status,'available') IN ('available','reserved')
                         AND component=? AND donor_group=? AND donor_rh=? AND series=? AND COALESCE(pack_no,'')=''
                         ORDER BY id DESC LIMIT 1""", (component, group, rh, series))
    except Exception:
        db_rollback_safe()
    return None

@app.post("/api/component-stock/issue")
@role_required("admin","transfusion")
def api_component_stock_issue_v6422():
    """V6.4.61: безпечна видача через «Компоненти і склад».

    Обов'язково перевіряє:
    - статус вимоги;
    - сумісність;
    - кількість видачі = amount у вимозі;
    - активні резерви інших вимог;
    - відповідність stock_entries ↔ blood_units.
    """
    u=current_user(); d=request.json or {}
    component=(d.get('component') or '').strip()
    donor_group=(d.get('donor_group') or d.get('group') or '').strip()
    donor_rh=(d.get('donor_rh') or d.get('rh') or '').strip()
    rid=int(d.get('request_id') or d.get('id') or 0)
    if not component or not donor_group or not donor_rh:
        return jsonify(ok=False,error="Компонент, група і Rh обов’язкові"),400
    if not rid:
        return jsonify(ok=False,error="Для видачі оберіть погоджену вимогу"),400
    req=row("SELECT * FROM requests WHERE id=?", (rid,))
    if not req:
        return jsonify(ok=False,error="Вимогу не знайдено"),404
    st=(req.get('status') or '').strip()
    if st not in _approved_request_statuses_for_component_actions():
        return jsonify(ok=False,error="Видача дозволена тільки для погодженої/зарезервованої вимоги"),403
    if not _component_matches_request(component, req.get('component','')):
        return jsonify(ok=False,error="Компонент складу не відповідає компоненту у вимозі"),409
    ok1,w1=abo_compatible(req.get('patient_group'), donor_group, req.get('component'))
    ok2,w2=rh_compatible(req.get('patient_rh'), donor_rh)
    if not (ok1 and ok2):
        return jsonify(ok=False,error="; ".join([x for x in [w1,w2] if x]) or "Несумісність"),409
    try:
        qty=float(str(d.get('qty') or req.get('amount') or 1).replace(',','.'))
        req_qty=float(str(req.get('amount') or 1).replace(',','.'))
    except Exception:
        return jsonify(ok=False,error="Кількість має бути числом"),400
    if qty<=0 or req_qty<=0:
        return jsonify(ok=False,error="Кількість має бути більше 0"),400
    if abs(qty-req_qty) > 0.000001:
        return jsonify(ok=False,error=f"Кількість видачі ({qty:g}) має дорівнювати кількості у вимозі ({req_qty:g}). Часткова видача в цій версії не підтримується."),409
    current=_component_current_qty(component, donor_group, donor_rh)
    if current < qty:
        return jsonify(ok=False,error=f"Недостатньо компонента на складі: доступно {current:g}"),409
    stock_entry_id=int(d.get('stock_entry_id') or 0)
    receipt, lot_qty, lot_err = _stock_find_receipt_for_outgoing(component, donor_group, donor_rh, qty, (d.get('pack_no') or '').strip(), (d.get('series') or '').strip(), stock_entry_id or None)
    if not receipt:
        return jsonify(ok=False,error=f"Недостатньо в конкретній партії/серії: доступно {lot_qty:g}, спроба видати {qty:g}"),409
    unit=_find_unit_for_stock_receipt(receipt)
    if not unit:
        return jsonify(ok=False,error="Для цієї партії немає коректно прив’язаної одиниці blood_units. Додайте/виправте надходження через журнал складу."),409
    # V6.4.61: не можна видати партію, яка зарезервована під іншу вимогу.
    ok_issue, msg_issue, avail_issue, need_issue = _can_issue_unit_for_request(req, unit)
    if not ok_issue:
        return jsonify(ok=False,error=msg_issue),409
    expiry=((receipt or {}).get('expiry') or (unit or {}).get('expiry') or '').strip()
    pack_no=((receipt or {}).get('pack_no') or (unit or {}).get('pack_no') or '').strip()
    series=((receipt or {}).get('series') or (unit or {}).get('series') or '').strip()
    actor=u.get('full_name') or u.get('username') or ''
    unit_id=(unit or {}).get('id')
    execute("""INSERT INTO stock_entries(created_at,user_id,type,component,donor_group,donor_rh,amount,pack_no,series,expiry,patient_name,note,qr_code,unit_id)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (now(),u.get('id'),"Видача",component,donor_group,donor_rh,qty,pack_no,series,expiry,req.get('patient_name',''),f"Видача через Компоненти на складі по вимозі №{rid}",(unit or {}).get('qr_code',''),unit_id))
    execute("""UPDATE requests SET status='Видано', issued_by=?, issued_at=?, donor_group=?, donor_rh=?, pack_no=?, series=?, expiry=?, unit_id=? WHERE id=?""",
            (actor, now(), donor_group, donor_rh, pack_no, series, expiry, unit_id, rid))
    if unit_id:
        execute("UPDATE blood_units SET updated_at=?, request_id=?, patient_name=?, issued_by=?, issued_at=? WHERE id=?",
                (now(), rid, req.get('patient_name',''), actor, now(), int(unit_id)))
        _recalc_blood_unit_status_for_id(unit_id, f"Видача через Компоненти на складі по вимозі №{rid}")
        unit_event(unit_id, 'issued_from_component_stock', f"Видано через Компоненти на складі по вимозі №{rid}", rid)
    for _code in [pack_no, series, (unit or {}).get('qr_code')]:
        if _code:
            log_traceability(_code, 'issued', req.get('patient_name',''), rid, req.get('department',''), f'Видано через Компоненти/Склад · {component} · к-сть {qty}')
    audit('component_stock_issue', f"request={rid}; {component} {donor_group} {donor_rh}; qty={qty}")
    return jsonify(ok=True, request_id=rid, qty=qty, unit_id=unit_id)

@app.post("/api/component-stock/writeoff")
@role_required("admin","transfusion")
def api_component_stock_writeoff_v6422():
    """V6.4.61: безпечне списання агрегованого компонента.

    Якщо списання прив'язане до вимоги, кількість має дорівнювати amount у вимозі;
    часткове списання із фінальним статусом «Списано» не допускається.
    """
    u=current_user(); d=request.json or {}
    component=(d.get('component') or '').strip()
    donor_group=(d.get('donor_group') or d.get('group') or '').strip()
    donor_rh=(d.get('donor_rh') or d.get('rh') or '').strip()
    if not component or not donor_group or not donor_rh:
        return jsonify(ok=False,error="Компонент, група і Rh обов’язкові"),400
    rid=int(d.get('request_id') or 0)
    req=None
    if rid:
        req=row("SELECT * FROM requests WHERE id=?", (rid,))
        if not req:
            return jsonify(ok=False,error="Вимогу не знайдено"),404
        if (req.get('status') or '').strip() not in _approved_request_statuses_for_component_actions() + ("Видано",):
            return jsonify(ok=False,error="Списання з прив’язкою дозволене тільки для погодженої/зарезервованої/виданої вимоги"),403
        if not _component_matches_request(component, req.get('component','')):
            return jsonify(ok=False,error="Компонент складу не відповідає компоненту у вимозі"),409
    try:
        qty=float(str(d.get('qty') or (req or {}).get('amount') or 1).replace(',','.'))
    except Exception:
        return jsonify(ok=False,error="Кількість має бути числом"),400
    if qty<=0:
        return jsonify(ok=False,error="Кількість має бути більше 0"),400
    if rid and req:
        try:
            req_qty=float(str(req.get('amount') or 1).replace(',','.'))
        except Exception:
            return jsonify(ok=False,error="Кількість у вимозі має бути числом"),400
        if abs(qty-req_qty) > 0.000001:
            return jsonify(ok=False,error=f"Кількість списання ({qty:g}) має дорівнювати кількості у вимозі ({req_qty:g}). Часткове списання з фінальним статусом «Списано» не підтримується."),409
    current=_component_current_qty(component, donor_group, donor_rh)
    if current < qty:
        return jsonify(ok=False,error=f"Недостатньо компонента на складі: доступно {current:g}"),409
    reason=(d.get('reason') or 'Списання компонента зі складу').strip()
    stock_entry_id=int(d.get('stock_entry_id') or 0)
    receipt, lot_qty, lot_err = _stock_find_receipt_for_outgoing(component, donor_group, donor_rh, qty, (d.get('pack_no') or '').strip(), (d.get('series') or '').strip(), stock_entry_id or None)
    if not receipt:
        return jsonify(ok=False,error=f"Недостатньо в конкретній партії/серії: доступно {lot_qty:g}, спроба списати {qty:g}"),409
    unit=_find_unit_for_stock_receipt(receipt)
    if not unit:
        return jsonify(ok=False,error="Для цієї партії немає коректно прив’язаної одиниці blood_units. Додайте/виправте надходження через журнал складу."),409
    locked, lock_msg = _request_is_locked_to_other_unit(req, unit) if rid and req and '_request_is_locked_to_other_unit' in globals() else (False, '')
    if locked:
        return jsonify(ok=False,error=lock_msg),409
    unit_id=(unit or {}).get('id')
    expiry=((receipt or {}).get('expiry') or (unit or {}).get('expiry') or '').strip()
    pack_no=((receipt or {}).get('pack_no') or (unit or {}).get('pack_no') or '').strip()
    series=((receipt or {}).get('series') or (unit or {}).get('series') or '').strip()
    if not (pack_no or series):
        return jsonify(ok=False,error="Для списання потрібно обрати конкретну партію/серію компонента"),400
    # V6.4.61: списання через компонентний екран не може забирати резерви інших вимог.
    exclude_rid = rid if rid else None
    available_qty, lot_balance, reserved_qty = _available_qty_for_stock_record(receipt, exclude_request_id=exclude_rid)
    if qty > available_qty + 0.000001:
        return jsonify(ok=False,error=_reserved_guard_message(pack_no, series, available_qty, reserved_qty, lot_balance, 'списати')),409
    patient=(req or {}).get('patient_name','')
    execute("""INSERT INTO stock_entries(created_at,user_id,type,component,donor_group,donor_rh,amount,pack_no,series,expiry,patient_name,note,qr_code,unit_id)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (now(),u.get('id'),"Списання",component,donor_group,donor_rh,qty,pack_no,series,expiry,patient,reason + (f"; вимога №{rid}" if rid else ''),(unit or {}).get('qr_code',''),unit_id))
    if rid:
        execute("""UPDATE requests SET status='Списано', writeoff_at=?, written_by=?, writeoff_reason=?, donor_group=?, donor_rh=?, pack_no=?, series=?, expiry=?, unit_id=? WHERE id=?""",
                (now(), u.get('full_name') or u.get('username') or '', reason, donor_group, donor_rh, pack_no, series, expiry, unit_id, rid))
    if unit_id:
        execute("UPDATE blood_units SET updated_at=?, written_off_at=?, writeoff_reason=?, request_id=? WHERE id=?",
                (now(), now(), reason, rid or (unit or {}).get('request_id'), int(unit_id)))
        _recalc_blood_unit_status_for_id(unit_id, 'Списання через Компоненти на складі')
        unit_event(unit_id, 'written_off_from_component_stock', reason, rid or None)
    for _code in [pack_no, series, (unit or {}).get('qr_code')]:
        if _code:
            log_traceability(_code, 'writeoff', patient, rid or None, (req or {}).get('department',''), f'{reason} · {component} · к-сть {qty}')
    audit('component_stock_writeoff', f"request={rid or '-'}; {component} {donor_group} {donor_rh}; qty={qty}; {reason}")
    return jsonify(ok=True, request_id=rid or None, qty=qty, unit_id=unit_id)

@app.post("/api/component-stock/update")
@role_required("admin","transfusion")
def api_component_stock_update_v6417():
    """V6.4.61: безпечне редагування агрегованої позиції «Компоненти на складі».

    Важливо: агрегований екран більше НЕ переписує історичні рухи складу
    (component/group/Rh/expiry). Такі зміни потрібно робити тільки через
    редагування конкретного запису надходження (/api/stock/update), де є
    контроль партії/серії. Цей endpoint дозволяє лише корекцію кількості
    окремим рухом з обов'язковою прив'язкою до пакета/серії.
    """
    try:
        u=current_user(); d=request.json or {}
        old_component=(d.get('old_component') or d.get('component') or '').strip() or '—'
        old_group=(d.get('old_donor_group') or d.get('old_group') or d.get('donor_group') or '').strip() or '—'
        old_rh=(d.get('old_donor_rh') or d.get('old_rh') or d.get('donor_rh') or '').strip() or '—'
        new_component=(d.get('new_component') or d.get('component') or old_component).strip() or old_component
        new_group=(d.get('new_donor_group') or d.get('new_group') or d.get('donor_group') or old_group).strip() or old_group
        new_rh=(d.get('new_donor_rh') or d.get('new_rh') or d.get('donor_rh') or old_rh).strip() or old_rh
        new_expiry=(d.get('expiry') or '').strip()
        qty_raw=d.get('qty') if 'qty' in d else d.get('amount')

        # Чи існує така агрегована позиція
        existing=row("""SELECT COUNT(*) AS c FROM stock_entries
                    WHERE COALESCE(NULLIF(component,''),'—')=?
                      AND COALESCE(NULLIF(donor_group,''),'—')=?
                      AND COALESCE(NULLIF(donor_rh,''),'—')=?""", (old_component, old_group, old_rh))
        if int((existing or {}).get('c') or 0) == 0:
            return jsonify(ok=False,error="Позицію компонента не знайдено"),404

        # V6.4.61: агреговане перейменування/зміна групи/Rh/терміну заборонені.
        # Інакше можна переписати історію надходжень/списань і порушити простежуваність.
        identity_changed = (
            not _stock_same_value(old_component, new_component) or
            not _stock_same_value(moz_norm_group(old_group), moz_norm_group(new_group)) or
            not _stock_same_value(moz_norm_rh(old_rh), moz_norm_rh(new_rh))
        )
        if identity_changed:
            return jsonify(ok=False,error="Зміна компонента, групи або Rh через агрегований екран заборонена. Редагуйте конкретне надходження у журналі складу."),409

        target=None
        if qty_raw not in (None, ""):
            try:
                target=float(str(qty_raw).replace(',','.'))
            except Exception:
                return jsonify(ok=False,error="Кількість має бути числом"),400
            if target < 0:
                return jsonify(ok=False,error="Залишок не може бути від’ємним"),400

        # Якщо користувач намагається змінити лише термін агрегованої позиції — блокуємо.
        # Термін придатності належить конкретній партії, а не агрегату.
        if target is None and new_expiry:
            return jsonify(ok=False,error="Зміна терміну через агрегований екран заборонена. Редагуйте конкретне надходження у журналі складу."),409

        # Якщо користувач не змінює кількість і не змінює ключові поля — нема що оновлювати.
        if target is None:
            audit("component_stock_update_noop", f"{old_component} {old_group} {old_rh}")
            return jsonify(ok=True, skipped=True)

        current=_component_current_qty(old_component, old_group, old_rh)
        diff=round(target-current, 6)
        if abs(diff) <= 0.000001:
            audit("component_stock_update_noop", f"{old_component} {old_group} {old_rh} qty={target}")
            return jsonify(ok=True, skipped=True, qty=current)

        corr_pack=(d.get('pack_no') or '').strip()
        corr_series=(d.get('series') or '').strip()
        corr_expiry=(d.get('expiry') or '').strip()
        stock_entry_id=d.get('stock_entry_id') or None

        # V6.4.61: будь-яка корекція кількості має бути прив'язана до конкретної партії.
        if not (corr_pack or corr_series or stock_entry_id):
            return jsonify(ok=False,error="Коригування кількості потрібно робити по конкретному пакету або серії"),409

        if diff < 0:
            receipt, lot_qty, lot_err = _stock_find_receipt_for_outgoing(old_component, old_group, old_rh, abs(diff), corr_pack, corr_series, stock_entry_id)
            if not receipt:
                return jsonify(ok=False,error=f"Недостатньо в конкретній партії/серії для коригування: доступно {lot_qty:g}, потрібно {abs(diff):g}"),409
            available_qty, lot_balance, reserved_qty = _available_qty_for_stock_record(receipt)
            if abs(diff) > available_qty + 0.000001:
                return jsonify(ok=False,error=_reserved_guard_message((receipt.get('pack_no') or corr_pack), (receipt.get('series') or corr_series), available_qty, reserved_qty, lot_balance, 'скоригувати/списати')),409
            corr_pack=(receipt.get('pack_no') or corr_pack or '').strip()
            corr_series=(receipt.get('series') or corr_series or '').strip()
            corr_expiry=(receipt.get('expiry') or corr_expiry or '').strip()
            stype="Списання"
            note=f"Коригування кількості компонента з {current} до {target}"
        else:
            # V6.4.61: агрегований endpoint більше не створює нові надходження.
            # Нові партії мають створюватися тільки через /api/stock/add, де одночасно
            # створюється stock_entries + blood_units і зберігається простежуваність.
            return jsonify(ok=False,error="Збільшення залишку робіть через «Надходження» у журналі складу, а не через агреговане редагування компонента."),409

        corr_unit=_find_unit_for_stock_receipt(receipt) if diff < 0 else None
        if diff < 0 and not corr_unit:
            return jsonify(ok=False,error="Коригування неможливе: для цієї партії немає коректно прив’язаної одиниці blood_units."),409

        execute("""INSERT INTO stock_entries(created_at,user_id,type,component,donor_group,donor_rh,amount,pack_no,series,expiry,patient_name,note,qr_code,unit_id)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (now(), u.get('id'), stype, old_component, old_group, old_rh, abs(diff), corr_pack, corr_series, corr_expiry, '', note, (corr_unit or {}).get('qr_code',''), (corr_unit or {}).get('id')))
        if corr_unit and corr_unit.get('id'):
            _recalc_blood_unit_status_for_id(corr_unit.get('id'), 'Коригування кількості компонента')
        audit("component_stock_update", f"{old_component} {old_group} {old_rh}: qty {current} -> {target}; lot {corr_pack}/{corr_series}")
        return jsonify(ok=True, qty=target, diff=diff, pack_no=corr_pack, series=corr_series)
    except Exception as e:
        db_rollback_safe()
        return jsonify(ok=False,error=f"Помилка редагування компонента: {str(e)}"),500

@app.post("/api/component-stock/delete")
@role_required("admin","transfusion")
def api_component_stock_delete_v6417():
    """Видалення агрегованої позиції компонента зі складу з перенесенням записів у кошик."""
    try:
        u=current_user(); d=request.json or {}
        component=(d.get('component') or d.get('old_component') or '').strip() or '—'
        donor_group=(d.get('donor_group') or d.get('group') or d.get('old_group') or '').strip() or '—'
        donor_rh=(d.get('donor_rh') or d.get('rh') or d.get('old_rh') or '').strip() or '—'
        reason=d.get('reason') or 'Видалено агреговану позицію компонента'
        recs=rows("""SELECT * FROM stock_entries
                    WHERE COALESCE(NULLIF(component,''),'—')=?
                      AND COALESCE(NULLIF(donor_group,''),'—')=?
                      AND COALESCE(NULLIF(donor_rh,''),'—')=?""", (component, donor_group, donor_rh))
        if not recs:
            return jsonify(ok=False,error="Позицію компонента не знайдено"),404
        # V6.4.61: не можна масово видалити компонент/партію, якщо по ній є активний резерв.
        locked_req=_request_lock_for_stock_records(recs) if '_request_lock_for_stock_records' in globals() else None
        if locked_req:
            return jsonify(ok=False,error=f"Неможливо видалити компонент: він пов’язаний із вимогою №{locked_req.get('id')}, статус «{locked_req.get('status')}». Спочатку скасуйте резерв/видачу або переведіть вимогу в безпечний статус."),409
        movement_count=sum(1 for rec in recs if _moz_type_kind(rec.get('type')) in ('issue','writeoff'))
        if movement_count:
            return jsonify(ok=False,error="Масове видалення агрегованого компонента заборонене, бо по ньому вже є видачі/списання. Видаляйте конкретні партії через журнал складу."),409
        for rec in recs:
            execute("INSERT INTO trash(created_at,source_table,source_id,data,deleted_by,reason) VALUES(?,?,?,?,?,?)",
                    (now(), "stock_entries", rec.get('id'), json.dumps(dict(rec),ensure_ascii=False), u.get('username'), reason))
        execute("""DELETE FROM stock_entries
                   WHERE COALESCE(NULLIF(component,''),'—')=?
                     AND COALESCE(NULLIF(donor_group,''),'—')=?
                     AND COALESCE(NULLIF(donor_rh,''),'—')=?""", (component, donor_group, donor_rh))
        try:
            execute("""UPDATE blood_units SET status='deleted', amount=0, updated_at=?, note=?
                       WHERE COALESCE(NULLIF(component,''),'—')=?
                         AND COALESCE(NULLIF(donor_group,''),'—')=?
                         AND COALESCE(NULLIF(donor_rh,''),'—')=?""", (now(), reason, component, donor_group, donor_rh))
        except Exception:
            db_rollback_safe()
        audit("component_stock_delete", f"{component} {donor_group} {donor_rh}; {len(recs)} records")
        return jsonify(ok=True, deleted=len(recs))
    except Exception as e:
        db_rollback_safe()
        return jsonify(ok=False,error=f"Помилка видалення компонента: {str(e)}"),500

@app.get("/api/version")
def api_version_v618():
    return jsonify(ok=True, version="V6.4.70", title=APP_TITLE)



# ================= V6.4.0 CORE UPGRADE API =================
@app.get("/api/security/role-matrix")
@role_required("admin")
def api_security_role_matrix_v640():
    return jsonify(ok=True, version=APP_VERSION, permissions={k: sorted(list(v)) for k,v in PERMISSIONS.items()})

@app.get("/api/units")
@role_required("admin","transfusion")
def api_units_v640():
    normalize_unit_statuses()
    status=request.args.get("status","").strip()
    component=request.args.get("component","").strip()
    q="SELECT * FROM blood_units WHERE 1=1"
    params=[]
    if status:
        q += " AND status=?"; params.append(status)
    if component:
        q += " AND component LIKE ?"; params.append(f"%{component}%")
    q += " ORDER BY CASE WHEN expiry IS NULL OR expiry='' THEN 1 ELSE 0 END, expiry ASC, id DESC LIMIT 1000"
    return jsonify(ok=True, items=rows(q, tuple(params)))

@app.get("/api/units/fefo")
@role_required("admin","transfusion")
def api_units_fefo_v640():
    normalize_unit_statuses()
    component=request.args.get("component","")
    group=request.args.get("group","") or request.args.get("donor_group","")
    rh=request.args.get("rh","") or request.args.get("donor_rh","")
    q="SELECT * FROM blood_units WHERE status IN ('available','reserved')"
    params=[]
    if component:
        q += " AND component=?"; params.append(component)
    if group:
        q += " AND donor_group=?"; params.append(group)
    if rh:
        q += " AND donor_rh=?"; params.append(rh)
    q += " ORDER BY CASE WHEN expiry IS NULL OR expiry='' THEN 1 ELSE 0 END, expiry ASC, id ASC LIMIT 100"
    out=[]
    for u in rows(q, tuple(params)):
        item=dict(u)
        try:
            avail=_available_qty_for_unit(u)
            reserved=_reserved_qty_for_unit(u.get('id'))
            balance=_stock_lot_balance_py(_stock_entry_lot_identity(u))
        except Exception:
            avail=float(u.get('amount') or 0); reserved=0.0; balance=float(u.get('amount') or 0)
        item['lot_balance']=balance
        item['reserved_qty']=reserved
        item['available_qty']=avail
        if avail > 0.000001:
            out.append(item)
        if len(out)>=50:
            break
    return jsonify(ok=True, rule="FEFO: першим пропонується компонент з найкоротшим терміном придатності; показано тільки доступний незарезервований залишок", items=out)

@app.post("/api/units/reserve")
@role_required("admin","transfusion")
def api_units_reserve_v640():
    d=request.json or {}
    rid=int(d.get("request_id") or d.get("id") or 0)
    req=row("SELECT * FROM requests WHERE id=?", (rid,))
    if not req:
        return jsonify(ok=False,error="Вимогу не знайдено"), 404
    unit=None
    if d.get("unit_id"):
        unit=row("SELECT * FROM blood_units WHERE id=?", (int(d.get("unit_id")),))
    if not unit:
        unit=find_fefo_unit_for_request(req, d.get("donor_group",""), d.get("donor_rh",""), d.get("pack_no",""), d.get("series",""), d.get("qr_code",""))
    if not unit:
        return jsonify(ok=False,error="Немає доступної одиниці для резервування"), 409
    ok1,w1=abo_compatible(req.get("patient_group"), unit.get("donor_group"), req.get("component"))
    ok2,w2=rh_compatible(req.get("patient_rh"), unit.get("donor_rh"))
    if not (ok1 and ok2):
        return jsonify(ok=False,error="; ".join([x for x in [w1,w2] if x]) or "Несумісність"), 409
    ok_res, msg_res, bal_res, qty_res = _can_reserve_unit_for_request(req, unit)
    if not ok_res:
        return jsonify(ok=False,error=msg_res),409
    actor=(current_user().get("full_name") or current_user().get("username") or "")
    if not reserve_unit_for_request(req, unit, actor):
        return jsonify(ok=False,error=msg_res or "Не вдалося зарезервувати одиницю"),409
    return jsonify(ok=True, unit=unit, lot_balance=bal_res, reserved_qty=qty_res)

@app.post("/api/units/writeoff")
@role_required("admin","transfusion")
def api_units_writeoff_v640():
    """V6.4.61: старий endpoint одиниць тепер використовує безпечне партійне списання."""
    d=request.json or {}
    code=(d.get("code") or d.get("pack_no") or d.get("qr_code") or d.get("series") or "").strip()
    unit=None
    if d.get("unit_id"):
        unit=row("SELECT * FROM blood_units WHERE id=?", (int(d.get("unit_id")),))
    elif code:
        unit=row("SELECT * FROM blood_units WHERE qr_code=? OR pack_no=? OR series=? ORDER BY id DESC LIMIT 1", (code,code,code))
    if not unit:
        return jsonify(ok=False,error="Одиницю не знайдено"), 404
    reason=d.get("reason") or "Списання"
    amount=d.get('amount') or unit.get('amount') or 1
    ok,msg,data=_safe_lot_writeoff(unit.get('component',''), unit.get('donor_group',''), unit.get('donor_rh',''), amount, unit.get('pack_no',''), unit.get('series',''), d.get('stock_entry_id'), reason, 'legacy /api/units/writeoff', unit.get('patient_name',''), unit.get('qr_code',''), unit.get('request_id'), True)
    if not ok:
        return jsonify(ok=False,error=msg),409 if 'Неможливо' in msg or 'Недостатньо' in msg else 400
    audit('unit_writeoff', f"unit={unit['id']}; {reason}")
    return jsonify(ok=True, unit_id=data.get('unit_id'), amount=data.get('amount'))

@app.post("/api/units/auto-expire")
@role_required("admin","transfusion")
def api_units_auto_expire_v640():
    before=rows("SELECT id FROM blood_units WHERE status IN ('available','reserved')")
    normalize_unit_statuses()
    expired=rows("SELECT * FROM blood_units WHERE status='expired' ORDER BY expiry ASC, id DESC LIMIT 500")
    return jsonify(ok=True, checked=len(before), expired=expired)

@app.get("/api/traceability/unit/<int:unit_id>")
@role_required("admin","transfusion")
def api_traceability_unit_v640(unit_id):
    unit=row("SELECT * FROM blood_units WHERE id=?", (unit_id,))
    if not unit:
        return jsonify(ok=False,error="Одиницю не знайдено"), 404
    events=rows("SELECT * FROM unit_events WHERE unit_id=? ORDER BY id ASC", (unit_id,))
    reqs=[]
    if unit.get("request_id"):
        reqs=rows("SELECT * FROM requests WHERE id=?", (unit.get("request_id"),))
    return jsonify(ok=True, unit=unit, events=events, requests=reqs)

@app.get("/api/temperature/summary")
@role_required("admin","transfusion")
def api_temperature_summary_v640():
    try:
        data=rows("""
            SELECT COALESCE(fridge_name,'Холодильник') fridge,
                   COUNT(*) count,
                   MIN(CAST(temperature AS REAL)) min_temp,
                   MAX(CAST(temperature AS REAL)) max_temp,
                   AVG(CAST(temperature AS REAL)) avg_temp
            FROM fridge_temperature_log
            GROUP BY COALESCE(fridge_name,'Холодильник')
            ORDER BY fridge
        """)
    except Exception:
        db_rollback_safe(); data=[]
    alerts=[]
    for x in data:
        mn=float(x.get('min_temp') or 0); mx=float(x.get('max_temp') or 0)
        if mn < 2 or mx > 6:
            alerts.append({"level":"critical","fridge":x.get('fridge'),"text":f"Температура поза межами 2–6°C: min {mn:g}, max {mx:g}"})
    return jsonify(ok=True, normal_range="2–6°C", items=data, alerts=alerts)


# ================= V6.4.40 PERIODIC MOZ BLOOD COMPONENT REPORTS + STOCK FULL COMPONENTS =================
MOZ_REPORT_COMPONENTS = [
    "Донорська кров",
    "Еритроцити",
    "Еритроцити в додатковому розчині",
    "Еритроцити з видаленим лейкотромбоцитарним шаром",
    "Еритроцити з видаленим лейкотромбоцитарним шаром, в додатковому розчині",
    "Еритроцити, збіднені на лейкоцити",
    "Еритроцити, збіднені на лейкоцити, в додатковому розчині",
    "Еритроцити відмиті",
    "Тромбоцити, аферез",
    "Тромбоцити, аферез, оброблені методом патогенредукції",
    "Тромбоцити, відновлені",
    "Плазма свіжозаморожена",
    "Плазма свіжозаморожена, збіднена на кріопреципітат",
    "Кріопреципітат",
    "Плазма свіжозаморожена, оброблена методом патогенредукції",
]
MOZ_GROUP_COLUMNS = [
    ("O(I)", "Rh(+)"), ("O(I)", "Rh(-)"),
    ("A(II)", "Rh(+)"), ("A(II)", "Rh(-)"),
    ("B(III)", "Rh(+)"), ("B(III)", "Rh(-)"),
    ("AB(IV)", "Rh(+)"), ("AB(IV)", "Rh(-)"),
]

def _txt(v):
    return str(v or '').strip()

def moz_norm_group(v):
    t=_txt(v).upper().replace(' ', '')
    if not t: return ''
    if t in ['0','O','O(I)','OI','I','1']: return 'O(I)'
    if t in ['A','A(II)','AII','II','2']: return 'A(II)'
    if t in ['B','B(III)','BIII','III','3']: return 'B(III)'
    if t in ['AB','AB(IV)','ABIV','IV','4']: return 'AB(IV)'
    return _txt(v)

def moz_norm_rh(v):
    t=_txt(v).lower().replace(' ', '')
    if not t: return ''
    if '+' in t or 'pos' in t or 'positive' in t or 'позит' in t: return 'Rh(+)'
    if '-' in t or 'neg' in t or 'negative' in t or 'негат' in t: return 'Rh(-)'
    if 'rh(+)' in t: return 'Rh(+)'
    if 'rh(-)' in t: return 'Rh(-)'
    return _txt(v)

def moz_component_category(name):
    n=_txt(name).lower().replace('ё','е')
    # Normalize common typos/spelling variants from manual entries.
    n=n.replace('лейкоцитарним шаром','лейкотромбоцитарним шаром').replace('лейкоцитарного шару','лейкотромбоцитарного шару')
    n=n.replace('лейко-тромбоцитар','лейкотромбоцитар')
    if not n: return None
    if 'патоген' in n and 'плазм' in n:
        return 'Плазма свіжозаморожена, оброблена методом патогенредукції'
    if 'збіднен' in n and 'кріо' in n and 'плазм' in n:
        return 'Плазма свіжозаморожена, збіднена на кріопреципітат'
    if 'кріопрецип' in n or 'криопрецип' in n:
        return 'Кріопреципітат'
    if 'плазм' in n:
        return 'Плазма свіжозаморожена'
    if 'тромбо' in n and 'патоген' in n:
        return 'Тромбоцити, аферез, оброблені методом патогенредукції'
    if 'тромбо' in n and ('віднов' in n or 'восстанов' in n or 'recovered' in n):
        return 'Тромбоцити, відновлені'
    if 'тромбо' in n:
        return 'Тромбоцити, аферез'
    if 'відмит' in n or 'отмит' in n or 'washed' in n:
        return 'Еритроцити відмиті'
    if ('видален' in n or 'лейкотромбоцитар' in n) and 'додатков' in n:
        return 'Еритроцити з видаленим лейкотромбоцитарним шаром, в додатковому розчині'
    if 'видален' in n or 'лейкотромбоцитар' in n:
        return 'Еритроцити з видаленим лейкотромбоцитарним шаром'
    if ('збіднен' in n or 'лейкодеплет' in n or 'лейкоцит' in n) and 'додатков' in n:
        return 'Еритроцити, збіднені на лейкоцити, в додатковому розчині'
    if 'збіднен' in n or 'лейкодеплет' in n or 'лейкоцит' in n:
        return 'Еритроцити, збіднені на лейкоцити'
    if 'ерит' in n and 'додатков' in n:
        return 'Еритроцити в додатковому розчині'
    if 'ерит' in n:
        return 'Еритроцити'
    if 'цільн' in n or 'ціла кров' in n or 'донорська кров' in n or 'whole blood' in n:
        return 'Донорська кров'
    return name

def moz_period_from_request():
    p=(request.args.get('period') or request.args.get('period_type') or 'month').lower()
    today=datetime.now().date()
    year=int(request.args.get('year') or today.year)
    if p == 'quarter':
        q=max(1,min(4,int(request.args.get('quarter') or ((today.month-1)//3+1))))
        m=(q-1)*3+1
        start=datetime(year,m,1).date()
        end=(datetime(year+1,1,1).date()-timedelta(days=1)) if q==4 else (datetime(year,m+3,1).date()-timedelta(days=1))
        label=f"{q} квартал {year} року"
    elif p == 'year':
        start=datetime(year,1,1).date(); end=datetime(year,12,31).date(); label=f"{year} рік"
    elif p == 'custom':
        start=datetime.strptime(request.args.get('date_from') or f'{year}-01-01','%Y-%m-%d').date()
        end=datetime.strptime(request.args.get('date_to') or today.strftime('%Y-%m-%d'),'%Y-%m-%d').date()
        label=f"з {start.strftime('%d.%m.%Y')} по {end.strftime('%d.%m.%Y')}"
    else:
        month=max(1,min(12,int(request.args.get('month') or today.month)))
        start=datetime(year,month,1).date()
        end=(datetime(year+1,1,1).date()-timedelta(days=1)) if month==12 else (datetime(year,month+1,1).date()-timedelta(days=1))
        months=['січні','лютому','березні','квітні','травні','червні','липні','серпні','вересні','жовтні','листопаді','грудні']
        label=f"у {months[month-1]} {year} року"
    if end < start: start,end=end,start
    return start,end,label,p

def _date_part(s):
    return _txt(s)[:10]

def _is_before(s, d):
    x=_date_part(s)
    return bool(x and x < d.strftime('%Y-%m-%d'))

def _in_period(s, start, end):
    x=_date_part(s)
    return bool(x and start.strftime('%Y-%m-%d') <= x <= end.strftime('%Y-%m-%d'))

def _zero_metrics():
    return {k:{'opening':0.0,'received':0.0,'used':0.0,'closing':0.0,'expired':0.0,'other_writeoff':0.0} for k in MOZ_GROUP_COLUMNS}

def moz_empty_summary():
    return {c:_zero_metrics() for c in MOZ_REPORT_COMPONENTS}

def _add_metric(summary, component, group, rh, metric, amount):
    cat=moz_component_category(component)
    if cat not in summary:
        summary[cat]=_zero_metrics()
    key=(moz_norm_group(group), moz_norm_rh(rh))
    if key not in summary[cat]:
        return
    try: val=float(amount or 0)
    except Exception: val=0.0
    summary[cat][key][metric]+=val


def _moz_amount(v):
    try:
        return abs(float(str(v or 0).replace(',', '.')))
    except Exception:
        return 0.0

def _moz_type_kind(t):
    x=_txt(t).lower()
    if x in ['надходження','in','income','add','plus','прихід','+']:
        return 'in'
    if x in ['видача','видано','issue','issued','out','-out']:
        return 'issue'
    if x in ['списання','списано','writeoff','written_off','write-off','minus','-']:
        return 'writeoff'
    return ''

def _moz_extract_request_id(text):
    import re
    m=re.search(r'(?:вимог[аеиі]?|request)\s*[№#:]?\s*(\d+)', _txt(text), flags=re.I)
    return int(m.group(1)) if m else 0

def _moz_request_for_stock(x):
    rid=_moz_extract_request_id(x.get('note') or '')
    if not rid and str(x.get('request_id') or '').isdigit():
        rid=int(x.get('request_id') or 0)
    if rid:
        r=row('SELECT * FROM requests WHERE id=?', (rid,))
        if r: return r
    uid=x.get('unit_id')
    if uid:
        r=row('SELECT * FROM requests WHERE unit_id=? ORDER BY id DESC LIMIT 1', (uid,))
        if r: return r
    pack=x.get('pack_no') or ''; ser=x.get('series') or ''
    if pack or ser:
        r=row("SELECT * FROM requests WHERE (pack_no=? AND pack_no<>'') OR (series=? AND series<>'') ORDER BY id DESC LIMIT 1", (pack, ser))
        if r: return r
    return None

def _moz_identity_from_code(code):
    code=_txt(code)
    if not code: return ('','','')
    queries=[
        "SELECT component,donor_group,donor_rh FROM stock_entries WHERE qr_code=? OR pack_no=? OR series=? ORDER BY id DESC LIMIT 1",
        "SELECT component,donor_group,donor_rh FROM blood_units WHERE qr_code=? OR pack_no=? OR series=? ORDER BY id DESC LIMIT 1",
        "SELECT component,donor_group,donor_rh FROM requests WHERE pack_no=? OR series=? ORDER BY id DESC LIMIT 1",
    ]
    for sql in queries:
        try:
            r=row(sql,(code,code,code) if sql.count('?')==3 else (code,code))
            if r:
                return (r.get('component') or '', r.get('donor_group') or '', r.get('donor_rh') or '')
        except Exception:
            db_rollback_safe()
    return ('','','')

def _moz_add_with_fallback(summary, component, group, rh, metric, amount, code=''):
    if (not group or not rh or not component) and code:
        c2,g2,r2=_moz_identity_from_code(code)
        component=component or c2; group=group or g2; rh=rh or r2
    _add_metric(summary, component, group, rh, metric, amount)


# ================= V6.4.40 EDITABLE REPORTS + PERIOD HISTORY CONTROL =================
def ensure_v6437_report_history_tables():
    try:
        if IS_POSTGRES:
            execute("""CREATE TABLE IF NOT EXISTS report_overrides(
                id SERIAL PRIMARY KEY, created_at TEXT, period_start TEXT, period_end TEXT,
                component TEXT, opening REAL, received REAL, used REAL, closing REAL,
                expired REAL, other_writeoff REAL, note TEXT, created_by TEXT, active INTEGER DEFAULT 1
            )""")
            execute("""CREATE TABLE IF NOT EXISTS history_period_actions(
                id SERIAL PRIMARY KEY, created_at TEXT, period_start TEXT, period_end TEXT,
                period_type TEXT, label TEXT, reason TEXT, created_by TEXT, active INTEGER DEFAULT 1, restored_at TEXT, restored_by TEXT
            )""")
        else:
            execute("""CREATE TABLE IF NOT EXISTS report_overrides(
                id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT, period_start TEXT, period_end TEXT,
                component TEXT, opening REAL, received REAL, used REAL, closing REAL,
                expired REAL, other_writeoff REAL, note TEXT, created_by TEXT, active INTEGER DEFAULT 1
            )""")
            execute("""CREATE TABLE IF NOT EXISTS history_period_actions(
                id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT, period_start TEXT, period_end TEXT,
                period_type TEXT, label TEXT, reason TEXT, created_by TEXT, active INTEGER DEFAULT 1, restored_at TEXT, restored_by TEXT
            )""")
    except Exception as e:
        db_rollback_safe(); log_suppressed_error('ensure_v6437_report_history_tables', e)

def _date_only(v):
    return _txt(v)[:10]

def _history_clears_active():
    ensure_v6437_report_history_tables()
    try:
        return rows("SELECT * FROM history_period_actions WHERE active=1 ORDER BY id DESC")
    except Exception:
        db_rollback_safe(); return []

def _is_history_period_hidden(created_at):
    d=_date_only(created_at)
    if not d: return False
    for h in _history_clears_active():
        if (h.get('period_start') or '') <= d <= (h.get('period_end') or ''):
            return True
    return False

def apply_report_overrides_v6437(summary, start, end):
    ensure_v6437_report_history_tables()
    try:
        ovs=rows("SELECT * FROM report_overrides WHERE active=1 AND period_start=? AND period_end=?", (start.isoformat(), end.isoformat()))
    except Exception:
        db_rollback_safe(); ovs=[]
    if not ovs:
        return []
    key=MOZ_GROUP_COLUMNS[0]
    applied=[]
    for o in ovs:
        comp=moz_component_category(o.get('component') or '')
        if comp not in summary:
            summary[comp]=_zero_metrics()
        # Manual report edits are total-level corrections. Put corrected total into first accounting column
        # so preview/PDF/Excel use the same edited official totals without touching real stock history.
        for k in summary[comp].keys():
            for metric in ['opening','received','used','closing','expired','other_writeoff']:
                summary[comp][k][metric]=0.0
        for metric,col in [('opening','opening'),('received','received'),('used','used'),('closing','closing'),('expired','expired'),('other_writeoff','other_writeoff')]:
            try: val=float(o.get(col) or 0)
            except Exception: val=0.0
            summary[comp][key][metric]=val
        applied.append({'component':comp,'note':o.get('note') or 'Ручна правка звіту'})
    return applied
# ================= END V6.4.40 =================

def _moz_lot_key(x):
    """Партія для звіту/залишків: компонент + група + Rh + пакет + серія + термін."""
    return (
        moz_component_category(x.get('component') or ''),
        moz_norm_group(x.get('donor_group') or x.get('group') or ''),
        moz_norm_rh(x.get('donor_rh') or x.get('rh') or ''),
        _txt(x.get('pack_no') or ''),
        _txt(x.get('series') or ''),
        _txt(x.get('expiry') or ''),
    )


def moz_build_report_data(start, end):
    """V6.4.40: звіт рахується партіями, без неможливих негативних залишків.

    Раніше звіт складав рухи по всій групі компонента. Якщо в базі був старий помилковий
    рух без надходження, він міг створити мінус або "фантомні" тромбоцити. Тепер кожна
    партія перевіряється окремо: якщо по пакету/серії виходить мінус, ця партія не псує
    офіційні підсумки, а потрапляє в попередження обліку.
    """
    summary=moz_empty_summary()
    issue_rows=[]
    writeoff_rows=[]
    audit_warnings=[]
    try:
        stock=rows('SELECT * FROM stock_entries ORDER BY created_at ASC, id ASC')
    except Exception:
        db_rollback_safe(); stock=[]

    def is_deleted_or_trash(x):
        st=_txt(x.get('status')).lower()
        note=_txt(x.get('note')).lower()
        return st in ('deleted','trash','removed') or ('видалено' in note and 'кошик' in note)

    lots={}
    for x in stock:
        if is_deleted_or_trash(x) or _is_history_period_hidden(x.get('created_at') or ''):
            continue
        kind=_moz_type_kind(x.get('type'))
        amount=_moz_amount(x.get('amount'))
        if amount<=0 or not kind:
            continue
        comp=x.get('component') or ''
        cat=moz_component_category(comp)
        ng=moz_norm_group(x.get('donor_group') or x.get('group') or '')
        nr=moz_norm_rh(x.get('donor_rh') or x.get('rh') or '')
        if not comp:
            audit_warnings.append({'id':x.get('id'), 'message':'Запис складу без компонента не врахований', 'type':x.get('type')})
            continue
        if cat not in MOZ_REPORT_COMPONENTS:
            audit_warnings.append({'id':x.get('id'), 'component':comp, 'message':'Компонент не входить до формуляра звіту і не врахований'})
            continue
        if not ng or not nr or (ng,nr) not in MOZ_GROUP_COLUMNS:
            audit_warnings.append({'id':x.get('id'), 'component':comp, 'group':x.get('donor_group',''), 'rh':x.get('donor_rh',''), 'message':'Запис не врахований: відсутня або неправильна група/Rh'})
            continue
        key=_moz_lot_key(x)
        lots.setdefault(key, []).append(x)

    for key, recs in lots.items():
        cat, ng, nr, pack, ser, expiry = key
        opening=received=used=expired=other=0.0
        lot_issues=[]; lot_writeoffs=[]
        for x in recs:
            kind=_moz_type_kind(x.get('type'))
            amount=_moz_amount(x.get('amount'))
            created=x.get('created_at') or ''
            if _is_before(created, start):
                if kind=='in': opening += amount
                elif kind in ('issue','writeoff'): opening -= amount
            if _in_period(created, start, end):
                if kind=='in':
                    received += amount
                elif kind=='issue':
                    used += amount
                    req=_moz_request_for_stock(x)
                    lot_issues.append({'Дата видачі': created, 'Пацієнт': (req or {}).get('patient_name','') if req else x.get('patient_name',''), 'Відділення': (req or {}).get('department','') if req else '', 'Компонент': cat, 'Група': ng, 'Rh': nr, 'Кількість': amount, '№ пакета': x.get('pack_no',''), 'Серія': x.get('series',''), 'Термін': x.get('expiry',''), 'Примітка': x.get('note','')})
                elif kind=='writeoff':
                    note=_txt(x.get('note')).lower()
                    if ('термін' in note or 'строк' in note or 'придат' in note or 'expire' in note or 'закін' in note):
                        expired += amount
                    else:
                        other += amount
                    req=_moz_request_for_stock(x)
                    lot_writeoffs.append({'Дата списання': created, 'Пацієнт': (req or {}).get('patient_name','') if req else x.get('patient_name',''), 'Відділення': (req or {}).get('department','') if req else '', 'Компонент': cat, 'Група': ng, 'Rh': nr, 'Кількість': amount, '№ пакета': x.get('pack_no',''), 'Серія': x.get('series',''), 'Термін': x.get('expiry',''), 'Причина': x.get('note','')})
        closing=opening+received-used-expired-other
        if closing < -0.000001:
            audit_warnings.append({'component':cat, 'group':ng, 'rh':nr, 'message':f'Партія {pack or "без пакета"}/{ser or "без серії"}: негативний залишок {closing:g}. Партію виключено з офіційних підсумків. Перевірте/видаліть пов’язаний рух партії.'})
            continue
        # Якщо партія повністю нульова і за період руху немає — не додаємо нічого.
        if abs(opening)+abs(received)+abs(used)+abs(expired)+abs(other)+abs(closing) <= 0.000001:
            continue
        summary[cat][(ng,nr)]['opening'] += opening
        summary[cat][(ng,nr)]['received'] += received
        summary[cat][(ng,nr)]['used'] += used
        summary[cat][(ng,nr)]['expired'] += expired
        summary[cat][(ng,nr)]['other_writeoff'] += other
        summary[cat][(ng,nr)]['closing'] += closing
        issue_rows.extend(lot_issues)
        writeoff_rows.extend(lot_writeoffs)

    applied_overrides=apply_report_overrides_v6437(summary, start, end)
    for o in applied_overrides:
        audit_warnings.append({'component':o.get('component',''), 'group':'', 'rh':'', 'message':'Застосовано ручну правку звіту. Реальна історія складу не змінена.'})
    return summary, issue_rows, writeoff_rows, audit_warnings

def moz_totals_for_component(metrics, metric_name):
    vals=[]
    for key in MOZ_GROUP_COLUMNS:
        vals.append(metrics.get(key,{}).get(metric_name,0) or 0)
    vals.append(sum(vals))
    return vals

def moz_flat_summary(summary):
    out=[]
    for comp in MOZ_REPORT_COMPONENTS:
        metrics=summary.get(comp) or _zero_metrics()
        out.append({
            'component': comp,
            'opening_total': sum(m.get('opening',0) for m in metrics.values()),
            'received_total': sum(m.get('received',0) for m in metrics.values()),
            'used_total': sum(m.get('used',0) for m in metrics.values()),
            'closing_total': sum(m.get('closing',0) for m in metrics.values()),
            'expired_total': sum(m.get('expired',0) for m in metrics.values()),
            'other_writeoff_total': sum(m.get('other_writeoff',0) for m in metrics.values()),
        })
    return out

@app.get('/api/reports/moz-components')
@role_required('admin','transfusion')
def api_reports_moz_components_v643():
    return jsonify(ok=True, components=MOZ_REPORT_COMPONENTS, groups=[{'group':g,'rh':rh} for g,rh in MOZ_GROUP_COLUMNS])

@app.get('/api/reports/moz-period-summary')
@role_required('admin','transfusion')
def api_reports_moz_period_summary_v643():
    start,end,label,ptype=moz_period_from_request()
    summary, issues, writeoffs, warnings=moz_build_report_data(start,end)
    return jsonify(ok=True, period={'type':ptype,'date_from':start.isoformat(),'date_to':end.isoformat(),'label':label,'issued_at':now(), 'units':'дози/од.'}, rows=moz_flat_summary(summary), issues=issues[:200], writeoffs=writeoffs[:200], warnings=warnings[:200], explanation={'opening':'Залишок на перший день періоду','received':'Надходження за період з розділу Склад','used':'Видача пацієнтам за період','closing':'Розрахований залишок на кінець періоду','expired':'Списання через закінчення терміну','other_writeoff':'Інші списання'})


@app.post('/api/reports/moz-period-overrides')
@role_required('admin','transfusion')
def api_reports_moz_period_overrides_v6437():
    ensure_v6437_report_history_tables()
    start,end,label,ptype=moz_period_from_request()
    data=request.get_json(silent=True) or {}
    items=data.get('rows') or []
    note=_txt(data.get('note') or 'Ручна правка звіту')
    u=current_user() or {}
    try:
        execute("UPDATE report_overrides SET active=0 WHERE period_start=? AND period_end=?", (start.isoformat(), end.isoformat()))
    except Exception:
        db_rollback_safe()
    saved=0
    for it in items:
        comp=moz_component_category(it.get('component') or '')
        if not comp: continue
        vals=[]
        for k in ['opening_total','received_total','used_total','closing_total','expired_total','other_writeoff_total']:
            try: vals.append(float(str(it.get(k,0)).replace(',','.')))
            except Exception: vals.append(0.0)
        execute("""INSERT INTO report_overrides(created_at,period_start,period_end,component,opening,received,used,closing,expired,other_writeoff,note,created_by,active)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,1)""", (now(), start.isoformat(), end.isoformat(), comp, vals[0], vals[1], vals[2], vals[3], vals[4], vals[5], note, u.get('username','')))
        saved+=1
    audit('report_override_save', f'{start.isoformat()}..{end.isoformat()} rows={saved}')
    return jsonify(ok=True, saved=saved, message='Ручні правки звіту збережено')

@app.post('/api/reports/moz-period-overrides/clear')
@role_required('admin','transfusion')
def api_reports_moz_period_overrides_clear_v6437():
    ensure_v6437_report_history_tables()
    start,end,label,ptype=moz_period_from_request()
    execute("UPDATE report_overrides SET active=0 WHERE period_start=? AND period_end=?", (start.isoformat(), end.isoformat()))
    audit('report_override_clear', f'{start.isoformat()}..{end.isoformat()}')
    return jsonify(ok=True, message='Ручні правки звіту скасовано. Звіт знову рахується з історії складу.')

@app.get('/api/history/period-actions')
@role_required('admin','transfusion')
def api_history_period_actions_list_v6437():
    ensure_v6437_report_history_tables()
    return jsonify(ok=True, rows=rows("SELECT * FROM history_period_actions ORDER BY id DESC LIMIT 200"))

@app.post('/api/history/clear-period')
@role_required('admin','transfusion')
def api_history_clear_period_v6437():
    ensure_v6437_report_history_tables()
    start,end,label,ptype=moz_period_from_request()
    data=request.get_json(silent=True) or {}
    reason=_txt(data.get('reason') or 'Очищення історії за період')
    u=current_user() or {}
    # Safe design: do not physically delete medical records. We mark the whole period as hidden from reports/history views.
    # Restore simply deactivates this mark.
    execute("""INSERT INTO history_period_actions(created_at,period_start,period_end,period_type,label,reason,created_by,active)
               VALUES(?,?,?,?,?,?,?,1)""", (now(), start.isoformat(), end.isoformat(), ptype, label, reason, u.get('username','')))
    audit('history_clear_period', f'{start.isoformat()}..{end.isoformat()} {reason}')
    return jsonify(ok=True, message=f'Історію за період приховано з розрахунків звітів: {label}. Дані не видалені фізично, їх можна відновити.')

@app.post('/api/history/restore-period')
@role_required('admin','transfusion')
def api_history_restore_period_v6437():
    ensure_v6437_report_history_tables()
    data=request.get_json(silent=True) or {}
    hid=safe_int(data.get('id'),0)
    if hid<=0:
        return jsonify(ok=False,error='ID очищення обов’язковий'),400
    u=current_user() or {}
    execute("UPDATE history_period_actions SET active=0, restored_at=?, restored_by=? WHERE id=?", (now(), u.get('username',''), hid))
    audit('history_restore_period', f'id={hid}')
    return jsonify(ok=True, message='Історію за період відновлено у звітах')

def moz_period_workbook(start,end,label):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    summary, issues, writeoffs, warnings=moz_build_report_data(start,end)
    wb=Workbook(); ws=wb.active; ws.title='Звіт МОЗ'
    thin=Side(style='thin', color='999999')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    header_fill=PatternFill('solid', fgColor='D9EAF7')
    title_fill=PatternFill('solid', fgColor='E2F0D9')
    ws.merge_cells('A1:K1'); ws['A1']='Додаток 2 до Порядку'
    ws.merge_cells('A2:K2'); ws['A2']='ЗВІТ про отримання, використання та списання донорської крові та компонентів крові'
    ws.merge_cells('A3:K3'); ws['A3']=label
    ws.merge_cells('A4:K4'); ws['A4']=f'Період: {start.strftime("%d.%m.%Y")} – {end.strftime("%d.%m.%Y")} · Сформовано: {now()}'
    for c in ['A1','A2','A3','A4']:
        ws[c].alignment=Alignment(horizontal='center'); ws[c].font=Font(bold=True, size=12)
    r=6
    for comp in MOZ_REPORT_COMPONENTS:
        metrics=summary.get(comp) or _zero_metrics()
        ws.cell(r,1,'№ з/п'); ws.cell(r,2,'Назва показника'); ws.cell(r,3,comp); ws.cell(r,11,'Разом')
        for col in range(1,12):
            ws.cell(r,col).fill=title_fill; ws.cell(r,col).font=Font(bold=True); ws.cell(r,col).border=border
        r+=1
        ws.cell(r,1,''); ws.cell(r,2,'')
        groups=['O(I)','', 'A(II)','', 'B(III)','', 'AB(IV)','']
        for i,g in enumerate(groups, start=3): ws.cell(r,i,g)
        ws.cell(r,11,'Разом')
        r+=1
        ws.cell(r,1,''); ws.cell(r,2,'')
        rhs=['Rh(+)','Rh(-)','Rh(+)','Rh(-)','Rh(+)','Rh(-)','Rh(+)','Rh(-)']
        for i,rh in enumerate(rhs, start=3): ws.cell(r,i,rh)
        ws.cell(r,11,'')
        for rr in [r-1,r]:
            for col in range(1,12):
                ws.cell(rr,col).fill=header_fill; ws.cell(rr,col).font=Font(bold=True); ws.cell(rr,col).alignment=Alignment(horizontal='center', vertical='center', wrap_text=True); ws.cell(rr,col).border=border
        r+=1
        lines=[('Залишок на перший день звітного періоду, доз','opening'),('Одержано у звітному періоді, доз','received'),('Використано/видано у звітному періоді, доз','used'),('Залишок на останній день звітного періоду, доз','closing'),('Списано у зв’язку із закінченням строку придатності, доз','expired'),('Списано з інших причин, доз','other_writeoff')]
        for idx,(name,metric) in enumerate(lines, start=1):
            vals=moz_totals_for_component(metrics, metric)
            ws.cell(r,1,idx); ws.cell(r,2,name)
            for j,v in enumerate(vals, start=3): ws.cell(r,j,v)
            for col in range(1,12): ws.cell(r,col).border=border; ws.cell(r,col).alignment=Alignment(wrap_text=True, vertical='center')
            r+=1
        r+=1
    ws.column_dimensions['A'].width=8; ws.column_dimensions['B'].width=48
    for col in range(3,12): ws.column_dimensions[get_column_letter(col)].width=12
    ws.freeze_panes='A6'
    ws2=wb.create_sheet('Видано за період')
    headers=['Дата видачі','Пацієнт','Відділення','Компонент','Група','Rh','Кількість','№ пакета','Серія','Термін','Примітка']
    ws2.append(headers)
    for row_ in issues: ws2.append([row_.get(h,'') for h in headers])
    for col in range(1,len(headers)+1):
        ws2.cell(1,col).font=Font(bold=True); ws2.cell(1,col).fill=header_fill; ws2.cell(1,col).border=border
        ws2.column_dimensions[get_column_letter(col)].width=18
    ws2.column_dimensions['B'].width=28; ws2.column_dimensions['D'].width=42; ws2.column_dimensions['K'].width=35
    wsw=wb.create_sheet('Списано за період')
    wh=['Дата списання','Пацієнт','Відділення','Компонент','Група','Rh','Кількість','№ пакета','Серія','Термін','Причина']
    wsw.append(wh)
    for row_ in writeoffs: wsw.append([row_.get(h,'') for h in wh])
    for col in range(1,len(wh)+1):
        wsw.cell(1,col).font=Font(bold=True); wsw.cell(1,col).fill=header_fill; wsw.cell(1,col).border=border
        wsw.column_dimensions[get_column_letter(col)].width=18

    wsa=wb.create_sheet('Попередження обліку')
    wsa.append(['Компонент','Група','Rh','Повідомлення'])
    for w in warnings:
        wsa.append([w.get('component',''), w.get('group',''), w.get('rh',''), w.get('message','')])
    wsa.column_dimensions['A'].width=45; wsa.column_dimensions['D'].width=80
    ws3=wb.create_sheet('Компоненти шаблону')
    ws3.append(['№','Компонент крові у завантаженому шаблоні'])
    for i,comp in enumerate(MOZ_REPORT_COMPONENTS,1): ws3.append([i, comp])
    ws3.column_dimensions['B'].width=70
    return wb

@app.get('/reports/moz-period.xlsx')
@role_required('admin','transfusion')
def report_moz_period_xlsx_v643():
    start,end,label,ptype=moz_period_from_request()
    wb=moz_period_workbook(start,end,label)
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"moz_blood_components_{start.isoformat()}_{end.isoformat()}.xlsx")

@app.get('/reports/moz-period.pdf')
@role_required('admin','transfusion')
def report_moz_period_pdf_v643():
    start,end,label,ptype=moz_period_from_request()
    summary,issues,writeoffs,warnings=moz_build_report_data(start,end)
    bio=BytesIO()
    doc=SimpleDocTemplate(bio, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=18, bottomMargin=18)
    styles=getSampleStyleSheet()
    styles.add(ParagraphStyle(name='UATitle', parent=styles['Title'], fontName='DejaVu', fontSize=11, leading=13, alignment=1))
    styles.add(ParagraphStyle(name='UA', parent=styles['Normal'], fontName='DejaVu', fontSize=7.2, leading=8.8))
    styles.add(ParagraphStyle(name='UASmall', parent=styles['Normal'], fontName='DejaVu', fontSize=6.2, leading=7.4))
    setup_pdf_font(canvas.Canvas(BytesIO()), 9)
    story=[]
    story.append(Paragraph('Звіт про отримання, використання та списання донорської крові та компонентів крові', styles['UATitle']))
    story.append(Paragraph(f'{label}. Період: {start.strftime("%d.%m.%Y")} – {end.strftime("%d.%m.%Y")}. Сформовано: {now()}. Усі числа — дози/одиниці.', styles['UA']))
    story.append(Spacer(1,6))
    data=[[Paragraph('Компонент',styles['UASmall']),Paragraph('Зал. поч.',styles['UASmall']),Paragraph('Одержано',styles['UASmall']),Paragraph('Використано/видано',styles['UASmall']),Paragraph('Зал. кін.',styles['UASmall']),Paragraph('Списано: строк',styles['UASmall']),Paragraph('Списано: інше',styles['UASmall'])]]
    for r in moz_flat_summary(summary):
        data.append([Paragraph(r['component'],styles['UASmall']), f"{r['opening_total']:g}", f"{r['received_total']:g}", f"{r['used_total']:g}", f"{r['closing_total']:g}", f"{r['expired_total']:g}", f"{r['other_writeoff_total']:g}"])
    tbl=Table(data, colWidths=[255,65,65,85,65,75,75], repeatRows=1)
    tbl.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),'DejaVu'),('FONTSIZE',(0,0),(-1,-1),6.6),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#D9EAF7')),('GRID',(0,0),(-1,-1),0.35,colors.grey),('VALIGN',(0,0),(-1,-1),'TOP'),('ALIGN',(1,1),(-1,-1),'CENTER'),('LEFTPADDING',(0,0),(-1,-1),3),('RIGHTPADDING',(0,0),(-1,-1),3)]))
    story.append(tbl)
    if warnings:
        story.append(Spacer(1,8)); story.append(Paragraph('Попередження обліку: якщо є негативні залишки, це означає старі неповні або помилкові рухи складу. У підсумку залишок показано як 0, але рядок треба перевірити.', styles['UA']))
        wdarn=[[Paragraph('Компонент',styles['UASmall']),Paragraph('Група/Rh',styles['UASmall']),Paragraph('Повідомлення',styles['UASmall'])]]
        for w in warnings[:30]:
            wdarn.append([Paragraph(_txt(w.get('component')),styles['UASmall']), f"{w.get('group','')} {w.get('rh','')}", Paragraph(_txt(w.get('message')),styles['UASmall'])])
        wtbl=Table(wdarn, colWidths=[220,70,405], repeatRows=1)
        wtbl.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),'DejaVu'),('FONTSIZE',(0,0),(-1,-1),6.2),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#FFF2CC')),('GRID',(0,0),(-1,-1),0.35,colors.grey),('VALIGN',(0,0),(-1,-1),'TOP')]))
        story.append(wtbl)
    if writeoffs:
        story.append(Spacer(1,8)); story.append(Paragraph('Списання за період', styles['UA']))
        wd=[[Paragraph('Дата',styles['UASmall']),Paragraph('Компонент',styles['UASmall']),Paragraph('Група/Rh',styles['UASmall']),Paragraph('К-сть',styles['UASmall']),Paragraph('Серія/пакет',styles['UASmall']),Paragraph('Причина',styles['UASmall'])]]
        for w in writeoffs[:60]:
            wd.append([_txt(w.get('Дата списання'))[:16], Paragraph(_txt(w.get('Компонент')),styles['UASmall']), f"{w.get('Група','')} {w.get('Rh','')}", f"{w.get('Кількість','')}", Paragraph((_txt(w.get('Серія')) or _txt(w.get('№ пакета'))),styles['UASmall']), Paragraph(_txt(w.get('Причина')),styles['UASmall'])])
        wt=Table(wd, colWidths=[75,210,65,45,90,210], repeatRows=1)
        wt.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),'DejaVu'),('FONTSIZE',(0,0),(-1,-1),6.2),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#FCE4D6')),('GRID',(0,0),(-1,-1),0.35,colors.grey),('VALIGN',(0,0),(-1,-1),'TOP')]))
        story.append(wt)
    doc.build(story)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"moz_blood_components_{start.isoformat()}_{end.isoformat()}.pdf")
# ================= END V6.4.40 PERIODIC MOZ BLOOD COMPONENT REPORTS + STOCK FULL COMPONENTS =================

@app.get("/reports/full.xlsx")
@role_required("admin","transfusion")
def report_full_xlsx_v640():
    wb=Workbook()
    sheets=[
        ("Залишки", "SELECT component, donor_group, donor_rh, status, COUNT(*) packs, SUM(amount) amount, MIN(expiry) nearest_expiry FROM blood_units GROUP BY component, donor_group, donor_rh, status ORDER BY component"),
        ("Одиниці", "SELECT id,component,donor_group,donor_rh,amount,pack_no,series,expiry,status,patient_name,request_id FROM blood_units ORDER BY id DESC"),
        ("Вимоги", "SELECT id,created_at,patient_name,department,component,amount,status,pack_no,series,expiry FROM requests ORDER BY id DESC"),
        ("Події одиниць", "SELECT * FROM unit_events ORDER BY id DESC"),
        ("Реакції", "SELECT id,created_at,patient_name,component,reaction_type,reaction_severity,reaction_description,reaction_result FROM requests WHERE reaction_present='Так' ORDER BY id DESC")
    ]
    first=True
    for title, sql in sheets:
        ws=wb.active if first else wb.create_sheet(title)
        first=False; ws.title=title
        try:
            data=rows(sql)
        except Exception:
            db_rollback_safe(); data=[]
        if data:
            ws.append(list(data[0].keys()))
            for r in data:
                ws.append([r.get(k) for k in data[0].keys()])
        else:
            ws.append(["Немає даних"])
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="blood_bank_full_report_v640.xlsx")

@app.get("/api/upgrade/v630-check")
@role_required("admin","transfusion")
def api_upgrade_v630_check():
    ensure_v630_schema(); normalize_unit_statuses()
    return jsonify(ok=True, version=APP_VERSION,
                   units=row("SELECT COUNT(*) c FROM blood_units"),
                   events=row("SELECT COUNT(*) c FROM unit_events"),
                   available=row("SELECT COUNT(*) c FROM blood_units WHERE status='available'"),
                   expired=row("SELECT COUNT(*) c FROM blood_units WHERE status='expired'"))
# ================= END V6.4.0 CORE UPGRADE API =================


# ================= V6.4.0 EXTENDED MODULES EXCEPT IMMUTABLE AUDIT =================
def _iddef_v640():
    return "SERIAL PRIMARY KEY" if IS_POSTGRES else "INTEGER PRIMARY KEY AUTOINCREMENT"

def ensure_v640_schema():
    # Do NOT make audit immutable here by user request: point 8 skipped.
    ensure_v630_schema()
    ensure_v59_tables()
    pk = _iddef_v640()
    ddl = [
        f"""CREATE TABLE IF NOT EXISTS app_settings(
            id {pk}, key TEXT UNIQUE, value TEXT, updated_at TEXT, updated_by TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS two_factor_codes(
            id {pk}, created_at TEXT, user_id INTEGER, code TEXT, expires_at TEXT, used INTEGER DEFAULT 0
        )""",
        f"""CREATE TABLE IF NOT EXISTS transfusion_protocols(
            id {pk}, created_at TEXT, request_id INTEGER, patient_name TEXT, before_state TEXT,
            first_15_min TEXT, after_state TEXT, bp TEXT, pulse TEXT, temperature REAL,
            complaints TEXT, reaction TEXT, controlled_by TEXT, signature TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS reaction_acts(
            id {pk}, created_at TEXT, request_id INTEGER, patient_name TEXT, reaction_type TEXT,
            severity TEXT, symptoms TEXT, actions TEXT, result TEXT, stopped INTEGER DEFAULT 1,
            reported_by TEXT, signature TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS temperature_signatures(
            id {pk}, created_at TEXT, log_id INTEGER, signed_by TEXT, signature TEXT, note TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS backup_schedule_log(
            id {pk}, created_at TEXT, status TEXT, filename TEXT, details TEXT
        )""",
        f"""CREATE TABLE IF NOT EXISTS migration_versions(
            id {pk}, created_at TEXT, version TEXT UNIQUE, note TEXT
        )"""
    ]
    for q in ddl:
        try: execute(q)
        except Exception: db_rollback_safe()
    try:
        execute("INSERT INTO migration_versions(created_at,version,note) VALUES(?,?,?)", (now(), "v6_4_0", "Extended reports, Telegram settings, 2FA optional, protocols, labels, backups"))
    except Exception:
        db_rollback_safe()
    # soft columns without breaking older DBs
    for table, name, definition in [
        ("fridge_temperature_log", "signed_by", "TEXT"),
        ("fridge_temperature_log", "signed_at", "TEXT"),
        ("fridge_temperature_log", "signature", "TEXT"),
        ("component_writeoffs", "act_no", "TEXT"),
        ("component_writeoffs", "approved_by", "TEXT"),
    ]:
        add_column_if_missing(table, name, definition)
    return True

def get_setting_v640(key, default=""):
    try:
        ensure_v640_schema()
        r = row("SELECT value FROM app_settings WHERE key=?", (key,))
        return r.get("value") if r else default
    except Exception:
        db_rollback_safe(); return default

def set_setting_v640(key, value):
    ensure_v640_schema()
    u = current_user() or {}
    if row("SELECT id FROM app_settings WHERE key=?", (key,)):
        execute("UPDATE app_settings SET value=?,updated_at=?,updated_by=? WHERE key=?", (str(value), now(), u.get("username",""), key))
    else:
        execute("INSERT INTO app_settings(key,value,updated_at,updated_by) VALUES(?,?,?,?)", (key, str(value), now(), u.get("username","")))

def totp_code_v640(secret, for_time=None, step=30, digits=6):
    for_time = int(for_time or time.time())
    try:
        key = base64.b32decode(str(secret).replace(" ", "").upper() + "=" * ((8-len(str(secret))%8)%8))
    except Exception:
        key = str(secret or SECRET_KEY).encode()
    msg = struct.pack(">Q", int(for_time // step))
    hs = hmac.new(key, msg, hashlib.sha1).digest()
    o = hs[-1] & 15
    code = (struct.unpack(">I", hs[o:o+4])[0] & 0x7fffffff) % (10 ** digits)
    return str(code).zfill(digits)

def verify_totp_v640(secret, code):
    code = str(code or "").strip()
    return any(hmac.compare_digest(totp_code_v640(secret, time.time()+off), code) for off in (-30,0,30))

def pdf_header_v640(c, title, subtitle=""):
    setup_pdf_font(c, 12)
    pdf_text(c, 40, 805, "КНР ХОР «Центр серцево-судинних та цереброваскулярних патологій»")
    setup_pdf_font(c, 16)
    pdf_text(c, 40, 780, title)
    setup_pdf_font(c, 10)
    pdf_text(c, 40, 762, subtitle or f"Сформовано: {now()}")
    c.line(40, 752, 555, 752)
    setup_pdf_font(c, 10)
    return 730

def pdf_kv_v640(c, y, k, v):
    if y < 60:
        c.showPage(); y = pdf_header_v640(c, "Продовження")
    pdf_text(c, 45, y, f"{k}: {v if v is not None else ''}")
    return y - 18

@app.before_request
def before_v640_schema():
    # Light lazy migration; skip static for speed.
    if request.path.startswith('/static/'):
        return None
    try:
        ensure_v640_schema()
    except Exception:
        db_rollback_safe()

@app.route('/two-factor', methods=['GET','POST'])
def two_factor_page_v640():
    uid = session.get('pending_2fa_user_id')
    if not uid:
        return redirect(url_for('index'))
    if request.method == 'GET':
        return render_template('two_factor.html', title=APP_TITLE, csrf=session.get('csrf',''), error=None)
    code = request.form.get('code','').strip()
    mode = get_setting_v640('2fa_mode','off')
    ok = False
    if mode == 'telegram':
        r = row("SELECT * FROM two_factor_codes WHERE user_id=? AND code=? AND used=0 ORDER BY id DESC LIMIT 1", (uid, code))
        if r:
            try: ok = datetime.strptime(r['expires_at'], "%Y-%m-%d %H:%M:%S") >= datetime.now()
            except Exception: ok = False
            if ok: execute("UPDATE two_factor_codes SET used=1 WHERE id=?", (r['id'],))
    elif mode == 'totp':
        secret = get_setting_v640('totp_secret','')
        ok = verify_totp_v640(secret, code)
    if not ok:
        execute("INSERT INTO login_attempts(created_at,username,ip,ok) VALUES(?,?,?,?)", (now(), '2FA', request.remote_addr or '', 0))
        return render_template('two_factor.html', title=APP_TITLE, csrf=session.get('csrf',''), error='Невірний або прострочений код')
    session.clear(); session['user_id'] = uid; session['csrf'] = secrets.token_hex(24); session['last_seen'] = time.time()
    audit('login_2fa', 'Вхід із двофакторним підтвердженням')
    return redirect(url_for('index'))

@app.get('/api/security/login-policy')
@role_required('admin')
def api_login_policy_v640():
    ensure_v640_schema()
    return jsonify(ok=True, lock_after=5, lock_minutes=15, critical_after=10,
                   two_factor_enabled=get_setting_v640('2fa_enabled','0'),
                   two_factor_mode=get_setting_v640('2fa_mode','off'))

@app.post('/api/security/2fa/setup')
@role_required('admin')
def api_2fa_setup_v640():
    ensure_v640_schema(); d=request.json or {}
    mode = d.get('mode','telegram')
    if mode not in ('telegram','totp','off'):
        return jsonify(ok=False,error='mode має бути telegram/totp/off'),400
    set_setting_v640('2fa_mode', mode)
    set_setting_v640('2fa_enabled', '0' if mode=='off' else '1')
    secret = ''
    if mode == 'totp':
        secret = base64.b32encode(secrets.token_bytes(10)).decode().rstrip('=')
        set_setting_v640('totp_secret', secret)
    audit('2fa_setup', mode)
    return jsonify(ok=True, mode=mode, secret=secret, note='2FA увімкнено. Для telegram коди надходять у Telegram; для TOTP додайте secret у додаток-автентифікатор.')

@app.post('/api/security/2fa/disable')
@role_required('admin')
def api_2fa_disable_v640():
    set_setting_v640('2fa_enabled','0'); set_setting_v640('2fa_mode','off')
    audit('2fa_disable','')
    return jsonify(ok=True)

@app.get('/api/telegram/settings')
@role_required('admin','transfusion')
def api_telegram_settings_get_v640():
    return jsonify(ok=True, enabled=TELEGRAM_ENABLED, bot_username=TELEGRAM_BOT_USERNAME,
                   chat_id_configured=bool(TELEGRAM_CHAT_ID), token_configured=bool(TELEGRAM_BOT_TOKEN),
                   silent_start=TELEGRAM_SILENT_START, silent_end=TELEGRAM_SILENT_END,
                   note='З міркувань безпеки Bot Token не повертається у відповідь API.')

@app.post('/api/telegram/settings')
@role_required('admin')
def api_telegram_settings_post_v640():
    # Runtime settings; env vars on Render still remain source for restart.
    d=request.json or {}
    if 'enabled' in d: set_setting_v640('telegram_enabled_ui', '1' if d.get('enabled') else '0')
    if d.get('chat_id'): set_setting_v640('telegram_chat_id_ui', d.get('chat_id'))
    if d.get('bot_token'): set_setting_v640('telegram_token_configured_ui', '1')
    audit('telegram_settings_update','UI settings saved')
    return jsonify(ok=True, warning='Для постійної роботи на Render Bot Token краще зберігати в Environment Variables.')

@app.get('/api/dashboard/role')
@login_required
def api_dashboard_role_v640():
    u=current_user(); role=u.get('role')
    ensure_v640_schema()
    if role in ('admin','transfusion'):
        return jsonify(ok=True, role=role, widgets={
            'new_requests': row("SELECT COUNT(*) c FROM requests WHERE status='Нова'"),
            'critical_stock': len(get_alerts_data().get('low',[])) if 'get_alerts_data' in globals() else 0,
            'temperature_alerts': row("SELECT COUNT(*) c FROM fridge_temperature_log WHERE alert_triggered=1"),
            'expired_units': row("SELECT COUNT(*) c FROM blood_units WHERE status='expired'")
        })
    if role == 'doctor':
        return jsonify(ok=True, role=role, widgets={
            'my_requests': row("SELECT COUNT(*) c FROM requests WHERE created_by=?", (u['id'],)),
            'pending': row("SELECT COUNT(*) c FROM requests WHERE created_by=? AND status IN ('Нова','Погоджено','Зарезервовано','Потребує перегляду')", (u['id'],)),
            'used': row("SELECT COUNT(*) c FROM requests WHERE created_by=? AND status LIKE ?", (u['id'], '%використ%'))
        })
    return jsonify(ok=True, role=role, widgets={
        'my_requests': row("SELECT COUNT(*) c FROM requests WHERE created_by=?", (u['id'],)),
        'to_confirm': row("SELECT COUNT(*) c FROM requests WHERE created_by=? AND status LIKE ?", (u['id'], '%Видано%'))
    })

@app.post('/api/request/wizard/validate')
@login_required
def api_request_wizard_validate_v640():
    d=request.json or {}
    fields=['patient_name','birth_date','patient_status','department','component','patient_group','patient_rh','amount','urgency','diagnosis']
    ok,msg,field=validate_required_ua(d, fields)
    if not ok: return jsonify(ok=False,error=msg,field=field),400
    try:
        if float(str(d.get('amount')).replace(',','.')) <= 0: return jsonify(ok=False,error='Кількість має бути більше 0',field='amount'),400
    except Exception: return jsonify(ok=False,error='Кількість має бути числом',field='amount'),400
    return jsonify(ok=True, steps=['Пацієнт','Діагноз','Компонент','Перевірка','Відправка'])

@app.post('/api/transfusion/protocol')
@login_required
def api_transfusion_protocol_save_v640():
    ensure_v640_schema(); u=current_user(); d=request.json or {}
    rid = int(d.get('request_id') or 0)
    if rid:
        req = row('SELECT * FROM requests WHERE id=?', (rid,))
        if req and u.get('role') in ('doctor','nurse') and int(req.get('created_by') or 0) != int(u.get('id') or 0):
            return jsonify(ok=False,error='Недостатньо прав'),403
    payload = '|'.join(str(d.get(k,'')) for k in ['request_id','patient_name','bp','pulse','temperature','controlled_by'])
    sig = signature_hash(payload)
    pid=insert_returning_id("""INSERT INTO transfusion_protocols(created_at,request_id,patient_name,before_state,first_15_min,after_state,bp,pulse,temperature,complaints,reaction,controlled_by,signature)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""", (now(), rid, d.get('patient_name',''), d.get('before_state',''), d.get('first_15_min',''), d.get('after_state',''), d.get('bp',''), d.get('pulse',''), d.get('temperature') or None, d.get('complaints',''), d.get('reaction',''), d.get('controlled_by') or u.get('full_name',''), sig))
    audit('transfusion_protocol_save', str(pid))
    return jsonify(ok=True, id=pid, signature=sig)

@app.get('/api/transfusion/protocols')
@role_required('admin','transfusion')
def api_transfusion_protocols_v640():
    ensure_v640_schema(); return jsonify(rows('SELECT * FROM transfusion_protocols ORDER BY id DESC LIMIT 500'))

@app.get('/reports/transfusion-protocol/<int:pid>.pdf')
@login_required
def report_transfusion_protocol_pdf_v640(pid):
    ensure_v640_schema(); p=row('SELECT * FROM transfusion_protocols WHERE id=?',(pid,))
    if not p: return 'Not found',404
    if not can_access_request_for_current_user(p.get('request_id')):
        return 'Недостатньо прав',403
    bio=BytesIO(); c=canvas.Canvas(bio,pagesize=A4); y=pdf_header_v640(c, f'Протокол трансфузії №{pid}')
    for k in ['created_at','request_id','patient_name','before_state','first_15_min','after_state','bp','pulse','temperature','complaints','reaction','controlled_by','signature']:
        y=pdf_kv_v640(c,y,k,p.get(k))
    c.save(); bio.seek(0); return send_file(bio, as_attachment=True, download_name=f'transfusion_protocol_{pid}.pdf')

@app.post('/api/reactions/advanced')
@login_required
def api_reaction_advanced_v640():
    ensure_v640_schema(); u=current_user(); d=request.json or {}
    rid=int(d.get('request_id') or 0)
    if u.get('role') in ('doctor','nurse') and not owns_request(u, rid):
        return forbid_json()
    sig=signature_hash(json.dumps(d, ensure_ascii=False)+str(u.get('id')))
    aid=insert_returning_id("""INSERT INTO reaction_acts(created_at,request_id,patient_name,reaction_type,severity,symptoms,actions,result,stopped,reported_by,signature)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?)""", (now(), rid, d.get('patient_name',''), d.get('reaction_type',''), d.get('severity',''), d.get('symptoms',''), d.get('actions',''), d.get('result',''), 1 if d.get('stopped', True) else 0, u.get('full_name') or u.get('username'), sig))
    try:
        telegram_broadcast_roles(f"🚨 <b>Трансфузійна реакція</b>\nПацієнт: {d.get('patient_name','')}\nТип: {d.get('reaction_type','')}\nТяжкість: {d.get('severity','')}", ('admin','transfusion'), 'reaction', True)
    except Exception as e: log_suppressed_error('suppressed', e)
    audit('reaction_act_create', str(aid))
    return jsonify(ok=True,id=aid, stopped=True, signature=sig)

@app.get('/reports/reaction-act/<int:aid>.pdf')
@login_required
def report_reaction_act_pdf_v640(aid):
    ensure_v640_schema(); a=row('SELECT * FROM reaction_acts WHERE id=?',(aid,))
    if not a: return 'Not found',404
    if not can_access_request_for_current_user(a.get('request_id')):
        return 'Недостатньо прав',403
    bio=BytesIO(); c=canvas.Canvas(bio,pagesize=A4); y=pdf_header_v640(c, f'Акт трансфузійної реакції №{aid}')
    for k in ['created_at','request_id','patient_name','reaction_type','severity','symptoms','actions','result','stopped','reported_by','signature']:
        y=pdf_kv_v640(c,y,k,a.get(k))
    c.save(); bio.seek(0); return send_file(bio, as_attachment=True, download_name=f'reaction_act_{aid}.pdf')

@app.post('/api/temperature/sign')
@role_required('admin','transfusion')
def api_temperature_sign_v640():
    ensure_v640_schema(); u=current_user(); d=request.json or {}; log_id=int(d.get('log_id') or 0)
    r=row('SELECT * FROM fridge_temperature_log WHERE id=?',(log_id,))
    if not r: return jsonify(ok=False,error='Запис температури не знайдено'),404
    sig=signature_hash(f"temp|{log_id}|{u.get('username')}|{r.get('temperature')}|{r.get('created_at')}")
    execute('INSERT INTO temperature_signatures(created_at,log_id,signed_by,signature,note) VALUES(?,?,?,?,?)', (now(),log_id,u.get('full_name') or u.get('username'),sig,d.get('note','')))
    try: execute('UPDATE fridge_temperature_log SET signed_by=?,signed_at=?,signature=? WHERE id=?',(u.get('full_name') or u.get('username'),now(),sig,log_id))
    except Exception: db_rollback_safe()
    return jsonify(ok=True,signature=sig)

@app.get('/api/temperature/graph-data')
@role_required('admin','transfusion')
def api_temperature_graph_v640():
    ensure_v640_schema(); fridge=request.args.get('fridge','')
    if fridge:
        data=rows('SELECT created_at,fridge_name,temperature,alert_triggered FROM fridge_temperature_log WHERE fridge_name=? ORDER BY id ASC LIMIT 1000',(fridge,))
    else:
        data=rows('SELECT created_at,fridge_name,temperature,alert_triggered FROM fridge_temperature_log ORDER BY id ASC LIMIT 1000')
    return jsonify(ok=True, normal_min=2, normal_max=6, items=data)

@app.get('/reports/temperature-journal.xlsx')
@role_required('admin','transfusion')
def report_temperature_journal_xlsx_v640():
    ensure_v640_schema(); data=rows('SELECT * FROM fridge_temperature_log ORDER BY id DESC LIMIT 5000')
    wb=Workbook(); ws=wb.active; ws.title='Температура';
    headers=list(data[0].keys()) if data else ['Немає даних']; ws.append(headers)
    for r in data: ws.append([r.get(k) for k in headers])
    bio=BytesIO(); wb.save(bio); bio.seek(0); return send_file(bio, as_attachment=True, download_name='temperature_journal.xlsx')

@app.get('/reports/temperature-journal.pdf')
@role_required('admin','transfusion')
def report_temperature_journal_pdf_v640():
    ensure_v640_schema(); data=rows('SELECT id,created_at,fridge_name,temperature,alert_triggered,signed_by FROM fridge_temperature_log ORDER BY id DESC LIMIT 120')
    bio=BytesIO(); c=canvas.Canvas(bio,pagesize=A4); y=pdf_header_v640(c,'Температурний журнал','Норма зберігання: 2–6°C')
    for r in data:
        y=pdf_kv_v640(c,y, f"#{r.get('id')} {r.get('created_at')} {r.get('fridge_name')}", f"{r.get('temperature')}°C | тривога={r.get('alert_triggered')} | підпис={r.get('signed_by') or ''}")
    c.save(); bio.seek(0); return send_file(bio, as_attachment=True, download_name='temperature_journal.pdf')

@app.get('/reports/moz.pdf')
@role_required('admin','transfusion')
def report_moz_pdf_v640():
    ensure_v640_schema(); bio=BytesIO(); c=canvas.Canvas(bio,pagesize=A4); y=pdf_header_v640(c,'Звіт банку крові','Форма внутрішнього контролю/МОЗ-стиль')
    summary=[
        ('Вимог всього', row('SELECT COUNT(*) c FROM requests').get('c')),
        ('Одиниць компонентів', row('SELECT COUNT(*) c FROM blood_units').get('c')),
        ('Доступні одиниці', row("SELECT COUNT(*) c FROM blood_units WHERE status='available'").get('c')),
        ('Зарезервовані', row("SELECT COUNT(*) c FROM blood_units WHERE status='reserved'").get('c')),
        ('Видані', row("SELECT COUNT(*) c FROM blood_units WHERE status='issued'").get('c')),
        ('Списані', row("SELECT COUNT(*) c FROM blood_units WHERE status='written_off'").get('c')),
        ('Протерміновані', row("SELECT COUNT(*) c FROM blood_units WHERE status='expired'").get('c')),
    ]
    for k,v in summary: y=pdf_kv_v640(c,y,k,v)
    y-=10; pdf_text(c,45,y,'Підпис відповідальної особи: ____________________')
    c.save(); bio.seek(0); return send_file(bio, as_attachment=True, download_name='moz_style_blood_bank_report.pdf')

@app.get('/reports/qr-labels.pdf')
@role_required('admin','transfusion')
def report_qr_labels_pdf_v640():
    ensure_v640_schema(); data=rows("SELECT id,component,donor_group,donor_rh,pack_no,series,expiry,status,qr_code FROM blood_units ORDER BY id DESC LIMIT 80")
    bio=BytesIO(); c=canvas.Canvas(bio,pagesize=A4); setup_pdf_font(c,8)
    x0,y0=40,790; w,h=170,82; x,y=x0,y0; count=0
    for u in data:
        if y < 60: c.showPage(); setup_pdf_font(c,8); x,y=x0,y0
        c.rect(x,y-h,w,h)
        label=f"#{u.get('id')} {u.get('component','')[:18]}\n{u.get('donor_group','')} {u.get('donor_rh','')}  до {u.get('expiry','')}\nПакет:{u.get('pack_no','')} Серія:{u.get('series','')}\nQR:{u.get('qr_code') or u.get('pack_no') or u.get('id')}\nСтатус:{u.get('status','')}"
        yy=y-14
        for line in label.split('\n'):
            pdf_text(c,x+6,yy,line); yy-=12
        count+=1; x+=w+10
        if x+w>555: x=x0; y-=h+12
    if not data: pdf_text(c,40,780,'Немає одиниць для друку етикеток')
    c.save(); bio.seek(0); return send_file(bio, as_attachment=True, download_name='qr_labels.pdf')


@app.post('/api/writeoff/act')
@role_required('admin','transfusion')
def api_writeoff_act_v641():
    """V6.4.61: акт списання також проходить через safe lot writeoff."""
    ensure_v640_schema(); d=request.json or {}
    unit=None
    if d.get('unit_id'):
        unit=row('SELECT * FROM blood_units WHERE id=?',(int(d.get('unit_id')),))
    code=(d.get('package_code') or d.get('code') or d.get('qr_code') or d.get('pack_no') or d.get('series') or '').strip()
    if not unit and code:
        unit=row('SELECT * FROM blood_units WHERE qr_code=? OR pack_no=? OR series=? ORDER BY id DESC LIMIT 1',(code,code,code))
    if not unit:
        return jsonify(ok=False,error='Одиницю не знайдено'),404
    reason=(d.get('reason') or unit.get('writeoff_reason') or 'Списання').strip()
    try:
        amount=float(str(d.get('amount') or unit.get('amount') or 1).replace(',','.'))
    except Exception:
        return jsonify(ok=False,error='Кількість має бути числом'),400
    ok,msg,data=_safe_lot_writeoff(unit.get('component',''), unit.get('donor_group',''), unit.get('donor_rh',''), amount, unit.get('pack_no',''), unit.get('series',''), d.get('stock_entry_id'), reason, 'writeoff act', unit.get('patient_name',''), unit.get('qr_code',''), unit.get('request_id'), False)
    if not ok:
        return jsonify(ok=False,error=msg),409 if 'Неможливо' in msg or 'Недостатньо' in msg else 400
    package_code = unit.get('qr_code') or unit.get('pack_no') or str(unit.get('id'))
    wid=insert_returning_id('INSERT INTO component_writeoffs(created_at,package_code,component,amount,reason,written_by,notes,act_no,approved_by) VALUES(?,?,?,?,?,?,?,?,?)',
                (now(), package_code, unit.get('component',''), amount, reason, current_user().get('full_name') or current_user().get('username'), d.get('notes',''), d.get('act_no',''), d.get('approved_by','')))
    unit_event(unit['id'], 'writeoff_act', reason, unit.get('request_id'))
    audit('writeoff_act_create', str(wid))
    return jsonify(ok=True,id=wid, unit_id=data.get('unit_id'))

@app.get('/reports/writeoff-act/<int:wid>.pdf')
@role_required('admin','transfusion')
def report_writeoff_act_pdf_v640(wid):
    ensure_v640_schema(); w=row('SELECT * FROM component_writeoffs WHERE id=?',(wid,))
    if not w: return 'Not found',404
    bio=BytesIO(); c=canvas.Canvas(bio,pagesize=A4); y=pdf_header_v640(c,f'Акт списання №{wid}')
    for k,v in w.items(): y=pdf_kv_v640(c,y,k,v)
    y-=15; pdf_text(c,45,y,'Комісія/відповідальна особа: ____________________')
    c.save(); bio.seek(0); return send_file(bio, as_attachment=True, download_name=f'writeoff_act_{wid}.pdf')

@app.get('/api/backups/schedule/status')
@role_required('admin')
def api_backup_schedule_status_v640():
    ensure_v640_schema(); return jsonify(ok=True, enabled=AUTO_BACKUP_ENABLED, hour=AUTO_BACKUP_HOUR, keep_days=BACKUP_KEEP_DAYS, last=row('SELECT * FROM backups ORDER BY id DESC LIMIT 1'), log=rows('SELECT * FROM backup_schedule_log ORDER BY id DESC LIMIT 20'))

@app.post('/api/backups/schedule/run')
@role_required('admin')
def api_backup_schedule_run_v640():
    ensure_v640_schema()
    try:
        path=make_backup('manual_schedule')
        execute('INSERT INTO backup_schedule_log(created_at,status,filename,details) VALUES(?,?,?,?)',(now(),'ok',os.path.basename(path),'manual run'))
        return jsonify(ok=True, filename=os.path.basename(path))
    except Exception as e:
        db_rollback_safe(); execute('INSERT INTO backup_schedule_log(created_at,status,filename,details) VALUES(?,?,?,?)',(now(),'error','',str(e)))
        return jsonify(ok=False,error=str(e)),500

@app.get('/api/migrations/status')
@role_required('admin')
def api_migrations_status_v640():
    ensure_v640_schema(); return jsonify(ok=True, engine='custom-compatible-alembic-style', versions=rows('SELECT * FROM migration_versions ORDER BY id DESC'), note='Alembic-подібний журнал версій без додаткової залежності. Повний Alembic можна підключити окремо при стабілізації схеми.')

@app.get('/api/system/postgres-readiness')
@role_required('admin')
def api_postgres_readiness_v640():
    return jsonify(ok=True, is_postgres=IS_POSTGRES, postgres_only_required=POSTGRES_ONLY, database_url_configured=bool(DATABASE_URL), recommendation='Для продакшену Render краще DATABASE_URL PostgreSQL, SQLite залишити тільки для локальних тестів.')

@app.get('/api/pwa/status')
@login_required
def api_pwa_status_v640():
    return jsonify(ok=True, manifest='/manifest.json', service_worker='/service-worker.js', safari_note='На iPhone: Поділитися → На початковий екран. Якщо зміни не видно — оновити застосунок/очистити кеш PWA.')

@app.get('/api/upgrade/v640-check')
@role_required('admin','transfusion')
def api_upgrade_v640_check():
    ensure_v640_schema()
    return jsonify(ok=True, version=APP_VERSION, implemented=[
        'MOZ-style PDF','temperature journal PDF/XLSX/signatures','temperature graph data','Telegram event/settings API',
        'optional Telegram/TOTP 2FA','login lockout policy','wizard validation','mobile-card CSS','transfusion protocol',
        'reaction act PDF','QR label PDF','writeoff act PDF','manual/scheduled backup API','migration version log',
        'PostgreSQL readiness','PWA status','role dashboards'
    ], skipped=['immutable audit log by user request'])
# ================= END V6.4.0 EXTENDED MODULES =================

@app.get("/manifest.json")
def manifest():
    return send_file("static/manifest.json", mimetype="application/manifest+json")

@app.get("/service-worker.js")
def sw():
    return send_file("static/service-worker.js", mimetype="application/javascript")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT",5000)), debug=DEBUG)
